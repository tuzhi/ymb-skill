#!/usr/bin/env python3
"""Inspect one bank-statement file against reader fingerprint YAML.

This script is intentionally read-only: it does not edit YAML or support_matrix.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import sys


def _add_core_path(repo_root: Path) -> None:
    core = repo_root / "ymb-standardization-core"
    if str(core) not in sys.path:
        sys.path.insert(0, str(core))


def _load_support_rows(matrix_path: Path) -> list[dict]:
    if not matrix_path.exists():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(matrix_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(value or "").strip() for value in rows[0]]
    records = []
    for values in rows[1:]:
        rec = {}
        for idx, name in enumerate(header):
            if not name:
                continue
            rec[name] = "" if idx >= len(values) or values[idx] is None else str(values[idx]).strip()
        if any(rec.values()):
            records.append(rec)
    return records


def _read_result(repo_root: Path, file_path: Path):
    _add_core_path(repo_root)
    from ymb_standardization_core import core
    from ymb_standardization_core.readers.input_router import ReadResult

    kind, preamble, rows, route_info = core.read_rows(str(file_path))
    return ReadResult(kind=kind, preamble=preamble, rows=rows, route_info=route_info)


def _row_signature(rows: list, limit: int = 5) -> dict:
    visible_rows = []
    for row in rows[:limit]:
        visible_rows.append([str(cell or "").strip() for cell in row])
    header = []
    for row in visible_rows:
        if sum(1 for cell in row if cell) >= 3:
            header = row
            break
    return {
        "row_count": len(rows),
        "first_rows": visible_rows,
        "header_guess": header,
    }


def _signature(repo_root: Path, path: Path) -> dict:
    result = _read_result(repo_root, path)
    route = result.route_info
    return {
        "kind": result.kind,
        "route_info": route,
        "signature": _row_signature(result.rows),
    }


def _relative_or_absolute(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return str(path)


def _support_peers(repo_root: Path, matrix_path: Path, fingerprint_id: str, suffix: str, limit: int = 20) -> list[dict]:
    rows = _load_support_rows(matrix_path)
    testdata_root = repo_root / "bank-statement-standardization" / "testdata"
    peers = []
    for row in rows:
        row_fingerprint_id = row.get("fingerprint_id") or row.get("router类")
        if row_fingerprint_id != fingerprint_id:
            continue
        if row.get("测试结果") and row.get("测试结果") != "PASS":
            continue
        rel = row.get("文件路径", "")
        if suffix and Path(rel).suffix.lower() != suffix:
            continue
        path = testdata_root / rel
        peers.append({
            "file_path": rel,
            "bank": row.get("银行", ""),
            "account_type_yaml": row.get("账户类型(YAML)", ""),
            "yaml_fingerprint": row.get("YAML指纹", ""),
            "exists": path.exists(),
        })
        if len(peers) >= limit:
            break
    return peers


def _recommendation(route_info: dict, peers: list[dict]) -> dict:
    decision = route_info.get("decision")
    fingerprint_id = route_info.get("fingerprint_id", "")
    if decision == "matched" and fingerprint_id:
        if peers:
            return {
                "status": "matched_unique_compare_peers",
                "action": "Compare target signature with support_matrix peers before updating support_matrix.",
            }
        return {
            "status": "matched_unique_no_peers",
            "action": "Treat as first maintained sample for this fingerprint; verify bank/account_type evidence manually.",
        }
    if decision == "ambiguous":
        return {
            "status": "ambiguous",
            "action": "Do not add a new fingerprint. Refine existing matched fingerprints until exactly one remains.",
        }
    return {
        "status": "unmatched",
        "action": "Inspect target signature and generic support_matrix peers. Draft a new fingerprint only if the format is stable and reusable.",
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Inspect a file against reader fingerprint YAML")
    parser.add_argument("--repo-root", default=".", help="Repository root containing ymb-standardization-core")
    parser.add_argument("--file", required=True, help="Target Excel/PDF file")
    parser.add_argument(
        "--support-matrix",
        default="bank-statement-standardization/testdata/support_matrix.xlsx",
        help="support_matrix.xlsx path relative to repo root or absolute",
    )
    args = parser.parse_args(argv)

    repo_root = Path(args.repo_root).resolve()
    file_path = Path(args.file)
    if not file_path.is_absolute():
        file_path = repo_root / file_path
    matrix_path = Path(args.support_matrix)
    if not matrix_path.is_absolute():
        matrix_path = repo_root / matrix_path

    result = _signature(repo_root, file_path)
    route_info = result["route_info"]
    peers = _support_peers(repo_root, matrix_path, route_info.get("fingerprint_id", ""), file_path.suffix.lower())
    report = {
        "file": _relative_or_absolute(file_path, repo_root),
        "target": result,
        "support_matrix": {
            "path": _relative_or_absolute(matrix_path, repo_root),
            "peer_count": len(peers),
            "peers": peers,
        },
        "recommendation": _recommendation(route_info, peers),
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
