#!/usr/bin/env python3
"""Compare one file with support_matrix peers for a parser fingerprint."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
import sys


def _add_core_path(repo_root: Path) -> None:
    core = repo_root / "ymb-standardization-core"
    if str(core) not in sys.path:
        sys.path.insert(0, str(core))


def _load_support_rows(matrix_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(matrix_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(value or "").strip() for value in rows[0]]
    records = []
    for values in rows[1:]:
        rec = {}
        for idx, name in enumerate(header):
            if name:
                rec[name] = "" if idx >= len(values) or values[idx] is None else str(values[idx]).strip()
        if any(rec.values()):
            records.append(rec)
    return records


def _read_rows(repo_root: Path, path: Path):
    _add_core_path(repo_root)
    from ymb_standardization_core import core
    from ymb_standardization_core.parsers import input_router

    input_router.configure_readers(core.read_rows_excel, core.read_rows_csv, core.NotABankStatement)
    result = input_router.read_rows(str(path))
    return result.kind, result.rows, result.route_info


def _signature(rows: list) -> dict:
    first_rows = [[str(cell or "").strip() for cell in row] for row in rows[:5]]
    header_guess = []
    for row in first_rows:
        if sum(1 for cell in row if cell) >= 3:
            header_guess = row
            break
    non_empty_counts = [sum(1 for cell in row if cell) for row in first_rows]
    return {
        "row_count": len(rows),
        "header_guess": header_guess,
        "first_rows": first_rows,
        "non_empty_counts": non_empty_counts,
    }


def _compatibility(target: dict, peer: dict) -> dict:
    target_header = target.get("header_guess") or []
    peer_header = peer.get("header_guess") or []
    same_header = target_header == peer_header
    shared = sorted(set(target_header) & set(peer_header))
    missing_from_peer = [item for item in target_header if item not in peer_header]
    extra_in_peer = [item for item in peer_header if item not in target_header]
    return {
        "same_header": same_header,
        "shared_header_count": len(shared),
        "missing_from_peer": missing_from_peer,
        "extra_in_peer": extra_in_peer,
    }


def _relative(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return str(path)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Compare a target file with support_matrix peers")
    parser.add_argument("--repo-root", default=".")
    parser.add_argument("--file", required=True)
    parser.add_argument("--fingerprint-id", required=True, help="Fingerprint id to compare")
    parser.add_argument("--limit", type=int, default=20)
    parser.add_argument("--support-matrix", default="bank-statement-standardization/testdata/support_matrix.xlsx")
    args = parser.parse_args(argv)

    repo_root = Path(args.repo_root).resolve()
    file_path = Path(args.file)
    if not file_path.is_absolute():
        file_path = repo_root / file_path
    matrix_path = Path(args.support_matrix)
    if not matrix_path.is_absolute():
        matrix_path = repo_root / matrix_path
    testdata_root = repo_root / "bank-statement-standardization" / "testdata"

    kind, rows, route_info = _read_rows(repo_root, file_path)
    target_sig = _signature(rows)
    peers = []
    for row in _load_support_rows(matrix_path):
        row_fingerprint_id = row.get("fingerprint_id") or row.get("router类")
        if row_fingerprint_id != args.fingerprint_id:
            continue
        if row.get("测试结果") and row.get("测试结果") != "PASS":
            continue
        peer_path = testdata_root / row.get("文件路径", "")
        if peer_path.suffix.lower() != file_path.suffix.lower() or not peer_path.exists():
            continue
        peer_kind, peer_rows, peer_route = _read_rows(repo_root, peer_path)
        peer_sig = _signature(peer_rows)
        peers.append({
            "file_path": row.get("文件路径", ""),
            "kind": peer_kind,
            "route_info": peer_route,
            "signature": peer_sig,
            "compatibility": _compatibility(target_sig, peer_sig),
        })
        if len(peers) >= args.limit:
            break

    report = {
        "file": _relative(file_path, repo_root),
        "parser": args.parser,
        "target": {
            "kind": kind,
            "route_info": route_info,
            "signature": target_sig,
        },
        "peer_count": len(peers),
        "peers": peers,
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
