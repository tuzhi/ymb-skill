#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""经营流水 BI 分析报告生成器。

串行于 bank-statement-standardization：输入其产物（已标准化/整合/打标的流水），
输出每个经营主体一份 .xlsx 报告，用 Excel 原生图表呈现余额波动、收支结构、
交易对手、上下游、流入流出趋势等审批分析要件。

输入可为：
  * 标准化交付物 .xlsx（含「标准化流水主表/主表/流水」工作表）；
  * `*__打标流水.csv` / `*__整合流水.csv` / 任意含标准字段的 CSV；
  * 一个文件夹（自动挑选其中的标准化产物，逐主体出报告）。

只读取标准字段，不改写、不重排明细（行序沿用标准化交付的「余额连续性」顺序，
绝不按交易时间重排——见 references/notes）。

用法：
  python build_bi_report.py --input "<文件或文件夹>" [--client 客户名] [--out-dir DIR]
"""
import argparse
import csv
import glob
import math
import os
import sys
from collections import OrderedDict, defaultdict
from datetime import datetime

try:
    import openpyxl
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("缺少 openpyxl，请先 `pip install openpyxl`。\n")
    raise

# ----------------------------------------------------------------------------
# 标准字段（与 bank-statement-standardization 交付物对齐）
# ----------------------------------------------------------------------------
F_TIME = "交易时间"
F_ACCT = "本方账户"
F_ACCT_TYPE = "账户类型"
F_SELF = "本方名称"
F_CP = "对手名称"
F_IN = "收入金额"
F_OUT = "支出金额"
F_AMT = "交易金额"
F_BAL = "账户余额"
F_DIR = "收支方向"
F_T1 = "一级标签"
F_T2 = "二级标签"
F_T3 = "三级标签"
F_CUST = "客户名称"
F_SUBJ = "主体名称"

SHEET_CANDIDATES = ["标准化流水主表", "主表", "流水", "打标流水", "整合流水"]

# 审批参考阈值（授信方可调；阈值仅作初判，不构成授信结论）
THRESHOLDS = {
    "近12月总流入下限": 20_000_000.0,   # 业务规模
    "月流入变异系数上限": 0.5,           # 稳定性
    "近12月零流入月数上限": 2,           # 连续性
    "最长连续零流入月上限": 1,           # 连续性
    "流入增长率下限": -0.20,             # 成长性
    "TOP1对手占比上限": 0.50,           # 集中度
    "TOP5对手占比上限": 0.80,           # 集中度
    "现金交易占比上限": 0.30,           # 现金依赖
    "往来款占比上限": 0.30,             # 资金往来/疑似空转
    "夜间交易占比上限": 0.20,           # 夜间交易（21:00-06:00）笔数/金额占比
    "收支比上下限": (0.6, 1.6),         # 收支匹配（流出/流入）
}

STRATEGY_VERSION = "BANKFLOW-BI-V1.0"

# ----------------------------------------------------------------------------
# 样式
# ----------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "2E5496"
LIGHT = "D9E1F2"
LIGHTER = "EAF0FA"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"
GREY = "808080"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def title_font(size=16, color="FFFFFF"):
    return Font(name="微软雅黑", size=size, bold=True, color=color)


def hdr_font(color="FFFFFF"):
    return Font(name="微软雅黑", size=10, bold=True, color=color)


def cell_font(bold=False):
    return Font(name="微软雅黑", size=10, bold=bold)


def fill(color):
    return PatternFill("solid", fgColor=color)


# ----------------------------------------------------------------------------
# 解析与加载
# ----------------------------------------------------------------------------
def to_num(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "")
    if s in ("", "-", "无", "None", "nan"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def month_of(s):
    """从交易时间取 YYYY-MM；失败返回空串。"""
    if not s:
        return ""
    s = str(s).strip()
    # 常见：YYYY-MM-DD..., YYYY/MM/DD, YYYYMMDD
    for sep in ("-", "/"):
        if sep in s and len(s) >= 7:
            parts = s.replace(" ", sep).split(sep)
            if len(parts) >= 2 and len(parts[0]) == 4 and parts[0].isdigit():
                mm = parts[1].zfill(2)[:2]
                return f"{parts[0]}-{mm}"
    digits = "".join(ch for ch in s if ch.isdigit())
    if len(digits) >= 6:
        return f"{digits[:4]}-{digits[4:6]}"
    return ""


def hour_of(s):
    """取交易时间的小时（0-23）。无时点信息或被补零的 00:00:00 视为「无时点」返回 None，
    避免把仅有日期的流水误判为午夜夜间交易。"""
    if not s:
        return None
    s = str(s).strip().replace("T", " ")
    if " " not in s:
        return None
    tm = s.split()[-1]
    bits = tm.split(":")
    if len(bits) < 2:
        return None
    try:
        h, mi = int(bits[0]), int(bits[1])
        se = int(bits[2]) if len(bits) > 2 else 0
    except ValueError:
        return None
    if not (0 <= h <= 23):
        return None
    if h == 0 and mi == 0 and se == 0:
        return None
    return h


def load_rows_from_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = None
    for name in SHEET_CANDIDATES:
        for sn in wb.sheetnames:
            if name in sn:
                ws = wb[sn]
                break
        if ws:
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = None
    for row in it:
        if row and any(c == F_TIME or c == F_BAL for c in row if c is not None):
            header = [str(c).strip() if c is not None else "" for c in row]
            break
    if header is None:
        wb.close()
        return []
    rows = []
    for row in it:
        if row is None or all(c is None or str(c).strip() == "" for c in row):
            continue
        rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
    wb.close()
    return rows


def load_rows_from_csv(path):
    for enc in ("utf-8-sig", "utf-8", "gbk"):
        try:
            with open(path, "r", encoding=enc, newline="") as f:
                return list(csv.DictReader(f))
        except (UnicodeDecodeError, UnicodeError):
            continue
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        return list(csv.DictReader(f))


def load_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return load_rows_from_xlsx(path)
    return load_rows_from_csv(path)


def pick_input_files(input_path):
    """文件夹 -> 找标准化产物；文件 -> 直接用。"""
    if os.path.isfile(input_path):
        return [input_path]
    pats = ["*__打标流水.csv", "*__整合流水.csv", "*标准化*.xlsx", "*流水*.xlsx", "*.csv"]
    seen, files = set(), []
    for pat in pats:
        for f in sorted(glob.glob(os.path.join(input_path, "**", pat), recursive=True)):
            if f not in seen and "_工作区" not in f:
                seen.add(f)
                files.append(f)
        if files:
            break
    return files


# ----------------------------------------------------------------------------
# 主体拆分与分析
# ----------------------------------------------------------------------------
def subject_key(r):
    for f in (F_CUST, F_SUBJ, F_SELF):
        v = (r.get(f) or "").strip() if isinstance(r.get(f), str) else r.get(f)
        if v:
            return str(v).strip()
    return "未知主体"


def split_subjects(rows):
    groups = OrderedDict()
    for r in rows:
        groups.setdefault(subject_key(r), []).append(r)
    return groups


def row_inflow_outflow(r):
    inc = to_num(r.get(F_IN))
    out = to_num(r.get(F_OUT))
    if inc == 0 and out == 0:
        amt = to_num(r.get(F_AMT))
        if amt > 0:
            inc = amt
        elif amt < 0:
            out = -amt
    return inc, out


def analyze(rows):
    """计算一个主体的全部分析结果。行序沿用输入顺序（余额连续性），不重排。"""
    a = {}
    accounts = OrderedDict()
    for r in rows:
        accounts.setdefault((r.get(F_ACCT) or "").strip(), []).append(r)
    a["账户数"] = len([k for k in accounts if k])

    total_in = total_out = 0.0
    months = OrderedDict()          # 月 -> dict
    cp_in = defaultdict(float)
    cp_out = defaultdict(float)
    cp_cnt = defaultdict(int)
    in_by_t2 = defaultdict(float)
    out_by_t2 = defaultdict(float)
    cash_amt = 0.0
    settle_amt = 0.0                # 往来款
    neg_bal_cnt = 0
    bals = []
    min_bal = math.inf
    max_bal = -math.inf
    time_min = time_max = None
    n_tx = 0
    # 夜间交易（21:00-06:00）监测——仅在「有时点」的交易上统计，避免日期补零 00:00:00 误判
    timed_cnt = 0
    timed_amt = 0.0
    night_cnt = 0
    night_amt = 0.0

    for r in rows:
        inc, out = row_inflow_outflow(r)
        if inc == 0 and out == 0:
            continue
        n_tx += 1
        total_in += inc
        total_out += out
        hh = hour_of(r.get(F_TIME))
        if hh is not None:
            timed_cnt += 1
            timed_amt += inc + out
            if hh >= 21 or hh < 6:
                night_cnt += 1
                night_amt += inc + out
        m = month_of(r.get(F_TIME))
        if m:
            mm = months.setdefault(m, {"in": 0.0, "out": 0.0, "cnt": 0})
            mm["in"] += inc
            mm["out"] += out
            mm["cnt"] += 1
        cp = (r.get(F_CP) or "").strip()
        if cp and cp not in ("现金", "ATM", "本人", "-"):
            cp_in[cp] += inc
            cp_out[cp] += out
            cp_cnt[cp] += 1
        t2 = (r.get(F_T2) or "其他").strip() or "其他"
        if inc > 0:
            in_by_t2[t2] += inc
        if out > 0:
            out_by_t2[t2] += out
        t3 = (r.get(F_T3) or "")
        t1 = (r.get(F_T1) or "")
        blob = f"{t1}{t2}{t3}"
        if "现金" in blob:
            cash_amt += inc + out
        if "往来" in blob:
            settle_amt += inc + out
        bal = r.get(F_BAL)
        if bal is not None and str(bal).strip() not in ("", "-"):
            b = to_num(bal)
            bals.append(b)
            if b < 0:
                neg_bal_cnt += 1
            min_bal = min(min_bal, b)
            max_bal = max(max_bal, b)
        t = (r.get(F_TIME) or "").strip() if isinstance(r.get(F_TIME), str) else str(r.get(F_TIME) or "")
        if t:
            time_min = t if time_min is None or t < time_min else time_min
            time_max = t if time_max is None or t > time_max else time_max

    # 月末/月均/月最低余额：按各账户输入顺序取该月最后一行余额，再跨账户求和
    month_end = defaultdict(float)
    month_min = defaultdict(lambda: 0.0)
    month_sum = defaultdict(float)
    month_balcnt = defaultdict(int)
    for acct, arows in accounts.items():
        last_by_month = {}
        for r in arows:
            m = month_of(r.get(F_TIME))
            b = r.get(F_BAL)
            if not m or b is None or str(b).strip() in ("", "-"):
                continue
            bv = to_num(b)
            last_by_month[m] = bv  # 输入顺序最后一行即月末
            month_sum[m] += bv
            month_balcnt[m] += 1
            cur = month_min.get(m)
            month_min[m] = bv if (m not in month_min or bv < cur) else cur
        for m, bv in last_by_month.items():
            month_end[m] += bv

    a["总流入"] = total_in
    a["总流出"] = total_out
    a["净流入"] = total_in - total_out
    a["交易笔数"] = n_tx
    a["收支比"] = (total_out / total_in) if total_in else 0.0
    a["期初时间"] = time_min or ""
    a["期末时间"] = time_max or ""
    a["最低余额"] = (min_bal if min_bal != math.inf else 0.0)
    a["最高余额"] = (max_bal if max_bal != -math.inf else 0.0)
    a["日均余额近似"] = (sum(bals) / len(bals)) if bals else 0.0
    a["负余额次数"] = neg_bal_cnt
    a["现金交易占比"] = (cash_amt / (total_in + total_out)) if (total_in + total_out) else 0.0
    a["往来款占比"] = (settle_amt / (total_in + total_out)) if (total_in + total_out) else 0.0

    # 夜间交易（21:00-06:00）——占比口径以「有时点交易」为分母
    a["夜间笔数"] = night_cnt
    a["夜间金额"] = night_amt
    a["有时点笔数"] = timed_cnt
    a["时点覆盖率"] = (timed_cnt / n_tx) if n_tx else 0.0
    a["夜间笔数占比"] = (night_cnt / timed_cnt) if timed_cnt else None
    a["夜间金额占比"] = (night_amt / timed_amt) if timed_amt else None

    # 月度时序表（按月份升序）
    mkeys = sorted(months.keys())
    monthly = []
    for m in mkeys:
        mm = months[m]
        monthly.append({
            "月份": m,
            "流入金额": mm["in"],
            "流出金额": mm["out"],
            "净流入": mm["in"] - mm["out"],
            "交易笔数": mm["cnt"],
            "月末余额": month_end.get(m, 0.0),
            "月均余额": (month_sum[m] / month_balcnt[m]) if month_balcnt.get(m) else 0.0,
            "月最低余额": month_min.get(m, 0.0),
        })
    a["月度"] = monthly
    a["采集时长月"] = len(mkeys)
    a["活跃月数"] = sum(1 for x in monthly if x["交易笔数"] > 0)

    # 标签结构
    a["收入结构"] = sorted(in_by_t2.items(), key=lambda kv: -kv[1])
    a["支出结构"] = sorted(out_by_t2.items(), key=lambda kv: -kv[1])

    # 交易对手 TOP（流入=上游/客户；流出=下游/供应商）
    top_in = sorted(cp_in.items(), key=lambda kv: -kv[1])[:10]
    top_out = sorted(cp_out.items(), key=lambda kv: -kv[1])[:10]
    a["十大流入对手"] = [(cp, amt, cp_cnt[cp], (amt / total_in if total_in else 0)) for cp, amt in top_in if amt > 0]
    a["十大流出对手"] = [(cp, amt, cp_cnt[cp], (amt / total_out if total_out else 0)) for cp, amt in top_out if amt > 0]

    # 净往来（上下游）：与同一对手 收-支 净额
    net = {cp: cp_in.get(cp, 0) - cp_out.get(cp, 0) for cp in set(cp_in) | set(cp_out)}
    up = sorted([(cp, v) for cp, v in net.items() if v > 0], key=lambda kv: -kv[1])[:10]
    down = sorted([(cp, v) for cp, v in net.items() if v < 0], key=lambda kv: kv[1])[:10]
    a["净上游"] = up
    a["净下游"] = down
    twoway = sorted([(cp, cp_in[cp], cp_out[cp]) for cp in set(cp_in) & set(cp_out)
                     if cp_in[cp] > 0 and cp_out[cp] > 0], key=lambda x: -(x[1] + x[2]))[:10]
    a["双向往来对手"] = twoway

    # 集中度
    a["TOP1流入占比"] = a["十大流入对手"][0][3] if a["十大流入对手"] else 0.0
    a["TOP5流入占比"] = sum(x[3] for x in a["十大流入对手"][:5])

    # 成长性：近12月 vs 远12月 流入
    a["流入增长率"] = growth_rate(monthly)
    a["月流入CV"] = cv([x["流入金额"] for x in monthly]) if monthly else 0.0
    a["近12月零流入月数"], a["最长连续零流入月"] = zero_gaps([x["流入金额"] for x in monthly[-12:]])
    a["近12月总流入"] = sum(x["流入金额"] for x in monthly[-12:])

    return a


def cv(xs):
    xs = [x for x in xs]
    if not xs:
        return 0.0
    mean = sum(xs) / len(xs)
    if mean == 0:
        return 0.0
    var = sum((x - mean) ** 2 for x in xs) / len(xs)
    return math.sqrt(var) / mean


def zero_gaps(xs):
    zero = sum(1 for x in xs if x <= 0)
    longest = cur = 0
    for x in xs:
        if x <= 0:
            cur += 1
            longest = max(longest, cur)
        else:
            cur = 0
    return zero, longest


def growth_rate(monthly):
    # 需满 24 个月才能做「近12月 vs 远12月」对比；否则远12月不完整，增长率会失真，按「数据不足」处理。
    if len(monthly) < 24:
        return None
    near = sum(x["流入金额"] for x in monthly[-12:])
    far = sum(x["流入金额"] for x in monthly[-24:-12])
    if far <= 0:
        return None
    return (near - far) / far


# ----------------------------------------------------------------------------
# 报告写出
# ----------------------------------------------------------------------------
def yi(x):
    return x / 10000.0  # 元 -> 万元


def cap_structure(items, top=8):
    """饼图切片过多会导致内部百分比标签互相挤叠：只保留金额最大的 top 项，其余并入「其他」。
    items 已按金额降序。"""
    items = [(k, v) for k, v in items if v]
    if len(items) <= top:
        return items
    rest = sum(v for _, v in items[top:])
    head = items[:top]
    return head + [("其他", rest)] if rest > 0 else head


def write_report(subject, a, client, out_path):
    wb = openpyxl.Workbook()
    _sheet_report(wb.active, subject, a, client)
    _sheet_monthly(wb.create_sheet("月度时序表"), a)
    _sheet_structure(wb.create_sheet("收支结构"), a)
    _sheet_counterparty(wb.create_sheet("交易对手"), a)
    _sheet_dashboard(wb.create_sheet("可视化看板"), a)
    wb.save(out_path)


def _band(ws, row, text, col_span=10, level=0):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=col_span)
    c = ws.cell(row, 2, text)
    c.fill = fill(NAVY if level == 0 else BLUE)
    c.font = hdr_font()
    c.alignment = LEFT
    ws.row_dimensions[row].height = 22
    return row + 1


def _kv_table(ws, row, pairs, per_row=2):
    """成对 KV 表：每行 per_row 组 (标签,值)。"""
    col0 = 2
    for i, (k, v) in enumerate(pairs):
        block = i % per_row
        if block == 0 and i > 0:
            row += 1
        base = col0 + block * 4
        kc = ws.cell(row, base, k)
        kc.fill = fill(LIGHT)
        kc.font = cell_font(bold=True)
        kc.alignment = LEFT
        kc.border = BORDER
        ws.merge_cells(start_row=row, start_column=base, end_row=row, end_column=base + 1)
        ws.cell(row, base + 1).border = BORDER
        vc = ws.cell(row, base + 2, v)
        vc.font = cell_font()
        vc.alignment = LEFT
        vc.border = BORDER
        ws.merge_cells(start_row=row, start_column=base + 2, end_row=row, end_column=base + 3)
        ws.cell(row, base + 3).border = BORDER
    return row + 2


def _fmt_money(x):
    return f"{x:,.2f}"


def _sheet_report(ws, subject, a, client):
    ws.title = "分析报告"
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 2.5, "B": 18, "C": 16, "D": 18, "E": 16, "F": 18,
                   "G": 16, "H": 18, "I": 16, "J": 18}.items():
        ws.column_dimensions[col].width = w

    # 抬头
    ws.merge_cells("B2:J2")
    t = ws.cell(2, 2, "经营流水 BI 分析报告")
    t.fill = fill(NAVY)
    t.font = title_font(18)
    t.alignment = CENTER
    ws.row_dimensions[2].height = 36
    row = _kv_table(ws, 4, [
        ("客户名称", client or subject),
        ("报告日期", datetime.now().strftime("%Y-%m-%d")),
        ("经营主体", subject),
        ("分析策略", STRATEGY_VERSION),
        ("流水区间", f"{a['期初时间']} ~ {a['期末时间']}"),
        ("账户数", a["账户数"]),
    ], per_row=2)

    # 模块一：经营流水观察
    row = _band(ws, row, "模块一  经营流水观察", 10, 0)
    row += 1
    row = _band(ws, row, "PART 1  流水整体概况", 10, 1)
    row = _kv_table(ws, row, [
        ("采集时长(月)", a["采集时长月"]),
        ("活跃月数", a["活跃月数"]),
        ("交易笔数", a["交易笔数"]),
        ("近12月总流入(元)", _fmt_money(a["近12月总流入"])),
        ("总流入(元)", _fmt_money(a["总流入"])),
        ("总流出(元)", _fmt_money(a["总流出"])),
        ("净流入(元)", _fmt_money(a["净流入"])),
        ("收支比(流出/流入)", f"{a['收支比']:.2f}"),
        ("日均余额(近似,元)", _fmt_money(a["日均余额近似"])),
        ("最低余额(元)", _fmt_money(a["最低余额"])),
        ("最高余额(元)", _fmt_money(a["最高余额"])),
        ("负余额次数", a["负余额次数"]),
        ("夜间交易笔数(21-6点)",
         f"{a['夜间笔数']}（占有时点 {a['夜间笔数占比']:.1%}）" if a["夜间笔数占比"] is not None
         else f"{a['夜间笔数']}（无时点数据）"),
        ("夜间交易金额(元)",
         f"{_fmt_money(a['夜间金额'])}（占有时点 {a['夜间金额占比']:.1%}）" if a["夜间金额占比"] is not None
         else f"{_fmt_money(a['夜间金额'])}（无时点数据）"),
    ], per_row=2)

    row = _band(ws, row, "PART 2  收支结构（按二级用途，金额单位：元）", 10, 1)
    # 收入结构 + 支出结构 并排
    hdr = row
    for j, txt in enumerate(["收入用途", "金额", "占比"]):
        c = ws.cell(hdr, 2 + j, txt)
        c.fill = fill(BLUE)
        c.font = hdr_font()
        c.alignment = CENTER
        c.border = BORDER
    for j, txt in enumerate(["支出用途", "金额", "占比"]):
        c = ws.cell(hdr, 6 + j, txt)
        c.fill = fill(BLUE)
        c.font = hdr_font()
        c.alignment = CENTER
        c.border = BORDER
    inc = a["收入结构"][:8]
    out = a["支出结构"][:8]
    for i in range(max(len(inc), len(out), 1)):
        r = hdr + 1 + i
        if i < len(inc):
            k, v = inc[i]
            ws.cell(r, 2, k).border = BORDER
            ws.cell(r, 3, round(v, 2)).border = BORDER
            ws.cell(r, 4, f"{v / a['总流入']:.1%}" if a["总流入"] else "-").border = BORDER
        if i < len(out):
            k, v = out[i]
            ws.cell(r, 6, k).border = BORDER
            ws.cell(r, 7, round(v, 2)).border = BORDER
            ws.cell(r, 8, f"{v / a['总流出']:.1%}" if a["总流出"] else "-").border = BORDER
        for cc in (2, 3, 4, 6, 7, 8):
            ws.cell(r, cc).font = cell_font()
    row = hdr + 1 + max(len(inc), len(out), 1) + 1

    # 模块二：审批要件指标看板
    row = _band(ws, row, "模块二  审批要件指标看板（阈值为参考口径，授信方可调）", 10, 0)
    row += 1
    row = _indicator_panel(ws, row, a)

    ws.cell(row + 1, 2, "说明：阈值仅为初筛参考，红/黄/绿判断不构成授信结论；明细见各数据表与「可视化看板」。").font = Font(
        name="微软雅黑", size=8, italic=True, color=GREY)


def _indicator_panel(ws, row, a):
    headers = ["指标类别", "指标", "口径", "参考阈值", "企业实际", "判断"]
    widths = [None]
    for j, h in enumerate(headers):
        c = ws.cell(row, 2 + j, h)
        c.fill = fill(BLUE)
        c.font = hdr_font()
        c.alignment = CENTER
        c.border = BORDER

    T = THRESHOLDS
    gr = a["流入增长率"]
    lo, hi = T["收支比上下限"]
    nzp = a["夜间笔数占比"]   # 夜间笔数占比
    nza = a["夜间金额占比"]   # 夜间金额占比
    cov = a["时点覆盖率"]
    items = [
        ("业务规模", "近12月总流入", "近12个月流入合计", f"≥{yi(T['近12月总流入下限']):.0f}万",
         f"{yi(a['近12月总流入']):,.1f}万", a["近12月总流入"] >= T["近12月总流入下限"]),
        ("收支匹配", "收支比", "总流出/总流入", f"{lo}~{hi}",
         f"{a['收支比']:.2f}", lo <= a["收支比"] <= hi),
        ("稳定性", "月流入变异系数", "月流入CV(σ/μ)", f"≤{T['月流入变异系数上限']}",
         f"{a['月流入CV']:.2f}", a["月流入CV"] <= T["月流入变异系数上限"]),
        ("连续性", "近12月零流入月数", "近12月无流入月份数", f"≤{T['近12月零流入月数上限']}",
         f"{a['近12月零流入月数']}", a["近12月零流入月数"] <= T["近12月零流入月数上限"]),
        ("连续性", "最长连续零流入月", "近12月最长连续无流入", f"≤{T['最长连续零流入月上限']}",
         f"{a['最长连续零流入月']}", a["最长连续零流入月"] <= T["最长连续零流入月上限"]),
        ("成长性", "流入增长率", "近12月vs远12月流入", f"≥{T['流入增长率下限']:.0%}",
         (f"{gr:.1%}" if gr is not None else "数据不足"), (gr is None or gr >= T["流入增长率下限"])),
        ("集中度", "TOP1对手流入占比", "最大流入对手/总流入", f"≤{T['TOP1对手占比上限']:.0%}",
         f"{a['TOP1流入占比']:.1%}", a["TOP1流入占比"] <= T["TOP1对手占比上限"]),
        ("集中度", "TOP5对手流入占比", "前5流入对手/总流入", f"≤{T['TOP5对手占比上限']:.0%}",
         f"{a['TOP5流入占比']:.1%}", a["TOP5流入占比"] <= T["TOP5对手占比上限"]),
        ("现金依赖", "现金交易占比", "现金类金额/总收支", f"≤{T['现金交易占比上限']:.0%}",
         f"{a['现金交易占比']:.1%}", a["现金交易占比"] <= T["现金交易占比上限"]),
        ("资金往来", "往来款占比", "往来类金额/总收支", f"≤{T['往来款占比上限']:.0%}",
         f"{a['往来款占比']:.1%}", a["往来款占比"] <= T["往来款占比上限"]),
        ("余额健康", "负余额次数", "账户余额<0笔数", "=0",
         f"{a['负余额次数']}", a["负余额次数"] == 0),
        ("夜间交易", "夜间笔数占比", f"21:00-06:00笔数/有时点笔数(覆盖{cov:.0%})",
         f"≤{T['夜间交易占比上限']:.0%}",
         (f"{nzp:.1%}" if nzp is not None else "无时点数据"),
         (nzp is None or nzp <= T["夜间交易占比上限"])),
        ("夜间交易", "夜间金额占比", f"21:00-06:00金额/有时点金额(覆盖{cov:.0%})",
         f"≤{T['夜间交易占比上限']:.0%}",
         (f"{nza:.1%}" if nza is not None else "无时点数据"),
         (nza is None or nza <= T["夜间交易占比上限"])),
    ]
    r = row + 1
    for cat, name, basis, thr, actual, ok in items:
        vals = [cat, name, basis, thr, actual, "通过" if ok else "关注"]
        for j, v in enumerate(vals):
            c = ws.cell(r, 2 + j, v)
            c.border = BORDER
            c.font = cell_font(bold=(j == 5))
            c.alignment = CENTER if j != 2 else LEFT
        ws.cell(r, 7).fill = fill(GREEN if ok else YELLOW)
        r += 1
    return r + 1


def _write_table(ws, start_row, headers, rows, money_cols=(), pct_cols=()):
    for j, h in enumerate(headers):
        c = ws.cell(start_row, 1 + j, h)
        c.fill = fill(BLUE)
        c.font = hdr_font()
        c.alignment = CENTER
        c.border = BORDER
    for i, rowvals in enumerate(rows):
        r = start_row + 1 + i
        for j, v in enumerate(rowvals):
            c = ws.cell(r, 1 + j, v)
            c.border = BORDER
            c.font = cell_font()
            if j in money_cols:
                c.number_format = "#,##0.00"
                c.alignment = RIGHT
            elif j in pct_cols:
                c.number_format = "0.0%"
                c.alignment = RIGHT
            else:
                c.alignment = CENTER
    return start_row + 1 + len(rows)


def _sheet_monthly(ws, a):
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 12, "B": 16, "C": 16, "D": 16, "E": 10,
                   "F": 16, "G": 16, "H": 16}.items():
        ws.column_dimensions[col].width = w
    headers = ["月份", "流入金额", "流出金额", "净流入", "交易笔数", "月末余额", "月均余额", "月最低余额"]
    rows = [[m["月份"], m["流入金额"], m["流出金额"], m["净流入"], m["交易笔数"],
             m["月末余额"], m["月均余额"], m["月最低余额"]] for m in a["月度"]]
    _write_table(ws, 1, headers, rows, money_cols=(1, 2, 3, 5, 6, 7))


def _sheet_structure(ws, a):
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 22, "B": 16, "D": 22, "E": 16}.items():
        ws.column_dimensions[col].width = w
    inc = [[k, round(v, 2)] for k, v in cap_structure(a["收入结构"])]
    out = [[k, round(v, 2)] for k, v in cap_structure(a["支出结构"])]
    ws.cell(1, 1, "收入用途结构").font = cell_font(bold=True)
    _write_table(ws, 2, ["收入用途", "金额"], inc, money_cols=(1,))
    ws.cell(1, 4, "支出用途结构").font = cell_font(bold=True)
    # 支出表放在右侧 D 列起
    start = 2
    for j, h in enumerate(["支出用途", "金额"]):
        c = ws.cell(start, 4 + j, h)
        c.fill = fill(BLUE)
        c.font = hdr_font()
        c.alignment = CENTER
        c.border = BORDER
    for i, (k, v) in enumerate(out):
        r = start + 1 + i
        ws.cell(r, 4, k).border = BORDER
        ws.cell(r, 4).font = cell_font()
        cc = ws.cell(r, 5, v)
        cc.border = BORDER
        cc.font = cell_font()
        cc.number_format = "#,##0.00"


def _sheet_counterparty(ws, a):
    ws.sheet_view.showGridLines = False
    for col in "ABCDEFGHIJ":
        ws.column_dimensions[col].width = 22 if col in "AEHJ" else 15
    r = 1
    ws.cell(r, 1, "十大流入对手（上游/客户）").font = cell_font(bold=True)
    r = _write_table(ws, r + 1, ["对手名称", "流入金额", "笔数", "占总流入"],
                     [[c, round(v, 2), n, p] for c, v, n, p in a["十大流入对手"]],
                     money_cols=(1,), pct_cols=(3,)) + 1
    ws.cell(r, 1, "十大流出对手（下游/供应商）").font = cell_font(bold=True)
    r = _write_table(ws, r + 1, ["对手名称", "流出金额", "笔数", "占总流出"],
                     [[c, round(v, 2), n, p] for c, v, n, p in a["十大流出对手"]],
                     money_cols=(1,), pct_cols=(3,)) + 1
    ws.cell(r, 1, "净上游（净收入对手，元）").font = cell_font(bold=True)
    r = _write_table(ws, r + 1, ["对手名称", "净流入"], [[c, round(v, 2)] for c, v in a["净上游"]],
                     money_cols=(1,)) + 1
    ws.cell(r, 1, "净下游（净支出对手，元）").font = cell_font(bold=True)
    r = _write_table(ws, r + 1, ["对手名称", "净流出"], [[c, round(v, 2)] for c, v in a["净下游"]],
                     money_cols=(1,)) + 1
    ws.cell(r, 1, "双向往来对手（同时有收有支，关注资金空转/互转）").font = cell_font(bold=True)
    _write_table(ws, r + 1, ["对手名称", "流入金额", "流出金额"],
                 [[c, round(i, 2), round(o, 2)] for c, i, o in a["双向往来对手"]],
                 money_cols=(1, 2))


def _pie_labels(pie):
    """饼图标签：只显示百分比、交由 Excel 最佳位置摆放（bestFit，带引线），图例靠右，避免内部文字挤叠。"""
    dl = DataLabelList()
    dl.showPercent = True
    dl.showVal = False
    dl.showCatName = False
    dl.showSerName = False
    dl.showLegendKey = False
    dl.dLblPos = "bestFit"
    pie.dataLabels = dl
    if pie.legend:
        pie.legend.position = "r"


def _sheet_dashboard(ws, a):
    """所有原生图表单列纵向排布（自上而下下拉查看），充分利用垂直空间、每图放大避免拥挤。"""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.cell(1, 1, "可视化看板").font = title_font(14, NAVY)
    ws.cell(2, 1, "（图表自上而下纵向排布，请下拉查看全部）").font = Font(
        name="微软雅黑", size=9, italic=True, color=GREY)

    # 纵向游标：每张图占用的行数 ≈ 高度(cm)*1.9，再留 3 行间距。
    state = {"row": 4}

    def place(chart, height_cm):
        ws.add_chart(chart, f"B{state['row']}")
        state["row"] += int(height_cm * 1.9) + 3

    n_month = len(a["月度"])
    monthly_ws = ws.parent["月度时序表"]
    cats = Reference(monthly_ws, min_col=1, min_row=2, max_row=n_month + 1) if n_month else None

    # 1) 月度收支趋势：柱(流入/流出) + 折线(净流入)
    if n_month >= 1:
        bar = BarChart()
        bar.title = "月度流入 / 流出趋势"
        bar.type = "col"
        bar.height, bar.width = 11, 26
        data = Reference(monthly_ws, min_col=2, max_col=3, min_row=1, max_row=n_month + 1)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        line = LineChart()
        net = Reference(monthly_ws, min_col=4, max_col=4, min_row=1, max_row=n_month + 1)
        line.add_data(net, titles_from_data=True)
        line.y_axis.axId = 200
        line.y_axis.title = "净流入"
        bar.y_axis.title = "金额(元)"
        bar.y_axis.crosses = "autoZero"
        line.y_axis.crosses = "max"
        bar += line
        place(bar, 11)

        # 2) 余额波动趋势：折线(月末/月最低)
        line2 = LineChart()
        line2.title = "余额波动趋势（月末 / 月最低）"
        line2.height, line2.width = 10, 26
        d2 = Reference(monthly_ws, min_col=6, max_col=6, min_row=1, max_row=n_month + 1)
        d3 = Reference(monthly_ws, min_col=8, max_col=8, min_row=1, max_row=n_month + 1)
        line2.add_data(d2, titles_from_data=True)
        line2.add_data(d3, titles_from_data=True)
        line2.set_categories(cats)
        line2.y_axis.title = "余额(元)"
        place(line2, 10)

    # 3) 收入用途结构 饼图（放大 + 图例靠右，避免标签拥挤）
    struct_ws = ws.parent["收支结构"]
    n_in = len(cap_structure(a["收入结构"]))   # 与「收支结构」表同口径（top8 + 其他）
    if n_in:
        pie = PieChart()
        pie.title = "收入用途结构"
        pie.height, pie.width = 12, 22
        d = Reference(struct_ws, min_col=2, min_row=2, max_row=n_in + 2)
        labels = Reference(struct_ws, min_col=1, min_row=3, max_row=n_in + 2)
        pie.add_data(d, titles_from_data=True)
        pie.set_categories(labels)
        _pie_labels(pie)
        place(pie, 12)

    # 4) 支出用途结构 饼图
    n_out = len(cap_structure(a["支出结构"]))
    if n_out:
        pie2 = PieChart()
        pie2.title = "支出用途结构"
        pie2.height, pie2.width = 12, 22
        d = Reference(struct_ws, min_col=5, min_row=2, max_row=n_out + 2)
        labels = Reference(struct_ws, min_col=4, min_row=3, max_row=n_out + 2)
        pie2.add_data(d, titles_from_data=True)
        pie2.set_categories(labels)
        _pie_labels(pie2)
        place(pie2, 12)

    # 5) 十大流入对手 条形图
    cp_ws = ws.parent["交易对手"]
    n_topin = len(a["十大流入对手"])
    if n_topin:
        b = BarChart()
        b.type = "bar"
        b.title = "十大流入对手（上游 / 客户）"
        b.height, b.width = 11, 24
        d = Reference(cp_ws, min_col=2, min_row=2, max_row=2 + n_topin)
        labels = Reference(cp_ws, min_col=1, min_row=3, max_row=2 + n_topin)
        b.add_data(d, titles_from_data=True)
        b.set_categories(labels)
        place(b, 11)

    # 6) 十大流出对手 条形图
    if a["十大流出对手"]:
        out_hdr = 2 + n_topin + 2  # 流入表(标题1 + 表头1 + n) + 空行 + 流出标题 -> 表头行
        n_topout = len(a["十大流出对手"])
        b2 = BarChart()
        b2.type = "bar"
        b2.title = "十大流出对手（下游 / 供应商）"
        b2.height, b2.width = 11, 24
        d = Reference(cp_ws, min_col=2, min_row=out_hdr, max_row=out_hdr + n_topout)
        labels = Reference(cp_ws, min_col=1, min_row=out_hdr + 1, max_row=out_hdr + n_topout)
        b2.add_data(d, titles_from_data=True)
        b2.set_categories(labels)
        place(b2, 11)


# ----------------------------------------------------------------------------
# 入口
# ----------------------------------------------------------------------------
def safe_name(s):
    return "".join(c for c in str(s) if c not in '\\/:*?"<>|').strip() or "主体"


def main():
    ap = argparse.ArgumentParser(description="经营流水 BI 分析报告生成器")
    ap.add_argument("--input", required=True, help="标准化产物文件或文件夹")
    ap.add_argument("--client", default="", help="授信客户归档名（用于命名）")
    ap.add_argument("--out-dir", default="", help="输出目录，默认输入同级")
    args = ap.parse_args()

    files = pick_input_files(args.input)
    if not files:
        print(f"【错误】未在 {args.input} 找到可用的标准化流水产物。")
        return 2

    rows = []
    for f in files:
        try:
            rows.extend(load_rows(f))
        except Exception as e:
            print(f"【警告】读取失败 {f}: {e}")
    if not rows:
        print("【错误】未读取到任何流水数据行。")
        return 2

    out_dir = args.out_dir or (args.input if os.path.isdir(args.input) else os.path.dirname(args.input))
    out_dir = out_dir or "."
    os.makedirs(out_dir, exist_ok=True)

    subjects = split_subjects(rows)
    print(f"读取 {len(rows)} 行，识别 {len(subjects)} 个经营主体。")
    written = []
    for subj, srows in subjects.items():
        a = analyze(srows)
        if a["交易笔数"] == 0:
            print(f"  跳过 {subj}：无有效交易。")
            continue
        cli = args.client or subj
        fname = f"{safe_name(cli)}__经营流水BI分析报告.xlsx" if len(subjects) == 1 \
            else f"{safe_name(cli)}__{safe_name(subj)}__经营流水BI分析报告.xlsx"
        out_path = os.path.join(out_dir, fname)
        write_report(subj, a, args.client, out_path)
        written.append(out_path)
        print(f"  ✓ {subj}: {out_path}  （{a['采集时长月']}个月 / {a['交易笔数']}笔 / 净流入 {a['净流入']:,.0f}）")

    print(f"完成，输出 {len(written)} 份报告。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
