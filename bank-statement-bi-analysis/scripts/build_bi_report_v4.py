#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""经营流水 BI 分析报告生成器 V4.0（BANKFLOW-BI-V4.0，对标见知现金流尽调系统）

V4.0 = V3.0 全部分析引擎 + 见知（Xencio）的呈现风格与账户级明细（设计见 references/design-v4.md）。
章节排序原则：事实类数据在前、分析判断类在后（主用户=客户经理/风险审批官，量化用户看末表）：
  一 核心速览（有效流入/流出口径 + TOP5 + 风险提示卡）
  二 指标明细（事实层：余额账户×月矩阵、借贷趋势与借贷方分型、收支账户×月与分类×月、
     对手方前10×月+重合方、关联方交易、需关注对手方、个人敏感支出夜间时段分布）
  三 数据验证（分析层：主体×账户完整性表、账户级质量评级、未提交账户账号级清单、主体名称核验）
  四 生意模式与审批要件（V3 平移：看板/风险信号/评分卡/交叉验证）
  报告头下附「数据处理概况（按主体）」：原始文件数/账户数/标准化笔数/打标命中率
  Sheet2 负债与现金流（Debt OS + 三情景推演 + 现金流量表，V3 平移）
  Sheet3 数据明细 / Sheet4 可视化看板（9 图）/ Sheet5 风险指标
  Sheet6 标准化流水明细（整合打标流水只读副本，带筛选，供手工分析）

样式规范（自见知导出 xlsx 逆向）：微软雅黑、章节编号(一/x.y/x.y.z)、A列窄编号列、
流入蓝头#99CCFF/流出红头#FF7E79、警示橙字#FFC000、无边框、末行总计。

红线沿用：不重排明细；只读不改；阈值是参考不是结论；缺数据如实标注。

用法：
  python build_bi_report_v4.py --input "<标准化产物文件或文件夹>" --client 客户名
    [--out-dir DIR] [--whitelist wl.json] [--new-loan 2000000,0.08,12] [--loans 借据表.csv]
"""
import argparse, json, math, os, re, sys
try:
    import pandas as pd
    import numpy as np
except ImportError:
    sys.stderr.write("缺少 pandas，请先 `pip install pandas openpyxl`。\n"); raise
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from build_bi_report_v3 import (pick_input, load, prep, spec_augment, daily_balance, nature,
                                norm_cp, analyze, THRESHOLDS, FIN_KW, JUD_KW, SPEC_KW,
                                PENDING, NEW_LOAN)
from generate_vars import compute_spec_metrics, load_loans, S_EMPTY, S_SHORT, S_NOEXT, S_NOCALC

SPEC_SENTINELS = {S_EMPTY, S_SHORT, S_NOEXT, S_NOCALC}
STRATEGY_VERSION = "BANKFLOW-BI-V4.0"
NONBANK_KW = "金融租赁|融资租赁|金租|保理|小额贷款|消费金融|信托|担保|典当"
MF, MF0, PF = "#,##0.00", "#,##0", "0.00%"
# 年销售额估算：补贴/退税剔除口径
SUBSIDY_KW = "补贴|退税|奖励|扶持"
SUBSIDY_CP = "财政|国库|农业农村局|税务局|人力资源和社会保障|社保"
THRESHOLDS.setdefault("过账重合比下限", 0.50)   # 对手双向重合比≥此值 → 视为过账/对倒
THRESHOLDS.setdefault("旺季指数下限", 1.20)     # 月流入/月均流入 ≥ → 旺季
THRESHOLDS.setdefault("淡季指数上限", 0.80)     # 月流入/月均流入 < → 淡季

def lender_type(name):
    if re.search("银行", name): return "银行"
    if re.search(NONBANK_KW, name): return "非银"
    return "民间" if nature(name) == "个人" else "其他"

ACCT_COLS = ("本方账户", "对手账户", "账号")

def _clean_acct(s):
    """账号列规范化：转字符串、去空白、去浮点尾巴（'…003.0'→'…003'）。"""
    s = s.fillna("").astype(str).str.strip()
    return s.str.replace(r"\.0+$", "", regex=True).replace("nan", "")

def load_v4(path):
    """账号列以字符串读入（pandas 默认转 float 会丢精度/出科学计数法，毁掉未提交账户匹配）。"""
    if path.lower().endswith(".csv"):
        return pd.read_csv(path, dtype={c: str for c in ACCT_COLS}), None, None
    df, dbal, vchk = load(path)
    for c in ACCT_COLS:
        if c in df.columns: df[c] = _clean_acct(df[c])
    return df, dbal, vchk

# ---------------------------------------------------------------- V4 增量分析
def analyze_v4(df, A, vchk):
    V = {}
    MONTHS = A["months"]
    members = sorted((set(df["主体名称"]) | set(df["本方名称"])) - {""})
    V["members"] = members
    acct_col = "本方账户" if "本方账户" in df.columns else None
    df["_acct"] = _clean_acct(df[acct_col]) if acct_col else ""
    df["_owner"] = df["本方名称"].where(df["本方名称"] != "", df["主体名称"])
    own_accts = set(df.loc[df["_acct"] != "", "_acct"])
    V["own_accts"] = own_accts

    cur_col = next((c for c in ("币种", "货币", "币别") if c in df.columns), None)
    V["currencies"] = sorted(set(df[cur_col].dropna().astype(str).str.strip()) - {""}) if cur_col else []
    bank_col = next((c for c in ("开户行", "本方银行", "所属银行") if c in df.columns), None)

    # 账户级断点（余额校验表里有账户列则用；否则整体数）
    brk_by_acct = {}
    if vchk is not None and "余额断点" in vchk.columns:
        acol = next((c for c in ("本方账户", "账户", "账号") if c in vchk.columns), None)
        if acol:
            brk_by_acct = {str(k).strip(): int(v) for k, v in
                           pd.to_numeric(vchk.set_index(acol)["余额断点"], errors="coerce").fillna(0).items()}

    # ---- 账户档案 + 账户×月余额矩阵 ----
    accounts, avg_rows, eom_rows, in_rows, out_rows = [], [], [], [], []
    daily_bal = {}
    for (owner, acct), g in df.groupby(["_owner", "_acct"], sort=True):
        if not acct and acct_col: continue
        key = (owner, acct or "(未提供账号)")
        tmin, tmax = g["交易时间"].min(), g["交易时间"].max()
        n_m = g["月份"].nunique()
        cp_fill = float((g["对手名称"] != "").mean()); memo_fill = float((g["摘要"].str.strip() != "").mean())
        jx = g[(g["对手名称"] == "") & (g["收入金额"] > 0) & (g["收入金额"] < 1000) &
               ((g["摘要"].str.contains("结息|利息")) |
                ((g["交易时间"].dt.day.isin([20, 21])) & (g["交易时间"].dt.month.isin([3, 6, 9, 12]))))]
        has_jx = len(jx[~jx["摘要"].str.contains("冲正|手续费")]) > 0
        brk = brk_by_acct.get(acct)
        fill = min(cp_fill, memo_fill)
        if len(g) < 30 or n_m < 3: grade = "数据不足"
        elif has_jx and cp_fill >= .8 and memo_fill >= .8 and not brk: grade = "优秀"
        elif fill >= .6 and not brk: grade = "良好"
        else: grade = "一般"
        detail = []
        if not has_jx: detail.append("无结息记录")
        if cp_fill < .8: detail.append(f"对手完整度{cp_fill:.0%}")
        if memo_fill < .8: detail.append(f"摘要完整度{memo_fill:.0%}")
        if brk: detail.append(f"余额断点{brk}处")
        accounts.append(dict(owner=owner, acct=acct or "(未提供账号)",
                             bank=(g[bank_col].dropna().astype(str).iloc[0] if bank_col and g[bank_col].notna().any() else ""),
                             cur=(g[cur_col].dropna().astype(str).iloc[0] if cur_col and g[cur_col].notna().any() else ""),
                             t0=tmin, t1=tmax, months=n_m, rows=len(g), breaks=brk,
                             cp="有" if cp_fill > 0 else "无", memo="有" if memo_fill > 0 else "无",
                             jx="有" if has_jx else "无", fill=fill, grade=grade,
                             detail="；".join(detail) or "—"))
        # 余额矩阵（日末余额前向填充）
        t = g.dropna(subset=["账户余额"])
        if len(t):
            last = t.groupby("日期")["账户余额"].last()
            idx = pd.date_range(min(t["日期"]), max(t["日期"]))
            s = last.reindex(idx.date).ffill(); s.index = idx
            mm = s.groupby(s.index.strftime("%Y-%m"))
            avg_rows.append((key, mm.mean())); eom_rows.append((key, mm.last()))
            daily_bal[key] = s.dropna()
        else:
            avg_rows.append((key, pd.Series(dtype=float))); eom_rows.append((key, pd.Series(dtype=float)))
        gm = g.groupby("月份").agg(i=("收入金额", "sum"), o=("支出金额", "sum"))
        in_rows.append((key, gm["i"])); out_rows.append((key, gm["o"]))
    V["accounts"] = accounts
    def matx(rows):
        return pd.DataFrame({k: s for k, s in rows}).T.reindex(columns=MONTHS)
    V["avg_mx"], V["eom_mx"] = matx(avg_rows), matx(eom_rows)
    V["in_mx"], V["out_mx"] = matx(in_rows).fillna(0.0), matx(out_rows).fillna(0.0)

    # ---- 借贷口径 ----
    fin_fix = (df["收入金额"] >= 500_000) & df["对手key"].str.contains(FIN_KW)
    loan_in = (df["一级标签"] == "筹资类") & (df["收入金额"] > 0)
    loan_in |= (df["二级标签"] == "票据及贸易融资") & (df["收入金额"] > 0)
    loan_in |= fin_fix
    loan_out = ((df["一级标签"] == "筹资类") & (df["支出金额"] > 0)) | (df["三级标签"] == "偿还票据及贸易融资")
    V["loan_in_mask"], V["loan_out_mask"] = loan_in, loan_out
    lm = pd.DataFrame({
        "借贷流入": df.loc[loan_in].groupby("月份")["收入金额"].sum(),
        "偿还借贷": df.loc[loan_out].groupby("月份")["支出金额"].sum()}).reindex(MONTHS).fillna(0.0)
    lm["净额"] = lm["借贷流入"] - lm["偿还借贷"]
    V["loan_mon"] = lm
    lend = df[(loan_in | loan_out) & (df["对手key"] != "") & (~df["内部互转"])]
    ld = lend.groupby("对手key").agg(名称=("对手名称", "first"),
                                     流入=("收入金额", lambda s: s[loan_in.reindex(s.index)].sum()),
                                     偿还=("支出金额", lambda s: s[loan_out.reindex(s.index)].sum()))
    ld = ld[(ld["流入"] > 0) | (ld["偿还"] > 0)]
    ld["类型"] = ld["名称"].map(lender_type)
    V["lenders"] = ld.sort_values(["流入", "偿还"], ascending=False)
    V["lender_counts"] = ld["类型"].value_counts().to_dict()

    # ---- 有效流入/流出（剔互转+筹资+票据+结息+退款）----
    eff = df[(~df["内部互转"]) & (df["一级标签"] != "筹资类") &
             (~df["二级标签"].isin(["票据及贸易融资", "存款结息", "退款交易"]))]
    V["eff_in"], V["eff_out"] = eff["收入金额"].sum(), eff["支出金额"].sum()
    ecp = eff[eff["对手名称"] != ""].groupby("对手名称").agg(流入=("收入金额", "sum"), 流出=("支出金额", "sum"))
    V["eff_top_in"] = ecp[ecp["流入"] > 0].sort_values("流入", ascending=False).head(5)
    V["eff_top_out"] = ecp[ecp["流出"] > 0].sort_values("流出", ascending=False).head(5)

    # ---- 全口径对手（含本方成员，供 2.5 前10大 ×月）----
    cpa = df[df["对手名称"] != ""].groupby("对手key").agg(
        名称=("对手名称", "first"), 流入=("收入金额", "sum"), 流出=("支出金额", "sum"),
        流入笔数=("收入金额", lambda s: int((s > 0).sum())), 流出笔数=("支出金额", lambda s: int((s > 0).sum())))
    V["cp_all"] = cpa
    def mix(key, col):
        g = df[(df["对手key"] == key) & (df[col] > 0)].groupby("二级标签")[col].sum().sort_values(ascending=False)
        tot = g.sum()
        return "，".join(f"{k or '未标注'}{v/tot:.0%}" for k, v in g.head(3).items()) if tot else ""
    V["cp_mix"] = mix
    both = cpa[(cpa["流入"] > 0) & (cpa["流出"] > 0)].copy()
    both["重合金额"] = both[["流入", "流出"]].min(axis=1)
    both["重合比"] = both["重合金额"] / both[["流入", "流出"]].max(axis=1)
    V["overlap"] = both.sort_values("重合金额", ascending=False).head(10)

    # ---- 未提交账户 ----
    oa_col = "对手账户" if "对手账户" in df.columns else None
    unsub = []
    it = df[df["内部互转"]]
    oa_cover = 0.0
    if oa_col and len(it):
        t = it.copy(); t["_oa"] = _clean_acct(t[oa_col])
        oa_cover = float((t["_oa"] != "").mean())
        t = t[(t["_oa"] != "") & (~t["_oa"].isin(own_accts))]
        for (nm, oa), g in t.groupby(["对手名称", "_oa"]):
            unsub.append([nm, oa, g["支出金额"].sum(), g["收入金额"].sum()])
    if oa_col and oa_cover > 0:
        unsub_mode = f"账号级，互转笔对手账号覆盖率{oa_cover:.0%}"
    else:
        unsub_mode = "无对手账户数据，无法账号级检测（规则反推见3.3.1）"
    V["unsub"], V["unsub_mode"] = sorted(unsub, key=lambda x: -(x[2] + x[3]))[:20], unsub_mode

    # ---- 关联方交易（成员间互转清单）----
    rel = it.groupby(["_owner", "对手名称"]).agg(流入=("收入金额", "sum"), 流出=("支出金额", "sum"), 笔数=("收入金额", "size"))
    V["rel_pairs"] = rel.sort_values(["流入", "流出"], ascending=False)

    # ---- 需关注对手方 ----
    ext = df[(~df["内部互转"]) & (df["对手key"] != "")]
    watch = []
    for k, g in ext.groupby("对手key"):
        name = g["对手名称"].iloc[0]; tags = []
        i, o = g["收入金额"].sum(), g["支出金额"].sum()
        amts = g["收入金额"].add(g["支出金额"])
        if nature(name) == "个人" and i + o >= 500_000: tags.append("个人大额往来")
        if ((amts >= 1_000_000) & (amts % 10_000 == 0)).any(): tags.append("大额整数交易")
        if re.search(JUD_KW, k): tags.append("司法对手")
        if re.search(SPEC_KW, k): tags.append("金融投机/赌博")
        if (g["二级标签"] == "民间借贷").any(): tags.append("民间借贷")
        if tags: watch.append([name, "、".join(tags), i, o, i - o])
    V["watch"] = sorted(watch, key=lambda x: -(x[2] + x[3]))[:20]

    # ---- 个人敏感支出：夜间大额时段分布（22:00–06:00 对私流出）----
    has_t = df["交易时间"].dt.strftime("%H:%M:%S") != "00:00:00"
    night = df[has_t & (df["支出金额"] > 0) & (df["对手名称"].map(nature) == "个人")]
    hh = night["交易时间"].dt.hour
    V["night_buckets"] = [(f"{h}:00~{(h+1)%24}:00",
                           float(night.loc[hh == h, "支出金额"].sum()),
                           int((hh == h).sum())) for h in [22, 23, 0, 1, 2, 3, 4, 5]]
    # ---- 年销售额估算（三档口径：逐层剔除，下限 ⊆ 中值 ⊆ 上限）----
    # 对过账型贸易，「有效流入」仍含个人拆借与对倒过账，直接当销售额会高估数倍，故再剥三层。
    eff_ns = eff[~((eff["二级标签"].str.contains(SUBSIDY_KW)) |
                   (eff["三级标签"].str.contains(SUBSIDY_KW)) |
                   (eff["对手key"].str.contains(SUBSIDY_CP)))]
    e0 = float(eff.loc[eff["收入金额"] > 0, "收入金额"].sum())
    e1 = e0 - float(eff_ns.loc[eff_ns["收入金额"] > 0, "收入金额"].sum())      # 补贴/退税/财政
    ec = eff_ns[eff_ns["对手key"] != ""].groupby("对手key").agg(
        名称=("对手名称", "first"), 流入=("收入金额", "sum"), 流出=("支出金额", "sum"))
    ecb = ec[(ec["流入"] > 0) & (ec["流出"] > 0)].copy()
    ecb["重合"] = ecb[["流入", "流出"]].min(axis=1)
    ecb["重合比"] = ecb["重合"] / ecb[["流入", "流出"]].max(axis=1)
    pt = ecb[ecb["重合比"] >= THRESHOLDS["过账重合比下限"]].sort_values("重合", ascending=False)
    e2 = float(pt["重合"].sum())                                             # 对倒/过账重合
    lend_kw = "往来|借贷|拆借|借款|还款"
    rest = ec[(~ec.index.isin(pt.index)) & (ec["流入"] > 0)]
    per_rest = rest[rest["名称"].map(nature) == "个人"]
    labeled = ((eff_ns["二级标签"].str.contains(lend_kw)) | (eff_ns["三级标签"].str.contains(lend_kw)))
    lend_in = eff_ns[(eff_ns["收入金额"] > 0) & (eff_ns["对手key"].isin(per_rest.index)) &
                     (labeled | (eff_ns["一级标签"] != "经营类"))]
    e3 = float(lend_in["收入金额"].sum())                                    # 个人往来/拆借
    # 拆分：明确带往来/借贷标签 vs 仅因未打标(非经营类)而被剔除——后者是标签覆盖不足的产物，非经济结论
    e3_lab = float(lend_in.loc[labeled.reindex(lend_in.index).fillna(False), "收入金额"].sum())
    e3_untag = e3 - e3_lab
    hi = max(e0 - e1, 0.0); mid = max(hi - e2, 0.0); lo = max(mid - e3, 0.0)
    ann = 12.0 / max(len(MONTHS), 1)
    V["sales"] = dict(eff_in=e0, e1=e1, e2=e2, e3=e3, e3_lab=e3_lab, e3_untag=e3_untag,
                      lo=lo, mid=mid, hi=hi,
                      lo_a=lo * ann, mid_a=mid * ann, hi_a=hi * ann, months=len(MONTHS),
                      pt=pt.head(10), lend_top=per_rest.sort_values("流入", ascending=False).head(10))

    # ---- 淡旺季（季节性）----
    mon = A["mon"]
    base_in = float(mon["流入"].mean()) or 1.0
    base_n = float(mon["笔数"].mean()) or 1.0
    seas = pd.DataFrame({"流入": mon["流入"], "笔数": mon["笔数"],
                         "活跃对手数": mon.get("活跃对手数", pd.Series(0, index=mon.index)),
                         "净流入": mon["净流入"]})
    seas["流入指数"] = seas["流入"] / base_in
    seas["笔数指数"] = seas["笔数"] / base_n
    HI_, LO_ = THRESHOLDS["旺季指数下限"], THRESHOLDS["淡季指数上限"]
    seas["分档"] = np.where(seas["流入指数"] >= HI_, "旺季",
                            np.where(seas["流入指数"] < LO_, "淡季", "平季"))
    # 边月不满整月、或当月在测账户少于全期最大值（后期才开户/未提交早期流水）→ 不参与旺淡定性。
    # 否则「早期只提交了1个账户」会被误判成淡季，把数据覆盖问题读成经营season。
    t0, t1 = df["交易时间"].min(), df["交易时间"].max()
    partial = set()
    if t0.day > 5: partial.add(MONTHS[0])
    if t1.day < 25: partial.add(MONTHS[-1])
    # 覆盖度按「账户数据区间是否跨越该月」计，而非「该月是否有交易」——否则账户当月休眠会被误判为缺数据
    spans = df[df["_acct"] != ""].groupby("_acct")["月份"].agg(["min", "max"])
    n_acct_m = pd.Series({m: int(((spans["min"] <= m) & (spans["max"] >= m)).sum()) for m in MONTHS})
    seas["在测账户数"] = n_acct_m
    max_acct = int(n_acct_m.max()) if len(n_acct_m) else 0
    cover_bad = {m for m in MONTHS if int(n_acct_m.get(m, 0)) < max_acct}
    seas["完整性"] = ["；".join(filter(None, [
        "不完整(边月)" if m in partial else "",
        f"账户覆盖不全({int(n_acct_m.get(m,0))}/{max_acct})" if m in cover_bad else ""])) for m in seas.index]
    partial = partial | cover_bad
    V["max_acct"] = max_acct
    # 自然月跨年合并（1–12 月），旺季规律看这张
    cm = seas[~seas.index.isin(partial)].copy()
    cm["自然月"] = [int(m[-2:]) for m in cm.index]
    cmg = cm.groupby("自然月").agg(月份数=("流入", "size"), 月均流入=("流入", "mean"),
                                    月均笔数=("笔数", "mean"), 月均净流入=("净流入", "mean"))
    if len(cmg):
        cmg["流入指数"] = cmg["月均流入"] / (cmg["月均流入"].mean() or 1.0)
        cmg["分档"] = np.where(cmg["流入指数"] >= HI_, "旺季",
                               np.where(cmg["流入指数"] < LO_, "淡季", "平季"))
    core = seas[seas["完整性"] == ""]
    peak_m = list(core.index[core["分档"] == "旺季"])
    low_m = list(core.index[core["分档"] == "淡季"])
    run = best = 0                                                            # 最长连续淡季
    for d in core["分档"]:
        run = run + 1 if d == "淡季" else 0
        best = max(best, run)
    V["seas"] = dict(mx=seas, cm=cmg, peak=peak_m, low=low_m, max_run=best,
                     low_cf=float(core.loc[core["分档"] == "淡季", "净流入"].mean()) if low_m else float("nan"),
                     peak_cf=float(core.loc[core["分档"] == "旺季", "净流入"].mean()) if peak_m else float("nan"),
                     partial=sorted(partial))

    # ---- 余额峰值 TOP10 + 成因归因 ----
    end_dt = t1
    peaks = []
    for key, s in daily_bal.items():
        picked = []
        cand = s.sort_values(ascending=False)
        for d, v in cand.items():                                            # 同一波峰只取一次：间隔≥7天且余额差>1%
            if all(abs((d - p).days) >= 7 and abs(float(v) - pv) > 0.01 * max(abs(pv), 1) for p, pv in picked):
                picked.append((d, float(v))); peaks.append((key, d, float(v)))
            if len(picked) >= 3: break
    peaks.sort(key=lambda x: -x[2])
    rows_pk = []
    for (owner, acct), d, v in peaks[:10]:
        day = df[(df["_owner"] == owner) & (df["_acct"].replace("", "(未提供账号)") == acct) &
                 (pd.to_datetime(df["日期"]) == d) & (df["收入金额"] > 0)]
        if len(day):
            top = day.loc[day["收入金额"].idxmax()]
            cause, cp_, amt = top["对手名称"] or "(未标注对手)", top["摘要"][:30], float(top["收入金额"])
        else:
            cause, cp_, amt = "—（当日无流入，系前期余额延续）", "", 0.0
        rows_pk.append([owner, acct, d.strftime("%Y-%m-%d"), v, cause, cp_, amt,
                        amt / v if v else 0.0,
                        "是" if (end_dt - d).days <= 90 else ""])
    V["bal_peaks"] = rows_pk
    V["n_peak_pre"] = sum(1 for x in rows_pk if x[8] == "是")

    # ---- 单笔极值卡（账户 × 收支 × 经营性/非经营性）----
    is_op = (df["一级标签"] == "经营类") & (~df["二级标签"].str.contains("往来")) & (~df["内部互转"])
    ext_rows = []
    for (owner, acct), g in df.groupby(["_owner", "_acct"], sort=True):
        if not acct and acct_col: continue
        a = acct or "(未提供账号)"
        for lbl, col, m in [("单笔最高收入", "收入金额", None), ("单笔最高支出", "支出金额", None),
                            ("单笔最大经营性收入", "收入金额", True), ("单笔最大经营性支出", "支出金额", True)]:
            gg = g[is_op.reindex(g.index).fillna(False)] if m else g
            gg = gg[gg[col] > 0]
            if not len(gg):
                ext_rows.append([owner, a, lbl, None, "", "", ""]); continue
            x = gg.loc[gg[col].idxmax()]
            ext_rows.append([owner, a, lbl, float(x[col]), str(x["交易时间"])[:10],
                             x["对手名称"] or "(未标注)", x["摘要"][:40]])
    V["extremes"] = ext_rows

    # 汇总计数（风险提示卡）
    V["n_risk_acct"] = sum(1 for a in accounts if a["grade"] in ("一般", "数据不足"))
    V["n_brk_acct"] = sum(1 for a in accounts if a["breaks"])
    V["n_night"] = int(sum(b[2] for b in V["night_buckets"]))
    V["daily_avg_bal"] = float(A["bal"].mean()) if A["bal"] is not None else float("nan")
    return V

# ---------------------------------------------------------------- 写报告（见知样式系统）
def build_workbook(A, V, df, client, out_path, whitelist, spec_list=None):
    wb = Workbook(); wb.remove(wb.active)
    def F(**kw):
        kw.setdefault("size", 11); kw.setdefault("name", "微软雅黑"); return Font(**kw)
    TITLE_F = F(size=22, bold=True, color="1F4E78")
    CHAP_F = F(size=16, bold=True, color="1F4E78")
    SECT_F = F(size=13, bold=True, color="2F75B5")
    SUB_F = F(size=11, bold=True, color="2F75B5")
    ORNG = F(color="FFC000", bold=True)
    NOTE_F = F(size=9, italic=True, color="808080")
    BLUEF = F(color="0000FF")
    INH = PatternFill("solid", start_color="99CCFF")   # 流入/中性表头（见知蓝）
    OUTH = PatternFill("solid", start_color="FF7E79")  # 流出表头（见知红）
    GRN = PatternFill("solid", start_color="E2EFDA"); WARN = PatternFill("solid", start_color="FFF2CC")
    REDF = PatternFill("solid", start_color="FCE4E4"); YEL = PatternFill("solid", start_color="FFFF00")
    def put(ws, r, c, v, bold=False, fillc=None, fmt=None, font=None, wrap=False, right=None):
        cell = ws.cell(r, c, v); cell.font = font or F(bold=bold)
        if fillc: cell.fill = fillc
        if fmt: cell.number_format = fmt
        ha = "right" if (right if right is not None else isinstance(v, (int, float)) and not isinstance(v, bool)) else None
        cell.alignment = Alignment(vertical="center", wrap_text=wrap, horizontal=ha)
        return cell
    def chap(ws, r, num, text):
        put(ws, r, 1, num, font=CHAP_F); put(ws, r, 2, text, font=CHAP_F); ws.row_dimensions[r].height = 26
    def sect(ws, r, num, text):
        put(ws, r, 1, num, font=SECT_F); put(ws, r, 2, text, font=SECT_F); ws.row_dimensions[r].height = 22
    def sub(ws, r, num, text):
        put(ws, r, 1, num, font=SUB_F); put(ws, r, 2, text, font=SUB_F)
    def colw(ws, c):
        d = ws.column_dimensions.get(get_column_letter(c))
        return d.width if d is not None and d.width else 8.43
    def est_lines(text, width, size=11):
        """按可用宽度估算换行行数（列宽单位≈半角字符宽，CJK 计 2）。"""
        usable = max((width - 1.5) * 11 / size, 4)
        n = 0
        for seg in str(text).split("\n"):
            units = sum(2 if ord(ch) > 0x2E7F else 1 for ch in seg)
            n += max(1, math.ceil(units / usable))
        return n
    def fit_row(ws, r, lines):
        if lines > 1: ws.row_dimensions[r].height = min(lines * 15.5 + 3, 210)
    def mtext(ws, r, text, span=6, col=2, font=None, size=11):
        """长文本行：col 起横向合并 span 列，换行显示并按内容撑高行。"""
        put(ws, r, col, text, font=font, wrap=True)
        if span > 1:
            ws.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col + span - 1)
        fit_row(ws, r, est_lines(text, sum(colw(ws, c) for c in range(col, col + span)), size))
    def foot(ws, r, text, span=6):
        mtext(ws, r, text, span=span, font=NOTE_F, size=9)
    def cards(ws, r, items, fills=None):
        """一行标签(填充)+一行数值。items=[(label,value,fmt|None,orange?)...] 从B列排。"""
        lines = 1
        for j, it in enumerate(items):
            label, value, fmt = it[0], it[1], (it[2] if len(it) > 2 else None)
            orange = it[3] if len(it) > 3 else False
            fill = (fills[j] if fills else INH)
            put(ws, r, 2 + j, label, bold=True, fillc=fill, wrap=True)
            put(ws, r + 1, 2 + j, value, fmt=fmt, font=ORNG if orange else None)
            lines = max(lines, est_lines(label, colw(ws, 2 + j)))
        fit_row(ws, r, lines)
        return r + 3
    def table(ws, r0, headers, rows, fmts=None, fill=INH, seq=True, total=None, total_label="总计", spans=None):
        """A列序号 + B起表头(填充,无边框) + 数据行 + 可选总计行。返回下一空行。
        spans={逻辑列号: 合并列数}：长文本列横向合并若干物理列，行高按换行内容估算。"""
        spans = spans or {}
        starts, c = [], 2
        for i in range(len(headers)):
            starts.append(c); c += spans.get(i, 1)
        def wcell(rr, i, v, **kw):
            c0, n = starts[i], spans.get(i, 1)
            cell = put(ws, rr, c0, v, **kw)
            if n > 1:
                if kw.get("fillc"):
                    for cc in range(c0 + 1, c0 + n): ws.cell(rr, cc).fill = kw["fillc"]
                ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0 + n - 1)
            return cell
        def width_of(i):
            return sum(colw(ws, cc) for cc in range(starts[i], starts[i] + spans.get(i, 1)))
        lines = 1
        for i, hx in enumerate(headers):
            wcell(r0, i, hx, bold=True, fillc=fill, wrap=True)
            lines = max(lines, est_lines(hx, width_of(i)))
        fit_row(ws, r0, lines)
        for k, row in enumerate(rows):
            rr = r0 + 1 + k
            if seq: put(ws, rr, 1, k + 1)
            lines = 1
            for i, v in enumerate(row):
                fmt = fmts[i] if fmts and i < len(fmts) else None
                isnum = isinstance(v, (int, float, np.floating)) and not isinstance(v, bool)
                if isnum and (v != v): v = None
                wcell(rr, i, float(v) if isnum and v is not None else v,
                      fmt=fmt if isnum else None, wrap=not isnum)
                if isinstance(v, str) and v:
                    lines = max(lines, est_lines(v, width_of(i)))
            fit_row(ws, rr, lines)
        r = r0 + 1 + len(rows)
        if total is not None:
            put(ws, r, starts[0], total_label, bold=True)
            for i, v in enumerate(total, 1):
                if i >= len(starts): break
                isnum = isinstance(v, (int, float, np.floating)) and not isinstance(v, bool)
                if isnum and v == v:
                    wcell(r, i, float(v), bold=True, fmt=fmts[i] if fmts and i < len(fmts) else MF)
                elif v is not None and v == v:
                    wcell(r, i, v, bold=True)
            r += 1
        return r
    def hdr_row(ws, r, labels, start=2, fill=INH):
        """手写表头行：换行+按最长表头估算行高（窄列长表头不再截断）。"""
        lines = 1
        for j, hx in enumerate(labels, start):
            put(ws, r, j, hx, bold=True, fillc=fill, wrap=True)
            lines = max(lines, est_lines(hx, colw(ws, j)))
        fit_row(ws, r, lines)
    def nxt(ws, gap=2): return ws.max_row + gap if ws.max_row > 1 else 1

    MONTHS, L12, mon, mon_op, cp = A["months"], A["l12"], A["mon"], A["mon_op"], A["cp"]
    ext_in, ext_out = A["ext_in"], A["ext_out"]
    ext = df[~df["内部互转"]]
    fwd = pd.period_range(pd.Period(MONTHS[-1]) + 1, periods=12, freq="M").astype(str).tolist()
    nm = len(MONTHS)

    S1 = wb.create_sheet("分析报告")
    S1.column_dimensions["A"].width = 6
    for i, w in enumerate([30, 22, 20, 22, 18] + [15] * (max(nm, 12) + 1), 2):
        S1.column_dimensions[get_column_letter(i)].width = w
    S2 = wb.create_sheet("负债与现金流")
    S2.column_dimensions["A"].width = 6
    for i, w in enumerate([26, 16, 16, 14, 13, 13, 16, 13, 13, 13, 30, 13, 13], 2):
        S2.column_dimensions[get_column_letter(i)].width = w
    S3 = wb.create_sheet("数据明细")
    S3.column_dimensions["A"].width = 6
    for i, w in enumerate([32] + [13] * (nm + 3), 2): S3.column_dimensions[get_column_letter(i)].width = w
    S4 = wb.create_sheet("可视化看板"); S4.column_dimensions["A"].width = 2
    S5 = wb.create_sheet("风险指标")
    S5.column_dimensions["A"].width = 6
    for i, w in enumerate([16, 32, 18, 74], 2): S5.column_dimensions[get_column_letter(i)].width = w
    S6 = wb.create_sheet("标准化流水明细")

    # ============ Sheet1 分析报告 ============
    ws = S1
    subjects = "、".join(V["members"])
    put(ws, 1, 2, f"{client} · 经营流水尽调分析报告", font=TITLE_F); ws.row_dimensions[1].height = 36
    hdr = [("金额单位", "元"), ("货币单位", "、".join(V["currencies"]) or "人民币(未提供币种列)"),
           ("客户名称", client), ("关联主体(合并口径)", subjects),
           ("流水区间", f"{df['交易时间'].min()} ~ {df['交易时间'].max()}（{nm}个月）"),
           ("分析策略", STRATEGY_VERSION), ("报告日期", pd.Timestamp.today().strftime("%Y-%m-%d"))]
    r = 2
    for k, v in hdr:
        put(ws, r, 2, k, bold=True); mtext(ws, r, v, span=4, col=3); r += 1
    if len(V["currencies"]) > 1:
        mtext(ws, r, f"⚠ 检测到多币种：{'、'.join(V['currencies'])}；本报告金额未做汇率折算，跨币种合计仅作量级参考。", span=6, font=ORNG); r += 1

    # ---- 数据处理概况（按主体：原始文件数 + 标准化/打标成功情况）----
    r += 1
    put(ws, r, 2, "数据处理概况（按主体）", bold=True, fillc=INH); r += 1
    src_ok = "来源文件名" in df.columns
    if "命中关键词" in df.columns:
        hit_mask = df["命中关键词"].fillna("").astype(str).str.strip() != ""
        hit_cap = "打标命中=规则库命中（命中关键词非空），其余为兜底其他类"
    else:
        hit_mask = df["一级标签"].isin(["经营类", "筹资类", "投资类"])
        hit_cap = "无命中信息列，打标命中按「一级标签∈经营/筹资/投资类」近似"
    ov_rows = []
    for owner, g in df.groupby("_owner", sort=True):
        hits = int(hit_mask.reindex(g.index).sum())
        ov_rows.append([owner or "(未识别主体)",
                        int(g["来源文件名"].nunique()) if src_ok else "—",
                        int(g["_acct"].replace("", np.nan).nunique()) if "_acct" in g.columns else g["本方账户"].nunique(),
                        len(g), f"{g['交易时间'].min():%Y.%m.%d}~{g['交易时间'].max():%Y.%m.%d}",
                        hits, hits / max(len(g), 1)])
    tot_hits = int(hit_mask.sum())
    r = table(ws, r, ["主体名称", "原始文件数", "账户数", "标准化笔数", "流水期间", "打标命中笔数", "打标命中率"],
              ov_rows, fmts=[None, "#,##0", "#,##0", "#,##0", None, "#,##0", PF],
              total=[int(df["来源文件名"].nunique()) if src_ok else None,
                     int(df["_acct"].replace("", np.nan).nunique()) if "_acct" in df.columns else None,
                     len(df), None, tot_hits, tot_hits / max(len(df), 1)])
    foot(ws, r, f"口径：原始文件数按「来源文件名」去重（同一文件多账户会重复计入各主体）；{hit_cap}；"
                "余额校验断点与未提交账户见「三 数据验证」；全部明细见末表「标准化流水明细」。"); r += 1

    # ---- 一 核心速览 ----
    r = nxt(ws); chap(ws, r, "一", "核心速览"); r += 1
    sect(ws, r, "1.1", "核心统计"); r += 1
    r = cards(ws, r, [("分析主体", len(V["members"]), "#,##0"), ("分析账户", len(V["accounts"]), "#,##0"),
                      ("数据时长(月)", nm, "#,##0")])
    r = cards(ws, r, [("日均余额", V["daily_avg_bal"], MF),
                      ("借贷流入", float(V["loan_mon"]["借贷流入"].sum()), MF),
                      ("偿还借贷", -float(V["loan_mon"]["偿还借贷"].sum()), MF)],
              fills=[INH, INH, OUTH])
    lc = V["lender_counts"]
    r = cards(ws, r, [("银行借贷方", lc.get("银行", 0), "#,##0"), ("非银借贷方", lc.get("非银", 0), "#,##0"),
                      ("民间借贷方", lc.get("民间", 0), "#,##0", lc.get("民间", 0) >= 3),
                      ("其他借贷方", lc.get("其他", 0), "#,##0")])
    r = cards(ws, r, [("有效流入", V["eff_in"], MF), ("有效流出", -V["eff_out"], MF)], fills=[INH, OUTH])
    foot(ws, r - 1, "有效口径=剔除 内部互转/筹资类/票据及贸易融资/存款结息/退款交易 后的对外流水。"); r += 1
    put(ws, r, 2, "核心有效流入方", bold=True, fillc=INH); put(ws, r, 3, "流入金额", bold=True, fillc=INH)
    put(ws, r, 4, "占有效流入", bold=True, fillc=INH)
    put(ws, r, 5, "核心有效流出方", bold=True, fillc=OUTH); put(ws, r, 6, "流出金额", bold=True, fillc=OUTH)
    put(ws, r, 7, "占有效流出", bold=True, fillc=OUTH)
    ti, to = V["eff_top_in"], V["eff_top_out"]
    for i in range(5):
        rr = r + 1 + i; put(ws, rr, 1, i + 1)
        if i < len(ti):
            put(ws, rr, 2, ti.index[i]); put(ws, rr, 3, float(ti["流入"].iloc[i]), fmt=MF)
            put(ws, rr, 4, float(ti["流入"].iloc[i] / max(V["eff_in"], 1)), fmt=PF)
        if i < len(to):
            put(ws, rr, 5, to.index[i]); put(ws, rr, 6, -float(to["流出"].iloc[i]), fmt=MF)
            put(ws, rr, 7, float(to["流出"].iloc[i] / max(V["eff_out"], 1)), fmt=PF)
    r += 7
    sect(ws, r, "1.2", "风险提示"); r += 1
    r = cards(ws, r, [("综合数据质量", f"{A['q_total']}分·{A['q_grade']}", None, A["q_total"] < 70),
                      ("余额断点账户", V["n_brk_acct"] or (A["bal_breaks"] or 0), "#,##0", bool(V["n_brk_acct"])),
                      ("命中质量风险账户", V["n_risk_acct"], "#,##0", V["n_risk_acct"] > 0),
                      ("未提交账户(疑似)", len(V["unsub"]), "#,##0", len(V["unsub"]) > 0)])
    n_susp = int(((A["both"]["平衡度"] >= .7) & (A["both"][["流入", "流出"]].min(axis=1) >= 500_000)).sum())
    r = cards(ws, r, [("关联方(集团成员)", len(V["members"]), "#,##0"),
                      ("疑似关联外部方", n_susp, "#,##0", n_susp > 0),
                      ("需关注对手方", len(V["watch"]), "#,##0", len(V["watch"]) >= 10),
                      ("个人敏感支出(夜间笔数)", V["n_night"], "#,##0", V["n_night"] >= 20)])

    # ---- 1.3 年销售额估算 ----
    S = V["sales"]
    r = nxt(ws, 1); sect(ws, r, "1.3", "年销售额估算（三档口径）"); r += 1
    mtext(ws, r, "「有效流入」是对外流水口径，对过账型贸易仍含个人拆借与对倒过账，直接当销售额会显著高估。"
                 "下表在有效流入基础上逐层剔除，得到下限/中值/上限三档；三档为嵌套关系（下限⊆中值⊆上限），"
                 "剔除项可逐笔追溯。机器不下最终结论，最终销售额须与销项发票、税务应税收入、出入库台账交叉核定。"); r += 1
    r = cards(ws, r, [(f"估算销售额·下限({S['months']}个月)", S["lo"], MF),
                      ("估算销售额·中值", S["mid"], MF),
                      ("估算销售额·上限", S["hi"], MF)])
    r = cards(ws, r, [("年化·下限", S["lo_a"], MF), ("年化·中值", S["mid_a"], MF),
                      ("年化·上限", S["hi_a"], MF)])
    steps = [["有效流入（起点）", S["eff_in"], "剔除 内部互转/筹资/票据/结息/退款 后的对外流入"],
             ["− 补贴/退税/财政性收入", -S["e1"], f"对手含「{SUBSIDY_CP}」或标签含「{SUBSIDY_KW}」"],
             ["= 上限（最宽口径）", S["hi"], "把全部未定性流入都算作销售，销售额的绝对上界"],
             ["− 对倒/过账重合", -S["e2"], f"同一对手双向往来且重合比≥{THRESHOLDS['过账重合比下限']:.0%}，按 min(流入,流出) 剔除"],
             ["= 中值（推荐口径）", S["mid"], "剔除明显过账后的销售额，审批建议以此为基准"],
             ["− 个人往来/拆借流入", -S["e3"],
              f"对手为自然人且标签为往来/借贷或非经营类；其中明确带往来/借贷标签 {S['e3_lab']:,.0f} 元、"
              f"仅因未打标(非经营类)而剔除 {S['e3_untag']:,.0f} 元"],
             ["= 下限（最严口径）", S["lo"], "仅保留可定性为经营性销售的流入，销售额的保守下界"]]
    r = table(ws, r, ["口径步骤", "金额", "说明"], steps, fmts=[None, MF, None],
              seq=False, spans={3: 4})
    if S["e3_untag"] > 0 and S["e3"] > 0:
        mtext(ws, r, f"⚠ 下限敏感性提示：剔除项③中 {S['e3_untag']:,.0f} 元（占该项 {S['e3_untag']/S['e3']:.0%}）"
                     f"是「仅因未打标」而被剔除，并非已证实的拆借。全量打标命中率仅 {tot_hits/max(len(df),1):.1%}，"
                     "标签覆盖不足会系统性压低下限。这部分对手若经发票/合同证实为购销，应回补至销售额——"
                     "因此下限应读作「可确证的销售底线」，不等于「真实销售额」。", font=F(color="C00000")); r += 1
    if len(S["pt"]):
        r = nxt(ws, 1); sub(ws, r, "1.3.1", "对倒/过账对手明细（剔除项②，TOP10）"); r += 1
        rows = [[x["名称"], float(x["流入"]), -float(x["流出"]), float(x["重合"]), float(x["重合比"])]
                for _, x in S["pt"].iterrows()]
        r = table(ws, r, ["对手名称", "流入", "流出", "重合(剔除额)", "重合比"], rows,
                  fmts=[None, MF, MF, MF, PF])
    if len(S["lend_top"]):
        r = nxt(ws, 1); sub(ws, r, "1.3.2", "个人往来/拆借流入对手（剔除项③，TOP10）"); r += 1
        rows = [[x["名称"], float(x["流入"]), -float(x["流出"])] for _, x in S["lend_top"].iterrows()]
        r = table(ws, r, ["对手名称", "流入", "流出"], rows, fmts=[None, MF, MF])
        foot(ws, r, "注：本表按对手列示其全部流入；剔除额仅计其中标签为往来/借贷或非经营类的部分。"); r += 1

    # ---- 二 指标明细 ----
    mcols = MONTHS
    r = nxt(ws); chap(ws, r, "二", "指标明细"); r += 1
    sect(ws, r, "2.1", "余额分析"); r += 1
    for num, name, mx in [("2.1.1", "日均余额（账户×月）", V["avg_mx"]), ("2.1.2", "月末余额（账户×月）", V["eom_mx"])]:
        sub(ws, r, num, name); r += 1
        rows = []
        for (owner, acct), s in mx.iterrows():
            rows.append([owner, acct] + [s.get(m) for m in mcols] + [float(s.dropna().mean()) if s.notna().any() else None])
        tot = [None] + [float(mx[m].sum()) if mx[m].notna().any() else None for m in mcols] + [None]
        r = table(ws, r, ["本方名称", "本方账号"] + mcols + ["均值"], rows,
                  fmts=[None, None] + [MF] * (len(mcols) + 1), total=tot) + 1
    sub(ws, r, "2.1.3", "余额峰值 TOP10 与成因归因（识别申贷前冲量）"); r += 1
    pk = V["bal_peaks"]
    rows = [[o, a, d, v, cause, cp_, amt, sh, pre] for o, a, d, v, cause, cp_, amt, sh, pre in pk]
    r = table(ws, r, ["本方名称", "本方账号", "峰值日", "峰值余额", "主因交易对手", "摘要",
                      "当日最大单笔流入", "占峰值", "申贷前90天内"],
              rows or [["无日余额数据", "", "", None, "", "", None, None, ""]],
              fmts=[None, None, None, MF, None, None, MF, PF, None])
    foot(ws, r, "口径：按账户日末余额取峰值，同一波峰只取一次（间隔≥7天），每账户最多3个，全局取前10。"
                "「主因交易」= 峰值当日该账户最大单笔流入；占峰值高且落在申贷前90天内 → 疑似冲量，"
                f"本次命中 {V['n_peak_pre']} 个（申贷日以流水末日近似）。"); r += 2
    sect(ws, r, "2.2", "借贷分析"); r += 1
    sub(ws, r, "2.2.1", "借贷趋势"); r += 1
    lm = V["loan_mon"]
    for j, hx in enumerate([""] + ["合计"] + mcols, 2): put(ws, r, j, hx, bold=True, fillc=INH)
    for i, (lbl, sign) in enumerate([("借贷流入", 1), ("偿还借贷", -1), ("净额", 1)]):
        rr = r + 1 + i
        put(ws, rr, 2, lbl, bold=True)
        col = lm[lbl] if lbl != "净额" else lm["净额"]
        put(ws, rr, 3, float(col.sum()) * (sign if lbl == "偿还借贷" else 1), fmt=MF)
        for j, m in enumerate(mcols, 4):
            put(ws, rr, j, float(col.get(m, 0)) * (sign if lbl == "偿还借贷" else 1), fmt=MF)
    r += 5
    sub(ws, r, "2.2.2", "借贷方明细"); r += 1
    rows = [[x["名称"], x["类型"], float(x["流入"]), -float(x["偿还"])] for _, x in V["lenders"].iterrows()]
    r = table(ws, r, ["借贷方", "对方类型", "借贷流入", "偿还借贷"], rows or [["未识别到借贷方", "—", 0, 0]],
              fmts=[None, None, MF, MF])
    foot(ws, r, "类型口径：对手含「银行」→银行；金租/租赁/保理/小贷/消金/信托/担保/典当→非银；个人→民间。金融机构≥50万放款已并入借贷口径（防错标）。"); r += 2
    sect(ws, r, "2.3", "收支分析-趋势（账户×月）"); r += 1
    for num, name, mx, fill, sign in [("2.3.1", "流入", V["in_mx"], INH, 1), ("2.3.2", "流出", V["out_mx"], OUTH, -1)]:
        sub(ws, r, num, name); r += 1
        rows = []
        for (owner, acct), s in mx.iterrows():
            rows.append([owner, acct, float(s.sum()) * sign] + [float(s.get(m, 0)) * sign for m in mcols] + [float(s.mean()) * sign])
        tot = [None, float(mx.values.sum()) * sign] + [float(mx[m].sum()) * sign for m in mcols] + [float(mx.sum(axis=1).sum() / max(len(mcols), 1)) * sign]
        r = table(ws, r, ["本方名称", "本方账号", "合计"] + mcols + ["月均"], rows,
                  fmts=[None, None] + [MF] * (len(mcols) + 2), fill=fill, total=tot) + 1
    sect(ws, r, "2.4", "收支分析-构成（二级标签×月，含内部互转）"); r += 1
    for num, name, col, fill, sign, base in [("2.4.1", "流入", "收入金额", INH, 1, A["tot_in"]),
                                             ("2.4.2", "流出", "支出金额", OUTH, -1, A["tot_out"])]:
        sub(ws, r, num, name); r += 1
        g = df[df[col] > 0].copy()
        g["分类"] = g["二级标签"].where(g["二级标签"] != "", "未标注")
        g.loc[g["内部互转"], "分类"] = "内部转账(关联互转)"
        piv = g.pivot_table(index="分类", columns="月份", values=col, aggfunc="sum", fill_value=0.0)
        piv = piv.reindex(columns=mcols, fill_value=0.0)
        piv["合计"] = piv.sum(axis=1); piv = piv.sort_values("合计", ascending=False)
        rows = [[k, float(s["合计"]) * sign, float(s["合计"] / max(base, 1))] +
                [float(s[m]) * sign for m in mcols] + [float(s["合计"] / max(len(mcols), 1)) * sign]
                for k, s in piv.iterrows()]
        tot = [float(piv["合计"].sum()) * sign, 1.0] + [float(piv[m].sum()) * sign for m in mcols] + [float(piv["合计"].sum() / max(len(mcols), 1)) * sign]
        r = table(ws, r, ["分类", "合计", "占比"] + mcols + ["月均"], rows,
                  fmts=[None, MF, PF] + [MF] * (len(mcols) + 1), fill=fill, total=tot) + 1
    sect(ws, r, "2.5", "对手方分析"); r += 1
    cpa = V["cp_all"]; wl_names = {norm_cp(k) for k in whitelist}
    def cptype(k, name):
        if name in V["members"]: return "本方"
        if k in wl_names: return "白名单"
        if any(w[0] == name for w in V["watch"]): return "需关注对手方"
        return ""
    for num, name, colv, fill, sign, base, nb in [("2.5.1", "前10大流入方", "流入", INH, 1, A["tot_in"], "流入笔数"),
                                                  ("2.5.2", "前10大流出方", "流出", OUTH, -1, A["tot_out"], "流出笔数")]:
        sub(ws, r, num, name); r += 1
        top = cpa.sort_values(colv, ascending=False).head(10)
        amtcol = "收入金额" if colv == "流入" else "支出金额"
        piv = df[df["对手key"].isin(top.index)].pivot_table(index="对手key", columns="月份", values=amtcol, aggfunc="sum", fill_value=0.0).reindex(top.index).reindex(columns=mcols, fill_value=0.0)
        rows = [[x["名称"], cptype(k, x["名称"]), float(x[colv]) * sign, float(x[colv] / max(base, 1)), int(x[nb]), V["cp_mix"](k, amtcol)] +
                [float(piv.loc[k, m]) * sign for m in mcols] for k, x in top.iterrows()]
        tot = [None, float(top[colv].sum()) * sign, float(top[colv].sum() / max(base, 1)), int(top[nb].sum()), None] + [float(piv[m].sum()) * sign for m in mcols]
        r = table(ws, r, ["对方名称", "对方类型", f"{colv}总额", f"占总{colv}比", "笔数", f"{colv}构成(按二级标签)"] + mcols, rows,
                  fmts=[None, None, MF, PF, "#,##0", None] + [MF] * len(mcols), fill=fill, total=tot) + 1
    sub(ws, r, "2.5.3", "前10大流入/流出重合方（资金双向对手）"); r += 1
    rows = [[x["名称"], cptype(k, x["名称"]), float(x["流入"]), -float(x["流出"]), float(x["重合金额"]), float(x["重合比"])]
            for k, x in V["overlap"].iterrows()]
    r = table(ws, r, ["对方名称", "对方类型", "流入总额", "流出总额", "重合金额", "重合比"],
              rows or [["无双向对手", "", 0, 0, 0, 0]], fmts=[None, None, MF, MF, MF, PF])
    foot(ws, r, "重合金额=min(流入,流出)；重合比=重合金额/max(流入,流出)。高重合+大额=资金空转/对倒候选，与2.6.2互印。"); r += 2
    sub(ws, r, "2.5.4", "对手结构指标"); r += 1
    r = table(ws, r, ["指标", "数值", "说明"], [
        ["活跃对手数(月均，对外)", float(mon["活跃对手数"].mean()), ""],
        ["流入HHI / 流出HHI", f"{A['hhi_in']:.3f} / {A['hhi_out']:.3f}", ">0.18 高集中"],
        ["老对手(≥6个月)留存流入占比", A["retain_in"], "客户粘性"],
        ["对私交易金额占比", A["priv_ratio"], "过高需核实"],
        ["白名单加权流入占比", A["wl_share"], "经 --whitelist 配置" if whitelist else "未配置白名单，按0计"]],
        fmts=[None, "#,##0.000", None], spans={2: 3})
    r = nxt(ws); sect(ws, r, "2.6", "关联方交易"); r += 1
    sub(ws, r, "2.6.1", "集团成员间互转"); r += 1
    rows = [[o, c, float(x["流入"]), -float(x["流出"]), int(x["笔数"])] for (o, c), x in V["rel_pairs"].iterrows()]
    r = table(ws, r, ["本方成员", "对方成员", "流入", "流出", "笔数"], rows or [["无成员间互转", "", 0, 0, 0]],
              fmts=[None, None, MF, MF, "#,##0"],
              total=[None, float(V["rel_pairs"]["流入"].sum()) if len(V["rel_pairs"]) else 0,
                     -float(V["rel_pairs"]["流出"].sum()) if len(V["rel_pairs"]) else 0, None])
    foot(ws, r, "全部关联主体已合并为一个集团分析；本表金额已从「净对外/有效」口径剔除。"); r += 2
    sub(ws, r, "2.6.2", "疑似关联外部方（双向对倒判定：平衡度≥0.7 且双边≥50万）"); r += 1
    rows = []
    for k, x in A["both"].sort_values("平衡度", ascending=False).head(10).iterrows():
        flag = "疑似关联外部方" if (x["平衡度"] >= .7 and min(x["流入"], x["流出"]) >= 500_000) else ("正常产业双向" if nature(x["对手"]) == "企业" else "关注")
        rows.append([x["对手"], float(x["流入"]), -float(x["流出"]), float(x["平衡度"]), nature(x["对手"]), flag])
    r = table(ws, r, ["对手", "流入", "流出", "平衡度", "性质", "判定"], rows or [["无", 0, 0, 0, "", ""]],
              fmts=[None, MF, MF, "0.00", None, None])
    r = nxt(ws, 1); sub(ws, r, "2.6.3", "工商关联方（法人/股东/主要人员/对外投资）"); r += 1
    r = table(ws, r, ["关联类型", "对方名称", "流入", "流出", "说明"],
              [["待外部数据", "接入工商图谱后自动匹配流水对手", None, None, "可用 tyc-* 技能查询后人工回填"]], spans={4: 3})
    r = nxt(ws); sect(ws, r, "2.7", "需关注对手方"); r += 1
    rows = [[w[0], w[1], float(w[2]), -float(w[3]), float(w[4])] for w in V["watch"]]
    r = table(ws, r, ["对方名称", "风险点", "流入金额", "流出金额", "差额"],
              rows or [["未命中需关注规则", "—", 0, 0, 0]], fmts=[None, None, MF, MF, MF],
              total=[None, sum(w[2] for w in V["watch"]), -sum(w[3] for w in V["watch"]), None] if V["watch"] else None)
    foot(ws, r, "风险点规则：个人大额往来(个人对手双边合计≥50万)／大额整数交易(单笔≥100万整数)／司法对手／金融投机/赌博／民间借贷。TOP20按金额降序。"); r += 2
    sect(ws, r, "2.8", "个人敏感支出"); r += 1
    sub(ws, r, "2.8.1", "夜间大额时段分布（22:00–06:00，对私流出，仅含有时点交易）"); r += 1
    rows = [[b, -amt, n] for b, amt, n in V["night_buckets"]]
    r = table(ws, r, ["时间", "流出金额", "笔数"], rows, fmts=[None, MF, "#,##0"],
              total=[-sum(b[1] for b in V["night_buckets"]), sum(b[2] for b in V["night_buckets"])])

    # ---- 2.9 淡旺季分析 ----
    SE = V["seas"]
    r = nxt(ws); sect(ws, r, "2.9", "淡旺季（季节性）分析"); r += 1
    r = cards(ws, r, [("旺季月数", len(SE["peak"]), "#,##0"), ("淡季月数", len(SE["low"]), "#,##0"),
                      ("最长连续淡季(月)", SE["max_run"], "#,##0", SE["max_run"] >= 3)])
    r = cards(ws, r, [("旺季月均净流入", SE["peak_cf"], MF),
                      ("淡季月均净流入", SE["low_cf"], MF, SE["low_cf"] == SE["low_cf"] and SE["low_cf"] < 0)])
    mtext(ws, r, f"分档口径：月流入 ÷ 全期月均流入 = 流入指数；≥{THRESHOLDS['旺季指数下限']:.1f} 旺季、"
                 f"<{THRESHOLDS['淡季指数上限']:.1f} 淡季、其余平季。"
                 f"⚠ 边月（首尾不满整月）与「在测账户数 < 全期最大 {V['max_acct']} 个」的月份一律标注并剔出定性——"
                 "否则早期只提交了部分账户的区间会被误读成淡季（把数据覆盖问题当成经营季节性）。"
                 + (f"本次剔出 {len(SE['partial'])} 个月：{'、'.join(SE['partial'])}。" if SE["partial"] else "")
                 + ("淡季月均净流入为负 → 淡季经营现金流无法自我覆盖，须评估该区间还款来源。"
                    if SE["low_cf"] == SE["low_cf"] and SE["low_cf"] < 0 else "")); r += 1
    sub(ws, r, "2.9.1", "月度季节指数（时序）"); r += 1
    sx = SE["mx"]
    rows = [[m, float(x["流入"]), float(x["流入指数"]), int(x["笔数"]), float(x["笔数指数"]),
             int(x["活跃对手数"]), int(x["在测账户数"]), float(x["净流入"]),
             x["分档"] if not x["完整性"] else "—", x["完整性"]] for m, x in sx.iterrows()]
    r = table(ws, r, ["月份", "流入金额", "流入指数", "笔数", "笔数指数", "活跃对手数", "在测账户数",
                      "净流入", "分档", "完整性"],
              rows, fmts=[None, MF, "0.00", "#,##0", "0.00", "#,##0", "#,##0", MF, None, None]) + 1
    cmg = SE["cm"]
    if len(cmg):
        sub(ws, r, "2.9.2", "自然月季节性（跨年合并 1–12 月，看行业周期规律）"); r += 1
        rows = [[f"{int(m)}月", int(x["月份数"]), float(x["月均流入"]), float(x["流入指数"]),
                 float(x["月均笔数"]), float(x["月均净流入"]), x["分档"]] for m, x in cmg.iterrows()]
        r = table(ws, r, ["自然月", "样本月数", "月均流入", "流入指数", "月均笔数", "月均净流入", "分档"],
                  rows, fmts=[None, "#,##0", MF, "0.00", "#,##0.0", MF, None])
        foot(ws, r, "样本月数=1 的自然月为单年观测，季节性结论证据较弱；需≥2年数据才能稳定判定行业周期。"); r += 1

    # ---- 三 数据验证 ----
    r = nxt(ws); chap(ws, r, "三", "数据验证"); r += 1
    sect(ws, r, "3.1", "数据完整性校验（按主体×账户）"); r += 1
    for owner in dict.fromkeys(a["owner"] for a in V["accounts"]):
        put(ws, r, 2, owner, bold=True); r += 1
        rows = [[a["acct"], a["bank"] or "—", a["cur"] or "—",
                 f"{a['t0']:%Y.%m.%d}~{a['t1']:%Y.%m.%d}", f"{a['months']}个月",
                 ("无" if not a["breaks"] else f"{a['breaks']}处") if a["breaks"] is not None else "未提供校验表",
                 a["cp"], a["memo"]] for a in V["accounts"] if a["owner"] == owner]
        r = table(ws, r, ["账号", "所属银行", "币种", "数据时间范围", "数据时长", "余额断点", "对方名称", "交易信息"], rows, spans={7: 2}) + 1
    sect(ws, r, "3.2", "数据质量验证（账户级评级 + 全局真实性校验）"); r += 1
    rows = [[a["acct"], a["owner"], f"{a['rows']}笔", "有" if a["jx"] == "有" else "无",
             a["fill"], a["grade"], a["detail"]] for a in V["accounts"]]
    r = table(ws, r, ["账号", "所属主体", "数据量", "活期结息记录", "字段完整度", "数据质量", "数据质量详情"],
              rows, fmts=[None, None, None, None, PF, None, None], spans={6: 3})
    for i, a in enumerate(V["accounts"]):  # 评级橙字
        if a["grade"] in ("一般", "数据不足"):
            ws.cell(r - len(V["accounts"]) + i, 7).font = ORNG
    foot(ws, r, "评级是数据可用性分级（优秀/良好/一般/数据不足），不是风险结论；「数据不足」=笔数<30或月数<3。"); r += 2
    sa, st, sf, sc_ = A["qscore"]
    sub(ws, r, "3.2.1", "数据质量四维评分（前提分）"); r += 1
    r = table(ws, r, ["维度", "权重", "得分", "扣分原因"], [
        ["账户覆盖", 30, sa, f"隐藏账户信号{len(A['hidden'])}组、未提交账户{len(V['unsub'])}个（见3.3）"],
        ["时段覆盖", 25, st, "近12月覆盖率与首月完整性"],
        ["字段完整性", 20, sf, f"对手非空率{A['cp_fill']:.0%}；摘要非空率{A['rem_fill']:.0%}"],
        ["一致性校验", 25, sc_, f"余额断点{A['bal_breaks'] if A['bal_breaks'] is not None else '未提供'}；冲正{len(A['chz'])}笔"]],
        fmts=[None, "#,##0", "#,##0", None], spans={3: 3},
        total=[None, A["q_total"], f"{A['q_grade']}（<70分先补件再审批）"], total_label="总分")
    r += 1
    sub(ws, r, "3.2.2", "真实性校验清单"); r += 1
    jxs = f"隐含年化{A['jx_mean']:.3f}%±CV {A['jx_cv']:.0%}" if A["jx_mean"] == A["jx_mean"] else "未识别到结息记录"
    r = table(ws, r, ["校验项", "结果", "数值", "说明"], [
        ["余额连续性", "通过" if not A["bal_breaks"] else "异常", f"断点{A['bal_breaks'] if A['bal_breaks'] is not None else '—'}", "沿用上游标准化校验"],
        ["结息复算(隐含利率)", "通过" if (A["jx_cv"] == A["jx_cv"] and A["jx_cv"] <= .5) else "关注", jxs, "伪造流水难以维持跨季一致的隐含利率"],
        ["Benford首位分布", "关注" if A["ben_mad"] > .015 else "通过", f"MAD={A['ben_mad']:.3f}", "固定金额(租金/月供)为可解释偏离"],
        ["整数金额占比(≥1万)", "关注" if A["int_ratio"] > THRESHOLDS["整数占比上限"] else "通过", f"{A['int_ratio']:.1%}", "过高需交叉发票"],
        ["快进快出", "关注" if A["fastio"] > THRESHOLDS["快进快出上限"] else "通过", f"{A['fastio']:.0%}", "大额入账日T+1流出≥80%天数占比"],
        ["夜间交易", "关注" if A["night_ratio"] > THRESHOLDS["夜间占比上限"] else "通过", f"{A['night_n']}笔/{A['night_ratio']:.1%}", "分母仅含有时点交易"]], spans={3: 3})
    for i in range(6):
        c = ws.cell(r - 6 + i, 3)
        if str(c.value) in ("关注", "异常"): c.font = ORNG
    if A["jx_list"]:
        r = nxt(ws, 1); sub(ws, r, "3.2.3", "结息复算明细"); r += 1
        r = table(ws, r, ["结息日", "实际结息(元)", "季度日积数(元·日)", "隐含年化(%)"],
                  [[a, b, c, d] for a, b, c, d in A["jx_list"]], fmts=[None, MF, MF0, "0.0000"])
    r = nxt(ws); sect(ws, r, "3.3", f"未提交账户（{V['unsub_mode']}）"); r += 1
    if V["unsub"]:
        tot_i = sum(x[2] for x in V["unsub"]); tot_o = sum(x[3] for x in V["unsub"])
        rows = [[nm_, oa, i, i / max(A["tot_out"], 1), o, o / max(A["tot_in"], 1), i - o]
                for nm_, oa, i, o in V["unsub"]]
        r = table(ws, r, ["成员名称", "对方账号", "流入金额(流向该账号)", "占总流出比", "流出金额(来自该账号)", "占总流入比", "差额"],
                  rows, fmts=[None, None, MF, PF, MF, PF, MF],
                  total=[None, tot_i, tot_i / max(A["tot_out"], 1), tot_o, tot_o / max(A["tot_in"], 1), tot_i - tot_o])
        foot(ws, r, "口径：集团成员作为交易对手、但其账号不在本次提交账户集内 → 疑似未提交账户；金额为已提交账户与该账号的往来。要求补充对应账户流水。"); r += 1
    else:
        put(ws, r, 2, "未发现成员名下未提交账号（或已全部提交）。"); r += 1
    r = nxt(ws, 1); sub(ws, r, "3.3.1", "规则型缺失账户反推"); r += 1
    hid = A["hidden"] or [["未发现隐藏账户信号", "—", "—", "—"]]
    r = table(ws, r, ["推断", "证据", "置信度", "建议动作"], hid, spans={1: 3, 3: 2})
    r = nxt(ws); sect(ws, r, "3.4", "主体名称核验（不正确的主体名称将影响关联方捕获与分类准确性）"); r += 1
    person_owners = [m for m in V["members"] if nature(m) == "个人"]
    rows = ([[m, "个人户主，建议人工确认其经营主体（工商数据待接入）", PENDING] for m in person_owners]
            or [["—", "全部主体为企业名称", "无需处理"]])
    r = table(ws, r, ["当前主体名称", "建议", "确认结果"], rows, spans={1: 4})

    # ---- 四 生意模式与审批要件（V3 平移）----
    r = nxt(ws); chap(ws, r, "四", "生意模式与审批要件"); r += 1
    sect(ws, r, "4.1", "生意模式（脚本汇集证据，定性结论待复核补写）"); r += 1
    ind = cp.sort_values("流入", ascending=False).head(5); outd = cp.sort_values("流出", ascending=False).head(5)
    r = table(ws, r, ["要素", "自动汇集证据", "定性结论"], [
        ["所属行业", "主体：" + subjects + "；TOP流入对手：" + "、".join(ind["对手"]), PENDING],
        ["经营品类(上游采购)", "TOP流出对手：" + "、".join(outd["对手"]), PENDING],
        ["盈利逻辑", f"经营流入(剔往来){mon_op['经流入'].sum():,.0f} vs 经营流出{mon_op['经流出'].sum():,.0f}元", PENDING],
        ["刚性成本锚点", f"薪酬{df.loc[df['二级标签'].str.contains('人力'), '支出金额'].sum():,.0f}／税费{df.loc[df['二级标签'].str.contains('税'), '支出金额'].sum():,.0f}／运营{df.loc[df['二级标签'].str.contains('运营'), '支出金额'].sum():,.0f}元", "锚点强弱判断：" + PENDING],
        ["核心风险", "自动信号见4.3；集中度/未提交账户/申贷前融资见看板", PENDING],
        ["准入反欺诈滤网", "需工商+征信+进件数据（法人任职/法代变更/代持特征/空壳）", "待外部数据"]], spans={1: 3, 2: 4})
    r = nxt(ws); sect(ws, r, "4.2", "审批要件指标看板（阈值为参考口径，红/黄/绿不构成授信结论）"); r += 1
    l12v = mon.loc[[m for m in L12 if m in mon.index]]
    b1y = A["bal"][A["bal"].index >= A["bal"].index.max() - pd.Timedelta(days=365)] if A["bal"] is not None else None
    cv12 = float(l12v["流入"].std() / max(l12v["流入"].mean(), 1))
    mad = float((b1y - b1y.mean()).abs().mean() / max(b1y.mean(), 1)) if b1y is not None else float("nan")
    above = float((b1y > b1y.mean()).mean()) if b1y is not None else float("nan")
    top_in = cp.sort_values("流入", ascending=False)
    svc = A["svc_base"]; opnet = mon_op.loc[[m for m in L12 if m in mon_op.index], "经净"]
    minbal3 = mon.loc[[m for m in L12[-3:] if m in mon.index], "月最低余额"].mean()
    kpis = [
        ["业务规模", "近12月对外流入", "剔互转", f"≥{THRESHOLDS['近12月流入下限']/1e4:,.0f}万", l12v["流入"].sum(), "通过" if l12v["流入"].sum() >= THRESHOLDS["近12月流入下限"] else "关注"],
        ["业务规模", "有效流入(全区间)", "见1.1口径", "—", V["eff_in"], "通过"],
        ["关联交易", "内部互转占比", "互转额/总收支", "≤30%", (A["int_in"] + A["int_out"]) / max(A["tot_in"] + A["tot_out"], 1), "通过" if (A["int_in"]+A["int_out"])/max(A["tot_in"]+A["tot_out"],1) <= .3 else "关注"],
        ["稳定性", "月流入变异系数", "近12月CV", "≤0.5", cv12, "关注" if cv12 > .5 else "通过"],
        ["稳定性", "余额离差率", "日余额MAD/日均", "≤1.0", mad, "关注" if mad == mad and mad > 1 else "通过"],
        ["集中度", "TOP1对手流入占比", "对外口径", "≤50%", float(top_in["流入"].iloc[0] / max(ext_in, 1)) if len(top_in) else 0, "关注" if len(top_in) and top_in["流入"].iloc[0] / max(ext_in, 1) > .5 else "通过"],
        ["集中度", "TOP5对手流入占比", "对外口径", "≤80%", float(top_in["流入"].head(5).sum() / max(ext_in, 1)), "关注" if top_in["流入"].head(5).sum() / max(ext_in, 1) > .8 else "通过"],
        ["集中度", "流入HHI", "Σ份额²", "≤0.18", A["hhi_in"], "关注" if A["hhi_in"] > .18 else "通过"],
        ["资金往来", "往来款收入占比", "", "≤30%", A["wl_in"] / max(A["tot_in"], 1), "关注" if A["wl_in"] / max(A["tot_in"], 1) > .3 else "通过"],
        ["余额健康", "余额上穿日均占比", "近一年", "≥35%", above, "关注" if above == above and above < .35 else "通过"],
        ["白名单", "白名单加权流入占比", "可配置", "≥30%", A["wl_share"], "通过" if A["wl_share"] >= .3 else "关注"],
        ["负债", "存量月债务服务(元)", "识别月供合计", "—", svc, "关注" if svc else "通过"],
        ["负债", "DSR还款负担率", "月债务服务/月均对外流入", "≤15%", svc / max(l12v["流入"].mean(), 1), "关注" if svc / max(l12v["流入"].mean(), 1) > .15 else "通过"],
        ["现金流", "本息覆盖倍数(存量)", "月均经营净CF/月债务服务", "≥1.5", (opnet.mean() / svc) if svc else float("nan"), "高危" if svc and opnet.mean() / svc < 1 else "关注"],
        ["现金流", "安全垫天数", "近3月月最低余额均值/日均对外流出", "≥7", minbal3 / max(l12v["流出"].sum() / 365, 1), "高危" if minbal3 / max(l12v["流出"].sum() / 365, 1) < 3 else ("关注" if minbal3 / max(l12v["流出"].sum() / 365, 1) < 7 else "通过")],
        ["可信度", "数据质量总分", "四维加权", "≥70", A["q_total"], "关注" if A["q_total"] < 70 else "通过"],
        ["可信度", "隐藏账户信号组数", "规则反推+未提交账户", "=0", len(A["hidden"]) + (1 if V["unsub"] else 0), "高危" if len(A["hidden"]) + (1 if V["unsub"] else 0) >= 2 else ("关注" if (A["hidden"] or V["unsub"]) else "通过")]]
    r = table(ws, r, ["类别", "指标", "口径", "参考阈值", "企业实际", "判断"], kpis,
              fmts=[None, None, None, None, MF, None])
    for i in range(len(kpis)):
        cell = ws.cell(r - len(kpis) + i, 7); v = str(cell.value)
        cell.fill = GRN if v == "通过" else (REDF if "高危" in v else WARN)
    r = nxt(ws); sect(ws, r, "4.3", "重大风险行为信号（疑似度+证据链；≥60分强制人工定性）"); r += 1
    r = table(ws, r, ["信号", "疑似度", "证据链", "建议核查动作"], A["signals"], fmts=[None, "#,##0", None, None], spans={2: 3, 3: 2})
    for i in range(len(A["signals"])):
        v = A["signals"][i][1]; ws.cell(r - len(A["signals"]) + i, 3).fill = REDF if v >= 60 else (WARN if v >= 30 else GRN)
    r = nxt(ws); sect(ws, r, "4.4", "行业专家评分（共性维度自动证据；得分待人工标定）"); r += 1
    auto = [("共性|收入真实性", 10, f"结息复算{('通过' if A['jx_cv']==A['jx_cv'] and A['jx_cv']<=0.5 else '关注')}；往来款占比{A['wl_in']/max(A['tot_in'],1):.0%}；隐藏账户{len(A['hidden'])}组"),
            ("共性|负债负担", 10, f"月债务服务{svc:,.0f}元；DSR {svc/max(l12v['流入'].mean(),1):.0%}"),
            ("共性|现金流健康", 10, f"近12月经营净CF {opnet.sum():,.0f}元；期末余额{A['closing_cash']:,.0f}元"),
            ("行业|订单与回款质量", 20, f"白名单加权流入占比{A['wl_share']:.0%}"),
            ("行业|采购规律性", 15, "TOP流出对手连续性见数据明细·对手月度矩阵"),
            ("行业|刚性成本可见度", 15, "薪酬/税费/运营支出规模见4.1"),
            ("行业|产能与投入", 10, "设备/租赁融资见负债清单"),
            ("行业|集中度容忍", 10, f"TOP1 {float(top_in['流入'].iloc[0]/max(ext_in,1)) if len(top_in) else 0:.0%}")]
    r = table(ws, r, ["维度", "满分", "得分(待标定)", "证据(自动)"], [[a, b, PENDING, d] for a, b, d in auto],
              fmts=[None, "#,##0", None, None], spans={2: 3, 3: 3})
    foot(ws, r, "评分卡分型（制造/批发/零售/服务/出口）与维度权重见 references/design-v3.md 模块七；由生意模式判定后路由。"); r += 2
    sect(ws, r, "4.5", "交叉验证矩阵（流水列自动填充；其余数据源待补）"); r += 1
    mx = [["收入规模", f"近12月对外流入{l12v['流入'].sum():,.0f}元", "⚪销项发票", "⚪应税收入", "—", "⚪中标数据", "C(单源)"],
          ["采购成本", f"主营采购{ext.loc[ext['二级标签']=='主营业务','支出金额'].sum():,.0f}元", "⚪进项发票", "⚪财报", "—", "—", "C(单源)"],
          ["负债结构", f"识别月供{svc:,.0f}元+隐藏信号{len(A['hidden'])}组", "—", "—", "⚪征信在贷", "⚪司法纠纷", "D(待征信)"],
          ["对手真实性", "TOP10清单见2.5", "⚪发票购销方", "—", "—", "⚪工商在营", "C(单源)"],
          ["行业判断", "证据见4.1", "⚪发票品目", "⚪税种结构", "—", "⚪登记行业", "C(单源)"]]
    r = table(ws, r, ["核心结论", "流水", "发票", "税务", "征信", "工商/司法", "可信度"], mx, spans={1: 2})

    # ============ Sheet2 负债与现金流（五/六，V3 平移+编号）============
    ws = S2
    put(ws, 1, 2, "负债与现金流（Debt OS）", font=TITLE_F); ws.row_dimensions[1].height = 30
    r = 2; chap(ws, r, "五", "负债结构（黄色=假设待确认）"); r += 1
    sect(ws, r, "5.1", "债务清单（从流水识别+反推）"); r += 1
    rows = [[d["lender"], d["prod"], d["principal"] if d["principal"] == d["principal"] else "待核",
             d["apr"] if d["apr"] == d["apr"] else "待核", d["pmt"], d["term"] or "—", d["balloon"], d["src"], d["note"]]
            for d in A["debts"]]
    r = table(ws, r, ["出借方", "产品", "本金(元)", "APR(反推)", "月供(元)", "假设期数", "气球尾款", "识别来源", "备注"],
              rows or [["未识别到债务", "", 0, 0, 0, "—", 0, "", ""]],
              fmts=[None, None, MF, "0.0%", MF, None, MF, None, None], spans={8: 3})
    for i, d in enumerate(A["debts"]):
        if "假设" in d["note"] or "估算" in d["note"]: ws.cell(r - len(A["debts"]) + i, 7).fill = YEL
    r = nxt(ws); sect(ws, r, "5.2", "债务组合KPI"); r += 1
    known = [d for d in A["debts"] if d["principal"] == d["principal"] and d["apr"] == d["apr"]]
    wair = sum(d["principal"] * d["apr"] for d in known) / max(sum(d["principal"] for d in known), 1) if known else float("nan")
    r = table(ws, r, ["KPI", "数值", "口径"], [
        ["总在贷本金(可见口径,元)", sum(d["principal"] for d in A["debts"] if d["principal"] == d["principal"]), "含估算项；他行待核不含"],
        ["WAIR加权平均利率", wair, "仅含可反推APR的债务"],
        ["月债务服务(元)", A["svc_base"], "识别出的月供合计"],
        ["12个月内气球尾款(元)", sum(d["balloon"] for d in A["debts"]), "到期一次还本类"]], fmts=[None, MF, None], spans={2: 2})
    r = nxt(ws); sect(ws, r, "5.3", "未来12个月偿付Schedule（元/月；气球计入到期月）"); r += 1
    for j, hx in enumerate(["债务"] + fwd, 2): put(ws, r, j, hx, bold=True, fillc=INH)
    r0 = r + 1
    sch_rows = [d for d in A["debts"] if d["pmt"] > 0 or d["balloon"] > 0]
    for i, d in enumerate(sch_rows):
        put(ws, r0 + i, 1, i + 1); put(ws, r0 + i, 2, d["lender"][:16])
        for j in range(12):
            v = d["pmt"] + (d["balloon"] if (d["balloon"] > 0 and j == 5) else 0)
            put(ws, r0 + i, 3 + j, float(v if d["pmt"] > 0 or j == 5 else 0), fmt=MF0)
    SVC_ROW = r0 + max(len(sch_rows), 1)
    put(ws, SVC_ROW, 2, "月合计", bold=True, fillc=WARN)
    for j in range(12):
        cl = get_column_letter(3 + j)
        put(ws, SVC_ROW, 3 + j, f"=SUM({cl}{r0}:{cl}{SVC_ROW-1})", fmt=MF0, fillc=WARN)
    r = SVC_ROW + 2
    chap(ws, r, "六", "现金流健康（蓝色=可修改假设）"); r += 1
    sect(ws, r, "6.1", "历史月度经营净现金流与实付债务服务"); r += 1
    hdr_row(ws, r, ["月份", "经营净CF(剔往来款)", "债务服务(实付)", "DSCR", "月最低余额", "月末余额"])
    svc_hist = df[df["三级标签"].str.contains("偿还") | (df["二级标签"] == "负债")].groupby("月份")["支出金额"].sum().reindex(MONTHS, fill_value=0)
    r0 = r + 1
    for i, m in enumerate(MONTHS):
        put(ws, r0 + i, 1, i + 1); put(ws, r0 + i, 2, m)
        put(ws, r0 + i, 3, float(mon_op.loc[m, "经净"]), fmt=MF); put(ws, r0 + i, 4, float(svc_hist[m]), fmt=MF)
        put(ws, r0 + i, 5, f'=IF(D{r0+i}=0,"-",C{r0+i}/D{r0+i})', fmt="0.00")
        put(ws, r0 + i, 6, float(mon.loc[m, "月最低余额"]) if mon.loc[m, "月最低余额"] == mon.loc[m, "月最低余额"] else "—", fmt=MF)
        put(ws, r0 + i, 7, float(mon.loc[m, "月末余额"]) if mon.loc[m, "月末余额"] == mon.loc[m, "月末余额"] else "—", fmt=MF)
    r = r0 + nm + 1
    sect(ws, r, "6.2", "未来12个月现金推演（三情景）"); r += 1
    A0 = r
    wl_net = float((mon.loc[[m for m in L12 if m in mon.index], "往来款流入"].sum() - df.loc[df["二级标签"].str.contains("往来") & df["月份"].isin(L12), "支出金额"].sum()) / max(len(L12), 1))
    P, RATE, N = A["new_loan"]
    inputs = [("期初现金(元)", A["closing_cash"] if A["closing_cash"] == A["closing_cash"] else 0.0),
              ("月经营流入基线(近12月均,剔往来款,元)", float(mon_op.loc[[m for m in L12 if m in mon_op.index], "经流入"].mean())),
              ("月经营流出基线(元)", float(mon_op.loc[[m for m in L12 if m in mon_op.index], "经流出"].mean())),
              ("往来款月净流入(元)", wl_net), ("最低安全垫(元)", 200000.0),
              ("拟新增授信本金(元)", P), ("新增授信年利率", RATE), ("新增授信期数(月)", N)]
    for i, (k, v) in enumerate(inputs):
        put(ws, A0 + i, 2, k, bold=True, wrap=True); fit_row(ws, A0 + i, est_lines(k, colw(ws, 2)))
        c = put(ws, A0 + i, 3, v, fmt=MF if abs(v) > 1 else "0.0%"); c.font = BLUEF
    IN = {k: f"$C${A0+i}" for i, (k, v) in enumerate(inputs)}
    r = A0 + len(inputs)
    put(ws, r, 2, "新增授信月供(元)", bold=True)
    put(ws, r, 3, f'=ROUND(-PMT({IN["新增授信年利率"]}/12,{IN["新增授信期数(月)"]},{IN["拟新增授信本金(元)"]}),0)', fmt=MF0, fillc=WARN)
    NEWPMT = f"$C${r}"; r += 2
    hdr_row(ws, r, ["月份", "存量债务服务", "总债务服务(含新增)", "基线:期末现金", "收入-20%:期末现金", "收入-30%:期末现金", "基线DSCR"])
    T0 = r + 1
    for i, m in enumerate(fwd):
        rr = T0 + i; put(ws, rr, 1, i + 1); put(ws, rr, 2, m)
        put(ws, rr, 3, f"={get_column_letter(3+i)}{SVC_ROW}", fmt=MF0)
        put(ws, rr, 4, f"=C{rr}+{NEWPMT}", fmt=MF0)
        add = f'+{IN["拟新增授信本金(元)"]}' if i == 0 else ""
        for cix, mult_i, mult_o in [(5, "", ""), (6, "*0.8", "*0.9"), (7, "*0.7", "*0.85")]:
            prev = IN["期初现金(元)"] if i == 0 else f"{get_column_letter(cix)}{rr-1}"
            put(ws, rr, cix, f'={prev}+{IN["月经营流入基线(近12月均,剔往来款,元)"]}{mult_i}-{IN["月经营流出基线(元)"]}{mult_o}+{IN["往来款月净流入(元)"]}-D{rr}{add}', fmt=MF0)
        put(ws, rr, 8, f'=({IN["月经营流入基线(近12月均,剔往来款,元)"]}-{IN["月经营流出基线(元)"]}+{IN["往来款月净流入(元)"]})/D{rr}', fmt="0.00")
    TN = T0 + 11; r = TN + 2
    r = table(ws, r, ["推演KPI", "基线", "收入-20%", "收入-30%", "说明"], [
        ["最低期末现金(元)", f"=MIN(E{T0}:E{TN})", f"=MIN(F{T0}:F{TN})", f"=MIN(G{T0}:G{TN})", ""],
        ["触底次数", f'=COUNTIF(E{T0}:E{TN},"<"&{IN["最低安全垫(元)"]})', f'=COUNTIF(F{T0}:F{TN},"<"&{IN["最低安全垫(元)"]})', f'=COUNTIF(G{T0}:G{TN},"<"&{IN["最低安全垫(元)"]})', ""],
        ["所需增量融资(元)", f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(E{T0}:E{TN}))', f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(F{T0}:F{TN}))', f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(G{T0}:G{TN}))', "直接回答「该批多少/缺口多大」"],
        ["DSCR<1.2月份数(基线)", f'=COUNTIF(H{T0}:H{TN},"<1.2")', "", "", "含新增月供口径"]],
        fmts=[None, MF0, MF0, MF0, None], seq=False, spans={4: 2})
    DYN_T0, DYN_TN = T0, TN
    foot(ws, r, "机械外推样稿。蓝色输入可改；经营口径受标签质量与账户覆盖影响，补数据后重算。"); r += 1
    r = nxt(ws); sect(ws, r, "6.3", "流水折算现金流量表（会小企03口径；金融机构大额放款已修正归入筹资）"); r += 1
    items = ["销售产成品、商品、提供劳务收到的现金", "收到其他与经营活动有关的现金", "购买原材料、商品、接受劳务支付的现金", "支付的职工薪酬", "支付的税费", "支付其他与经营活动有关的现金",
             "收回投资收到的现金", "投资支付的现金", "取得借款收到的现金", "偿还借款本金及利息支付的现金"]
    put(ws, r, 2, "项目", bold=True, fillc=INH)
    for cix, (label, _) in enumerate(A["cfs_cols"], 3): put(ws, r, cix, label, bold=True, fillc=INH)
    lay = [("一、经营活动", None), (items[0], 0), (items[1], 1), (items[2], 2), (items[3], 3), (items[4], 4), (items[5], 5), ("  经营净额", "OP"),
           ("二、投资活动", None), (items[6], 6), (items[7], 7), ("  投资净额", "INV"),
           ("三、筹资活动", None), (items[8], 8), (items[9], 9), ("  筹资净额", "FIN"), ("四、现金净增加额", "NET")]
    rowmap = {}; r0 = r + 1
    for i, (name, idx) in enumerate(lay):
        rr = r0 + i; rowmap[idx] = rr
        put(ws, rr, 2, name, bold=isinstance(idx, str) or idx is None, fillc=INH if idx is None else None, wrap=True)
        fit_row(ws, rr, est_lines(name, colw(ws, 2)))
        if isinstance(idx, int):
            for cix, (_, vals) in enumerate(A["cfs_cols"], 3): put(ws, rr, cix, float(vals[idx]), fmt=MF)
    for cix in range(3, 3 + len(A["cfs_cols"])):
        cl = get_column_letter(cix)
        put(ws, rowmap["OP"], cix, f"={cl}{rowmap[0]}+{cl}{rowmap[1]}-{cl}{rowmap[2]}-{cl}{rowmap[3]}-{cl}{rowmap[4]}-{cl}{rowmap[5]}", fmt=MF, bold=True)
        put(ws, rowmap["INV"], cix, f"={cl}{rowmap[6]}-{cl}{rowmap[7]}", fmt=MF, bold=True)
        put(ws, rowmap["FIN"], cix, f"={cl}{rowmap[8]}-{cl}{rowmap[9]}", fmt=MF, bold=True)
        put(ws, rowmap["NET"], cix, f"={cl}{rowmap['OP']}+{cl}{rowmap['INV']}+{cl}{rowmap['FIN']}", fmt=MF, bold=True, fillc=WARN)
    foot(ws, rowmap["NET"] + 2, f"勾稽：区间净增加额应≈期末-期初（{A['closing_cash']:,.2f} - {A['opening_cash']:,.2f}，含互转自轧差）。")

    # ============ Sheet3 数据明细（七）============
    ws = S3
    put(ws, 1, 2, "数据明细", font=TITLE_F); ws.row_dimensions[1].height = 30
    r = 2; chap(ws, r, "七", "流水数据明细"); r += 1
    sect(ws, r, "7.1", "月度时序表"); r += 1
    TS_H = r
    hdr = ["月份", "流入金额", "流出金额", "净流入", "笔数", "月末余额", "月均余额", "月最低余额",
           "经营流入(剔往来)", "融资流入(修正)", "往来款流入", "活跃对手数", "借贷流入", "偿还借贷", "借贷净额"]
    hdr_row(ws, r, hdr)
    lm = V["loan_mon"]
    for i, m in enumerate(MONTHS):
        rr = r + 1 + i; put(ws, rr, 1, i + 1)
        vals = [m, mon.loc[m, "流入"], mon.loc[m, "流出"], mon.loc[m, "净流入"], int(mon.loc[m, "笔数"]),
                mon.loc[m, "月末余额"], mon.loc[m, "月均余额"], mon.loc[m, "月最低余额"],
                mon_op.loc[m, "经流入"], mon.loc[m, "融资流入(修正)"], mon.loc[m, "往来款流入"],
                int(mon.loc[m, "活跃对手数"]), lm.loc[m, "借贷流入"], lm.loc[m, "偿还借贷"], lm.loc[m, "净额"]]
        for j, v in enumerate(vals, 2):
            if isinstance(v, float) and v != v: v = None
            put(ws, rr, j, float(v) if isinstance(v, (int, float, np.floating)) and v is not None and j > 2 else v,
                fmt=MF if j in (3, 4, 5, 7, 8, 9, 10, 11, 12, 14, 15, 16) else None)
    TS_N = TS_H + nm
    r = nxt(ws); sect(ws, r, "7.2", "对手月度金额时序（TOP12×月；看突然出现/消失/仅申贷前活跃）"); r += 1
    def cp_mx(direction):
        colv = "收入金额" if direction == "in" else "支出金额"
        tops = cp.sort_values("流入" if direction == "in" else "流出", ascending=False).head(12).index
        piv = df[df["对手key"].isin(tops) & ~df["内部互转"]].pivot_table(index="对手key", columns="月份", values=colv, aggfunc="sum", fill_value=0)
        return piv.reindex(tops).reindex(columns=MONTHS, fill_value=0)
    for tag_, mx, fill in [("流入对手", cp_mx("in"), INH), ("流出对手", cp_mx("out"), OUTH)]:
        put(ws, r, 2, tag_, bold=True, fillc=fill)
        for j, m in enumerate(MONTHS, 3): put(ws, r, j, m, bold=True, fillc=fill)
        r0 = r + 1
        for i, (k, row) in enumerate(mx.iterrows()):
            put(ws, r0 + i, 1, i + 1); put(ws, r0 + i, 2, cp.loc[k, "对手"])
            for j, m in enumerate(MONTHS, 3): put(ws, r0 + i, j, float(row[m]), fmt=MF0)
        r = r0 + len(mx) + 1
    r = nxt(ws); sect(ws, r, "7.3", "单笔极值卡（账户 × 收支方向 × 经营性/全口径）"); r += 1
    r = table(ws, r, ["本方名称", "本方账号", "极值类型", "金额", "日期", "对手", "摘要"],
              V["extremes"] or [["无数据", "", "", None, "", "", ""]],
              fmts=[None, None, None, MF, None, None, None], spans={7: 3})
    foot(ws, r, "「经营性」= 一级标签为经营类、二级标签不含往来、且非内部互转；"
                "「单笔最高收入/支出」为该账户全口径极值（含往来与内部互转）。"
                "两者差距大 → 账户最大额资金动作并非经营行为，需核实其性质。"); r += 2
    sect(ws, r, "7.4", f"大额交易明细（单笔≥50万，{len(A['big_tx'])}笔）"); r += 1
    rows = [[str(x["交易时间"]), "收入" if x["收入金额"] > 0 else "支出", x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["二级标签"], x["三级标签"], x["摘要"][:60]] for _, x in A["big_tx"].iterrows()]
    r = table(ws, r, ["时间", "方向", "对手", "收入", "支出", "二级标签", "三级标签", "摘要"], rows,
              fmts=[None, None, None, MF, MF, None, None, None], spans={2: 2, 7: 4})
    r = nxt(ws); sect(ws, r, "7.5", "疑点交易清单（民间借贷/冲正；供人工核查工单）"); r += 1
    sus = [["民间借贷", str(x["交易时间"])[:10], x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["三级标签"]] for _, x in A["mj"].iterrows()]
    sus += [["冲正/退票", str(x["交易时间"])[:10], x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["摘要"][:40]] for _, x in A["chz"].head(15).iterrows()]
    r = table(ws, r, ["类型", "日期", "对手", "收入", "支出", "说明"], sus or [["无", "", "", 0, 0, ""]],
              fmts=[None, None, None, MF, MF, None], spans={2: 2, 5: 3})

    # ============ Sheet4 可视化看板（9 图）============
    ws = S4
    put(ws, 1, 2, "可视化看板（自上而下下拉查看）", font=TITLE_F)
    def add(ch, anchor, t, h=9, w=22):
        ch.title = t; ch.height = h; ch.width = w; ws.add_chart(ch, anchor)
    c1 = BarChart(); c1.type = "col"
    c1.add_data(Reference(S3, min_col=3, max_col=4, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c1.set_categories(Reference(S3, min_col=2, min_row=TS_H + 1, max_row=TS_N))
    l1 = LineChart(); l1.add_data(Reference(S3, min_col=5, max_col=5, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c1.y_axis.majorGridlines = None; c1 += l1
    add(c1, "B2", "月度流入/流出与净流入")
    c2 = LineChart(); c2.add_data(Reference(S3, min_col=7, max_col=9, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c2.set_categories(Reference(S3, min_col=2, min_row=TS_H + 1, max_row=TS_N)); add(c2, "B22", "余额波动（月末/月均/月最低）")
    c3 = LineChart(); c3.add_data(Reference(S3, min_col=10, max_col=12, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c3.set_categories(Reference(S3, min_col=2, min_row=TS_H + 1, max_row=TS_N)); add(c3, "B42", "经营 vs 融资 vs 往来款 月度流入")
    c9 = BarChart(); c9.type = "col"
    c9.add_data(Reference(S3, min_col=14, max_col=15, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c9.set_categories(Reference(S3, min_col=2, min_row=TS_H + 1, max_row=TS_N))
    l9 = LineChart(); l9.add_data(Reference(S3, min_col=16, max_col=16, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c9.y_axis.majorGridlines = None; c9 += l9
    add(c9, "B62", "借贷趋势（借贷流入/偿还借贷/净额）")
    inc = ext[ext["收入金额"] > 0].groupby("二级标签")["收入金额"].sum().sort_values(ascending=False)
    exp = ext[ext["支出金额"] > 0].groupby("二级标签")["支出金额"].sum().sort_values(ascending=False)
    for i, (k, v) in enumerate(inc.head(8).items()): ws.cell(1 + i, 50, k); ws.cell(1 + i, 51, float(v))
    c4 = PieChart(); c4.add_data(Reference(ws, min_col=51, min_row=1, max_row=max(1, min(8, len(inc)))))
    c4.set_categories(Reference(ws, min_col=50, min_row=1, max_row=max(1, min(8, len(inc))))); add(c4, "B82", "收入用途结构(对外)", h=9, w=13)
    for i, (k, v) in enumerate(exp.head(8).items()): ws.cell(11 + i, 50, k); ws.cell(11 + i, 51, float(v))
    c5 = PieChart(); c5.add_data(Reference(ws, min_col=51, min_row=11, max_row=max(11, 10 + min(8, len(exp)))))
    c5.set_categories(Reference(ws, min_col=50, min_row=11, max_row=max(11, 10 + min(8, len(exp))))); add(c5, "L82", "支出用途结构(对外)", h=9, w=13)
    ti2 = cp.sort_values("流入", ascending=False).head(10); to2 = cp.sort_values("流出", ascending=False).head(10)
    for i, (_, x) in enumerate(ti2.iterrows()): ws.cell(25 + i, 50, x["对手"]); ws.cell(25 + i, 51, float(x["流入"]))
    c6 = BarChart(); c6.type = "bar"; c6.add_data(Reference(ws, min_col=51, min_row=25, max_row=24 + len(ti2)))
    c6.set_categories(Reference(ws, min_col=50, min_row=25, max_row=24 + len(ti2))); c6.legend = None
    add(c6, "B102", "十大流入对手（外部上游/客户）")
    for i, (_, x) in enumerate(to2.iterrows()): ws.cell(40 + i, 50, x["对手"]); ws.cell(40 + i, 51, float(x["流出"]))
    c7 = BarChart(); c7.type = "bar"; c7.add_data(Reference(ws, min_col=51, min_row=40, max_row=39 + len(to2)))
    c7.set_categories(Reference(ws, min_col=50, min_row=40, max_row=39 + len(to2))); c7.legend = None
    add(c7, "B122", "十大流出对手（外部下游/供应商）")
    c8 = LineChart(); c8.add_data(Reference(S2, min_col=5, max_col=7, min_row=DYN_T0 - 1, max_row=DYN_TN), titles_from_data=True)
    c8.set_categories(Reference(S2, min_col=2, min_row=DYN_T0, max_row=DYN_TN))
    add(c8, "B142", "未来12个月期末现金推演（三情景）")

    # ============ Sheet5 风险指标 ============
    ws = S5
    put(ws, 1, 2, f"风险指标数据集 — {client}", font=TITLE_F); ws.row_dimensions[1].height = 30
    mtext(ws, 2, "依据 metrics-processing-spec-v1.md：数据清洗前置 → 反欺诈(AF) → 营销响应(MK) → 信用风险(LS)。"
                 "哨兵值 -9999表空 / -9986不足6月 / -9998外部数据缺失 / -9997无法复算，已黄色标注，"
                 "入模前须缺失化处理，切勿当数值参与统计。", span=4)
    GROUPS = [("CLN", "数据清洗质检"), ("AF", "反欺诈指标"), ("MK", "营销响应指标"), ("LS", "信用风险指标")]
    def group_of(code):
        for pfx, _ in GROUPS[:3]:
            if str(code).startswith(pfx): return pfx
        return "LS"
    r = 4
    for pfx, gname in GROUPS:
        items = [(c, n, v, lg) for (c, n, v, lg) in (spec_list or []) if group_of(c) == pfx]
        if not items: continue
        n_sent = sum(1 for c, n, v, lg in items if isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS)
        sect(ws, r, "", f"{gname}（{len(items)} 项，其中 {n_sent} 项因数据缺失/不足降级）"); r += 1
        for j, hx in enumerate(["指标编码", "指标名称", "指标值", "加工逻辑说明"], 2): put(ws, r, j, hx, bold=True, fillc=INH)
        r += 1
        for c, n, v, lg in items:
            is_sent = isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS
            put(ws, r, 2, c); put(ws, r, 3, n, wrap=True)
            put(ws, r, 4, v, fillc=YEL if is_sent else None, fmt="#,##0.0000" if isinstance(v, float) else None)
            put(ws, r, 5, lg, wrap=True, font=F(size=9, color="808080"))
            fit_row(ws, r, max(est_lines(n, colw(ws, 3)), est_lines(lg or "", colw(ws, 5), size=9)))
            r += 1
        r += 1

    # ============ Sheet6 标准化（合并）流水数据明细 ============
    ws = S6
    DETAIL_COLS = ["交易唯一编号", "客户名称", "主体名称", "账户类型", "本方名称", "本方账户", "开户行",
                   "交易时间", "对手名称", "对手账户", "收入金额", "支出金额", "交易金额", "账户余额",
                   "虚拟账户余额", "银行备注", "账户方附言", "收支方向", "一级标签", "二级标签", "三级标签",
                   "标签来源", "标签置信度", "命中关键词", "交易渠道", "来源文件名", "来源行号"]
    dcols = [c for c in DETAIL_COLS if c in df.columns]
    AMT_COLS = {"收入金额", "支出金额", "交易金额", "账户余额", "虚拟账户余额"}
    DW = {"交易唯一编号": 16, "交易时间": 19, "本方名称": 22, "对手名称": 26, "开户行": 22,
          "本方账户": 20, "对手账户": 20, "银行备注": 28, "账户方附言": 20, "来源文件名": 30, "命中关键词": 14}
    put(ws, 1, 1, "标准化（合并）流水数据明细表", font=TITLE_F); ws.row_dimensions[1].height = 30
    mtext(ws, 2, "标准化交付物「整合打标流水」的只读副本（含内部互转标记），供手工筛选分析，无需翻看上游产物；"
                 "行序沿用余额连续性还原结果，切勿按时间重排后回写。", span=8, col=1)
    hr = 4
    for j, c in enumerate(dcols + ["内部互转"], 1):
        put(ws, hr, j, c, bold=True, fillc=INH, wrap=True)
        ws.column_dimensions[get_column_letter(j)].width = DW.get(c, 13)
    ws.row_dimensions[hr].height = 28
    NUMF = Font(name="微软雅黑", size=10)
    amt_idx = {j for j, c in enumerate(dcols, 1) if c in AMT_COLS}
    for k, (_, x) in enumerate(df[dcols].iterrows()):
        rr = hr + 1 + k
        for j, c in enumerate(dcols, 1):
            v = x[c]
            if c == "交易时间": v = None if pd.isna(v) else str(v)
            elif isinstance(v, float) and v != v: v = None
            elif isinstance(v, (np.integer, np.floating)): v = float(v)
            cell = ws.cell(rr, j, v); cell.font = NUMF
            if j in amt_idx: cell.number_format = MF
        ws.cell(rr, len(dcols) + 1, "是" if bool(df["内部互转"].iloc[k]) else "").font = NUMF
    ws.freeze_panes = ws.cell(hr + 1, 1)
    ws.auto_filter.ref = f"A{hr}:{get_column_letter(len(dcols) + 1)}{hr + len(df)}"

    wb.save(out_path)
    return out_path

def main():
    ap = argparse.ArgumentParser(description="经营流水 BI 分析报告生成器 V4.0（见知风格）")
    ap.add_argument("--input", required=True); ap.add_argument("--client", default="客户")
    ap.add_argument("--out-dir", default=""); ap.add_argument("--whitelist", default="")
    ap.add_argument("--new-loan", default="", help="拟授信情景: 本金,年利率,期数 如 2000000,0.08,12")
    ap.add_argument("--loans", default="", help="借据表文件(csv/xlsx)，启用借据类风险指标；不传则降级为-9998")
    args = ap.parse_args()
    path = pick_input(args.input)
    whitelist = json.load(open(args.whitelist, encoding="utf-8")) if args.whitelist else {}
    whitelist = {k: tuple(v) for k, v in whitelist.items()}
    new_loan = NEW_LOAN
    if args.new_loan:
        p, r_, n = args.new_loan.split(","); new_loan = (float(p), float(r_), int(n))
    df, dbal, vchk = load_v4(path)
    df = prep(df)
    A = analyze(df, daily_balance(df, dbal), vchk, whitelist, new_loan)
    V = analyze_v4(df, A, vchk)
    loans = load_loans(args.loans) if args.loans else None
    wl_names = set(whitelist.keys()) if whitelist else None
    spec_list = compute_spec_metrics(spec_augment(df), loans=loans, whitelist=wl_names)
    out_dir = args.out_dir or os.path.dirname(os.path.abspath(path))
    out = os.path.join(out_dir, f"{args.client}__经营流水分析报告_V4.0.xlsx")
    build_workbook(A, V, df, args.client, out, whitelist, spec_list)
    n_sent = sum(1 for c, n, v, lg in spec_list if isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS)
    print(f"[OK] {out}")
    print(f"     质量分{A['q_total']}({A['q_grade']}) 账户{len(V['accounts'])}个(风险{V['n_risk_acct']}) "
          f"未提交账户{len(V['unsub'])}个 需关注对手{len(V['watch'])}个")
    print(f"     有效流入{V['eff_in']:,.0f}元 借贷流入{V['loan_mon']['借贷流入'].sum():,.0f}元 "
          f"识别债务{len(A['debts'])}笔 月债务服务{A['svc_base']:,.0f}元")
    print(f"     风险指标 {len(spec_list)} 项（{n_sent} 项降级）")
    print("     提示：①用 recalc 重算公式 ②补写「3.4主体核验/4.1生意模式/4.4评分卡」中的【待填】 ③口头汇报核心结论")

if __name__ == "__main__":
    main()
