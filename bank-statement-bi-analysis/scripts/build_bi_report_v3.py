#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""经营流水 BI 分析报告生成器 V3.0（BANKFLOW-BI-V3.0）

串行于 bank-statement-standardization：输入其标准化/整合/打标产物，输出一份
4-sheet 纵排版授信分析报告（分析报告 / 负债与现金流 / 数据明细 / 可视化看板）。

V3.0 相对 V1 的扩展（设计依据见 references/design-v3.md）：
  模块〇 数据可信度：四维质量评分、余额勾稽、结息复算(隐含利率跨季一致性)、
         Benford、整数占比、快进快出、缺失账户反推（租金无放款/担保费无贷款/薪酬失配）
  模块一 生意模式：行业证据自动汇集 + 定性结论待填（脚本产底稿，Claude/分析师补写）
  模块二 概览 + 审批看板 V3 扩展行（DSR/本息覆盖/安全垫/白名单/隐藏账户/申贷前集中融资）
  模块三 对手：名称归一化、白名单加权、双向对倒判定、HHI/留存/新增流失
  模块四 交易结构：经营/融资(标签修正)/往来三线 + 重大风险行为信号规则引擎
  模块五 负债(Debt OS)：债务清单(放款+月供反推APR) → 12月偿付Schedule → KPI → 气球观察
  模块六 现金流：历史DSCR + 未来12月三情景公式推演(蓝色输入可改) + 流水折算现金流量表
  模块七 行业评分卡：共性维度自动计分 + 行业维度证据填充、得分待人工标定
  模块八 交叉验证矩阵：流水列自动填充，其余数据源待补(占位)
  模块九 数据明细：月度时序/对手月度矩阵/大额交易/疑点清单

红线（沿用 V1）：不重排明细；只读不改；阈值是参考不是结论；缺数据如实标注。
脚本产出后建议用 LibreOffice/Excel 重算公式（如 xlsx 技能的 recalc.py）。

用法：
  python build_bi_report_v3.py --input "<标准化产物文件或文件夹>" --client 客户名
    [--out-dir DIR] [--whitelist wl.json] [--new-loan 2000000,0.08,12]
whitelist json: {"某某公司": ["L1", 1.0], ...}
"""
import argparse, glob, json, math, os, re, sys
try:
    import pandas as pd
    import numpy as np
except ImportError:
    sys.stderr.write("缺少 pandas，请先 `pip install pandas openpyxl`。\n"); raise
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference

# 复用 generate_vars 的指标加工层（AF 反欺诈 / MK 营销 / LS 信用风险，详见 metrics-processing-spec-v1.md）
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_vars import compute_spec_metrics, load_loans, nature as spec_nature, S_EMPTY, S_SHORT, S_NOEXT, S_NOCALC

SPEC_SENTINELS = {S_EMPTY, S_SHORT, S_NOEXT, S_NOCALC}  # 哨兵值：表空/不足6月/外部数据缺失/无法复算

STRATEGY_VERSION = "BANKFLOW-BI-V3.0"
SHEET_CANDIDATES = ["整合打标流水", "标准化流水主表", "主表", "打标流水", "整合流水", "流水"]
DAILY_BAL_SHEETS = ["组合日余额(虚拟账户)", "组合日余额", "日余额"]

THRESHOLDS = {
    "近12月流入下限": 20_000_000.0, "月流入CV上限": 0.5, "余额离差率上限": 1.0,
    "TOP1上限": 0.50, "TOP5上限": 0.80, "HHI上限": 0.18, "往来款占比上限": 0.30,
    "余额上穿日均下限": 0.35, "白名单加权占比下限": 0.30, "DSR上限": 0.15,
    "本息覆盖下限": 1.5, "安全垫天数下限": 7.0, "质量分下限": 70,
    "整数占比上限": 0.40, "快进快出上限": 0.10, "夜间占比上限": 0.20,
}
FIN_KW = "银行|金融租赁|融资租赁|小额贷款|保理|消费金融|信托|金租"
GUAR_KW = "担保"
JUD_KW = "法院|执行局|仲裁|国家税务稽查"
SPEC_KW = "期货|证券交易所|数字货币|虚拟币|博彩|彩票"
LEASE_HINT = "租金|租赁费|贷款|还款|月供"
DEFAULT_TERM = {"融资租赁": 24, "银行贷款": 24, "其他": 12}
NEW_LOAN = (2_000_000.0, 0.08, 12)   # 拟授信情景默认：本金/年利率/期数
PENDING = "【待填：定性结论，由分析师/Claude 复核补写】"

# ---------------------------------------------------------------- 输入
def pick_input(p):
    if os.path.isdir(p):
        cands = [f for f in glob.glob(os.path.join(p, "*")) if re.search(r"(打标|整合|标准化|待分析).*\.(xlsx|csv)$", f)]
        if not cands:
            cands = glob.glob(os.path.join(p, "*.xlsx")) + glob.glob(os.path.join(p, "*.csv"))
        if not cands:
            raise SystemExit(f"输入文件夹中未找到标准化产物: {p}")
        return sorted(cands)[0]
    return p

def load(path):
    """返回 (明细df, 日余额df|None, 余额校验df|None)"""
    if path.lower().endswith(".csv"):
        return pd.read_csv(path), None, None
    xl = pd.ExcelFile(path)
    sheet = next((s for c in SHEET_CANDIDATES for s in xl.sheet_names if c in s), xl.sheet_names[0])
    df = xl.parse(sheet)
    dbal = next((xl.parse(s) for c in DAILY_BAL_SHEETS for s in xl.sheet_names if c in s), None)
    vchk = next((xl.parse(s) for s in xl.sheet_names if "余额校验" in s), None)
    return df, dbal, vchk

def norm_cp(s):
    s = re.sub(r"\s+", "", str(s or ""))
    s = re.sub(r"[（(].*?[)）]", lambda m: m.group(0) if len(m.group(0)) > 8 else "", s)
    return s

def prep(df):
    for c in ["收入金额", "支出金额", "交易金额", "账户余额"]:
        if c in df.columns: df[c] = pd.to_numeric(df[c], errors="coerce")
    df["收入金额"] = df.get("收入金额", pd.Series(0, index=df.index)).fillna(0.0)
    df["支出金额"] = df.get("支出金额", pd.Series(0, index=df.index)).fillna(0.0)
    if (df["收入金额"] == 0).all() and (df["支出金额"] == 0).all() and "交易金额" in df.columns:
        df["收入金额"] = df["交易金额"].clip(lower=0); df["支出金额"] = (-df["交易金额"]).clip(lower=0)
    df["交易时间"] = pd.to_datetime(df["交易时间"], errors="coerce")
    df["月份"] = df["交易时间"].dt.strftime("%Y-%m")
    df["日期"] = df["交易时间"].dt.date
    df["对手名称"] = df.get("对手名称", "").fillna("").astype(str).str.strip()
    # 归一化：融资租赁公司常见「金融租赁/融资租赁」混写 → 统一 key
    df["对手key"] = df["对手名称"].map(norm_cp).str.replace("融资租赁股份有限公司", "金融租赁股份有限公司")
    for c in ["一级标签", "二级标签", "三级标签", "主体名称", "本方名称", "客户名称"]:
        if c not in df.columns: df[c] = ""
        df[c] = df[c].fillna("").astype(str)
    df["摘要"] = (df.get("银行备注", pd.Series("", index=df.index)).fillna("").astype(str) + " " +
                  df.get("账户方附言", pd.Series("", index=df.index)).fillna("").astype(str))
    # 关联主体互转：对手命中集团成员（主体名称/本方名称）
    members = (set(df["主体名称"]) | set(df["本方名称"])) - {""}
    df["内部互转"] = df["对手名称"].isin(members)
    return df

def spec_augment(df):
    """补齐 compute_spec_metrics 所需、而 v3 prep 未产出的派生列：
    距截止月数（近N月窗口）、对手性质（企业/个人判定）、对手账户（AF05，缺则空串降级）。"""
    d = df.copy()
    md = d["交易时间"].max()
    d["距截止月数"] = (md.year - d["交易时间"].dt.year) * 12 + md.month - d["交易时间"].dt.month + 1
    d["对手性质"] = d["对手名称"].apply(spec_nature)
    if "对手账户" not in d.columns:
        d["对手账户"] = ""
    return d

def daily_balance(df, dbal):
    if dbal is not None and len(dbal.columns) >= 2:
        d = dbal.copy(); d.columns = [str(c) for c in d.columns]
        d["日期"] = pd.to_datetime(d[d.columns[0]], errors="coerce")
        col = "合计余额" if "合计余额" in d.columns else d.columns[-1]
        return d.dropna(subset=["日期"]).set_index("日期")[col].astype(float)
    # 回退：按(账户,日)末行余额 → 前向填充 → 跨账户求和
    if "账户余额" not in df.columns: return None
    t = df.dropna(subset=["账户余额"]).copy()
    t["acct"] = t.get("本方账户", "ALL").fillna("ALL")
    last = t.groupby(["acct", "日期"])["账户余额"].last().unstack(0)
    idx = pd.date_range(min(t["日期"]), max(t["日期"]))
    last.index = pd.to_datetime(last.index)
    return last.reindex(idx).ffill().fillna(0).sum(axis=1)

def solve_apr(pv, pmt, n):
    if not pv or not pmt or pmt * n <= pv: return float("nan")
    lo, hi = 1e-6, 0.1
    for _ in range(100):
        mid = (lo + hi) / 2
        if pv * mid / (1 - (1 + mid) ** -n) > pmt: hi = mid
        else: lo = mid
    return mid * 12

def nature(n):
    if re.search(FIN_KW, n): return "金融机构"
    if re.search(r"局|国库|支库|财政", n): return "政府/国库"
    if re.search(r"公司|厂|中心|商行|合作社|集团|医院|学校", n): return "企业"
    return "个人" if 0 < len(n) <= 4 else "企业"

# ---------------------------------------------------------------- 分析
def analyze(df, dbal_series, vchk, whitelist, new_loan):
    A = {}
    ext = df[~df["内部互转"]]
    A["months"] = MONTHS = sorted(m for m in df["月份"].dropna().unique() if m)
    A["l12"] = L12 = MONTHS[-12:]
    A["tot_in"], A["tot_out"] = df["收入金额"].sum(), df["支出金额"].sum()
    A["int_in"] = df.loc[df["内部互转"], "收入金额"].sum(); A["int_out"] = df.loc[df["内部互转"], "支出金额"].sum()
    A["ext_in"], A["ext_out"] = ext["收入金额"].sum(), ext["支出金额"].sum()
    mon = ext.groupby("月份").agg(流入=("收入金额", "sum"), 流出=("支出金额", "sum"), 笔数=("收入金额", "size")).reindex(MONTHS, fill_value=0)
    mon["净流入"] = mon["流入"] - mon["流出"]
    bal = dbal_series
    if bal is not None:
        mb = bal.groupby(bal.index.strftime("%Y-%m")).agg(["last", "mean", "min"])
        mon = mon.join(mb.rename(columns={"last": "月末余额", "mean": "月均余额", "min": "月最低余额"}))
    else:
        for c in ["月末余额", "月均余额", "月最低余额"]: mon[c] = np.nan
    wl_mask = ext["二级标签"].str.contains("往来")
    A["wl_in"], A["wl_out"] = ext.loc[wl_mask, "收入金额"].sum(), ext.loc[wl_mask, "支出金额"].sum()
    op = ext[(ext["一级标签"] == "经营类") & (~ext["二级标签"].str.contains("往来"))]
    mon_op = op.groupby("月份").agg(经流入=("收入金额", "sum"), 经流出=("支出金额", "sum")).reindex(MONTHS, fill_value=0.0)
    mon_op["经净"] = mon_op["经流入"] - mon_op["经流出"]
    # 融资流入（标签 + 大额金融机构放款修正）
    fin_in = ext[(ext["一级标签"].isin(["筹资类", "投资类"])) | ((ext["收入金额"] >= 500_000) & ext["对手key"].str.contains(FIN_KW))]
    mon["融资流入(修正)"] = fin_in.groupby("月份")["收入金额"].sum().reindex(MONTHS, fill_value=0)
    mon["融资流出"] = ext[ext["一级标签"].isin(["筹资类", "投资类"])].groupby("月份")["支出金额"].sum().reindex(MONTHS, fill_value=0)
    mon["往来款流入"] = ext[wl_mask].groupby("月份")["收入金额"].sum().reindex(MONTHS, fill_value=0)
    mon["活跃对手数"] = ext[ext["对手key"] != ""].groupby("月份")["对手key"].nunique().reindex(MONTHS, fill_value=0)
    A["mon"], A["mon_op"] = mon, mon_op

    cp = ext[ext["对手key"] != ""].groupby("对手key").agg(
        对手=("对手名称", "first"), 流入=("收入金额", "sum"), 流出=("支出金额", "sum"),
        笔数=("收入金额", "size"), 首次=("月份", "min"), 末次=("月份", "max"))
    A["cp"] = cp
    A["hhi_in"] = ((cp["流入"] / max(cp["流入"].sum(), 1)) ** 2).sum()
    A["hhi_out"] = ((cp["流出"] / max(cp["流出"].sum(), 1)) ** 2).sum()
    both = cp[(cp["流入"] > 0) & (cp["流出"] > 0)].copy()
    both["平衡度"] = both[["流入", "流出"]].min(axis=1) / both[["流入", "流出"]].max(axis=1)
    A["both"] = both
    cpm = ext[(ext["对手key"] != "") & (ext["收入金额"] > 0)].groupby("对手key")["月份"].nunique()
    A["retain_in"] = cp.loc[cp.index.intersection(cpm[cpm >= 6].index), "流入"].sum() / max(A["ext_in"], 1)
    A["priv_ratio"] = ext[ext["对手名称"].map(nature) == "个人"][["收入金额", "支出金额"]].sum().sum() / max(A["tot_in"] + A["tot_out"], 1)
    wlm = {norm_cp(k): v for k, v in whitelist.items()}
    cp["白名单"] = [wlm.get(k, ("L0", 0))[0] for k in cp.index]
    A["wl_flow"] = sum(cp.loc[k, "流入"] * wlm[k][1] for k in cp.index if k in wlm)
    A["wl_share"] = A["wl_flow"] / max(A["ext_in"], 1)

    # ---- 真实性 ----
    A["bal_breaks"] = None
    if vchk is not None and "余额断点" in vchk.columns:
        A["bal_breaks"] = int(pd.to_numeric(vchk["余额断点"], errors="coerce").fillna(0).sum())
    jx = df[(df["对手名称"] == "") & (df["收入金额"] > 0) & (df["收入金额"] < 1000) &
            ((df["摘要"].str.contains("结息|利息")) | ((df["交易时间"].dt.day.isin([20, 21])) & (df["交易时间"].dt.month.isin([3, 6, 9, 12]))))]
    jx = jx[~jx["摘要"].str.contains("冲正|手续费")]
    jx_list = []
    if bal is not None:
        for _, r in jx.iterrows():
            end = r["交易时间"].normalize(); start = end - pd.DateOffset(months=3) + pd.Timedelta(days=1)
            acc = bal[(bal.index >= start) & (bal.index <= end)].sum()
            jx_list.append((end.strftime("%Y-%m-%d"), float(r["收入金额"]), round(float(acc)), round(r["收入金额"] * 360 / acc * 100, 4) if acc > 0 else float("nan")))
    A["jx_list"] = jx_list
    imp = [x[3] for x in jx_list if x[3] == x[3]]
    A["jx_mean"] = float(np.mean(imp)) if imp else float("nan")
    A["jx_cv"] = float(np.std(imp) / np.mean(imp)) if len(imp) >= 2 else float("nan")
    amts = df["收入金额"].add(df["支出金额"]).abs(); amts = amts[amts >= 10]
    fdg = amts.astype(str).str.lstrip("0.").str[0]
    fdg = fdg[fdg.str.isdigit()].astype(int)
    ben = fdg.value_counts(normalize=True).reindex(range(1, 10), fill_value=0)
    A["ben_mad"] = float((ben - pd.Series({d: math.log10(1 + 1 / d) for d in range(1, 10)})).abs().mean())
    A["int_ratio"] = float(((amts[amts >= 10000] % 10000) == 0).mean()) if (amts >= 10000).any() else 0.0
    day = df.groupby("日期").agg(i=("收入金额", "sum"), o=("支出金额", "sum"))
    day = day.reindex(pd.date_range(day.index.min(), day.index.max()).date, fill_value=0.0)
    onx = day["o"].shift(-1).fillna(0); bid = day[day["i"] >= 100_000]
    A["fastio"] = float((((bid["o"] + onx.loc[bid.index]) / bid["i"]) >= 0.8).mean()) if len(bid) else 0.0
    hh = df["交易时间"].dt.hour
    has_t = df["交易时间"].dt.strftime("%H:%M:%S") != "00:00:00"
    nightmask = ((hh >= 21) | (hh < 6)) & has_t
    A["night_n"] = int(nightmask.sum()); A["night_ratio"] = float(nightmask.sum() / max(has_t.sum(), 1))
    A["chz"] = df[df["摘要"].str.contains("冲正|退票")]
    A["cp_fill"] = float((df["对手名称"] != "").mean()); A["rem_fill"] = float((df["摘要"].str.strip() != "").mean())

    # ---- 负债识别（Debt OS）----
    debts, hidden = [], []
    fin_rows = ext[ext["对手key"].str.contains(FIN_KW)]
    for k, g in fin_rows.groupby("对手key"):
        name = g["对手名称"].iloc[0]
        loans = g[g["收入金额"] >= 200_000]
        pays = g[(g["支出金额"] > 0) & ((g["摘要"].str.contains(LEASE_HINT)) | (g["三级标签"].str.contains("偿还")) | (g["二级标签"] == "负债"))]
        pmt = 0.0
        if len(pays) >= 2 and pays["支出金额"].std() / max(pays["支出金额"].mean(), 1) < 0.1:
            pmt = float(pays["支出金额"].median())
        elif len(pays) == 1:
            pmt = float(pays["支出金额"].iloc[0])
        principal = float(loans["收入金额"].sum())
        if "票据" in "".join(g["二级标签"].unique()):
            debts.append(dict(lender=name, prod="票据贴现", principal=principal, apr=float("nan"), pmt=0.0, term=0,
                              balloon=0.0, src=f"{len(loans)}笔票据入账", note="贴现融资，无后续偿付；关注票据池"))
            continue
        prod = "融资租赁" if re.search("租赁", name) else "银行贷款"
        term = DEFAULT_TERM.get(prod, 12)
        if principal > 0 and pmt > 0:
            debts.append(dict(lender=name, prod=prod, principal=principal, apr=solve_apr(principal, pmt, term), pmt=pmt,
                              term=term, balloon=0.0, src="流水放款+月供实测", note=f"期限{term}期为假设(黄)"))
        elif pmt > 0:
            debts.append(dict(lender=name, prod=prod + "(隐藏账户)", principal=pmt * 12, apr=float("nan"), pmt=pmt,
                              term=12, balloon=0.0, src="仅见月供、放款不在本账户", note="本金=月供×12估算；触发缺失账户反推"))
            hidden.append([f"存在他行账户承接「{name}」放款", f"每期支付{pmt:,.0f}元共{len(pays)}笔，无对应放款入账", "高", "要求补充合同与放款账户流水"])
        elif principal > 0:
            debts.append(dict(lender=name, prod=prod, principal=principal, apr=float("nan"), pmt=0.0, term=term,
                              balloon=0.0, src="仅见放款、未见还款", note="还款方式待核（或经他行归还）"))
    guar = ext[(ext["支出金额"] > 0) & ext["对手key"].str.contains(GUAR_KW)]
    if len(guar):
        fee = guar["支出金额"].sum()
        hidden.append([f"存在他行贷款（担保公司佐证，{len(guar)}笔担保费）",
                       f"担保费合计{fee:,.0f}元且无对应放款入账；按~1%费率估对应贷款约{fee*100:,.0f}元", "高", "调征信核实；补他行流水"])
        debts.append(dict(lender="他行贷款(担保公司佐证)", prod="银行贷款(待核)", principal=float("nan"), apr=float("nan"),
                          pmt=0.0, term=0, balloon=0.0, src=f"担保费{fee:,.0f}元无对应放款", note=f"估对应贷款约{fee*100:,.0f}元(待核)"))
    mj = ext[ext["二级标签"].str.contains("民间借贷")]
    if len(mj):
        pos = (mj.groupby("对手key").agg(借入=("收入金额", "sum"), 归还=("支出金额", "sum")))
        outstanding = float((pos["借入"] - pos["归还"]).clip(lower=0).sum())
        if outstanding > 0:
            names = "、".join(pos[(pos["借入"] - pos["归还"]) > 0].index[:4])
            debts.append(dict(lender=f"民间借贷({names})", prod="民间借贷", principal=outstanding, apr=float("nan"),
                              pmt=0.0, term=0, balloon=outstanding, src="借入-归还累计净敞口", note="假设6个月后到期一次还本(黄)"))
    salary = ext[(ext["二级标签"].str.contains("人力")) & (ext["支出金额"] > 0)]["支出金额"].sum()
    priv_hi = A["cp"][(A["cp"]["笔数"] >= 100) & (A["cp"].index.map(nature) == "个人")]
    if salary < A["ext_out"] * 0.02 and len(priv_hi):
        hidden.append(["存在个人卡收付通道", f"账面薪酬{salary:,.0f}元与经营规模不匹配；对私高频对手：" +
                       "、".join(f"{r['对手']}({int(r['笔数'])}笔/流出{r['流出']:,.0f})" for _, r in priv_hi.head(2).iterrows()),
                       "中", "要求提供相关个人卡流水"])
    A["debts"], A["hidden"] = debts, hidden
    A["svc_base"] = float(sum(d["pmt"] for d in debts))
    A["mj"] = mj

    # ---- 风险信号 ----
    sig = []
    m_last = pd.Period(MONTHS[-1]); recent3 = [str(m_last - i) for i in range(3)]
    new_fin = float(mon.loc[[m for m in recent3 if m in mon.index], "融资流入(修正)"].sum())
    if new_fin > 0:
        n_fin = len(fin_rows[(fin_rows["收入金额"] >= 200_000) & (fin_rows["月份"].isin(recent3))])
        score = 75 if (new_fin >= 1_000_000 and n_fin >= 2) else 40
        sig.append(["多头借贷/申贷前集中融资", score, f"近3个月新增融资{new_fin:,.0f}元({n_fin}笔机构放款)", "要求出具全部融资清单与征信授权"])
    if len(mj):
        sig.append(["民间借贷", 70 if mj[["收入金额", "支出金额"]].sum().sum() >= 1_000_000 else 40,
                    f"借入{mj['收入金额'].sum():,.0f}/归还{mj['支出金额'].sum():,.0f}元，对手{mj['对手key'].nunique()}人", "倒推利率；逐笔核实用途"])
    if A["fastio"] > THRESHOLDS["快进快出上限"]:
        sig.append(["资金过桥/快进快出", 55, f"大额入账日T+1流出≥80%占比{A['fastio']:.0%}", "核对过桥时点与他行贷款到期日"])
    if len(priv_hi):
        r0 = priv_hi.iloc[0]
        sig.append(["对私大额通道", 50, f"{r0['对手']}：{int(r0['笔数'])}笔/流出{r0['流出']:,.0f}元", "核实身份与个人卡流水"])
    jud = ext[ext["对手key"].str.contains(JUD_KW)]
    sig.append(["司法执行", 60 if len(jud) else 0, f"司法类对手交易{len(jud)}笔" if len(jud) else "无法院/执行局/仲裁对手", "逐笔核实" if len(jud) else "—"])
    spec = ext[ext["对手key"].str.contains(SPEC_KW)]
    sig.append(["金融投机/赌博", 60 if len(spec) else 5, f"命中{len(spec)}笔" if len(spec) else "无期货/数字货币/博彩类对手", "一票否决候选" if len(spec) else "—"])
    A["signals"] = sig

    # ---- 质量评分 ----
    s_acct = max(5, 30 - 5 * len(hidden) - (0 if len(df.get("本方账户", pd.Series()).unique()) > 1 else 5))
    cover = len([m for m in L12 if m in mon.index and mon.loc[m, "笔数"] > 0]) / max(len(L12), 1)
    s_time = round(25 * cover) - (1 if df["交易时间"].min().day > 5 else 0)
    s_field = round(20 * (0.4 * A["cp_fill"] + 0.3 * A["rem_fill"] + 0.3))
    s_cons = 25 - (5 if (A["bal_breaks"] or 0) > 0 else 0) - (2 if len(A["chz"]) else 0) - (3 if (A["jx_cv"] == A["jx_cv"] and A["jx_cv"] > 0.5) else 0)
    A["qscore"] = (s_acct, s_time, s_field, s_cons)
    A["q_total"] = int(s_acct + s_time + s_field + s_cons)
    A["q_grade"] = "优" if A["q_total"] >= 85 else ("良" if A["q_total"] >= 70 else ("中" if A["q_total"] >= 50 else "差"))

    # ---- CFS ----
    def cfs_vals(dfx):
        dfx = dfx[~dfx["内部互转"]]
        zy_in = dfx.loc[dfx["二级标签"] == "主营业务", "收入金额"].sum()
        jz_fk = dfx.loc[(dfx["收入金额"] >= 500_000) & dfx["对手key"].str.contains(FIN_KW), "收入金额"].sum()
        borrow = dfx.loc[dfx["二级标签"].str.contains("民间借贷|票据|筹资|负债"), "收入金额"].sum() + jz_fk
        invb = dfx.loc[dfx["二级标签"].str.contains("出借|投资"), "收入金额"].sum()
        o_in = dfx["收入金额"].sum() - zy_in - borrow - invb
        cg = dfx.loc[dfx["二级标签"] == "主营业务", "支出金额"].sum()
        xz = dfx.loc[dfx["二级标签"].str.contains("人力|保险"), "支出金额"].sum()
        tax = dfx.loc[dfx["二级标签"].str.contains("税"), "支出金额"].sum()
        rp = dfx.loc[dfx["二级标签"].str.contains("民间借贷|负债|筹资"), "支出金额"].sum()
        invo = dfx.loc[dfx["二级标签"].str.contains("出借|投资"), "支出金额"].sum()
        o_out = dfx["支出金额"].sum() - cg - xz - tax - rp - invo
        return [zy_in, o_in, cg, xz, tax, o_out, invb, invo, borrow, rp]
    years = sorted({m[:4] for m in MONTHS})
    A["cfs_cols"] = [(y, cfs_vals(df[df["月份"].str.startswith(y)])) for y in years] + [("区间合计", cfs_vals(df))]
    first = df.iloc[0]
    A["opening_cash"] = float(first.get("账户余额", 0) or 0) + float(first["支出金额"]) - float(first["收入金额"]) if "账户余额" in df.columns else float("nan")
    A["closing_cash"] = float(df["账户余额"].dropna().iloc[-1]) if "账户余额" in df.columns and df["账户余额"].notna().any() else float("nan")
    A["big_tx"] = ext[(ext["收入金额"] >= 500_000) | (ext["支出金额"] >= 500_000)].sort_values("交易时间")
    A["new_loan"] = new_loan
    A["bal"] = bal
    return A

# ---------------------------------------------------------------- 写报告
MF, PF = "#,##0.00", "0.0%"
def build_workbook(A, df, client, out_path, whitelist, spec_list=None):
    wb = Workbook(); wb.remove(wb.active)
    def F(**kw):
        kw.setdefault("size", 10); kw.setdefault("name", "微软雅黑"); return Font(**kw)
    HDRF = PatternFill("solid", start_color="1F4E79"); SECF = PatternFill("solid", start_color="D9E1F2")
    WARN = PatternFill("solid", start_color="FFF2CC"); REDF = PatternFill("solid", start_color="FCE4E4")
    GRN = PatternFill("solid", start_color="E2EFDA"); YEL = PatternFill("solid", start_color="FFFF00")
    BLUEF = Font(name="微软雅黑", size=10, color="0000FF")
    thin = Border(*[Side("thin", color="BFBFBF")] * 4)
    def put(ws, r, c, v, bold=False, fillc=None, fmt=None, font=None, wrap=False):
        cell = ws.cell(r, c, v); cell.font = font or F(bold=bold)
        if fillc: cell.fill = fillc
        if fmt: cell.number_format = fmt
        cell.alignment = Alignment(vertical="center", wrap_text=wrap); cell.border = thin
        return cell
    def title(ws, r, text, ncol=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        put(ws, r, 1, text, fillc=HDRF, font=F(bold=True, color="FFFFFF", size=12))
    def section(ws, r, text, ncol=8):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        put(ws, r, 1, text, bold=True, fillc=SECF)
    def table(ws, r0, headers, rows, fmts=None):
        for j, hx in enumerate(headers, 1): put(ws, r0, j, hx, bold=True, fillc=SECF)
        for i, row in enumerate(rows):
            for j, v in enumerate(row, 1):
                fmt = fmts[j - 1] if fmts else None
                put(ws, r0 + 1 + i, j, v, fmt=fmt if isinstance(v, (int, float)) and not isinstance(v, bool) else None, wrap=True)
        return r0 + 1 + len(rows)
    def nxt(ws, gap=2): return ws.max_row + gap if ws.max_row > 1 else 1

    MONTHS, L12, mon, mon_op, cp = A["months"], A["l12"], A["mon"], A["mon_op"], A["cp"]
    ext_in, ext_out = A["ext_in"], A["ext_out"]
    fwd = pd.period_range(pd.Period(MONTHS[-1]) + 1, periods=12, freq="M").astype(str).tolist()

    S1 = wb.create_sheet("分析报告")
    for i, w in enumerate([26, 22, 18, 17, 17, 17, 13, 42], 1): S1.column_dimensions[get_column_letter(i)].width = w
    S2 = wb.create_sheet("负债与现金流")
    for i, w in enumerate([26, 16, 16, 14, 12, 12, 16, 12, 12, 12, 30, 12, 12], 1): S2.column_dimensions[get_column_letter(i)].width = w
    S3 = wb.create_sheet("数据明细")
    for i, w in enumerate([34] + [12] * (len(MONTHS)), 1): S3.column_dimensions[get_column_letter(i)].width = w
    S4 = wb.create_sheet("可视化看板"); S4.column_dimensions["A"].width = 2
    S5 = wb.create_sheet("风险指标")
    for i, w in enumerate([16, 32, 18, 74], 1): S5.column_dimensions[get_column_letter(i)].width = w

    # ===== S1 =====
    ws = S1
    subjects = "、".join(sorted(set(df["主体名称"]) - {""})) or "、".join(sorted(set(df["本方名称"]) - {""}))
    title(ws, 1, f"{client} · 经营流水分析报告 V3.0（关联主体合并 · 单表纵排版）")
    r = 2
    for a, b, c, d in [("客户名称", client, "报告日期", pd.Timestamp.today().strftime("%Y-%m-%d")),
                       ("关联主体", subjects, "分析策略", STRATEGY_VERSION),
                       ("流水区间", f"{df['交易时间'].min()} ~ {df['交易时间'].max()}（{len(MONTHS)}个月）", "内部互转",
                        f"流入{A['int_in']:,.0f}/流出{A['int_out']:,.0f}元（已从对外口径剔除）")]:
        put(ws, r, 1, a, bold=True); ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4); put(ws, r, 2, b, wrap=True)
        put(ws, r, 5, c, bold=True); ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8); put(ws, r, 6, d, wrap=True); r += 1

    r = nxt(ws); title(ws, r, "模块〇 数据可信度与流水真实性"); r += 1
    section(ws, r, "〇-1 数据质量四维评分（前提分：分数过低时全报告结论降级）"); r += 1
    sa, st, sf, sc_ = A["qscore"]
    r = table(ws, r, ["维度", "权重", "得分", "扣分原因", "", "", "", ""], [
        ["账户覆盖", 30, sa, f"隐藏账户信号{len(A['hidden'])}组（见〇-3）", "", "", "", ""],
        ["时段覆盖", 25, st, "近12月覆盖率与首月完整性", "", "", "", ""],
        ["字段完整性", 20, sf, f"对手非空率{A['cp_fill']:.0%}；摘要非空率{A['rem_fill']:.0%}", "", "", "", ""],
        ["一致性校验", 25, sc_, f"余额断点{A['bal_breaks'] if A['bal_breaks'] is not None else '未提供校验表'}；冲正{len(A['chz'])}笔；结息复算见下", "", "", "", ""]])
    put(ws, r + 1, 1, "总分", bold=True, fillc=WARN); put(ws, r + 1, 2, f"=SUM(C{r-4}:C{r-1})", bold=True, fillc=WARN)
    put(ws, r + 1, 3, f"{A['q_total']} 分 → {A['q_grade']}", bold=True, fillc=WARN)
    put(ws, r + 1, 4, "分数过低时：收入/负债结论按「已见账户口径」理解，先补件再审批", fillc=WARN, wrap=True)
    r = nxt(ws); section(ws, r, "〇-2 真实性校验结果"); r += 1
    jxs = f"隐含年化{A['jx_mean']:.3f}%±CV {A['jx_cv']:.0%}" if A["jx_mean"] == A["jx_mean"] else "未识别到结息记录"
    r = table(ws, r, ["校验项", "结果", "数值", "说明", "", "", "", ""], [
        ["余额连续性", "通过" if not A["bal_breaks"] else "异常", f"断点{A['bal_breaks'] if A['bal_breaks'] is not None else '—'}", "沿用上游标准化校验", "", "", "", ""],
        ["结息复算(隐含利率)", "通过" if (A["jx_cv"] == A["jx_cv"] and A["jx_cv"] <= 0.5) else "关注", jxs, "伪造流水难以维持跨季一致的隐含利率", "", "", "", ""],
        ["Benford首位分布", "关注" if A["ben_mad"] > 0.015 else "通过", f"MAD={A['ben_mad']:.3f}", "固定金额(租金/月供)为可解释偏离", "", "", "", ""],
        ["整数金额占比(≥1万)", "关注" if A["int_ratio"] > THRESHOLDS["整数占比上限"] else "通过", f"{A['int_ratio']:.1%}", "过高需交叉发票", "", "", "", ""],
        ["快进快出", "关注" if A["fastio"] > THRESHOLDS["快进快出上限"] else "通过", f"{A['fastio']:.0%}", "大额入账日T+1流出≥80%天数占比", "", "", "", ""],
        ["夜间交易", "关注" if A["night_ratio"] > THRESHOLDS["夜间占比上限"] else "通过", f"{A['night_n']}笔/{A['night_ratio']:.1%}", "分母仅含有时点交易", "", "", "", ""]])
    if A["jx_list"]:
        r = nxt(ws); section(ws, r, "结息复算明细"); r += 1
        r = table(ws, r, ["结息日", "实际结息(元)", "季度日积数(元·日)", "隐含年化(%)", "", "", "", ""],
                  [[a, b, c, d, "", "", "", ""] for a, b, c, d in A["jx_list"]], fmts=[None, MF, "#,##0", "0.0000", None, None, None, None])
    r = nxt(ws); section(ws, r, "〇-3 缺失账户反推"); r += 1
    hid = A["hidden"] or [["未发现隐藏账户信号", "—", "—", "—"]]
    r = table(ws, r, ["推断", "证据", "置信度", "建议动作", "", "", "", ""], [h + ["", "", "", ""] for h in hid])

    r = nxt(ws); title(ws, r, "模块一 生意模式（脚本汇集证据，定性结论待复核补写）"); r += 1
    ind = cp.sort_values("流入", ascending=False).head(5)
    outd = cp.sort_values("流出", ascending=False).head(5)
    r = table(ws, r, ["要素", "自动汇集证据", "定性结论", "", "", "", "", ""], [
        ["所属行业", "主体：" + subjects + "；TOP流入对手：" + "、".join(ind["对手"]), PENDING, "", "", "", "", ""],
        ["经营品类(上游采购)", "TOP流出对手：" + "、".join(outd["对手"]), PENDING, "", "", "", "", ""],
        ["盈利逻辑", f"经营流入(剔往来){mon_op['经流入'].sum():,.0f} vs 经营流出{mon_op['经流出'].sum():,.0f}；现金依赖0核对；票据/收单占比见标签", PENDING, "", "", "", "", ""],
        ["刚性成本锚点", f"薪酬{df.loc[df['二级标签'].str.contains('人力'), '支出金额'].sum():,.0f}／税费{df.loc[df['二级标签'].str.contains('税'), '支出金额'].sum():,.0f}／运营{df.loc[df['二级标签'].str.contains('运营'), '支出金额'].sum():,.0f}元", "锚点强弱判断：" + PENDING, "", "", "", "", ""],
        ["核心风险", "自动信号见模块四；集中度/隐藏账户/申贷前融资见看板", PENDING, "", "", "", "", ""],
        ["准入反欺诈滤网", "需工商+征信+进件数据（法人任职/法代变更/代持特征/空壳）", "待外部数据", "", "", "", "", ""]])

    r = nxt(ws); title(ws, r, "模块二 流水概览与审批要件看板"); r += 1
    section(ws, r, "PART 1 流水整体概况（对外口径=剔除关联互转）"); r += 1
    l12v = mon.loc[[m for m in L12 if m in mon.index]]
    b1y = A["bal"][A["bal"].index >= A["bal"].index.max() - pd.Timedelta(days=365)] if A["bal"] is not None else None
    ov = [("总交易笔数", int(len(df)), "对私交易金额占比", A["priv_ratio"]),
          ("总流入(含互转,元)", A["tot_in"], "总流出(含互转,元)", A["tot_out"]),
          ("内部互转流入(元)", A["int_in"], "内部互转流出(元)", A["int_out"]),
          ("净对外流入(元)", ext_in, "净对外流出(元)", ext_out),
          ("近12月对外流入(元)", l12v["流入"].sum(), "收支比(对外)", ext_out / max(ext_in, 1)),
          ("近12月经营流入-剔往来款(元)", mon_op.loc[[m for m in L12 if m in mon_op.index], "经流入"].sum(), "往来款收入占比", A["wl_in"] / max(A["tot_in"], 1)),
          ("期初余额(元)", A["opening_cash"], "期末余额(元)", A["closing_cash"]),
          ("近一年日均余额(元)", float(b1y.mean()) if b1y is not None else float("nan"), "最高余额(元)", float(A["bal"].max()) if A["bal"] is not None else float("nan"))]
    for a, b, c, d in ov:
        put(ws, r, 1, a, bold=True); put(ws, r, 2, b, fmt=MF if isinstance(b, float) else "#,##0")
        put(ws, r, 5, c, bold=True); put(ws, r, 6, d, fmt=PF if isinstance(d, float) and abs(d) < 10 else MF); r += 1
    r = nxt(ws); section(ws, r, "PART 2 收支结构（按二级用途，对外口径）"); r += 1
    ext = df[~df["内部互转"]]
    inc = ext[ext["收入金额"] > 0].groupby("二级标签")["收入金额"].sum().sort_values(ascending=False)
    exp = ext[ext["支出金额"] > 0].groupby("二级标签")["支出金额"].sum().sort_values(ascending=False)
    for j, hx in enumerate(["收入用途", "金额", "占比", "", "支出用途", "金额", "占比", ""], 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    r0 = r + 1
    for i in range(max(len(inc), len(exp))):
        if i < len(inc): put(ws, r0 + i, 1, inc.index[i]); put(ws, r0 + i, 2, float(inc.iloc[i]), fmt=MF); put(ws, r0 + i, 3, float(inc.iloc[i] / max(ext_in, 1)), fmt=PF)
        if i < len(exp): put(ws, r0 + i, 5, exp.index[i]); put(ws, r0 + i, 6, float(exp.iloc[i]), fmt=MF); put(ws, r0 + i, 7, float(exp.iloc[i] / max(ext_out, 1)), fmt=PF)
    r = r0 + max(len(inc), len(exp)) + 1
    section(ws, r, "PART 3 审批要件指标看板（V3.0；阈值为参考口径，红/黄/绿不构成授信结论）"); r += 1
    cv12 = float(l12v["流入"].std() / max(l12v["流入"].mean(), 1))
    mad = float((b1y - b1y.mean()).abs().mean() / max(b1y.mean(), 1)) if b1y is not None else float("nan")
    above = float((b1y > b1y.mean()).mean()) if b1y is not None else float("nan")
    top_in = cp.sort_values("流入", ascending=False)
    svc = A["svc_base"]; opnet = mon_op.loc[[m for m in L12 if m in mon_op.index], "经净"]
    minbal3 = mon.loc[[m for m in L12[-3:] if m in mon.index], "月最低余额"].mean()
    kpis = [
        ["业务规模", "近12月对外流入", "剔互转", f"≥{THRESHOLDS['近12月流入下限']/1e4:,.0f}万", l12v["流入"].sum(), "通过" if l12v["流入"].sum() >= THRESHOLDS["近12月流入下限"] else "关注"],
        ["关联交易", "内部互转占比", "互转额/总收支", "≤30%", (A["int_in"] + A["int_out"]) / max(A["tot_in"] + A["tot_out"], 1), "通过" if (A["int_in"]+A["int_out"])/max(A["tot_in"]+A["tot_out"],1) <= .3 else "关注"],
        ["稳定性", "月流入变异系数", "近12月CV", "≤0.5", cv12, "关注" if cv12 > .5 else "通过"],
        ["稳定性", "余额离差率", "日余额MAD/日均", "≤1.0", mad, "关注" if mad == mad and mad > 1 else "通过"],
        ["集中度", "TOP1对手流入占比", "", "≤50%", float(top_in["流入"].iloc[0] / max(ext_in, 1)) if len(top_in) else 0, "关注" if len(top_in) and top_in["流入"].iloc[0] / max(ext_in, 1) > .5 else "通过"],
        ["集中度", "TOP5对手流入占比", "", "≤80%", float(top_in["流入"].head(5).sum() / max(ext_in, 1)), "关注" if top_in["流入"].head(5).sum() / max(ext_in, 1) > .8 else "通过"],
        ["集中度", "流入HHI", "Σ份额²", "≤0.18", A["hhi_in"], "关注" if A["hhi_in"] > .18 else "通过"],
        ["资金往来", "往来款收入占比", "", "≤30%", A["wl_in"] / max(A["tot_in"], 1), "关注" if A["wl_in"] / max(A["tot_in"], 1) > .3 else "通过"],
        ["余额健康", "余额上穿日均占比", "近一年", "≥35%", above, "关注" if above == above and above < .35 else "通过"],
        ["★V3 白名单", "白名单加权流入占比", "样例名单可配置", "≥30%", A["wl_share"], "通过" if A["wl_share"] >= .3 else "关注"],
        ["★V3 负债", "存量月债务服务(元)", "识别月供合计", "—", svc, "关注" if svc else "通过"],
        ["★V3 负债", "DSR还款负担率", "月债务服务/月均对外流入", "≤15%", svc / max(l12v["流入"].mean(), 1), "关注" if svc / max(l12v["流入"].mean(), 1) > .15 else "通过"],
        ["★V3 现金流", "本息覆盖倍数(存量)", "月均经营净CF/月债务服务", "≥1.5", (opnet.mean() / svc) if svc else float("nan"), "高危" if svc and opnet.mean() / svc < 1 else "关注"],
        ["★V3 现金流", "安全垫天数", "近3月月最低余额均值/日均对外流出", "≥7", minbal3 / max(l12v["流出"].sum() / 365, 1), "高危" if minbal3 / max(l12v["流出"].sum() / 365, 1) < 3 else ("关注" if minbal3 / max(l12v["流出"].sum() / 365, 1) < 7 else "通过")],
        ["★V3 可信度", "数据质量总分", "四维加权", "≥70", A["q_total"], "关注" if A["q_total"] < 70 else "通过"],
        ["★V3 可信度", "隐藏账户信号组数", "缺失账户反推", "=0", len(A["hidden"]), "高危" if len(A["hidden"]) >= 2 else ("关注" if A["hidden"] else "通过")]]
    r = table(ws, r, ["类别", "指标", "口径", "参考阈值", "企业实际", "判断", "", ""], [k + ["", ""] for k in kpis],
              fmts=[None, None, None, None, "#,##0.00", None, None, None])
    for i in range(len(kpis)):
        cell = ws.cell(r - len(kpis) + i + 1, 6); v = str(cell.value)
        cell.fill = GRN if v == "通过" else (REDF if "高危" in v else WARN)

    r = nxt(ws); title(ws, r, "模块三 交易对手（名称已归一化）"); r += 1
    section(ws, r, "一、十大流入对手（上游/客户）"); r += 1
    rows = [[x["对手"], float(x["流入"]), float(x["流入"] / max(ext_in, 1)), int(x["笔数"]), nature(x["对手"]), x["首次"], x["末次"], x["白名单"]]
            for _, x in cp.sort_values("流入", ascending=False).head(10).iterrows()]
    r = table(ws, r, ["对手", "流入金额", "占比", "笔数", "性质", "首次月", "末次月", "白名单"], rows, fmts=[None, MF, PF, "#,##0", None, None, None, None])
    r = nxt(ws); section(ws, r, "二、十大流出对手（下游/供应商）"); r += 1
    rows = [[x["对手"], float(x["流出"]), float(x["流出"] / max(ext_out, 1)), int(x["笔数"]), nature(x["对手"]), x["首次"], x["末次"], ""]
            for _, x in cp.sort_values("流出", ascending=False).head(10).iterrows()]
    r = table(ws, r, ["对手", "流出金额", "占比", "笔数", "性质", "首次月", "末次月", ""], rows, fmts=[None, MF, PF, "#,##0", None, None, None, None])
    r = nxt(ws); section(ws, r, "三、白名单匹配（名单可经 --whitelist 配置；负面名单筛查待外部API）"); r += 1
    wrow = [[k, v[0], v[1], float(cp.loc[norm_cp(k), "流入"]) if norm_cp(k) in cp.index else 0.0, "", "", "", ""] for k, v in whitelist.items()]
    wrow.append(["合计（加权流入占比）", "", "", A["wl_flow"], f"{A['wl_share']:.1%}", "", "", ""])
    r = table(ws, r, ["白名单对手", "等级", "权重", "流入", "备注", "", "", ""], wrow, fmts=[None, None, "0.0", MF, None, None, None, None])
    r = nxt(ws); section(ws, r, "四、双向往来对手与对倒判定（平衡度≥0.7且双边≥50万→疑似关联外部方）"); r += 1
    rows = []
    for k, x in A["both"].sort_values("平衡度", ascending=False).head(10).iterrows():
        flag = "疑似关联外部方" if (x["平衡度"] >= 0.7 and min(x["流入"], x["流出"]) >= 500_000) else ("正常产业双向" if nature(x["对手"]) == "企业" else "关注")
        rows.append([x["对手"], float(x["流入"]), float(x["流出"]), float(x["平衡度"]), nature(x["对手"]), flag, "", ""])
    r = table(ws, r, ["对手", "流入", "流出", "平衡度", "性质", "判定", "", ""], rows, fmts=[None, MF, MF, "0.00", None, None, None, None])
    r = nxt(ws); section(ws, r, "五、对手结构与留存"); r += 1
    r = table(ws, r, ["指标", "数值", "说明", "", "", "", "", ""], [
        ["活跃对手数(月均)", float(mon["活跃对手数"].mean()), "", "", "", "", "", ""],
        ["流入HHI / 流出HHI", f"{A['hhi_in']:.3f} / {A['hhi_out']:.3f}", ">0.18 高集中", "", "", "", "", ""],
        ["老对手(≥6个月)留存流入占比", A["retain_in"], "客户粘性", "", "", "", "", ""],
        ["对私交易金额占比", A["priv_ratio"], "过高需核实", "", "", "", "", ""]], fmts=[None, "#,##0.000", None, None, None, None, None, None])

    r = nxt(ws); title(ws, r, "模块四 交易结构与重大风险行为信号"); r += 1
    section(ws, r, "一、经营/融资/往来三线月度趋势"); r += 1
    for j, hx in enumerate(["月份", "经营流入(剔往来款)", "融资流入(修正)", "往来款流入", "经营流出", "融资流出", "净流入", ""], 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    r0 = r + 1
    for i, m in enumerate(MONTHS):
        put(ws, r0 + i, 1, m)
        for j, v in enumerate([mon_op.loc[m, "经流入"], mon.loc[m, "融资流入(修正)"], mon.loc[m, "往来款流入"], mon_op.loc[m, "经流出"], mon.loc[m, "融资流出"], mon.loc[m, "净流入"]], 2):
            put(ws, r0 + i, j, float(v), fmt=MF)
    r = r0 + len(MONTHS) + 1
    section(ws, r, "二、重大风险行为信号（疑似度+证据链；≥60分强制人工定性）"); r += 1
    r = table(ws, r, ["信号", "疑似度", "证据链", "建议核查动作", "", "", "", ""], [s + ["", "", "", ""] for s in A["signals"]],
              fmts=[None, "#,##0", None, None, None, None, None, None])
    for i in range(len(A["signals"])):
        v = A["signals"][i][1]; ws.cell(r - len(A["signals"]) + i + 1, 2).fill = REDF if v >= 60 else (WARN if v >= 30 else GRN)

    r = nxt(ws); title(ws, r, "模块七 行业专家评分（共性维度自动计分；行业维度证据已填、得分待人工标定）"); r += 1
    auto = [("共性|收入真实性", 10, None, f"结息复算{('通过' if A['jx_cv']==A['jx_cv'] and A['jx_cv']<=0.5 else '关注')}；往来款占比{A['wl_in']/max(A['tot_in'],1):.0%}；隐藏账户{len(A['hidden'])}组；整数占比{A['int_ratio']:.0%}"),
            ("共性|负债负担", 10, None, f"月债务服务{svc:,.0f}元；DSR {svc/max(l12v['流入'].mean(),1):.0%}"),
            ("共性|现金流健康", 10, None, f"近12月经营净CF {opnet.sum():,.0f}元；期末余额{A['closing_cash']:,.0f}元"),
            ("行业|订单与回款质量", 20, None, f"白名单加权流入占比{A['wl_share']:.0%}"),
            ("行业|采购规律性", 15, None, "TOP流出对手连续性见对手月度矩阵"),
            ("行业|刚性成本可见度", 15, None, "薪酬/税费/运营支出规模见模块一"),
            ("行业|产能与投入", 10, None, "设备/租赁融资见负债清单"),
            ("行业|集中度容忍", 10, None, f"TOP1 {float(top_in['流入'].iloc[0]/max(ext_in,1)) if len(top_in) else 0:.0%}")]
    r = table(ws, r, ["维度", "满分", "得分(待标定)", "证据(自动)", "", "", "", ""], [[a, b, PENDING if c is None else c, d, "", "", "", ""] for a, b, c, d in auto],
              fmts=[None, "#,##0", None, None, None, None, None, None])
    put(ws, r + 1, 1, "评分卡分型（制造/批发/零售/服务/出口）与维度权重见 references/design-v3.md；由生意模式判定后路由。", font=F(italic=True), wrap=True)

    r = nxt(ws); title(ws, r, "模块八 交叉验证矩阵（流水列自动填充；其余数据源待补）"); r += 1
    mx = [["收入规模", f"近12月对外流入{l12v['流入'].sum():,.0f}元", "⚪销项发票", "⚪应税收入", "—", "⚪中标数据", "C(单源)"],
          ["采购成本", f"主营采购{ext.loc[ext['二级标签']=='主营业务','支出金额'].sum():,.0f}元", "⚪进项发票", "⚪财报", "—", "—", "C(单源)"],
          ["负债结构", f"识别月供{svc:,.0f}元+隐藏信号{len(A['hidden'])}组", "—", "—", "⚪征信在贷", "⚪司法纠纷", "D(待征信)"],
          ["对手真实性", "TOP10清单见模块三", "⚪发票购销方", "—", "—", "⚪工商在营", "C(单源)"],
          ["行业判断", "证据见模块一", "⚪发票品目", "⚪税种结构", "—", "⚪登记行业", "C(单源)"]]
    r = table(ws, r, ["核心结论", "流水", "发票", "税务", "征信", "工商/司法", "可信度", ""], [m + [""] for m in mx])

    # ===== S2 负债与现金流 =====
    ws = S2
    title(ws, 1, "模块五 负债结构（Debt OS：清单→Schedule→KPI；黄色=假设待确认）", 13)
    r = 2; section(ws, r, "一、债务清单（从流水识别+反推）", 13); r += 1
    for j, hx in enumerate(["出借方", "产品", "本金(元)", "APR(反推)", "月供(元)", "假设期数", "气球尾款", "识别来源", "备注"], 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    r0 = r + 1
    for i, d in enumerate(A["debts"]):
        put(ws, r0 + i, 1, d["lender"]); put(ws, r0 + i, 2, d["prod"])
        put(ws, r0 + i, 3, d["principal"] if d["principal"] == d["principal"] else "待核", fmt=MF)
        put(ws, r0 + i, 4, d["apr"] if d["apr"] == d["apr"] else "待核", fmt="0.0%")
        put(ws, r0 + i, 5, d["pmt"], fmt=MF); put(ws, r0 + i, 6, d["term"] or "—")
        put(ws, r0 + i, 7, d["balloon"], fmt=MF); put(ws, r0 + i, 8, d["src"], wrap=True); put(ws, r0 + i, 9, d["note"], wrap=True)
        if "假设" in d["note"] or "估算" in d["note"]: ws.cell(r0 + i, 6).fill = YEL
    r = r0 + max(len(A["debts"]), 1) + 1
    section(ws, r, "二、债务组合KPI", 13); r += 1
    known = [d for d in A["debts"] if d["principal"] == d["principal"] and d["apr"] == d["apr"]]
    wair = sum(d["principal"] * d["apr"] for d in known) / max(sum(d["principal"] for d in known), 1) if known else float("nan")
    balloon12 = sum(d["balloon"] for d in A["debts"])
    r = table(ws, r, ["KPI", "数值", "口径"] + [""] * 10, [
        ["总在贷本金(可见口径,元)", sum(d["principal"] for d in A["debts"] if d["principal"] == d["principal"]), "含估算项；他行待核不含"] + [""] * 10,
        ["WAIR加权平均利率", wair, "仅含可反推APR的债务"] + [""] * 10,
        ["月债务服务(元)", A["svc_base"], "识别出的月供合计"] + [""] * 10,
        ["12个月内气球尾款(元)", balloon12, "到期一次还本类"] + [""] * 10],
        fmts=[None, MF, None] + [None] * 10)
    r = nxt(ws); section(ws, r, "三、未来12个月偿付Schedule（元/月；气球计入到期月）", 13); r += 1
    for j, hx in enumerate(["债务"] + fwd, 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    r0 = r + 1
    sch_rows = [d for d in A["debts"] if d["pmt"] > 0 or d["balloon"] > 0]
    for i, d in enumerate(sch_rows):
        put(ws, r0 + i, 1, d["lender"][:16])
        for j in range(12):
            v = d["pmt"] + (d["balloon"] if (d["balloon"] > 0 and j == 5) else 0)
            put(ws, r0 + i, 2 + j, float(v if d["pmt"] > 0 or j == 5 else 0), fmt="#,##0")
    SVC_ROW = r0 + max(len(sch_rows), 1)
    put(ws, SVC_ROW, 1, "月合计", bold=True, fillc=WARN)
    for j in range(12):
        cl = get_column_letter(2 + j)
        put(ws, SVC_ROW, 2 + j, f"=SUM({cl}{r0}:{cl}{SVC_ROW-1})", fmt="#,##0", fillc=WARN)
    r = SVC_ROW + 2

    title(ws, r, "模块六 现金流健康（历史+未来12个月三情景推演；蓝色=可修改假设）", 13); r += 1
    section(ws, r, "一、历史月度经营净现金流与实付债务服务", 13); r += 1
    for j, hx in enumerate(["月份", "经营净CF(剔往来款)", "债务服务(实付)", "DSCR", "月最低余额", "月末余额"], 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    svc_hist = df[df["三级标签"].str.contains("偿还") | (df["二级标签"] == "负债")].groupby("月份")["支出金额"].sum().reindex(MONTHS, fill_value=0)
    r0 = r + 1
    for i, m in enumerate(MONTHS):
        put(ws, r0 + i, 1, m); put(ws, r0 + i, 2, float(mon_op.loc[m, "经净"]), fmt=MF); put(ws, r0 + i, 3, float(svc_hist[m]), fmt=MF)
        put(ws, r0 + i, 4, f'=IF(C{r0+i}=0,"-",B{r0+i}/C{r0+i})', fmt="0.00")
        put(ws, r0 + i, 5, float(mon.loc[m, "月最低余额"]) if mon.loc[m, "月最低余额"] == mon.loc[m, "月最低余额"] else "—", fmt=MF)
        put(ws, r0 + i, 6, float(mon.loc[m, "月末余额"]) if mon.loc[m, "月末余额"] == mon.loc[m, "月末余额"] else "—", fmt=MF)
    r = r0 + len(MONTHS) + 1
    section(ws, r, "二、未来12个月现金推演（Debt OS引擎）", 13); r += 1
    A0 = r
    wl_net = float((mon.loc[[m for m in L12 if m in mon.index], "往来款流入"].sum() - df.loc[df["二级标签"].str.contains("往来") & df["月份"].isin(L12), "支出金额"].sum()) / max(len(L12), 1))
    P, RATE, N = A["new_loan"]
    inputs = [("期初现金(元)", A["closing_cash"] if A["closing_cash"] == A["closing_cash"] else 0.0),
              ("月经营流入基线(近12月均,剔往来款,元)", float(mon_op.loc[[m for m in L12 if m in mon_op.index], "经流入"].mean())),
              ("月经营流出基线(元)", float(mon_op.loc[[m for m in L12 if m in mon_op.index], "经流出"].mean())),
              ("往来款月净流入(元)", wl_net), ("最低安全垫(元)", 200000.0),
              ("拟新增授信本金(元)", P), ("新增授信年利率", RATE), ("新增授信期数(月)", N)]
    for i, (k, v) in enumerate(inputs):
        put(ws, A0 + i, 1, k, bold=True); c = put(ws, A0 + i, 2, v, fmt=MF if abs(v) > 1 else "0.0%"); c.font = BLUEF
    IN = {k: f"$B${A0+i}" for i, (k, v) in enumerate(inputs)}
    r = A0 + len(inputs)
    put(ws, r, 1, "新增授信月供(元)", bold=True)
    put(ws, r, 2, f'=ROUND(-PMT({IN["新增授信年利率"]}/12,{IN["新增授信期数(月)"]},{IN["拟新增授信本金(元)"]}),0)', fmt="#,##0", fillc=WARN)
    NEWPMT = f"$B${r}"; r += 2
    for j, hx in enumerate(["月份", "存量债务服务", "总债务服务(含新增)", "基线:期末现金", "收入-20%:期末现金", "收入-30%:期末现金", "基线DSCR"], 1): put(ws, r, j, hx, bold=True, fillc=SECF)
    T0 = r + 1
    for i, m in enumerate(fwd):
        rr = T0 + i; put(ws, rr, 1, m)
        put(ws, rr, 2, f"={get_column_letter(2+i)}{SVC_ROW}", fmt="#,##0")
        put(ws, rr, 3, f"=B{rr}+{NEWPMT}", fmt="#,##0")
        add = f'+{IN["拟新增授信本金(元)"]}' if i == 0 else ""
        for cix, mult_i, mult_o in [(4, "", ""), (5, "*0.8", "*0.9"), (6, "*0.7", "*0.85")]:
            prev = IN["期初现金(元)"] if i == 0 else f"{get_column_letter(cix)}{rr-1}"
            put(ws, rr, cix, f'={prev}+{IN["月经营流入基线(近12月均,剔往来款,元)"]}{mult_i}-{IN["月经营流出基线(元)"]}{mult_o}+{IN["往来款月净流入(元)"]}-C{rr}{add}', fmt="#,##0")
        put(ws, rr, 7, f'=({IN["月经营流入基线(近12月均,剔往来款,元)"]}-{IN["月经营流出基线(元)"]}+{IN["往来款月净流入(元)"]})/C{rr}', fmt="0.00")
    TN = T0 + 11; r = TN + 2
    r = table(ws, r, ["推演KPI", "基线", "收入-20%", "收入-30%", "说明"] + [""] * 8, [
        ["最低期末现金(元)", f"=MIN(D{T0}:D{TN})", f"=MIN(E{T0}:E{TN})", f"=MIN(F{T0}:F{TN})", ""] + [""] * 8,
        ["触底次数", f'=COUNTIF(D{T0}:D{TN},"<"&{IN["最低安全垫(元)"]})', f'=COUNTIF(E{T0}:E{TN},"<"&{IN["最低安全垫(元)"]})', f'=COUNTIF(F{T0}:F{TN},"<"&{IN["最低安全垫(元)"]})', ""] + [""] * 8,
        ["所需增量融资(元)", f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(D{T0}:D{TN}))', f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(E{T0}:E{TN}))', f'=MAX(0,{IN["最低安全垫(元)"]}-MIN(F{T0}:F{TN}))', "直接回答「该批多少/缺口多大」"] + [""] * 8,
        ["DSCR<1.2月份数(基线)", f'=COUNTIF(G{T0}:G{TN},"<1.2")', "", "", "含新增月供口径"] + [""] * 8],
        fmts=[None, "#,##0", "#,##0", "#,##0", None] + [None] * 8)
    DYN_T0, DYN_TN = T0, TN
    put(ws, r + 1, 1, "注：机械外推样稿。蓝色输入可改；经营口径受标签质量与账户覆盖影响，补数据后重算。", font=F(italic=True), wrap=True)

    r = nxt(ws); title(ws, r, "附表 流水折算现金流量表（会小企03口径；金融机构大额放款已修正归入筹资）", 13); r += 1
    items = ["销售产成品、商品、提供劳务收到的现金", "收到其他与经营活动有关的现金", "购买原材料、商品、接受劳务支付的现金", "支付的职工薪酬", "支付的税费", "支付其他与经营活动有关的现金",
             "收回投资收到的现金", "投资支付的现金", "取得借款收到的现金", "偿还借款本金及利息支付的现金"]
    put(ws, r, 1, "项目", bold=True, fillc=SECF)
    for cix, (label, _) in enumerate(A["cfs_cols"], 2): put(ws, r, cix, label, bold=True, fillc=SECF)
    lay = [("一、经营活动", None), (items[0], 0), (items[1], 1), (items[2], 2), (items[3], 3), (items[4], 4), (items[5], 5), ("  经营净额", "OP"),
           ("二、投资活动", None), (items[6], 6), (items[7], 7), ("  投资净额", "INV"),
           ("三、筹资活动", None), (items[8], 8), (items[9], 9), ("  筹资净额", "FIN"), ("四、现金净增加额", "NET")]
    rowmap = {}; r0 = r + 1
    for i, (name, idx) in enumerate(lay):
        rr = r0 + i; rowmap[idx] = rr
        put(ws, rr, 1, name, bold=isinstance(idx, str) or idx is None, fillc=SECF if idx is None else None)
        if isinstance(idx, int):
            for cix, (_, vals) in enumerate(A["cfs_cols"], 2): put(ws, rr, cix, float(vals[idx]), fmt=MF)
    for cix in range(2, 2 + len(A["cfs_cols"])):
        cl = get_column_letter(cix)
        put(ws, rowmap["OP"], cix, f"={cl}{rowmap[0]}+{cl}{rowmap[1]}-{cl}{rowmap[2]}-{cl}{rowmap[3]}-{cl}{rowmap[4]}-{cl}{rowmap[5]}", fmt=MF, bold=True)
        put(ws, rowmap["INV"], cix, f"={cl}{rowmap[6]}-{cl}{rowmap[7]}", fmt=MF, bold=True)
        put(ws, rowmap["FIN"], cix, f"={cl}{rowmap[8]}-{cl}{rowmap[9]}", fmt=MF, bold=True)
        put(ws, rowmap["NET"], cix, f"={cl}{rowmap['OP']}+{cl}{rowmap['INV']}+{cl}{rowmap['FIN']}", fmt=MF, bold=True, fillc=WARN)
    put(ws, rowmap["NET"] + 2, 1, f"勾稽：区间净增加额应≈期末-期初（{A['closing_cash']:,.2f} - {A['opening_cash']:,.2f}，含互转自轧差）。", font=F(italic=True), wrap=True)

    # ===== S3 数据明细 =====
    ws = S3
    ncol = 1 + len(MONTHS)
    title(ws, 1, "模块九 流水数据明细（月度时序 / 对手月度矩阵 / 大额与疑点清单）", min(ncol, 19))
    r = 3; TS_H = r
    hdr = ["月份", "流入金额", "流出金额", "净流入", "笔数", "月末余额", "月均余额", "月最低余额", "经营流入(剔往来)", "融资流入(修正)", "往来款流入", "活跃对手数"]
    for j, hx in enumerate(hdr, 1): put(ws, r, j, hx, bold=True, fillc=HDRF, font=F(bold=True, color="FFFFFF"))
    for i, m in enumerate(MONTHS):
        rr = r + 1 + i
        vals = [m, mon.loc[m, "流入"], mon.loc[m, "流出"], mon.loc[m, "净流入"], int(mon.loc[m, "笔数"]), mon.loc[m, "月末余额"], mon.loc[m, "月均余额"],
                mon.loc[m, "月最低余额"], mon_op.loc[m, "经流入"], mon.loc[m, "融资流入(修正)"], mon.loc[m, "往来款流入"], int(mon.loc[m, "活跃对手数"])]
        for j, v in enumerate(vals, 1):
            if isinstance(v, float) and v != v: v = None
            put(ws, rr, j, float(v) if isinstance(v, (int, float, np.floating)) and j > 1 else v, fmt=MF if j in (2, 3, 4, 6, 7, 8, 9, 10, 11) else None)
    TS_N = TS_H + len(MONTHS)
    def cp_mx(direction):
        colv = "收入金额" if direction == "in" else "支出金额"
        tops = cp.sort_values("流入" if direction == "in" else "流出", ascending=False).head(12).index
        piv = df[df["对手key"].isin(tops) & ~df["内部互转"]].pivot_table(index="对手key", columns="月份", values=colv, aggfunc="sum", fill_value=0)
        return piv.reindex(tops).reindex(columns=MONTHS, fill_value=0)
    r = nxt(ws); section(ws, r, "对手月度金额时序（TOP12×月；看突然出现/消失/仅申贷前活跃）", min(ncol, 19)); r += 1
    for tag, mx in [("流入对手", cp_mx("in")), ("流出对手", cp_mx("out"))]:
        put(ws, r, 1, tag, bold=True, fillc=SECF)
        for j, m in enumerate(MONTHS, 2): put(ws, r, j, m, bold=True, fillc=SECF)
        r0 = r + 1
        for i, (k, row) in enumerate(mx.iterrows()):
            put(ws, r0 + i, 1, cp.loc[k, "对手"])
            for j, m in enumerate(MONTHS, 2): put(ws, r0 + i, j, float(row[m]), fmt="#,##0")
        r = r0 + len(mx) + 1
    section(ws, r, f"大额交易明细（单笔≥50万，{len(A['big_tx'])}笔）", min(ncol, 19)); r += 1
    rows = [[str(x["交易时间"]), "收入" if x["收入金额"] > 0 else "支出", x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["二级标签"], x["三级标签"], x["摘要"][:60]] for _, x in A["big_tx"].iterrows()]
    r = table(ws, r, ["时间", "方向", "对手", "收入", "支出", "二级标签", "三级标签", "摘要"], rows, fmts=[None, None, None, MF, MF, None, None, None])
    r = nxt(ws); section(ws, r, "疑点交易清单（民间借贷/冲正；供人工核查工单）", min(ncol, 19)); r += 1
    sus = [["民间借贷", str(x["交易时间"])[:10], x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["三级标签"]] for _, x in A["mj"].iterrows()]
    sus += [["冲正/退票", str(x["交易时间"])[:10], x["对手名称"], float(x["收入金额"]), float(x["支出金额"]), x["摘要"][:40]] for _, x in A["chz"].head(15).iterrows()]
    r = table(ws, r, ["类型", "日期", "对手", "收入", "支出", "说明"], sus or [["无", "", "", "", "", ""]], fmts=[None, None, None, MF, MF, None])

    # ===== S4 图表 =====
    ws = S4
    put(ws, 1, 1, "可视化看板（自上而下下拉查看）", bold=True, font=F(bold=True, size=13))
    def add(ch, anchor, t, h=9, w=22):
        ch.title = t; ch.height = h; ch.width = w; ws.add_chart(ch, anchor)
    c1 = BarChart(); c1.type = "col"
    c1.add_data(Reference(S3, min_col=2, max_col=3, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c1.set_categories(Reference(S3, min_col=1, min_row=TS_H + 1, max_row=TS_N))
    l1 = LineChart(); l1.add_data(Reference(S3, min_col=4, max_col=4, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c1.y_axis.majorGridlines = None; c1 += l1
    add(c1, "B2", "月度流入/流出与净流入")
    c2 = LineChart(); c2.add_data(Reference(S3, min_col=6, max_col=8, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c2.set_categories(Reference(S3, min_col=1, min_row=TS_H + 1, max_row=TS_N)); add(c2, "B22", "余额波动（月末/月均/月最低）")
    c3 = LineChart(); c3.add_data(Reference(S3, min_col=9, max_col=11, min_row=TS_H, max_row=TS_N), titles_from_data=True)
    c3.set_categories(Reference(S3, min_col=1, min_row=TS_H + 1, max_row=TS_N)); add(c3, "B42", "经营 vs 融资 vs 往来款 月度流入")
    for i, (k, v) in enumerate(inc.head(8).items()): ws.cell(1 + i, 50, k); ws.cell(1 + i, 51, float(v))
    c4 = PieChart(); c4.add_data(Reference(ws, min_col=51, min_row=1, max_row=max(1, min(8, len(inc)))))
    c4.set_categories(Reference(ws, min_col=50, min_row=1, max_row=max(1, min(8, len(inc))))); add(c4, "B62", "收入用途结构", h=9, w=13)
    for i, (k, v) in enumerate(exp.head(8).items()): ws.cell(1 + i, 53, k); ws.cell(1 + i, 54, float(v))
    c5 = PieChart(); c5.add_data(Reference(ws, min_col=54, min_row=1, max_row=max(1, min(8, len(exp)))))
    c5.set_categories(Reference(ws, min_col=53, min_row=1, max_row=max(1, min(8, len(exp))))); add(c5, "L62", "支出用途结构", h=9, w=13)
    ti = cp.sort_values("流入", ascending=False).head(10); to = cp.sort_values("流出", ascending=False).head(10)
    for i, (_, x) in enumerate(ti.iterrows()): ws.cell(20 + i, 50, x["对手"]); ws.cell(20 + i, 51, float(x["流入"]))
    c6 = BarChart(); c6.type = "bar"; c6.add_data(Reference(ws, min_col=51, min_row=20, max_row=19 + len(ti)))
    c6.set_categories(Reference(ws, min_col=50, min_row=20, max_row=19 + len(ti))); c6.legend = None
    add(c6, "B82", "十大流入对手（外部上游/客户）")
    for i, (_, x) in enumerate(to.iterrows()): ws.cell(32 + i, 50, x["对手"]); ws.cell(32 + i, 51, float(x["流出"]))
    c7 = BarChart(); c7.type = "bar"; c7.add_data(Reference(ws, min_col=51, min_row=32, max_row=31 + len(to)))
    c7.set_categories(Reference(ws, min_col=50, min_row=32, max_row=31 + len(to))); c7.legend = None
    add(c7, "B102", "十大流出对手（外部下游/供应商）")
    c8 = LineChart(); c8.add_data(Reference(S2, min_col=4, max_col=6, min_row=DYN_T0 - 1, max_row=DYN_TN), titles_from_data=True)
    c8.set_categories(Reference(S2, min_col=1, min_row=DYN_T0, max_row=DYN_TN))
    add(c8, "B122", "未来12个月期末现金推演（三情景）")

    # ===== S5 风险指标（AF 反欺诈 / MK 营销响应 / LS 信用风险）=====
    ws = S5
    title(ws, 1, f"风险指标数据集 — {client}", ncol=4)
    put(ws, 2, 1, "依据 metrics-processing-spec-v1.md：数据清洗前置 → 反欺诈(AF) → 营销响应(MK) → 信用风险(LS)。"
                  "哨兵值 -9999表空 / -9986不足6月 / -9998外部数据缺失(需借据表·白名单·币种等) / -9997无法复算，"
                  "已黄色标注，入模前须缺失化处理，切勿当数值参与统计。", wrap=True)
    ws.merge_cells("A2:D2"); ws.row_dimensions[2].height = 42
    GROUPS = [("CLN", "数据清洗质检"), ("AF", "反欺诈指标"), ("MK", "营销响应指标"), ("LS", "信用风险指标")]
    def group_of(code):
        for pfx, _ in GROUPS[:3]:
            if str(code).startswith(pfx):
                return pfx
        return "LS"  # LS* 及中文码（隐债/年均/余额/借据类）归信用风险
    r = 4
    for pfx, gname in GROUPS:
        items = [(c, n, v, lg) for (c, n, v, lg) in (spec_list or []) if group_of(c) == pfx]
        if not items:
            continue
        n_sent = sum(1 for c, n, v, lg in items if isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS)
        section(ws, r, f"{gname}（{len(items)} 项，其中 {n_sent} 项因数据缺失/不足降级）", ncol=4); r += 1
        for j, hx in enumerate(["指标编码", "指标名称", "指标值", "加工逻辑说明"], 1):
            put(ws, r, j, hx, bold=True, fillc=SECF)
        r += 1
        for c, n, v, lg in items:
            is_sent = isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS
            put(ws, r, 1, c)
            put(ws, r, 2, n, wrap=True)
            put(ws, r, 3, v, fillc=WARN if is_sent else None,
                fmt="#,##0.0000" if isinstance(v, float) else None)
            put(ws, r, 4, lg, wrap=True, font=F(size=9, color="808080"))
            r += 1
        r += 1

    wb.save(out_path)
    return out_path

def main():
    ap = argparse.ArgumentParser(description="经营流水 BI 分析报告生成器 V3.0")
    ap.add_argument("--input", required=True); ap.add_argument("--client", default="客户")
    ap.add_argument("--out-dir", default=""); ap.add_argument("--whitelist", default="")
    ap.add_argument("--new-loan", default="", help="拟授信情景: 本金,年利率,期数 如 2000000,0.08,12")
    ap.add_argument("--loans", default="", help="借据表文件(csv/xlsx)，启用风险指标 MK09/12/14/15 及借据类；不传则降级为-9998")
    args = ap.parse_args()
    path = pick_input(args.input)
    whitelist = json.load(open(args.whitelist, encoding="utf-8")) if args.whitelist else {}
    whitelist = {k: tuple(v) for k, v in whitelist.items()}
    new_loan = NEW_LOAN
    if args.new_loan:
        p, r_, n = args.new_loan.split(","); new_loan = (float(p), float(r_), int(n))
    df, dbal, vchk = load(path)
    df = prep(df)
    A = analyze(df, daily_balance(df, dbal), vchk, whitelist, new_loan)
    # 风险指标层：白名单 dict → 名称集合；借据表可选
    loans = load_loans(args.loans) if args.loans else None
    wl_names = set(whitelist.keys()) if whitelist else None
    spec_list = compute_spec_metrics(spec_augment(df), loans=loans, whitelist=wl_names)
    out_dir = args.out_dir or os.path.dirname(os.path.abspath(path))
    out = os.path.join(out_dir, f"{args.client}__经营流水分析报告_V3.0.xlsx")
    build_workbook(A, df, args.client, out, whitelist, spec_list)
    n_sent = sum(1 for c, n, v, lg in spec_list if isinstance(v, (int, float)) and not isinstance(v, bool) and v in SPEC_SENTINELS)
    print(f"[OK] {out}")
    print(f"     质量分{A['q_total']}({A['q_grade']}) 隐藏账户信号{len(A['hidden'])}组 识别债务{len(A['debts'])}笔 月债务服务{A['svc_base']:,.0f}元")
    print(f"     风险指标 {len(spec_list)} 项（{n_sent} 项因数据缺失/不足降级；借据类需 --loans，白名单类需 --whitelist）")
    print("     提示：①用 recalc 重算公式 ②补写「生意模式/评分卡」中的【待填】定性单元格 ③白名单经 --whitelist 配置")

if __name__ == "__main__":
    main()
