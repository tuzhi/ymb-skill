"""统一 ChartBundle 的 Excel 与 ECharts 渲染适配器。"""

from __future__ import annotations

from collections.abc import Mapping
import json
from typing import Any

from openpyxl.chart import BarChart, LineChart, PieChart, Reference  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from .models import ChartBundle, ChartSpec


_COLORS = {
    "inflow": "5B9BD5",
    "outflow": "ED7D31",
    "net": "70AD47",
    "operating": "5B9BD5",
    "financing": "FFC000",
    "related_party": "A5A5A5",
    "forecast": "4472C4",
    "neutral": "5B9BD5",
}

_EXCEL_LAYOUT = {
    "monthly_cash_flow": ("B2", 9, 22),
    "monthly_balance": ("B22", 9, 22),
    "monthly_inflow_composition": ("B42", 9, 22),
    "monthly_loan_trend": ("B62", 9, 22),
    "external_income_structure": ("B82", 9, 13),
    "external_expense_structure": ("L82", 9, 13),
    "top_inflow_counterparties": ("B102", 9, 22),
    "top_outflow_counterparties": ("B122", 9, 22),
    "future_cash_projection": ("B142", 9, 22),
}


def render_echarts(
    bundle: ChartBundle,
    *,
    source_statement_run_id: str = "",
) -> dict[str, Any]:
    """将统一图表模型转成不含函数、可严格 JSON 序列化的 ECharts 5 配置。"""
    charts = []
    for chart in bundle.charts:
        if chart.kind == "pie":
            option = _echarts_pie(chart)
        elif chart.kind == "horizontal_bar":
            option = _echarts_horizontal_bar(chart)
        else:
            option = _echarts_cartesian(chart)
        charts.append({
            "chart_id": chart.chart_id,
            "title": chart.title,
            "unit": chart.unit,
            "metadata": dict(chart.metadata),
            "option": option,
        })
    payload = {
        "schema_version": bundle.schema_version,
        "strategy_version": bundle.strategy_version,
        "echarts_version": "5",
        "source_statement_run_id": source_statement_run_id,
        "currency": bundle.currency,
        "charts": charts,
        "omitted_charts": [
            {"chart_id": item.chart_id, "reason": item.reason}
            for item in bundle.omitted_charts
        ],
        "warnings": list(bundle.warnings),
    }
    json.dumps(payload, ensure_ascii=False, allow_nan=False)
    return payload


def render_excel_charts(
    worksheet: Any,
    bundle: ChartBundle,
    *,
    live_references: Mapping[str, Mapping[str, Any]] | None = None,
) -> int:
    """从同一 ChartBundle 创建 Excel 原生图表，并返回生成数量。"""
    references = live_references or {}
    hidden_column = 50
    generated = 0
    for chart in bundle.charts:
        source = references.get(chart.chart_id)
        if source:
            source_sheet = source["worksheet"]
            header_row = int(source["header_row"])
            first_row = int(source["first_row"])
            last_row = int(source["last_row"])
            category_column = int(source["category_col"])
            series_columns = tuple(int(value) for value in source["series_cols"])
        else:
            source_sheet = worksheet
            header_row = 1
            first_row = 2
            last_row = first_row + len(chart.categories) - 1
            category_column = hidden_column
            series_columns = tuple(
                hidden_column + index + 1 for index in range(len(chart.series))
            )
            _write_chart_source(
                source_sheet,
                chart,
                category_column,
                series_columns,
                header_row,
                first_row,
            )
            hidden_column += len(chart.series) + 2

        if len(series_columns) != len(chart.series) or last_row < first_row:
            continue
        excel_chart = _excel_chart(
            chart,
            source_sheet,
            category_column,
            series_columns,
            header_row,
            first_row,
            last_row,
        )
        if excel_chart is None:
            continue
        anchor, height, width = _EXCEL_LAYOUT.get(chart.chart_id, ("B2", 9, 22))
        excel_chart.title = chart.title
        excel_chart.height = height
        excel_chart.width = width
        worksheet.add_chart(excel_chart, anchor)
        generated += 1
    return generated


def _echarts_cartesian(chart: ChartSpec) -> dict[str, Any]:
    return {
        "title": {"text": chart.title},
        "tooltip": {"trigger": "axis"},
        "legend": {"data": [series.name for series in chart.series]},
        "xAxis": {"type": "category", "data": list(chart.categories)},
        "yAxis": {"type": "value", "name": chart.unit},
        "series": [
            {
                "name": series.name,
                "type": series.chart_type,
                "data": list(series.values),
                "itemStyle": {"color": _COLORS.get(series.role, _COLORS["neutral"])},
                **({"smooth": False} if series.chart_type == "line" else {}),
            }
            for series in chart.series
        ],
    }


def _echarts_pie(chart: ChartSpec) -> dict[str, Any]:
    series = chart.series[0]
    return {
        "title": {"text": chart.title},
        "tooltip": {"trigger": "item"},
        "legend": {"type": "scroll", "orient": "vertical", "right": 0},
        "series": [{
            "name": chart.title,
            "type": "pie",
            "radius": ["35%", "65%"],
            "data": [
                {"name": name, "value": value}
                for name, value in zip(chart.categories, series.values)
            ],
        }],
    }


def _echarts_horizontal_bar(chart: ChartSpec) -> dict[str, Any]:
    series = chart.series[0]
    return {
        "title": {"text": chart.title},
        "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
        "xAxis": {"type": "value", "name": chart.unit},
        "yAxis": {"type": "category", "data": list(chart.categories)},
        "series": [{
            "name": series.name,
            "type": "bar",
            "data": list(series.values),
            "itemStyle": {"color": _COLORS.get(series.role, _COLORS["neutral"])},
        }],
    }


def _write_chart_source(
    worksheet: Any,
    chart: ChartSpec,
    category_column: int,
    series_columns: tuple[int, ...],
    header_row: int,
    first_row: int,
) -> None:
    worksheet.cell(header_row, category_column, "分类")
    worksheet.column_dimensions[get_column_letter(category_column)].hidden = True
    for index, name in enumerate(chart.categories):
        worksheet.cell(first_row + index, category_column, name)
    for column, series in zip(series_columns, chart.series):
        worksheet.cell(header_row, column, series.name)
        worksheet.column_dimensions[get_column_letter(column)].hidden = True
        for index, value in enumerate(series.values):
            worksheet.cell(first_row + index, column, value)


def _excel_chart(
    chart: ChartSpec,
    worksheet: Any,
    category_column: int,
    series_columns: tuple[int, ...],
    header_row: int,
    first_row: int,
    last_row: int,
) -> Any:
    categories = Reference(
        worksheet,
        min_col=category_column,
        min_row=first_row,
        max_row=last_row,
    )
    if chart.kind == "pie":
        result = PieChart()
        result.add_data(Reference(
            worksheet,
            min_col=series_columns[0],
            min_row=header_row,
            max_row=last_row,
        ), titles_from_data=True)
        result.set_categories(categories)
        return result
    if chart.kind == "horizontal_bar":
        result = BarChart()
        result.type = "bar"
        result.add_data(Reference(
            worksheet,
            min_col=series_columns[0],
            min_row=header_row,
            max_row=last_row,
        ), titles_from_data=True)
        result.set_categories(categories)
        result.legend = None
        return result

    grouped: list[tuple[str, list[int]]] = []
    for index, series in enumerate(chart.series):
        if grouped and grouped[-1][0] == series.chart_type:
            grouped[-1][1].append(series_columns[index])
        else:
            grouped.append((series.chart_type, [series_columns[index]]))
    result = None
    for chart_type, columns in grouped:
        current = LineChart() if chart_type == "line" else BarChart()
        if isinstance(current, BarChart):
            current.type = "col"
        current.add_data(Reference(
            worksheet,
            min_col=min(columns),
            max_col=max(columns),
            min_row=header_row,
            max_row=last_row,
        ), titles_from_data=True)
        current.set_categories(categories)
        if result is None:
            result = current
        else:
            result += current
    if isinstance(result, BarChart):
        result.y_axis.majorGridlines = None
    return result
