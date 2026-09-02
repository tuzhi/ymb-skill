import json
from pathlib import Path
import sys

from openpyxl import Workbook, load_workbook
import pandas as pd


BI_ROOT = Path(__file__).resolve().parents[1]
if str(BI_ROOT) not in sys.path:
    sys.path.insert(0, str(BI_ROOT))

from bank_statement_bi_analysis.chart_renderers import (  # noqa: E402
    render_echarts,
    render_excel_charts,
)
from bank_statement_bi_analysis.chart_spec import build_chart_bundle  # noqa: E402


def _chart_inputs():
    months = ["2026-01", "2026-02"]
    frame = pd.DataFrame({
        "内部互转": [False, False, False],
        "二级标签": ["主营业务", "采购", "往来款"],
        "收入金额": [60.125, 0.0, 40.0],
        "支出金额": [0.0, 50.0, 30.0],
        "月份": ["2026-01", "2026-01", "2026-02"],
    })
    monthly = pd.DataFrame({
        "流入": [60.125, 40.0],
        "流出": [50.0, 30.0],
        "净流入": [10.125, 10.0],
        "月末余额": [110.0, 120.0],
        "月均余额": [105.0, 115.0],
        "月最低余额": [90.0, 100.0],
        "融资流入(修正)": [0.0, 5.0],
        "往来款流入": [0.0, 40.0],
    }, index=months)
    operating = pd.DataFrame({
        "经流入": [60.125, 35.0],
        "经流出": [50.0, 30.0],
    }, index=months)
    counterparties = pd.DataFrame({
        "对手": ["甲公司", "乙公司"],
        "流入": [60.125, 40.0],
        "流出": [10.0, 70.0],
    })
    loan_monthly = pd.DataFrame({
        "借贷流入": [0.0, 5.0],
        "偿还借贷": [2.0, 2.0],
        "净额": [-2.0, 3.0],
    }, index=months)
    analysis = {
        "months": months,
        "l12": months,
        "mon": monthly,
        "mon_op": operating,
        "cp": counterparties,
        "closing_cash": 120.0,
        "new_loan": (1200.0, 0.0, 12),
        "debts": [],
    }
    detail = {"loan_mon": loan_monthly, "currencies": ["人民币"]}
    return analysis, detail, frame


def test_chart_bundle_is_single_source_for_excel_and_echarts(tmp_path):
    analysis, detail, frame = _chart_inputs()
    bundle = build_chart_bundle(
        analysis,
        detail,
        frame,
        strategy_version="BANKFLOW-BI-V4.0",
    )

    assert len(bundle.charts) == 9
    assert bundle.omitted_charts == ()
    assert bundle.charts[0].series[0].values[0] == 60.12
    assert bundle.charts[6].title == "十大流入对手（客户/下游）"
    assert bundle.charts[7].title == "十大流出对手（供应商/上游）"
    assert bundle.charts[8].series[1].name == "压力情景1（收入-20%、支出-10%）"
    assert bundle.charts[8].series[0].values[0] == 1232.56

    payload = render_echarts(bundle, source_statement_run_id="run-real")
    json.dumps(payload, ensure_ascii=False, allow_nan=False)
    assert [item["chart_id"] for item in payload["charts"]] == [
        chart.chart_id for chart in bundle.charts
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "可视化看板"
    assert render_excel_charts(worksheet, bundle) == 9
    assert len(worksheet._charts) == len(bundle.charts)
    output = tmp_path / "charts.xlsx"
    workbook.save(output)
    reopened = load_workbook(output)
    assert len(reopened["可视化看板"]._charts) == 9


def test_chart_bundle_reports_missing_and_multi_currency_data():
    bundle = build_chart_bundle(
        {},
        {"currencies": ["人民币", "美元"]},
        None,
        strategy_version="BANKFLOW-BI-V4.0",
    )

    assert bundle.currency == "MULTI"
    assert bundle.charts == ()
    assert len(bundle.omitted_charts) == 9
    assert "多币种" in bundle.warnings[0]
