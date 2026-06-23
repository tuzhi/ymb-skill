from __future__ import annotations

import csv
import json
import shutil
import zipfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .tokenization import TokenizationResult
from .vault_service import TokenVaultService


@dataclass(frozen=True)
class FileTokenizationResult:
    """文件级 Token 化产物描述。

    只记录输出路径和非敏感统计。Token Vault 文件会被打包返回，
    但其内容不进入日志和服务端持久化存储。
    """

    tokenized_path: Path
    vault_path: Path
    vault_ref_path: Path | None
    summary_path: Path
    archive_path: Path
    span_count: int
    by_label: dict[str, int]
    token_vault: dict[str, dict[str, str]]
    client_alias_counts: Counter[str]


class StandardizedFileTokenizer:
    """标准化文件 Token 化适配器。

    输入必须是 ymb-skill 标准化成功后的 CSV/XLSX。
    本类只替换单元格内容，不改变 sheet、行数、列数、列名和金额/日期字段结构。
    """

    def __init__(self, token_service: TokenVaultService) -> None:
        self._token_service = token_service

    def tokenize_file(
        self,
        standardized_path: Path,
        output_dir: Path,
        *,
        enabled_labels: list[str] | None = None,
        token_vault: dict[str, dict[str, str]] | None = None,
        vault_ref: dict[str, Any] | None = None,
    ) -> FileTokenizationResult:
        # 文件级入口只接受标准化后的结构化产物。
        # 原始 PDF/Excel 的读取和字段映射已经在前置标准化阶段完成。
        suffix = standardized_path.suffix.lower()
        if suffix == ".csv":
            result = self._tokenize_csv(
                standardized_path,
                output_dir,
                enabled_labels=enabled_labels,
                token_vault=token_vault,
                vault_ref=vault_ref,
            )
        elif suffix == ".xlsx":
            result = self._tokenize_xlsx(
                standardized_path,
                output_dir,
                enabled_labels=enabled_labels,
                token_vault=token_vault,
                vault_ref=vault_ref,
            )
        else:
            raise ValueError("unsupported_standardized_file_type")

        archive_path = output_dir / "tokenized_output.zip"
        with zipfile.ZipFile(archive_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.write(result.tokenized_path, result.tokenized_path.name)
            archive.write(result.vault_path, result.vault_path.name)
            if result.vault_ref_path is not None:
                archive.write(result.vault_ref_path, result.vault_ref_path.name)
            archive.write(result.summary_path, result.summary_path.name)
        return FileTokenizationResult(
            tokenized_path=result.tokenized_path,
            vault_path=result.vault_path,
            vault_ref_path=result.vault_ref_path,
            summary_path=result.summary_path,
            archive_path=archive_path,
            span_count=result.span_count,
            by_label=result.by_label,
            token_vault=result.token_vault,
            client_alias_counts=result.client_alias_counts,
        )

    def _tokenize_csv(
        self,
        standardized_path: Path,
        output_dir: Path,
        *,
        enabled_labels: list[str] | None,
        token_vault: dict[str, dict[str, str]] | None,
        vault_ref: dict[str, Any] | None,
    ) -> FileTokenizationResult:
        with standardized_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.reader(handle)
            rows = list(reader)
        if not rows:
            raise ValueError("empty_standardized_file")

        columns = rows[0]
        data_rows = rows[1:]
        # CSV 输出仍使用原列名，确保 WorkBuddy 标准化输入模式可以直接识别。
        output = self._token_service.tokenize_standardized(
            columns,
            data_rows,
            enabled_labels,
            token_vault,
        )
        tokenized_path = output_dir / f"{standardized_path.stem}_tokenized.csv"
        with tokenized_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(columns)
            writer.writerows(output.rows)
        return self._write_sidecars(tokenized_path, output_dir, output, vault_ref)

    def _tokenize_xlsx(
        self,
        standardized_path: Path,
        output_dir: Path,
        *,
        enabled_labels: list[str] | None,
        token_vault: dict[str, dict[str, str]] | None,
        vault_ref: dict[str, Any] | None,
    ) -> FileTokenizationResult:
        # XLSX 通过复制后原地替换字符串单元格，尽量保留原工作簿结构。
        tokenized_path = output_dir / f"{standardized_path.stem}_tokenized.xlsx"
        shutil.copy2(standardized_path, tokenized_path)
        workbook = load_workbook(tokenized_path)

        aggregate_counter: Counter[str] = Counter()
        client_alias_counts: Counter[str] = Counter()
        span_count = 0
        vault = token_vault
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows())
            if not rows:
                continue
            columns = ["" if cell.value is None else str(cell.value) for cell in rows[0]]
            raw_rows: list[list[Any]] = [
                [cell.value for cell in row]
                for row in rows[1:]
            ]
            if not raw_rows:
                continue
            output = self._token_service.tokenize_standardized(
                columns,
                raw_rows,
                enabled_labels,
                vault,
            )
            vault = output.mapping
            span_count += output.span_count
            aggregate_counter.update(output.by_label)
            client_alias_counts.update(_client_alias_counts_from_tokenized_result(output))
            for row_index, row in enumerate(rows[1:]):
                for cell_index, cell in enumerate(row):
                    if isinstance(cell.value, str):
                        cell.value = output.rows[row_index][cell_index]

        workbook.save(tokenized_path)
        output = TokenizationResult(
            mapping=vault or {},
            span_count=span_count,
            by_label=dict(aggregate_counter),
        )
        return self._write_sidecars(
            tokenized_path,
            output_dir,
            output,
            vault_ref,
            client_alias_counts=client_alias_counts,
        )

    @staticmethod
    def _write_sidecars(
        tokenized_path: Path,
        output_dir: Path,
        output: TokenizationResult,
        vault_ref: dict[str, Any] | None,
        client_alias_counts: Counter[str] | None = None,
    ) -> FileTokenizationResult:
        # Token Vault 是可逆映射文件，只写入返回包，不写入日志或服务端持久库。
        vault_path = output_dir / f"{tokenized_path.stem}_token_vault.json"
        vault_ref_path = (
            output_dir / f"{tokenized_path.stem}_token_vault_ref.json"
            if vault_ref is not None
            else None
        )
        summary_path = output_dir / f"{tokenized_path.stem}_summary.json"
        vault_path.write_text(
            json.dumps(output.mapping, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        if vault_ref_path is not None:
            vault_ref_path.write_text(
                json.dumps(vault_ref, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
        summary = {
            "span_count": output.span_count,
            "by_label": output.by_label,
            "tokenized_file": tokenized_path.name,
            "token_vault_file": vault_path.name,
            "token_vault_ref_file": vault_ref_path.name if vault_ref_path else None,
        }
        summary_path.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return FileTokenizationResult(
            tokenized_path=tokenized_path,
            vault_path=vault_path,
            vault_ref_path=vault_ref_path,
            summary_path=summary_path,
            archive_path=output_dir / "tokenized_output.zip",
            span_count=output.span_count,
            by_label=dict(output.by_label),
            token_vault=dict(output.mapping),
            client_alias_counts=client_alias_counts or _client_alias_counts_from_tokenized_result(output),
        )


CLIENT_ALIAS_COLUMNS = ("本方名称", "客户名称", "主体名称", "客户姓名")


def _client_alias_counts_from_tokenized_result(
    output: TokenizationResult,
) -> Counter[str]:
    target_indexes = [
        index
        for index, column in enumerate(output.columns)
        if column in CLIENT_ALIAS_COLUMNS
    ]
    counts: Counter[str] = Counter()
    if not target_indexes:
        return counts
    for row in output.rows:
        for index in target_indexes:
            if index >= len(row):
                continue
            alias = str(row[index] or "").strip()
            if _is_tokenized_client_alias_candidate(alias):
                counts[alias] += 1
    return counts


def _is_tokenized_client_alias_candidate(alias: str) -> bool:
    if not alias or alias.lower() in {"nan", "none", "null"}:
        return False
    if any(mark in alias for mark in ("/", "\\", "*", "?", "<", ">", "|", ":")):
        return False
    if len(alias) > 40:
        return False
    return any(char.isdigit() for char in alias)


