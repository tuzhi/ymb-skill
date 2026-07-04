"""输入文件路由。

本模块只负责按文件类型选择 reader，并返回统一 ReadResult；
字段映射、金额方向、账户识别仍由 core.standardize 处理。
"""

from dataclasses import dataclass
import inspect
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET

from ymb_standardization_core.parsers.router import read_pdf_rows
from ymb_standardization_core.parsers.routing.rule_loader import infer_parser_id, load_excel_route_rules


@dataclass
class ReadResult:
    kind: str
    preamble: str
    rows: list
    route_info: dict


_excel_reader = None
_unsupported_error = RuntimeError


def configure_readers(excel_reader, csv_reader=None, unsupported_error=RuntimeError):
    """注册 core 中已有的 reader，避免 router 反向依赖标准化主流程。"""
    global _excel_reader, _unsupported_error
    _excel_reader = excel_reader
    _unsupported_error = unsupported_error


def _require_reader(reader, name):
    if reader is None:
        raise RuntimeError(f"{name} reader is not configured")
    return reader


def _excel_candidate(rule, match):
    return {
        "id": rule.id,
        "fingerprint_id": rule.id,
        "parser": rule.parser,
        "format_id": rule.format_id or rule.parser,
        "parser_id": rule.parser_id or infer_parser_id(rule.format_id or rule.parser, rule.file_type),
        "decision": "matched",
        "route_status": "matched",
        "file_type": rule.file_type,
        "bank": rule.bank,
        "account_type": rule.account_type,
        "identity_evidence": match["identity_evidence"],
        "layout_evidence": match["layout_evidence"],
        "metadata_evidence": match.get("metadata_evidence", {}),
        "style_evidence": match.get("style_evidence", []),
        "data_evidence": match.get("data_evidence", []),
        "date_format_evidence": match.get("date_format_evidence", []),
    }


def _excel_fallback(sheet, candidate_fingerprints=None):
    return {
        "parser": "generic_excel",
        "format_id": "generic_excel",
        "parser_id": "excel_grid",
        "decision": "unmatched",
        "route_status": "unmatched",
        "file_type": "excel",
        "fingerprint_id": "",
        "bank": "",
        "account_type": "",
        "identity_evidence": [],
        "layout_evidence": [sheet],
        "candidate_fingerprints": candidate_fingerprints or [],
    }


def route_excel(rows, sheet, context=None):
    candidates = []
    candidate_fingerprints = []
    for rule in load_excel_route_rules():
        candidate = rule.fingerprint_candidate(rows, context=context)
        if candidate:
            candidate_fingerprints.append(candidate)
        match = rule.match(rows, context=context)
        if match:
            candidates.append(_excel_candidate(rule, match))
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        return _excel_fallback(sheet, candidate_fingerprints=candidate_fingerprints)
    return {
        "parser": "ambiguous_router_match",
        "format_id": "ambiguous_router_match",
        "parser_id": "none",
        "decision": "ambiguous",
        "route_status": "ambiguous",
        "file_type": "excel",
        "fingerprint_id": "",
        "candidates": candidates,
        "candidate_fingerprints": candidate_fingerprints,
    }


def _rows_lines(rows):
    return [" ".join(str(c or "") for c in row) for row in rows[:500]]


def _date_patterns_from_text(text):
    patterns = set()
    if re.search(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", text):
        patterns.add("yyyy-mm-dd hh:mm:ss")
    if re.search(r"\d{4}-\d{2}-\d{2}(?!\s+\d{2}:\d{2}:\d{2})", text):
        patterns.add("yyyy-mm-dd")
    if re.search(r"\d{8}", text):
        patterns.add("yyyymmdd")
    return sorted(patterns)


def _xlsx_application(path):
    try:
        with zipfile.ZipFile(path) as z:
            data = z.read("docProps/app.xml")
    except Exception:
        return ""
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return ""
    for child in root:
        if child.tag.rsplit("}", 1)[-1] == "Application":
            return child.text or ""
    return ""


def _excel_context(path, rows, sheet, open_password=None):
    text = "\n".join(_rows_lines(rows))
    context = {
        "metadata": {"sheet": sheet},
        "styles": [],
        "lines": _rows_lines(rows),
        "date_patterns": _date_patterns_from_text(text),
    }
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        _add_xlsx_context(path, context, open_password=open_password)
    elif ext == ".xls":
        _add_xls_context(path, context, open_password=open_password)
    return context


def _add_xlsx_context(path, context, open_password=None):
    try:
        import openpyxl

        with _maybe_decrypted_office_file(path, open_password) as source:
            wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
    except Exception:
        return

    props = wb.properties
    context["metadata"].update({
        "creator": props.creator or "",
        "last_modified_by": props.lastModifiedBy or "",
        "application": _xlsx_application(path),
    })
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 60)):
        for cell in row:
            if cell.value in (None, ""):
                continue
            context["styles"].append({
                "text": str(cell.value).strip(),
                "font": cell.font.name or "",
                "size": cell.font.sz,
                "bold": bool(cell.font.bold),
                "row": cell.row,
                "col": cell.column,
                "number_format": cell.number_format,
            })


def _add_xls_context(path, context, open_password=None):
    try:
        import xlrd

        with _maybe_decrypted_office_file(path, open_password) as source:
            book = xlrd.open_workbook(source, formatting_info=True)
    except Exception:
        return

    context["metadata"].update({
        "creator": getattr(book, "user_name", "") or "",
        "application": "BIFF/XLS",
    })
    sheet = book.sheet_by_index(0)
    for r in range(min(sheet.nrows, 60)):
        for c in range(min(sheet.ncols, 60)):
            value = sheet.cell_value(r, c)
            if value in (None, ""):
                continue
            try:
                xf = book.xf_list[sheet.cell_xf_index(r, c)]
                font = book.font_list[xf.font_index]
                fmt = book.format_map.get(xf.format_key)
                font_name = font.name
                font_size = font.height / 20
                bold = bool(font.bold)
                number_format = fmt.format_str if fmt else str(xf.format_key)
            except Exception:
                font_name = ""
                font_size = None
                bold = False
                number_format = ""
            context["styles"].append({
                "text": str(value).strip(),
                "font": font_name,
                "size": font_size,
                "bold": bold,
                "row": r + 1,
                "col": c + 1,
                "number_format": number_format,
            })


class _OfficeSource:
    def __init__(self, path, open_password=None):
        self.path = path
        self.open_password = open_password
        self.tmp = None

    def __enter__(self):
        if not self.open_password:
            return self.path
        try:
            import msoffcrypto
        except ImportError as exc:
            raise RuntimeError("加密 Excel 需要安装 msoffcrypto-tool 才能使用 open_password") from exc
        self.tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(self.path)[1])
        self.tmp.close()
        with open(self.path, "rb") as src, open(self.tmp.name, "wb") as dst:
            office_file = msoffcrypto.OfficeFile(src)
            office_file.load_key(password=self.open_password)
            office_file.decrypt(dst)
        return self.tmp.name

    def __exit__(self, exc_type, exc, tb):
        if self.tmp:
            try:
                os.unlink(self.tmp.name)
            except OSError:
                pass
        return False


def _maybe_decrypted_office_file(path, open_password=None):
    return _OfficeSource(path, open_password=open_password)


def _call_excel_reader(path, open_password=None):
    reader = _require_reader(_excel_reader, "excel")
    try:
        params = inspect.signature(reader).parameters
    except (TypeError, ValueError):
        params = {}
    if "open_password" in params:
        return reader(path, open_password=open_password)
    return reader(path)


def read_rows(path, hints=None):
    hints = hints or {}
    open_password = hints.get("open_password") or None
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        sheet, rows = _call_excel_reader(path, open_password=open_password)
        context = _excel_context(path, rows, sheet, open_password=open_password)
        return ReadResult(
            kind="excel",
            preamble="",
            rows=rows,
            route_info=route_excel(rows, sheet, context=context),
        )
    if ext in (".csv", ".txt", ".tsv"):
        raise _unsupported_error("CSV/TXT/TSV 当前不作为原始流水支持格式")
    if ext == ".pdf":
        preamble, rows, route_info = read_pdf_rows(path, open_password=open_password)
        return ReadResult(kind="pdf", preamble=preamble, rows=rows, route_info=route_info)
    raise _unsupported_error(f"不支持的文件类型：{ext}")
