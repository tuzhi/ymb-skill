"""输入文件路由。

本模块只负责按文件类型选择 reader，并返回统一 ReadResult；
字段映射、金额方向、账户识别仍由 core.standardize 处理。
"""

from dataclasses import dataclass
import base64
import inspect
import os
import re
import tempfile
import zipfile
import xml.etree.ElementTree as ET

from ymb_standardization_core.readers.router import read_pdf_rows
from ymb_standardization_core.readers.routing.rule_loader import load_excel_route_rules
from ymb_standardization_core.contracts import RouteDecision


@dataclass
class ReadResult:
    kind: str
    preamble: str
    rows: list
    route_info: RouteDecision


_excel_reader = None
_unsupported_error = RuntimeError


def pdf_to_wps_rejection_reason(path):
    """Return a user-facing rejection reason for spreadsheets converted from PDF by WPS."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        return ""
    try:
        with zipfile.ZipFile(path) as archive:
            custom_properties = archive.read("docProps/custom.xml")
        root = ET.fromstring(custom_properties)
    except (OSError, KeyError, zipfile.BadZipFile, ET.ParseError):
        return ""

    for prop in root:
        if prop.attrib.get("name") != "CRO":
            continue
        encoded = next((str(child.text or "").strip() for child in prop if child.text), "")
        if not encoded:
            continue
        try:
            padded = encoded + "=" * (-len(encoded) % 4)
            marker = base64.b64decode(padded, validate=True).decode("utf-8", errors="replace")
        except (ValueError, UnicodeError):
            marker = encoded
        if "Kingsoft PDF to WPS" in marker:
            return (
                f"WPS PDF 转 Excel 文件（检测到 {marker} 元数据），不作为原始流水接收；"
                "请提供银行原始 Excel 或可抽取文本的原始 PDF"
            )
    return ""


def configure_readers(excel_reader, csv_reader=None, unsupported_error=RuntimeError):
    """注册 core 中已有的 reader，避免 router 反向依赖标准化主流程。"""
    global _excel_reader, _unsupported_error
    _excel_reader = excel_reader
    _unsupported_error = unsupported_error


def _require_reader(reader, name):
    if reader is None:
        raise RuntimeError(f"{name} reader is not configured")
    return reader


def _excel_candidate(rule, match):
    return {
        "id": rule.id,
        "fingerprint_id": rule.id,
        "reader_id": rule.reader_id,
        "decision": match["decision"],
        "file_type": rule.file_type,
        "bank": rule.bank,
        "account_type": rule.account_type,
        "series_family": rule.series_family,
        "source_order": rule.source_order,
        "date_order": rule.date_order,
        "multi_sheet_same_layout": rule.multi_sheet_same_layout,
        "header_merge": rule.header_merge,
        "column_mapping": rule.column_mapping,
        "preamble_mapping": rule.preamble_mapping,
        "preamble_extractors": rule.preamble_extractors,
        "conditional_mapping": rule.conditional_mapping,
        "extract_mapping": rule.extract_mapping,
        "require_monetary_value": rule.require_monetary_value,
        "identity_evidence": match["identity_evidence"],
        "columns_evidence": match["columns_evidence"],
        "required_columns_evidence": match.get("required_columns_evidence", []),
        "optional_columns_evidence": match.get("optional_columns_evidence", []),
        "missing_required_columns": match.get("missing_required_columns", []),
        "missing_hints": match.get("missing_hints", []),
        "metadata_evidence": match.get("metadata_evidence", {}),
        "style_evidence": match.get("style_evidence", []),
        "date_format_evidence": match.get("date_format_evidence", []),
    }


def _excel_fallback(sheet, candidate_fingerprints=None):
    return {
        "reader_id": "openpyxl_grid",
        "decision": "unmatched",
        "file_type": "excel",
        "fingerprint_id": "",
        "bank": "",
        "account_type": "",
        "column_mapping": {},
        "identity_evidence": [],
        "columns_evidence": [sheet],
        "candidate_fingerprints": candidate_fingerprints or [],
    }


def _choose_specific_candidate(candidates):
    if not candidates:
        return None
    def score(item):
        return (
            1 if item.get("decision") == "matched" else 0,
            len(item.get("columns_evidence", []))
            + len(item.get("metadata_evidence", {})) * 2
            + len(item.get("style_evidence", []))
            + len(item.get("date_format_evidence", [])),
        )

    by_score = sorted(candidates, key=score, reverse=True)
    if len(by_score) == 1:
        return by_score[0]
    if score(by_score[0]) > score(by_score[1]):
        return by_score[0]

    identified = [item for item in candidates if item.get("bank") and item.get("bank") != "未识别"]
    unidentified = [item for item in candidates if not item.get("bank") or item.get("bank") == "未识别"]
    if len(identified) == 1 and unidentified:
        return identified[0]
    return None


def route_excel(rows, sheet, context=None):
    candidates = []
    candidate_fingerprints = []
    for rule in load_excel_route_rules():
        candidate = rule.fingerprint_candidate(rows, context=context)
        if candidate:
            candidate_fingerprints.append(candidate)
        match = rule.match(rows, context=context)
        if match:
            candidates.append(_excel_candidate(rule, match))
    if len(candidates) == 1:
        return candidates[0]
    if not candidates:
        return _excel_fallback(sheet, candidate_fingerprints=candidate_fingerprints)
    specific = _choose_specific_candidate(candidates)
    if specific:
        return specific
    return {
        "reader_id": "none",
        "decision": "ambiguous",
        "file_type": "excel",
        "fingerprint_id": "",
        "column_mapping": {},
        "candidates": candidates,
        "candidate_fingerprints": candidate_fingerprints,
    }


def _rows_lines(rows):
    return [" ".join(str(c or "") for c in row) for row in rows[:500]]


def _date_patterns_from_text(text):
    patterns = set()
    if re.search(r"\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2}", text):
        patterns.add("yyyy-mm-dd hh:mm:ss")
    if re.search(r"\d{4}-\d{2}-\d{2}(?!\s+\d{2}:\d{2}:\d{2})", text):
        patterns.add("yyyy-mm-dd")
    if re.search(r"\d{8}", text):
        patterns.add("yyyymmdd")
    return sorted(patterns)


def _xlsx_application(path):
    try:
        with zipfile.ZipFile(path) as z:
            data = z.read("docProps/app.xml")
    except Exception:
        return ""
    try:
        root = ET.fromstring(data)
    except ET.ParseError:
        return ""
    for child in root:
        if child.tag.rsplit("}", 1)[-1] == "Application":
            return child.text or ""
    return ""


def _excel_context(path, rows, sheet, open_password=None):
    text = "\n".join(_rows_lines(rows))
    context = {
        "metadata": {"sheet": sheet},
        "styles": [],
        "lines": _rows_lines(rows),
        "date_patterns": _date_patterns_from_text(text),
    }
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        _add_xlsx_context(path, context, open_password=open_password)
    elif ext == ".xls":
        _add_xls_context(path, context, open_password=open_password)
    return context


def _add_xlsx_context(path, context, open_password=None):
    try:
        import openpyxl

        with _maybe_decrypted_office_file(path, open_password) as source:
            wb = openpyxl.load_workbook(source, read_only=False, data_only=True)
    except Exception:
        return

    props = wb.properties
    context["metadata"].update({
        "creator": props.creator or "",
        "last_modified_by": props.lastModifiedBy or "",
        "application": _xlsx_application(path),
    })
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60), max_col=min(ws.max_column, 60)):
        for cell in row:
            if cell.value in (None, ""):
                continue
            context["styles"].append({
                "text": str(cell.value).strip(),
                "font": cell.font.name or "",
                "size": cell.font.sz,
                "bold": bool(cell.font.bold),
                "row": cell.row,
                "col": cell.column,
                "number_format": cell.number_format,
            })


def _add_xls_context(path, context, open_password=None):
    try:
        import xlrd

        with _maybe_decrypted_office_file(path, open_password) as source:
            book = xlrd.open_workbook(source, formatting_info=True)
    except Exception:
        return

    context["metadata"].update({
        "creator": getattr(book, "user_name", "") or "",
        "application": "BIFF/XLS",
    })
    sheet = book.sheet_by_index(0)
    for r in range(min(sheet.nrows, 60)):
        for c in range(min(sheet.ncols, 60)):
            value = sheet.cell_value(r, c)
            if value in (None, ""):
                continue
            try:
                xf = book.xf_list[sheet.cell_xf_index(r, c)]
                font = book.font_list[xf.font_index]
                fmt = book.format_map.get(xf.format_key)
                font_name = font.name
                font_size = font.height / 20
                bold = bool(font.bold)
                number_format = fmt.format_str if fmt else str(xf.format_key)
            except Exception:
                font_name = ""
                font_size = None
                bold = False
                number_format = ""
            context["styles"].append({
                "text": str(value).strip(),
                "font": font_name,
                "size": font_size,
                "bold": bold,
                "row": r + 1,
                "col": c + 1,
                "number_format": number_format,
            })


class _OfficeSource:
    def __init__(self, path, open_password=None):
        self.path = path
        self.open_password = open_password
        self.tmp = None

    def __enter__(self):
        if not self.open_password:
            return self.path
        try:
            import msoffcrypto
        except ImportError as exc:
            raise RuntimeError("加密 Excel 需要安装 msoffcrypto-tool 才能使用 open_password") from exc
        self.tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(self.path)[1])
        self.tmp.close()
        with open(self.path, "rb") as src, open(self.tmp.name, "wb") as dst:
            office_file = msoffcrypto.OfficeFile(src)
            office_file.load_key(password=self.open_password)
            office_file.decrypt(dst)
        return self.tmp.name

    def __exit__(self, exc_type, exc, tb):
        if self.tmp:
            try:
                os.unlink(self.tmp.name)
            except OSError:
                pass
        return False


def _maybe_decrypted_office_file(path, open_password=None):
    return _OfficeSource(path, open_password=open_password)


def _call_excel_reader(path, open_password=None, all_sheets_same_layout=False):
    reader = _require_reader(_excel_reader, "excel")
    try:
        params = inspect.signature(reader).parameters
    except (TypeError, ValueError):
        params = {}
    kwargs = {}
    if "open_password" in params:
        kwargs["open_password"] = open_password
    if "all_sheets_same_layout" in params:
        kwargs["all_sheets_same_layout"] = all_sheets_same_layout
    if kwargs:
        return reader(path, **kwargs)
    return reader(path)


def _merge_configured_excel_header(rows, route_info):
    """按 fingerprint 显式配置合并多层表头，不改变原始数据行号。"""
    config = (route_info or {}).get("header_merge") or {}
    if not rows or not config:
        return rows, route_info
    row_count = int(config.get("rows") or 0)
    route_columns = {
        str(column or "").strip()
        for column in (route_info.get("column_mapping") or {})
        if str(column or "").strip()
    }
    if row_count < 2 or not route_columns:
        return rows, route_info
    header_index = max(
        range(min(30, len(rows))),
        key=lambda index: sum(
            1 for value in rows[index]
            if str(value or "").strip() in route_columns
        ),
    )
    if header_index + row_count > len(rows):
        return rows, route_info

    width = max(len(rows[header_index + offset]) for offset in range(row_count))
    separator = str(config.get("separator") or "")
    merged = []
    parent = ""
    for column_index in range(width):
        top = str(
            rows[header_index][column_index]
            if column_index < len(rows[header_index]) else ""
        ).strip()
        if top:
            parent = top
        parts = []
        for offset in range(1, row_count):
            row = rows[header_index + offset]
            value = str(row[column_index] if column_index < len(row) else "").strip()
            if value:
                parts.append(value)
        if parts:
            merged.append(separator.join([value for value in (top or parent, *parts) if value]))
        else:
            merged.append(top)

    output = [list(row) for row in rows]
    output[header_index] = merged
    for offset in range(1, row_count):
        output[header_index + offset] = [None] * width
    updated_route = dict(route_info)
    updated_mapping = dict(updated_route.get("column_mapping") or {})
    updated_mapping.update(config.get("columns") or {})
    updated_route["column_mapping"] = updated_mapping
    return output, updated_route


def _parse_cmb_compact_row(row):
    """解析招行 Excel 首页把日期/币种/金额压在一个单元格的行。"""
    left = str(row[0] or "").strip() if row else ""
    middle = str(row[4] or "").strip() if len(row) > 4 else ""
    left_match = re.match(
        r"^(20\d{2}-\d{2}-\d{2})\s+([A-Z]{3})\s+([+-]?[\d,]+\.\d{2})$", left
    )
    middle_match = re.match(r"^([+-]?[\d,]+\.\d{2})\s+(.+)$", middle)
    if not left_match or not middle_match:
        return None
    return [
        left_match.group(1), left_match.group(2), left_match.group(3),
        middle_match.group(1), middle_match.group(2),
        str(row[7] or "").strip() if len(row) > 7 else "",
    ]


def _amount(value):
    try:
        return float(str(value).replace(",", ""))
    except (TypeError, ValueError):
        return None


def _repair_cmb_amount_scale(records):
    """按逐笔余额恒等式修复招行导出中偶发丢失两位小数的数值单元格。"""
    previous_balance = None
    for record in records:
        amount = _amount(record[2])
        balance = _amount(record[3])
        if amount is None or balance is None:
            continue
        if previous_balance is not None:
            candidates = []
            for amount_scale in (1, 100):
                for balance_scale in (1, 100):
                    fixed_amount = amount / amount_scale
                    fixed_balance = balance / balance_scale
                    residual = abs(previous_balance + fixed_amount - fixed_balance)
                    scaled_fields = (amount_scale != 1) + (balance_scale != 1)
                    candidates.append((round(residual, 6), scaled_fields, fixed_amount, fixed_balance))
            residual, _scaled, amount, balance = min(candidates)
            # 只有余额方程能在分币精度内闭合时才修正，不对缺行场景作猜测。
            if residual <= 0.02:
                record[2], record[3] = round(amount, 2), round(balance, 2)
        previous_balance = _amount(record[3])
    return records


def _read_cmb_mixed_grid(rows):
    """统一招行 Excel 首页压缩布局与后续普通网格，同时保留原始行号位置。"""
    normalized = [list(row) for row in rows]
    if len(normalized) < 31:
        return normalized
    header = ["记账日期", "货币", "交易金额", "联机余额", "交易摘要", "对手信息"]
    normalized[10] = header
    records = []
    positions = []
    for index in range(11, min(28, len(normalized))):
        record = _parse_cmb_compact_row(normalized[index])
        if record:
            records.append(record)
            positions.append(index)
    normalized[28] = [None] * len(header)
    normalized[29] = [None] * len(header)
    for index in range(30, len(normalized)):
        row = normalized[index]
        date = row[0] if row else None
        if not (hasattr(date, "year") or re.match(r"^20\d{2}-\d{2}-\d{2}(?:\s|$)", str(date or ""))):
            normalized[index] = [None] * len(header)
            continue
        record = [
            date,
            row[1] if len(row) > 1 else "",
            row[2] if len(row) > 2 else "",
            row[4] if len(row) > 4 else "",
            row[5] if len(row) > 5 else "",
            row[7] if len(row) > 7 else "",
        ]
        records.append(record)
        positions.append(index)
    _repair_cmb_amount_scale(records)
    for index, record in zip(positions, records):
        normalized[index] = record
    return normalized


def read_rows(path, hints=None):
    hints = hints or {}
    open_password = hints.get("open_password") or None
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        rejection_reason = pdf_to_wps_rejection_reason(path)
        if rejection_reason:
            raise _unsupported_error(rejection_reason)
        sheet, rows = _call_excel_reader(path, open_password=open_password)
        context = _excel_context(path, rows, sheet, open_password=open_password)
        route_info = route_excel(rows, sheet, context=context)
        if route_info.get("multi_sheet_same_layout"):
            sheet, rows = _call_excel_reader(
                path,
                open_password=open_password,
                all_sheets_same_layout=True,
            )
        if route_info.get("reader_id") == "openpyxl_cmb_mixed_grid":
            rows = _read_cmb_mixed_grid(rows)
        rows, route_info = _merge_configured_excel_header(rows, route_info)
        return ReadResult(
            kind="excel",
            preamble="",
            rows=rows,
            route_info=RouteDecision.from_mapping(route_info),
        )
    if ext in (".csv", ".txt", ".tsv"):
        raise _unsupported_error("CSV/TXT/TSV 当前不作为原始流水支持格式")
    if ext == ".pdf":
        preamble, rows, route_info = read_pdf_rows(path, open_password=open_password)
        return ReadResult(
            kind="pdf",
            preamble=preamble,
            rows=rows,
            route_info=RouteDecision.from_mapping(route_info),
        )
    raise _unsupported_error(f"不支持的文件类型：{ext}")
