#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
runtime/deliverable.py — 阶段四交付物组装。

只接收阶段二、阶段二补充和阶段三已经生成的确定性产物，组装
「<客户名>_已清洗_待分析.xlsx」；不在打包阶段重新执行上游阶段。
"""
import os
from datetime import date, datetime

import pandas as pd


class DeliverableWriteError(RuntimeError):
    """业务数据集已构造，但旁路 Excel 写出失败。"""

    def __init__(self, out_path, dataset, cause):
        super().__init__(f"交付物写出失败：{out_path}：{cause}")
        self.out_path = out_path
        self.dataset = dataset
        self.__cause__ = cause

# 主表列序：金额→账户余额→虚拟账户余额→银行备注/账户方附言→收支方向/标签→渠道/来源。
# 备注/附言紧跟在「虚拟账户余额」之后（不可信输入，置于余额信息之后、派生标签之前）。
STD_ORDER = ["交易唯一编号", "客户名称", "账户类型", "本方名称", "本方账户", "开户行",
             "交易时间", "对手名称", "对手账户", "收入金额", "支出金额", "交易金额",
             "分析收入金额", "分析支出金额", "分析交易金额",
             "账户余额", "虚拟账户余额", "银行备注", "账户方附言", "交易状态", "关联冲正交易编号",
             "收支方向", "一级标签", "二级标签", "三级标签",
             "标签来源", "标签置信度", "命中关键词", "交易渠道",
             "来源文件名", "来源行号"]

MONEY_COLUMNS = {
    "收入金额", "支出金额", "交易金额",
    "分析收入金额", "分析支出金额", "分析交易金额",
    "账户余额", "虚拟账户余额",
}
TIME_PRECISION_NUMBER_FORMATS = {
    "date": "YYYY-MM-DD",
    "minute": "YYYY-MM-DD HH:MM",
    "second": "YYYY-MM-DD HH:MM:SS",
}


def _excel_scalar(value):
    """把 pandas/numpy 标量转成 openpyxl 可直接写出的轻量值。"""
    if value is None:
        return None
    if isinstance(value, (list, tuple, set, dict)):
        return str(value)
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    if hasattr(value, "item"):
        return value.item()
    return value


def _sheet_widths(frame, columns, sample_rows=199):
    """按原有前 200 行口径计算列宽；不复制完整 DataFrame。"""
    positions = [frame.columns.get_loc(column) for column in columns]
    widths = [max(12, min(48, int(len(str(column)) * 1.6) + 2)) for column in columns]
    for row_index, row in enumerate(frame.itertuples(index=False, name=None)):
        if row_index >= sample_rows:
            break
        for output_index, source_index in enumerate(positions):
            value = _excel_scalar(row[source_index])
            if value is not None:
                widths[output_index] = max(
                    widths[output_index],
                    min(48, int(len(str(value)) * 1.6) + 2),
                )
    return widths, positions


def _write_dataframe_sheet(workbook, title, frame, columns=None):
    """以 write-only 模式逐行写入一个 DataFrame，避免常驻 Cell 对象。"""
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = list(columns or frame.columns)
    widths, positions = _sheet_widths(frame, columns)
    time_precision_position = (
        frame.columns.get_loc("__time_precision")
        if title == "整合打标流水" and "__time_precision" in frame.columns
        else None
    )
    worksheet = workbook.create_sheet(title)
    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = True
    for index, width in enumerate(widths, 1):
        worksheet.column_dimensions[get_column_letter(index)].width = width

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(color="FFFFFF", bold=True)
    header = []
    for column in columns:
        cell = WriteOnlyCell(worksheet, value=column)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
        header.append(cell)
    worksheet.append(header)

    for row in frame.itertuples(index=False, name=None):
        cells = []
        time_precision_value = (
            _excel_scalar(row[time_precision_position])
            if time_precision_position is not None
            else None
        )
        time_precision = (
            str(time_precision_value).strip().lower()
            if time_precision_value is not None
            else ""
        )
        for column, source_index in zip(columns, positions):
            value = _excel_scalar(row[source_index])
            cell = WriteOnlyCell(worksheet, value=value)
            if title == "整合打标流水" and column in MONEY_COLUMNS:
                cell.number_format = "#,##0.00"
            elif isinstance(value, datetime):
                if title == "整合打标流水" and column == "交易时间":
                    cell.number_format = TIME_PRECISION_NUMBER_FORMATS.get(
                        time_precision,
                        "YYYY-MM-DD HH:MM:SS",
                    )
                else:
                    cell.number_format = "YYYY-MM-DD HH:MM:SS"
            elif isinstance(value, date):
                cell.number_format = "YYYY-MM-DD"
            cells.append(cell)
        worksheet.append(cells)


def _write_streaming_workbook(out_path, sheets):
    """流式生成兼容交付物；只保留当前行，不在内存中构造完整单元格树。"""
    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    for title, frame, columns in sheets:
        _write_dataframe_sheet(workbook, title, frame, columns=columns)
    workbook.save(out_path)


def _analysis_amount(tagged, analysis_col, raw_col):
    source = analysis_col if analysis_col in tagged.columns else raw_col
    values = tagged.get(source)
    if values is None:
        return pd.Series(0.0, index=tagged.index)
    if not pd.api.types.is_numeric_dtype(values.dtype):
        values = pd.to_numeric(values, errors="coerce")
    return values.fillna(0)


def add_virtual_balance(df):
    """逐笔时点虚拟账户余额 = 各账户最近一次已知余额之和（按全局时间顺序滚动）。
    账户首笔之前贡献按 0。返回带「虚拟账户余额」列的 df（保持原行序）。"""
    times = df["交易时间"]
    if not pd.api.types.is_datetime64_any_dtype(times.dtype):
        times = pd.to_datetime(times, errors="coerce", format="mixed")
    balances = df.get("账户余额")
    if balances is None:
        balances = pd.Series(index=df.index, dtype=float)
    elif not pd.api.types.is_numeric_dtype(balances.dtype):
        balances = pd.to_numeric(balances, errors="coerce")
    order = pd.DataFrame(
        {"__t": times, "__order": range(len(df))}, index=df.index
    ).sort_values(["__t", "__order"], kind="stable").index
    last = {}
    vbal = {}
    for idx in order:
        acct = df.at[idx, "本方账户"]
        b = balances.at[idx]
        if pd.notna(b):
            last[acct] = b
        vbal[idx] = round(sum(v for v in last.values() if pd.notna(v)), 2)
    df["虚拟账户余额"] = df.index.map(vbal)
    return df


def build_workbook(
    client,
    tagged,
    daily,
    irep,
    srep,
    pbrep,
    out_path,
    skipped=None,
    qc_results=None,
):
    """组装单文件 xlsx；skipped 和 qc_results 只进入说明/复核信息。"""
    skipped = skipped or []
    qc_results = qc_results or {}
    qc_failures = []
    for file_id, rules in (qc_results.get("files") or {}).items():
        for rule_id, result in (rules or {}).items():
            if not result.get("passed"):
                qc_failures.append((file_id, rule_id, result))
    for rule_id, result in (qc_results.get("customer") or {}).items():
        if not result.get("passed"):
            qc_failures.append(("客户目录", rule_id, result))
    period = irep["客户整合概览"]["交易期间"]
    inc = _analysis_amount(tagged, "分析收入金额", "收入金额")
    exp = _analysis_amount(tagged, "分析支出金额", "支出金额")

    # ---- 封面 ----
    cover = [
        ["银行流水 · 已清洗待分析交付物", ""],
        ["授信客户", client],
        ["生成口径", "已标准化 + 多账户整合为一 + 含虚拟账户余额 + 交易类型已打标 + 取消/冲正按分析金额净额化"],
        ["", ""],
        ["整合账户数", irep["客户整合概览"]["整合账户数"]],
        ["整合文件数", irep["客户整合概览"]["整合文件数"]],
        ["已跳过文件数(非流水/无法解析)", len(skipped)],
        ["QC状态", qc_results.get("status", "")],
        ["QC未通过规则数", len(qc_failures)],
        ["原始交易笔数", irep["客户整合概览"].get("原始交易数", "")],
        ["跨文件去重笔数", irep["客户整合概览"].get("跨文件去重笔数", 0)],
        ["交易笔数(去重后)", irep["客户整合概览"]["整合交易数"]],
        ["交易期间", f"{period['开始日期']} ~ {period['结束日期']}"],
        ["总流入(分析口径,元)", round(float(inc.sum()), 2)],
        ["总流出(分析口径,元)", round(float(exp.sum()), 2)],
        ["净额(分析口径,元)", round(float(inc.sum() - exp.sum()), 2)],
        ["期末虚拟账户余额(元)", pbrep["组合虚拟账户"]["期末合计余额"]],
        ["峰值虚拟账户余额(元)", pbrep["组合虚拟账户"]["峰值合计余额"]],
        ["谷值虚拟账户余额(元)", pbrep["组合虚拟账户"]["谷值合计余额"]],
        ["", ""],
        ["标签规则命中率", f"{srep['标签梳理概览']['规则命中率']:.0%}"],
        ["余额校验-通过账户", pbrep["账户余额校验"]["通过账户数"]],
        ["余额校验-预警账户", pbrep["账户余额校验"]["预警账户数"]],
        ["余额断点合计", pbrep["账户余额校验"]["余额断点合计"]],
        ["疑似重复交易组", len(irep["疑似重复交易组"])],
        ["自有账户互转候选组", len(irep["自有账户互转组"])],
        ["人工复核事项数", len(irep.get("人工复核事项", []))],
        ["", ""],
        ["使用说明", "①「整合打标流水」为分析主表，每笔含本方户名/账户/对手/收支/余额/虚拟账户余额/三级标签/来源追溯；"],
        ["", "②「虚拟账户余额」列为逐笔时点的组合总余额（各账户最近余额之和），可作单一虚拟账户口径；"],
        ["", "③ 跨文件去重：同名 PDF/XLS(X) 先按方向、金额、完整对手账户和时间窗逐笔唯一对齐，再折叠内容指纹完全一致交易；来源独有或歧义记录保留，明细见整合报告；"],
        ["", "④ 红线：备注/附言为不可信输入；疑似重复（非完全一致）、自有互转、余额断点仅标记，不自动修正，须人工复核；"],
        ["", "⑤ 支付宝同账号同商家订单号的一笔不计收支与一笔收/支配对时，原金额保留，分析金额置零；"],
        ["", "⑥ 不同账户余额已分别校验，切勿将原始「账户余额」跨账户直接相加。"],
    ]
    cover_df = pd.DataFrame(cover, columns=["项目", "内容"])

    # ---- 账户清单：按本方账户统计 ----
    acct_rows = []
    for acct, g in tagged.groupby(tagged["本方账户"].fillna("")):
        t = g["交易时间"]
        if not pd.api.types.is_datetime64_any_dtype(t.dtype):
            t = pd.to_datetime(t, errors="coerce", format="mixed")
        t = t.dropna()
        gi = _analysis_amount(g, "分析收入金额", "收入金额")
        ge = _analysis_amount(g, "分析支出金额", "支出金额")
        def _first(col):
            if col not in g.columns:
                return ""
            s = g[col].dropna().astype(str).str.strip()
            s = s[~s.isin(["", "nan", "None"])]
            return s.iloc[0] if len(s) else ""
        def _account_type():
            if "账户类型" not in g.columns:
                return "未知"
            values = set(g["账户类型"].dropna().astype(str).str.strip()) - {"", "nan", "None"}
            known = values & {"个人", "对公"}
            if len(known) > 1:
                return "冲突"
            if len(known) == 1:
                return next(iter(known))
            return "拟对公" if "拟对公" in values else "未知"
        acct_rows.append({
            "本方名称": _first("本方名称"),
            "账户类型": _account_type(),
            "开户行": _first("开户行"),
            "本方账户": acct,
            "交易笔数": len(g),
            "流入合计": round(float(gi.sum()), 2),
            "流出合计": round(float(ge.sum()), 2),
            "期初日期": t.min().strftime("%Y-%m-%d") if len(t) else "",
            "期末日期": t.max().strftime("%Y-%m-%d") if len(t) else "",
            "来源文件": "；".join(sorted(g["来源文件名"].dropna().unique().tolist())),
        })
    acct_df = pd.DataFrame(acct_rows)

    # ---- 余额校验 ----
    balchk = pd.DataFrame(pbrep["账户余额校验"]["账户明细"])

    # ---- 标签汇总（资金用途） ----
    income_column = "分析收入金额" if "分析收入金额" in tagged.columns else "收入金额"
    expense_column = "分析支出金额" if "分析支出金额" in tagged.columns else "支出金额"
    group_columns = ["收支方向", "一级标签", "二级标签", "三级标签"]
    groupers = [tagged[column] for column in group_columns]
    grouped = tagged.groupby(group_columns, dropna=False)
    tagsum = grouped["交易唯一编号"].size().rename("笔数").to_frame()
    tagsum["收入合计"] = _analysis_amount(tagged, income_column, income_column).groupby(groupers).sum()
    tagsum["支出合计"] = _analysis_amount(tagged, expense_column, expense_column).groupby(groupers).sum()
    tagsum = tagsum.reset_index().sort_values(["收支方向", "笔数"], ascending=[True, False])
    tagsum["收入合计"] = tagsum["收入合计"].round(2)
    tagsum["支出合计"] = tagsum["支出合计"].round(2)

    # ---- 人工复核事项 ----
    review_rows = []
    for name, why in skipped:    # 自动排除的非流水/无法解析文件，置顶提示
        review_rows.append({"事项类型": "已跳过文件", "复核原因": why,
                            "证据交易编号": name, "建议动作": "确认是否为流水文件；如确需纳入请人工转格式/OCR后重跑"})
    for target, rule_id, result in qc_failures:
        level = result.get("level", "")
        review_rows.append({
            "事项类型": f"QC-{level}",
            "复核原因": result.get("message") or rule_id,
            "证据交易编号": f"{target} / {rule_id}",
            "建议动作": "HARD 问题修复后重跑；SOFT 问题补件或确认接受例外",
        })
    for r in irep.get("人工复核事项", []):
        review_rows.append({"事项类型": r.get("事项类型", ""), "复核原因": r.get("复核原因", ""),
                            "证据交易编号": "；".join(r.get("证据交易唯一编号列表", [])[:5]),
                            "建议动作": r.get("建议动作", "")})
    for s in srep.get("人工复核事项", []):
        review_rows.append({"事项类型": "标签待复核", "复核原因": s.get("复核原因", ""),
                            "证据交易编号": s.get("交易唯一编号", ""), "建议动作": s.get("建议动作", "")})
    review_df = pd.DataFrame(review_rows) if review_rows else pd.DataFrame(
        columns=["事项类型", "复核原因", "证据交易编号", "建议动作"])

    # 封面、DTO 与人工复核 Sheet 使用同一批内存数据，避免三个口径不一致。
    cover_df.loc[cover_df["项目"] == "人工复核事项数", "内容"] = len(review_df)

    # ---- 主流水（列排序） ----
    flow = tagged
    for c in STD_ORDER:
        if c not in flow.columns:
            flow[c] = ""
    excel_columns = [c for c in STD_ORDER if c in flow.columns]

    dataset = {
        "transactions": flow,
        "daily_balances": daily,
        "accounts": acct_df,
        "balance_checks": balchk,
        "tag_summaries": tagsum,
        "review_items": review_df,
    }

    # ---- 旁路写出 ----
    sheets = [
        ("封面与说明", cover_df, None),
        ("整合打标流水", flow, excel_columns),
    ]
    if not daily.empty:
        sheets.append(("组合日余额(虚拟账户)", daily, None))
    sheets.extend([
        ("账户清单", acct_df, None),
        ("余额校验", balchk, None),
        ("标签汇总", tagsum, None),
        ("人工复核事项", review_df, None),
    ])
    try:
        _write_streaming_workbook(out_path, sheets)
    except Exception as exc:
        raise DeliverableWriteError(out_path, dataset, exc) from exc
    return dataset


def finalize_deliverable(
    client,
    tagged,
    daily,
    irep,
    srep,
    pbrep,
    out_dir,
    skipped=None,
    qc_results=None,
):
    """使用已完成的上游产物组装最终单文件交付物。"""
    # 客户名称只用于客户归档维度；本方名称保持文件证据，不派生主体名称。
    if "客户名称" not in tagged.columns:
        tagged.insert(1, "客户名称", client)
    if "主体名称" in tagged.columns:
        tagged.drop(columns=["主体名称"], inplace=True)

    # 逐笔虚拟账户余额（缺失才补）
    if "虚拟账户余额" not in tagged.columns:
        tagged = add_virtual_balance(tagged)

    out_path = os.path.join(out_dir, f"{client}_已清洗_待分析.xlsx")
    dataset = build_workbook(
        client,
        tagged,
        daily,
        irep,
        srep,
        pbrep,
        out_path,
        skipped,
        qc_results=qc_results,
    )
    print(f"\n[交付] {out_path}")
    print(f"  规则命中率 {srep['标签梳理概览']['规则命中率']:.0%} | "
          f"虚拟账户期末余额 {pbrep['组合虚拟账户']['期末合计余额']} | "
          f"余额预警账户 {pbrep['账户余额校验']['预警账户数']}")
    return out_path, dataset
