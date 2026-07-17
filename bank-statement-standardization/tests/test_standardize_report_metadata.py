import csv
import importlib.util
import json
import re
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
STANDARDIZE_PATH = REPO_ROOT / "bank-statement-standardization" / "scripts" / "standardize.py"
spec = importlib.util.spec_from_file_location("standardize", STANDARDIZE_PATH)
standardize = importlib.util.module_from_spec(spec)
spec.loader.exec_module(standardize)


class StandardizeReportMetadataTest(unittest.TestCase):
    def test_icbc_personal_pdfium_extracts_owner_without_watermark_accounts(self):
        samples = (
            (
                "吕建光", "工商银行历史明细.pdf", "吕建光", "1502211001208140653",
                "md5:12073bf82ed836a1dcba3d4bb8aa2047", 399,
            ),
        )
        for folder, filename, expected_name, expected_account, expected_fp, expected_rows in samples:
            source = REPO_ROOT / "bank-statement-standardization" / "testdata" / folder / filename
            if not source.exists():
                self.skipTest(f"本地未提供工行个人 PDF 样本：{filename}")
            with self.subTest(filename=filename), tempfile.TemporaryDirectory() as tmp:
                csv_path, _json_path, report = standardize.standardize(str(source), out_dir=tmp)
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                image = report["文件画像"]
                self.assertEqual(image["fingerprint_id"], expected_fp)
                self.assertEqual(image["本方名称"], expected_name)
                self.assertEqual(len(rows), expected_rows)
                self.assertEqual({row["本方名称"] for row in rows}, {expected_name})
                self.assertEqual({row["本方账户"] for row in rows}, {expected_account})

    def test_alipay_pdf_extracts_owner_and_account_from_preamble(self):
        source = (
            REPO_ROOT / "bank-statement-standardization" / "testdata" / "徐育发"
            / "支付宝交易明细(20250501-20260430).pdf"
        )
        if not source.exists():
            self.skipTest("本地未提供支付宝 PDF 样本")
        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(source), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        image = report["文件画像"]
        self.assertEqual(image["fingerprint_id"], "md5:d6adc02bebd0a0b6f6ee4af2bae3f5a6")
        self.assertEqual(image["本方名称"], "付丽翠")
        self.assertEqual(image["本方账户"], "19979389877")
        cases = (
            ("余额宝-自动转入", "支出金额"),
            ("余额宝-转出到余额", "收入金额"),
            ("余额宝-2026.04.29-收益", "收入金额"),
            ("提现-实时提现", "支出金额"),
            ("花呗自动还款", "支出金额"),
            ("花呗主动还款", "支出金额"),
            ("余额升级服务收益发放", "收入金额"),
        )
        for keyword, amount_field in cases:
            matched = [row for row in rows if keyword in row["银行备注"]]
            self.assertTrue(matched, keyword)
            self.assertTrue(all(row[amount_field] for row in matched), keyword)

    def test_datetime_preserves_row_level_date_and_minute_precision(self):
        self.assertEqual(standardize.parse_datetime("2026-05-05", ""), "2026-05-05")
        self.assertEqual(standardize.parse_datetime("2026-05-05", "09:30"), "2026-05-05 09:30")
        self.assertEqual(
            standardize.parse_datetime("2026-05-05", "09:30:12"),
            "2026-05-05 09:30:12",
        )
        self.assertEqual(
            standardize.parse_datetime("01-12-25", "08:23", "dmy"),
            "2025-12-01 08:23",
        )
        rows = [
            ["2026-05-05", ""],
            ["2026-05-06", "09:30:12"],
        ]
        self.assertEqual(
            standardize.infer_transaction_time_precision(rows, [0], [1]),
            "mixed",
        )

    def test_abc_account_query_extracts_preamble_identity_and_mixed_time_precision(self):
        source = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "廖诗川"
            / "廖诗川农行流水.xlsx"
        )
        if not source.exists():
            self.skipTest("本地未提供廖诗川农行流水样本")
        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(source), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual({row["本方名称"] for row in rows}, {"廖贵平"})
        self.assertEqual({row["本方账户"] for row in rows}, {"622845****3315"})
        self.assertEqual(report["文件画像"]["transaction_time_precision"], "mixed")
        self.assertEqual(sum(len(row["交易时间"]) == 10 for row in rows), 20)

    def test_nanjing_statement_reports_source_time_precision_and_series_family(self):
        root = REPO_ROOT / "bank-statement-standardization" / "testdata" / "金鼎"
        samples = (
            ("金鼎南京2022年10月对账明细.xls", "date"),
            ("金鼎南京2023年1月对账明细.xls", "second"),
        )
        for filename, precision in samples:
            source = root / filename
            if not source.exists():
                self.skipTest(f"本地未提供南京银行样本：{filename}")
            with self.subTest(filename=filename), tempfile.TemporaryDirectory() as tmp:
                _csv_path, _json_path, report = standardize.standardize(str(source), out_dir=tmp)

            image = report["文件画像"]
            self.assertEqual(image["series_family"], "nanjing_transaction_detail_biff_v1")
            self.assertEqual(image["transaction_time_precision"], precision)

    def test_icbc_openpdf_without_counterparty_columns_filters_rotated_watermark(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "江西赣驰"
            / "夏侯军刚流水161827.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供江西赣驰工行无对手列 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 2500)
        self.assertEqual(report["文件画像"]["fingerprint_id"], "md5:1ac2b953b8acacc332c7e2e0544eb1f6")
        self.assertEqual({row["本方名称"] for row in rows}, {"夏侯军刚"})
        self.assertEqual({row["本方账户"] for row in rows}, {"6222081505000091789"})
        self.assertEqual({row["开户行"] for row in rows}, {"中国工商银行"})
        self.assertTrue(all(row["交易时间"].startswith("20") for row in rows))
        self.assertTrue(all(row["交易金额"] for row in rows))
        self.assertEqual({row["对手名称"] for row in rows}, {""})
        self.assertEqual({row["对手账户"] for row in rows}, {""})

    def test_rural_account_detail_xls_combines_same_layout_sheets(self):
        samples = [
            ("江西赣驰2025年流水(1).xls", 936, "江西赣驰10-12月流水!"),
            ("江西赣驰2026年流水(2).xls", 718, "2026年4-5月!"),
        ]
        for filename, expected_rows, expected_last_sheet in samples:
            excel = REPO_ROOT / "bank-statement-standardization" / "testdata" / "江西赣驰" / filename
            if not excel.exists():
                self.skipTest(f"本地未提供多 Sheet 样本：{filename}")
            with self.subTest(filename=filename), tempfile.TemporaryDirectory() as tmp:
                csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                self.assertEqual(len(rows), expected_rows)
                self.assertTrue(report["文件画像"]["series_family"].startswith("rural_account_detail_query"))
                self.assertTrue(any(row["来源行号"].startswith(expected_last_sheet) for row in rows))

    def test_yaml_source_order_precedes_auto_order_for_icbc_card_detail(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "广州沛瑾家具"
            / "广州沛瑾家具有限公司@李果红_中国工商银行_TF_1.xlsx"
        )
        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 1474)
        self.assertEqual(report["标准化统计"]["行序整理策略"], "YAML配置：整体翻转")
        self.assertEqual(rows[0]["交易时间"], "2018-01-22")
        self.assertEqual(rows[-1]["交易时间"], "2019-01-20")
        self.assertAlmostEqual(sum(float(row["收入金额"] or 0) for row in rows), 2276959.33, places=2)
        self.assertAlmostEqual(sum(float(row["支出金额"] or 0) for row in rows), 2302721.48, places=2)

        breaks = 0
        previous = None
        for row in rows:
            balance = float(row["账户余额"])
            income = float(row["收入金额"] or 0)
            expense = float(row["支出金额"] or 0)
            if previous is not None and abs(balance - (previous + income - expense)) >= 0.01:
                breaks += 1
            previous = balance
        self.assertEqual(breaks, 0)

    def test_missing_yaml_source_order_keeps_automatic_balance_ordering(self):
        with tempfile.TemporaryDirectory() as tmp:
            excel = Path(tmp) / "auto_order.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["交易日期", "收入金额", "支出金额", "余额"])
            sheet.append(["2026-01-03", "", "5", "105"])
            sheet.append(["2026-01-02", "10", "", "110"])
            sheet.append(["2026-01-01", "", "", "100"])
            workbook.save(excel)

            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(report["标准化统计"]["行序整理策略"], "整体翻转")
        self.assertEqual([row["交易时间"] for row in rows], [
            "2026-01-01",
            "2026-01-02",
            "2026-01-03",
        ])

    def test_precise_timestamp_order_can_repair_intraday_mixed_rows(self):
        rows = [
            (38464.34, None, 24006.00, "2025-11-30 10:00:00"),
            (62470.34, 50000.00, None, "2025-11-24 10:00:00"),
            (12470.34, 100.00, None, "2025-11-21 15:00:00"),
            (12470.34, 0.10, None, "2025-11-21 16:35:00"),
            (12470.24, None, 0.10, "2025-11-21 16:07:00"),
        ]

        order, strategy = standardize.best_continuity_order(rows)

        self.assertEqual(strategy, "按完整时间升序")
        self.assertEqual(order, [2, 4, 3, 1, 0])

    def test_boc_payer_payee_columns_resolve_against_inquirer_account(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "斑马商业对公流水"
            / "商业中国银行7920(青山路支行)流水.xls"
        )
        if not excel.exists():
            self.skipTest("本地未提供中国银行查询账号流水样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(report["文件画像"]["conditional_mapping"]), 2)
        self.assertEqual({row["本方账户"] for row in rows}, {"202264057920"})
        normalize_name = lambda value: str(value or "").replace(" ", "").replace("（", "(").replace("）", ")")
        self.assertEqual({normalize_name(row["本方名称"]) for row in rows}, {"斑马(南昌)商业有限公司"})

        self.assertTrue(any(row["对手账户"] for row in rows))
        self.assertTrue(any(row["对手名称"] for row in rows))
        self.assertTrue(any(row["交易渠道"] for row in rows))

    def test_bank_aliases_are_loaded_from_yaml_without_python_patterns(self):
        self.assertFalse(hasattr(standardize, "BANK_PATTERNS"))
        self.assertEqual(standardize.infer_bank("开户行：中国工商银行"), "中国工商银行")
        self.assertEqual(standardize.infer_bank("开户行：工行南昌支行"), "中国工商银行")

    def test_internal_transaction_profile_infers_bank_without_changing_router_bank(self):
        records = []
        for idx, memo in enumerate(("非分期贷款放款", "非分期贷款扣款", "非分期贷款扣款")):
            records.append({
                "交易唯一编号": f"TX-{idx}",
                "对手账户": f"10010{idx} 上饶银行",
                "银行备注": memo,
                "账户方附言": "",
            })

        bank, profile = standardize.infer_bank_from_internal_transactions(records)

        self.assertEqual(bank, "上饶银行")
        self.assertEqual(profile["candidate_count"], 3)
        self.assertEqual(profile["candidate_ratio"], 1.0)
        self.assertEqual(profile["evidence_transaction_ids"], ["TX-0", "TX-1", "TX-2"])

    def test_regular_counterparty_bank_does_not_infer_self_bank(self):
        records = [
            {
                "交易唯一编号": "TX-1",
                "对手账户": "6217000000000000 中国建设银行",
                "银行备注": "采购货款",
                "账户方附言": "",
            }
            for _ in range(5)
        ]

        bank, profile = standardize.infer_bank_from_internal_transactions(records)

        self.assertEqual(bank, "")
        self.assertEqual(profile["candidate_count"], 0)

    def test_bank_fees_do_not_infer_self_bank(self):
        records = [
            {
                "交易唯一编号": f"TX-{idx}",
                "对手账户": "100101 上饶银行",
                "银行备注": "银行手续费",
                "账户方附言": "",
            }
            for idx in range(3)
        ]

        bank, profile = standardize.infer_bank_from_internal_transactions(records)

        self.assertEqual(bank, "")
        self.assertEqual(profile["candidate_count"], 0)

    def test_srbank_corporate_pdf_confirms_router_and_transaction_profile(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "斑马商业对公流水"
            / "斑马商业上饶一般户（南昌县支行）-8259流水........pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供斑马商业上饶 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            _csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)

        image = report["文件画像"]
        self.assertEqual(image["router_bank"], "上饶银行")
        self.assertEqual(image["inferred_bank"], "上饶银行")
        self.assertEqual(image["bank_status"], "confirmed")
        self.assertEqual(image["bank_source"], "router")
        self.assertEqual(image["确认银行"], "上饶银行")
        self.assertEqual(image["internal_transaction_profile"]["candidate_count"], 11)

    def test_wps_converted_pdfs_are_marked_and_skipped(self):
        samples = (
            ("广源流水", "熊亮流水.pdf_2"),
            ("金伟", "2063891248809000962_1.pdf_2"),
            ("徐长河", "工商银行历史明细（申请单号：26060513375455134973）.pdf_2"),
            ("宁聚&付亮亮&徐美琴", "付亮亮建行3763.pdf_2"),
        )
        paths = [
            str(REPO_ROOT / "bank-statement-standardization" / "testdata" / folder / filename)
            for folder, filename in samples
        ]

        candidates, skipped = standardize.screen_files(paths)

        self.assertEqual(candidates, [])
        self.assertEqual({name for name, _reason in skipped}, {filename for _folder, filename in samples})
        self.assertTrue(all("转换文件或非原始文件" in reason for _name, reason in skipped))

    def test_icbc_timestamp_pdf_extracts_corporate_owner_and_account(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "江西奥飞科技"
            / "25年3月1-25年7月31流水.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供江西奥飞科技工行 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 444)
        self.assertEqual(report["文件画像"]["router_bank"], "中国工商银行")
        self.assertEqual(report["文件画像"]["账户类型"], "对公")
        self.assertEqual(report["文件画像"]["本方名称"], "江西奥飞科技有限公司")
        self.assertEqual(report["文件画像"]["本方账户"], "1502209509300265378")
        self.assertEqual({row["本方名称"] for row in rows}, {"江西奥飞科技有限公司"})
        self.assertEqual({row["本方账户"] for row in rows}, {"1502209509300265378"})

    def test_account_metadata_is_not_guessed_without_route_configuration(self):
        info = standardize.sniff_account_info(
            [["交易时间", "交易金额", "余额"]],
            0,
            "企业名称：张三 账号：6222000000000001 参考号：9988776655443322",
        )

        self.assertEqual(info["本方名称"], "")
        self.assertEqual(info["本方账户"], "")
        self.assertEqual(info["账户类型线索"], "")

    def test_account_metadata_uses_route_configured_extractors(self):
        info = standardize.sniff_account_info(
            [["交易时间", "交易金额", "余额"]],
            0,
            "微信昵称：[刘伟兰]",
            preamble_extractors=[
                {
                    "field": "本方名称",
                    "pattern": r"微信昵称[:：]?\s*[［\[]\s*([^］\]\s]+)\s*[］\]]",
                },
                {
                    "field": "本方账户",
                    "pattern": r"微信昵称[:：]?\s*[［\[]\s*([^］\]\s]+)\s*[］\]]",
                    "template": "微信支付#{value}",
                },
            ],
        )

        self.assertEqual(info["本方名称"], "刘伟兰")
        self.assertEqual(info["本方账户"], "微信支付#刘伟兰")

    def test_extract_mapping_extracts_one_field(self):
        rules = [{
            "source": "对方账号与户名",
            "field": "对手账户",
            "pattern": r"^(\d{8,})/.*$",
        }, {
            "source": "对方账号与户名",
            "field": "对手名称",
            "pattern": r"^\d{8,}/",
            "replacement": "",
        }]

        split = standardize.apply_extract_mapping(
            ["6217002020026242362/邓俊英/附加信息"],
            ["对方账号与户名"],
            rules,
        )
        fallback = standardize.apply_extract_mapping(
            ["浙江民禾南昌律师事务所"],
            ["对方账号与户名"],
            rules,
        )

        self.assertEqual(split, {
            "对手账户": "6217002020026242362",
            "对手名称": "邓俊英/附加信息",
        })
        self.assertEqual(fallback, {})

    def test_extract_mapping_maps_account_and_name_separately(self):
        rules = [{
            "source": "对手信息",
            "field": "对手账户",
            "pattern": r"^.*?(?<!\d)(\d{8,})(?!\d).*$",
        }, {
            "source": "对手信息",
            "field": "对手名称",
            "pattern": r"(?<!\d)\d{8,}(?!\d)",
            "replacement": "",
        }]

        combined = standardize.apply_extract_mapping(
            ["曾小园 6217002020025481698 招商银行第三方平台交易资金"],
            ["对手信息"],
            rules,
        )
        account_only = standardize.apply_extract_mapping(
            ["12591713522210004"],
            ["对手信息"],
            rules,
        )

        self.assertEqual(combined, {
            "对手账户": "6217002020025481698",
            "对手名称": "曾小园 招商银行第三方平台交易资金",
        })
        self.assertEqual(account_only, {
            "对手账户": "12591713522210004",
            "对手名称": "",
        })

    def test_abc_personal_extracts_only_long_numeric_counterparty_accounts(self):
        rules = [{
            "source": "对手信息",
            "field": "对手账户",
            "pattern": r"^\s*(\d{12,})\s*$",
        }, {
            "source": "对手信息",
            "field": "对手名称",
            "pattern": r"^\s*\d{12,}\s*$",
            "replacement": "",
        }]

        account = standardize.apply_extract_mapping(
            ["14012157750000836"], ["对手信息"], rules)
        short_identifier = standardize.apply_extract_mapping(
            ["243300133"], ["对手信息"], rules)

        self.assertEqual(account, {
            "对手账户": "14012157750000836",
            "对手名称": "",
        })
        self.assertEqual(short_identifier, {})

    def test_dai_jinwang_pdfs_extract_owner_and_real_accounts(self):
        folder = REPO_ROOT / "bank-statement-standardization" / "testdata" / "戴金旺"
        samples = [
            (
                folder / "微信支付交易明细证明(20250520-20260519)_20260520094427.pdf",
                3404,
                "戴金旺",
                "wxid_q7atqd3k4x2822",
                "微信支付",
            ),
            (
                folder / "江西·农商银行(2026年06月09日20时43分53秒).pdf",
                463,
                "戴金旺",
                "6226820031001918612",
                "江西农商银行",
            ),
        ]
        if not all(path.exists() for path, *_ in samples):
            self.skipTest("本地未提供戴金旺 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            for index, (path, expected_rows, owner, account, bank) in enumerate(samples):
                csv_path, _json_path, _report = standardize.standardize(
                    str(path), out_dir=str(Path(tmp) / str(index))
                )
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                self.assertEqual(len(rows), expected_rows, path.name)
                self.assertEqual({row["本方名称"] for row in rows}, {owner}, path.name)
                self.assertEqual({row["本方账户"] for row in rows}, {account}, path.name)
                self.assertEqual({row["开户行"] for row in rows}, {bank}, path.name)
                if bank == "微信支付":
                    repayment_rows = [row for row in rows if row["银行备注"].startswith("分付还款")]
                    self.assertTrue(repayment_rows, path.name)
                    self.assertTrue(all(not row["收入金额"] for row in repayment_rows), path.name)
                    self.assertTrue(all(row["支出金额"] for row in repayment_rows), path.name)
                    self.assertTrue(all(float(row["交易金额"]) < 0 for row in repayment_rows), path.name)

    def test_zeng_xiaoyuan_cmb_pdf_extracts_owner_account_and_counterparty_account(self):
        pdf = (
            REPO_ROOT / "bank-statement-standardization" / "testdata" / "曾小园"
            / "招商银行交易流水(申请时间2026年06月05日13时47分40秒).pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供曾小园招商银行 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 337)
        self.assertEqual({row["本方名称"] for row in rows}, {"宋志鹏"})
        self.assertEqual({row["本方账户"] for row in rows}, {"6214853380229907"})
        self.assertEqual(sum(not row["对手账户"] for row in rows), 0)
        self.assertFalse(any(re.search(r"\d{8,}", row["对手名称"]) for row in rows))
        self.assertEqual(rows[0]["对手账户"], "6227002022070397612")
        self.assertEqual(rows[0]["对手名称"], "宋志鹏")
        multi_number = rows[2]
        self.assertEqual(multi_number["对手账户"], "12591713522210004")
        self.assertEqual(multi_number["对手名称"], "招商银行第三方平台交易资金")
        self.assertEqual(report["字段映射"]["对手账户"]["原始字段"], "对手信息")
        self.assertEqual(report["文件画像"]["extract_mapping"][0]["field"], "对手账户")

    def test_zeng_xiaoyuan_ccb_excels_split_counterparty_account_and_name(self):
        folder = REPO_ROOT / "bank-statement-standardization" / "testdata" / "曾小园"
        samples = [
            (folder / "hqmx_20260605134404.xls", 3315, "曾小园", "6217002020025481698"),
            (folder / "hqmx_20260605135123.xls", 1036, "宋志鹏", "6227002022070397612"),
        ]
        if not all(path.exists() for path, *_ in samples):
            self.skipTest("本地未提供曾小园建设银行 XLS 样本")

        with tempfile.TemporaryDirectory() as tmp:
            for path, expected_rows, owner, account in samples:
                csv_path, _json_path, report = standardize.standardize(str(path), out_dir=tmp)
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                self.assertEqual(len(rows), expected_rows)
                self.assertEqual({row["本方名称"] for row in rows}, {owner})
                self.assertEqual({row["本方账户"] for row in rows}, {account})
                self.assertFalse(any("/" in row["对手账户"] for row in rows))
                self.assertEqual(report["字段映射"]["对手账户"]["原始字段"], "对方账号与户名")
                self.assertEqual(report["字段映射"]["对手名称"]["原始字段"], "对方账号与户名")

    def test_zeng_yao_xia_weipeng_owner_and_counterparty_fields(self):
        folder = (
            REPO_ROOT / "bank-statement-standardization" / "testdata"
            / "曾耀夏伟鹏个人流水"
        )
        samples = [
            ("夏伟鹏的交易明细20260422120619.pdf", 1325, "夏伟鹏", "622908 **** 2028"),
            ("曾耀招商密码350142.pdf", 1731, "曾耀", "6215581502003868833"),
            ("曾耀建行.xls", 2984, "曾耀", "6217002020031660616"),
            ("曾耀建行-0616.pdf", 3000, "曾耀", "6217002020031660616"),
            ("曾耀建行0616.pdf", 474, "曾耀", "6217002020031660616"),
        ]
        if not all((folder / name).exists() for name, *_ in samples):
            self.skipTest("本地未提供曾耀、夏伟鹏个人流水样本")

        ccb_accounts = set()
        with tempfile.TemporaryDirectory() as tmp:
            for index, (name, expected_rows, owner, account) in enumerate(samples):
                out_dir = Path(tmp) / str(index)
                csv_path, _json_path, _report = standardize.standardize(
                    str(folder / name),
                    out_dir=str(out_dir),
                )
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                self.assertEqual(len(rows), expected_rows, name)
                self.assertEqual({row["本方名称"] for row in rows}, {owner}, name)
                self.assertEqual({row["本方账户"] for row in rows}, {account}, name)
                self.assertFalse(
                    any("/" in row["对手账户"] for row in rows),
                    name,
                )
                self.assertFalse(
                    any("\n" in row["对手名称"] for row in rows),
                    name,
                )
                self.assertFalse(
                    any("户          名" in row["交易时间"] for row in rows),
                    name,
                )
                if name == "曾耀建行.xls":
                    ccb_accounts = {row["对手账户"] for row in rows}

        self.assertIn("4******9202", ccb_accounts)
        self.assertIn("Z******0010", ccb_accounts)

    def test_abc_xlsx_extracts_owner_and_account_from_preamble(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "丰城市利华金属制品有限公司"
            / "2025.5.1-2025.5.31农行.xlsx"
        )
        if not excel.exists():
            self.skipTest("本地未提供丰城市利华农行 XLSX 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 644)
        self.assertEqual({row["本方名称"] for row in rows}, {"丰城市利华金属制品有限公司"})
        self.assertEqual({row["本方账户"] for row in rows}, {"14-081501040001694"})
        self.assertEqual(report["文件画像"]["本方名称"], "丰城市利华金属制品有限公司")
        self.assertEqual(report["文件画像"]["本方账户"], "14-081501040001694")

    def test_changhao_bank_owner_account_and_counterparty_fields(self):
        folder = REPO_ROOT / "bank-statement-standardization" / "testdata" / "昌浩公司流水"
        samples = [
            (
                "北京银行2025.4.1-2026.3.31号流水.xlsx", 490,
                "江西昌浩实业有限公司", "20000080375000116971117", "北京银行",
            ),
            (
                "北京银行流水明细2025.5-2026.4.xlsx", 270,
                "上饶昌浩玻璃有限公司", "20000089851200169748190", "北京银行",
            ),
            (
                "九江银行流水明细2025.5-2026.4.xlsx", 792,
                "", "337059300000010618", "九江银行",
            ),
            (
                "农行2025.4.1-10.31号流水.xlsx", 231,
                "江西昌浩实业有限公司", "14-382401040007179", "中国农业银行",
            ),
        ]
        if not all((folder / name).exists() for name, *_ in samples):
            self.skipTest("本地未提供昌浩公司流水样本")

        with tempfile.TemporaryDirectory() as tmp:
            for index, (name, expected_rows, owner, account, bank) in enumerate(samples):
                csv_path, _json_path, _report = standardize.standardize(
                    str(folder / name), out_dir=str(Path(tmp) / str(index)),
                )
                with open(csv_path, encoding="utf-8-sig", newline="") as f:
                    rows = list(csv.DictReader(f))

                self.assertEqual(len(rows), expected_rows, name)
                self.assertEqual({row["本方账户"] for row in rows}, {account}, name)
                self.assertEqual({row["开户行"] for row in rows}, {bank}, name)
                if owner:
                    self.assertEqual({row["本方名称"] for row in rows}, {owner}, name)
                self.assertTrue(all(row["交易时间"] for row in rows), name)
                if name.startswith("北京银行"):
                    self.assertTrue(all(row["对手名称"] for row in rows), name)
                    self.assertTrue(all(row["对手账户"] for row in rows), name)

    def test_cao_jian_pdfs_extract_owner_and_split_ccb_counterparty(self):
        folder = REPO_ROOT / "bank-statement-standardization" / "testdata" / "曹吉安"
        abc = folder / "26060309491107220299.pdf"
        ccb = folder / "hqmx_20260603095339(9).pdf"
        if not abc.exists() or not ccb.exists():
            self.skipTest("本地未提供曹吉安 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            abc_csv, _abc_json, _abc_report = standardize.standardize(str(abc), out_dir=tmp)
            with open(abc_csv, encoding="utf-8-sig", newline="") as f:
                abc_rows = list(csv.DictReader(f))

            ccb_csv, _ccb_json, ccb_report = standardize.standardize(str(ccb), out_dir=tmp)
            with open(ccb_csv, encoding="utf-8-sig", newline="") as f:
                ccb_rows = list(csv.DictReader(f))

        self.assertEqual({row["本方名称"] for row in abc_rows}, {"吴春梅"})
        self.assertEqual({row["本方账户"] for row in abc_rows}, {"6230520920052630479"})
        self.assertEqual({row["本方名称"] for row in ccb_rows}, {"吴春梅"})
        self.assertEqual({row["本方账户"] for row in ccb_rows}, {"6217002020093758837"})
        self.assertEqual(ccb_rows[0]["对手账户"], "6217002020026242362")
        self.assertEqual(ccb_rows[0]["对手名称"], "邓俊英")
        fallback_rows = [row for row in ccb_rows if row["对手名称"] == "浙江民禾南昌律师事务所"]
        self.assertTrue(fallback_rows)
        self.assertEqual({row["对手账户"] for row in fallback_rows}, {"", "14983101040029131"})
        self.assertEqual(ccb_report["字段映射"]["对手账户"]["原始字段"], "对方账号与户名")
        self.assertEqual(ccb_report["字段映射"]["对手名称"]["原始字段"], "对方账号与户名")
        self.assertFalse(any(item["字段"] == "对手名称" for item in ccb_report["人工复核事项"]))

    def test_report_account_metadata_uses_data_columns_when_header_sniff_is_empty(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "statement.xlsx"
            out_dir = tmp_path / "out"
            rows = [
                {
                    "客户账号": "36050183115000001593",
                    "账户名称": "江西省鹏达石业有限公司",
                    "交易时间": "2026-01-02 09:00:00",
                    "借方发生额(支取)": "",
                    "贷方发生额(收入)": "100.00",
                    "余额": "1100.00",
                    "对方户名": "付款方",
                    "对方账号": "10001",
                    "摘要": "收款",
                },
                {
                    "客户账号": "36050183115000001593",
                    "账户名称": "江西省鹏达石业有限公司",
                    "交易时间": "2026-01-03 10:00:00",
                    "借方发生额(支取)": "30.00",
                    "贷方发生额(收入)": "",
                    "余额": "1070.00",
                    "对方户名": "收款方",
                    "对方账号": "10002",
                    "摘要": "付款",
                },
            ]
            wb = Workbook()
            ws = wb.active
            ws.append(list(rows[0].keys()))
            for row in rows:
                ws.append([row[key] for key in rows[0].keys()])
            wb.save(src)

            csv_path, json_path, report = standardize.standardize(str(src), out_dir=str(out_dir))

            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))
            self.assertEqual(out_rows[0]["本方名称"], "江西省鹏达石业有限公司")
            self.assertEqual(out_rows[0]["本方账户"], "36050183115000001593")

            self.assertEqual(report["文件画像"]["本方名称"], "江西省鹏达石业有限公司")
            self.assertEqual(report["文件画像"]["本方账户"], "36050183115000001593")
            self.assertNotIn("_账号未识别", json.dumps(report, ensure_ascii=False))

            with open(json_path, encoding="utf-8") as f:
                saved_report = json.load(f)
            self.assertEqual(saved_report["文件画像"]["本方名称"], "江西省鹏达石业有限公司")
            self.assertEqual(saved_report["文件画像"]["本方账户"], "36050183115000001593")

    def test_card_bin_overrides_corporate_template_default_to_personal(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "statement.xlsx"
            out_dir = tmp_path / "out"
            headers = [
                "客户账号", "账户名称", "交易时间", "借方发生额（支取）", "贷方发生额（收入）",
                "余额", "币种", "对方户名", "对方账号", "对方开户机构", "记账日期", "摘要",
            ]
            row = [
                "6212263602003903457", "张三", "2026-01-02 09:00:00", "", "100.00",
                "1100.00", "人民币", "付款方", "10001", "开户行", "2026-01-02", "收款",
            ]
            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            ws.append(row)
            wb.save(src)

            csv_path, _json_path, report = standardize.standardize(str(src), out_dir=str(out_dir))

            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))
            self.assertEqual(out_rows[0]["账户类型"], "个人")
            self.assertEqual(report["文件画像"]["账户类型"], "个人")
            self.assertEqual(report["文件画像"]["account_type_source"], "card_bin")
            self.assertEqual(report["文件画像"]["开户行"], "中国工商银行")
            self.assertEqual(report["文件画像"]["开户行识别来源"], "card_bin")
            self.assertEqual(out_rows[0]["开户行"], "中国工商银行")

    def test_card_bin_bank_name_uses_external_mapping_config(self):
        self.assertEqual(
            standardize.bank_name_from_card_bin({"bank": "CEB"}),
            "中国光大银行",
        )

    def test_mybank_pdf_uses_enterprise_name_and_router_bank(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "程旭"
            / "江西嘟咔熊网商银行对账单2025.1.1-2025.12.31.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供网商银行 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))

        self.assertEqual(report["文件画像"]["本方名称"], "江西嘟咔熊电子商务有限公司")
        self.assertEqual(report["文件画像"]["本方账户"], "8888888826100206")
        self.assertEqual(report["文件画像"]["开户行"], "浙江网商银行")
        self.assertEqual(report["文件画像"]["开户行识别来源"], "router")
        self.assertEqual(out_rows[0]["本方名称"], "江西嘟咔熊电子商务有限公司")
        self.assertEqual(out_rows[0]["本方账户"], "8888888826100206")
        self.assertEqual(out_rows[0]["开户行"], "浙江网商银行")

    def test_wechat_bill_excel_uses_nickname_as_payment_account(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "万建平"
            / "微信支付账单流水文件(20250301-20260301)——【.xlsx"
        )
        if not excel.exists():
            self.skipTest("本地未提供微信支付账单 Excel 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))

        self.assertEqual(report["文件画像"]["本方名称"], "刘伟兰")
        self.assertEqual(report["文件画像"]["本方账户"], "微信支付#刘伟兰")
        self.assertEqual(out_rows[0]["本方名称"], "刘伟兰")
        self.assertEqual(out_rows[0]["本方账户"], "微信支付#刘伟兰")
        self.assertFalse(out_rows[0]["本方账户"].startswith("未识别账户#"))

        internal_rows = [
            row for row in out_rows
            if row["银行备注"].startswith(("零钱充值", "零钱提现", "转入零钱通", "零钱通转出"))
        ]
        self.assertEqual(len(internal_rows), 11)
        self.assertTrue(all(bool(row["收入金额"]) != bool(row["支出金额"]) for row in internal_rows))
        self.assertAlmostEqual(
            sum(abs(float(row["交易金额"])) for row in internal_rows),
            215782.67,
            places=2,
        )

    def test_wechat_installment_repayments_override_slash_direction_as_expense(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "万建平"
            / "微信支付账单流水文件(20250301-20260301)——【 2.xlsx"
        )
        if not excel.exists():
            self.skipTest("本地未提供微信支付账单 Excel 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, _report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))

        repayment_rows = [row for row in out_rows if row["银行备注"].startswith("分付还款")]
        self.assertEqual(len(repayment_rows), 15)
        self.assertTrue(all(row["收入金额"] == "" for row in repayment_rows))
        self.assertTrue(all(row["支出金额"] != "" for row in repayment_rows))
        self.assertTrue(all(float(row["交易金额"]) < 0 for row in repayment_rows))
        self.assertAlmostEqual(
            sum(float(row["支出金额"]) for row in repayment_rows),
            3907.48,
            places=2,
        )

    def test_cmb_corporate_pdf_treats_payee_account_as_counterparty_account(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "斑马商业对公流水"
            / "斑马商业招行一般户（青山湖支行）-1221流水.1.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供斑马商业招行 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))

        self.assertEqual(len(out_rows), 1377)
        self.assertEqual({row["本方账户"] for row in out_rows}, {
            "未识别账户#斑马商业招行一般户（青山湖支行）-1221流水.1"
        })
        self.assertEqual({row["本方名称"] for row in out_rows}, {"斑马（南昌）商业有限公司"})
        self.assertEqual(out_rows[1]["对手账户"], "979154850070019810")
        self.assertEqual(report["文件画像"]["本方名称"], "斑马（南昌）商业有限公司")
        self.assertEqual(report["文件画像"]["账户类型"], "对公")
        self.assertEqual(report["文件画像"]["开户行"], "招商银行")
        self.assertEqual(report["文件画像"]["router_bank"], "招商银行")
        self.assertEqual(report["文件画像"]["bank_status"], "confirmed")

    def test_cmb_corporate_excel_uses_statement_account_and_owner(self):
        excel = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "斑马商业对公流水"
            / "斑马商业招行一般户（青山湖支行）-1221流水.xlsx"
        )
        if not excel.exists():
            self.skipTest("本地未提供斑马商业招行 Excel 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(excel), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))

        self.assertEqual(len(out_rows), 1377)
        self.assertEqual({row["本方账户"] for row in out_rows}, {"791912215110008"})
        self.assertEqual({row["本方名称"] for row in out_rows}, {"斑马（南昌）商业有限公司"})
        self.assertEqual(out_rows[1]["对手账户"], "979154850070019810")
        self.assertEqual(report["文件画像"]["本方账户"], "791912215110008")
        self.assertEqual(report["文件画像"]["本方名称"], "斑马（南昌）商业有限公司")

    def test_cmbc_personal_pdf_extracts_customer_name_and_account_from_preamble(self):
        pdf = (
            REPO_ROOT
            / "bank-statement-standardization"
            / "testdata"
            / "范新春"
            / "20260527134259699999991324503110064813999998417140.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供民生银行个人账户对账单 PDF 样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        self.assertEqual(len(rows), 976)
        self.assertEqual({row["本方名称"] for row in rows}, {"范新春"})
        self.assertEqual({row["本方账户"] for row in rows}, {"6216917800007827"})
        alipay_rows = [row for row in rows if row["对手账户"] == "20884029356388680156"]
        self.assertEqual(len(alipay_rows), 53)
        self.assertTrue(any(row["对手名称"] == "范新春" for row in alipay_rows))
        self.assertFalse(any("20884029356388680156" in row["对手名称"] for row in rows))
        self.assertEqual(report["文件画像"]["本方名称"], "范新春")
        self.assertEqual(report["文件画像"]["本方账户"], "6216917800007827")

    def test_enterprise_counterparty_ratio_marks_probable_corporate(self):
        with tempfile.TemporaryDirectory() as tmp:
            tmp_path = Path(tmp)
            src = tmp_path / "statement.xlsx"
            out_dir = tmp_path / "out"
            headers = ["交易时间", "收入金额", "支出金额", "账户余额", "对方户名", "对方账号"]
            enterprise_names = [
                "江西新怡光学仪器有限公司",
                "上饶伟力达塑胶制品有限公司",
                "上海付费通信息服务有限公司",
                "连连银通电子支付有限公司",
                "南京汇合堂文化传播中心",
                "中国民生银行股份有限公司上饶分行营业部",
            ]
            personal_names = [f"张三{i}" for i in range(14)]

            wb = Workbook()
            ws = wb.active
            ws.append(headers)
            for idx, name in enumerate(enterprise_names + personal_names, start=1):
                ws.append([
                    f"2026-01-{idx:02d} 09:00:00",
                    "100.00" if idx % 2 else "",
                    "" if idx % 2 else "50.00",
                    str(1000 + idx),
                    name,
                    f"10{idx:04d}",
                ])
            wb.save(src)

            csv_path, _json_path, report = standardize.standardize(str(src), out_dir=str(out_dir))

            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                out_rows = list(csv.DictReader(f))
            self.assertEqual(len(out_rows), 20)
            self.assertEqual(out_rows[0]["账户类型"], "拟对公")
            self.assertEqual(report["文件画像"]["账户类型"], "拟对公")
            self.assertEqual(report["文件画像"]["account_type_source"], "counterparty_profile")
            self.assertEqual(report["文件画像"]["counterparty_profile"]["enterprise_counterparty_count"], 6)
            self.assertEqual(report["文件画像"]["counterparty_profile"]["valid_counterparty_count"], 20)

    def test_wechat_yaml_direction_mapping_restores_other_amounts(self):
        pdf = REPO_ROOT / "bank-statement-standardization" / "testdata" / "邓子威" / "邓微信.pdf"
        if not pdf.exists():
            self.skipTest("本地未提供邓子威微信流水样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, _report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        cases = {
            "零钱提现": ("支出", 19),
            "零钱充值": ("收入", 8),
            "转入零钱通": ("支出", 3),
        }
        for keyword, (direction, expected_count) in cases.items():
            matched = [row for row in rows if keyword in row["银行备注"]]
            self.assertEqual(len(matched), expected_count)
            self.assertTrue(all(row["交易金额"] for row in matched))
            amount_field = "收入金额" if direction == "收入" else "支出金额"
            opposite_field = "支出金额" if direction == "收入" else "收入金额"
            self.assertTrue(all(row[amount_field] for row in matched))
            self.assertTrue(all(not row[opposite_field] for row in matched))

    def test_wechat_pdf_removes_wrap_space_before_private_use_nickname_symbol(self):
        pdf = (
            REPO_ROOT / "bank-statement-standardization" / "testdata" / "熊全子"
            / "微信支付交易明细证明(20250101-20251231)_20260606101535.pdf"
        )
        if not pdf.exists():
            self.skipTest("本地未提供熊全子微信流水样本")

        with tempfile.TemporaryDirectory() as tmp:
            csv_path, _json_path, _report = standardize.standardize(str(pdf), out_dir=tmp)
            with open(csv_path, encoding="utf-8-sig", newline="") as f:
                rows = list(csv.DictReader(f))

        matched = [row for row in rows if row["对手名称"].startswith("AAAAA\ue513峰")]
        self.assertEqual(len(matched), 9)
        self.assertEqual(
            {row["对手名称"] for row in matched},
            {"AAAAA\ue513峰\ue513尚\ue513五金建材"},
        )


if __name__ == "__main__":
    unittest.main()
