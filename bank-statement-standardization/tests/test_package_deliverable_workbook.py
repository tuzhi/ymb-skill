import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import pandas as pd


SKILL_ROOT = Path(__file__).resolve().parents[1]
if str(SKILL_ROOT) not in sys.path:
    sys.path.insert(0, str(SKILL_ROOT))

from runtime import deliverable as package_deliverable


class PackageDeliverableWorkbookTest(unittest.TestCase):
    def test_build_workbook_styles_without_reloading_saved_xlsx(self):
        tagged = pd.DataFrame([{
            "交易唯一编号": "tx-1",
            "客户名称": "测试客户",
            "账户类型": "对公",
            "本方名称": "测试主体",
            "本方账户": "10001",
            "开户行": "测试银行",
            "交易时间": pd.Timestamp("2026-01-01 10:00:00"),
            "__time_precision": "second",
            "对手名称": "对手",
            "对手账户": "20001",
            "收入金额": "100.00",
            "支出金额": "0.00",
            "交易金额": "100.00",
            "账户余额": "100.00",
            "虚拟账户余额": "100.00",
            "银行备注": "备注",
            "账户方附言": "",
            "收支方向": "收入",
            "一级标签": "经营",
            "二级标签": "销售",
            "三级标签": "销售回款",
            "标签来源": "规则库",
            "标签置信度": "1.0",
            "命中关键词": "销售",
            "交易渠道": "网银",
            "来源文件名": "sample.csv",
            "来源行号": "2",
            "关联冲正交易编号": [],
        }])
        daily = pd.DataFrame([{"日期": "2026-01-01", "10001": 100.0, "合计余额": 100.0}])
        irep = {
            "客户整合概览": {
                "交易期间": {"开始日期": "2026-01-01", "结束日期": "2026-01-01"},
                "整合账户数": 1,
                "整合文件数": 1,
                "原始交易数": 1,
                "跨文件去重笔数": 0,
                "整合交易数": 1,
            },
            "疑似重复交易组": [],
            "自有账户互转组": [],
            "人工复核事项": [],
        }
        srep = {
            "标签梳理概览": {"规则命中率": 1.0},
            "人工复核事项": [],
        }
        pbrep = {
            "组合虚拟账户": {
                "期末合计余额": 100.0,
                "峰值合计余额": 100.0,
                "谷值合计余额": 100.0,
            },
            "账户余额校验": {
                "通过账户数": 1,
                "预警账户数": 0,
                "余额断点合计": 0,
                "账户明细": [{"账户": "10001", "交易数": 1, "校验状态": "未校验"}],
            },
        }
        qc_results = {
            "status": "PASS_WITH_WARNINGS",
            "files": {},
            "customer": {
                "customer.coverage_two_years": {
                    "level": "SOFT",
                    "passed": False,
                    "message": "全部有效文件覆盖不足两年",
                }
            },
        }

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "deliverable.xlsx"
            # 性能约束：写出后不允许再 load_workbook 读回整份 xlsx 做样式处理。
            with patch("openpyxl.load_workbook", side_effect=AssertionError("不应二次加载已写出的 xlsx")):
                dataset = package_deliverable.build_workbook(
                    "测试客户",
                    tagged,
                    daily,
                    irep,
                    srep,
                    pbrep,
                    out_path,
                    [],
                    qc_results=qc_results,
                )

            self.assertTrue(out_path.exists())
            self.assertEqual(set(dataset), {
                "transactions",
                "daily_balances",
                "accounts",
                "balance_checks",
                "tag_summaries",
                "review_items",
            })
            self.assertEqual(len(dataset["transactions"]), 1)
            self.assertIs(dataset["transactions"], tagged)
            self.assertIs(dataset["daily_balances"], daily)
            cover = pd.read_excel(out_path, sheet_name="封面与说明", dtype=str)
            review = pd.read_excel(out_path, sheet_name="人工复核事项", dtype=str)
            self.assertEqual(
                cover.loc[cover["项目"] == "QC状态", "内容"].iloc[0],
                "PASS_WITH_WARNINGS",
            )
            self.assertIn("QC-SOFT", set(review["事项类型"]))

            from openpyxl import load_workbook

            workbook = load_workbook(out_path, read_only=False, data_only=True)
            self.assertEqual(workbook.sheetnames, [
                "封面与说明",
                "整合打标流水",
                "组合日余额(虚拟账户)",
                "账户清单",
                "余额校验",
                "标签汇总",
                "人工复核事项",
            ])
            flow_sheet = workbook["整合打标流水"]
            self.assertEqual(flow_sheet.freeze_panes, "A2")
            self.assertEqual(flow_sheet["A1"].fill.fgColor.rgb, "001F4E78")
            self.assertTrue(flow_sheet["A1"].font.bold)
            headers = {cell.value: cell.column for cell in flow_sheet[1]}
            self.assertEqual(
                flow_sheet.cell(2, headers["收入金额"]).number_format,
                "#,##0.00",
            )
            self.assertEqual(
                flow_sheet.cell(2, headers["交易时间"]).number_format,
                "YYYY-MM-DD HH:MM:SS",
            )
            self.assertGreaterEqual(flow_sheet.column_dimensions["A"].width, 12)
            workbook.close()

    def test_transaction_time_uses_original_precision_for_excel_display(self):
        frame = pd.DataFrame([
            {"交易时间": pd.Timestamp("2026-01-01 00:00:00"), "__time_precision": "date"},
            {"交易时间": pd.Timestamp("2026-01-02 10:20:00"), "__time_precision": "minute"},
            {"交易时间": pd.Timestamp("2026-01-03 10:20:30"), "__time_precision": "second"},
            {"交易时间": pd.Timestamp("2026-01-04 00:00:00"), "__time_precision": "unknown"},
        ])

        with tempfile.TemporaryDirectory() as tmp:
            out_path = Path(tmp) / "time-precision.xlsx"
            package_deliverable._write_streaming_workbook(
                out_path,
                [("整合打标流水", frame, ["交易时间"])],
            )

            from openpyxl import load_workbook

            workbook = load_workbook(out_path, read_only=False, data_only=True)
            sheet = workbook["整合打标流水"]
            self.assertEqual(
                [sheet.cell(row, 1).number_format for row in range(2, 6)],
                [
                    "YYYY-MM-DD",
                    "YYYY-MM-DD HH:MM",
                    "YYYY-MM-DD HH:MM:SS",
                    "YYYY-MM-DD HH:MM:SS",
                ],
            )
            self.assertEqual(sheet.max_column, 1)
            self.assertEqual(sheet["A2"].value, pd.Timestamp("2026-01-01").to_pydatetime())
            workbook.close()


if __name__ == "__main__":
    unittest.main()
