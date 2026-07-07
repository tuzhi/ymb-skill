"""生成 testdata 支持矩阵和当前标准化基准。

输出文件默认写入 testdata 目录：
- support_matrix.xlsx：唯一维护的支持矩阵事实源
- baseline_summary.json：仅在 --write-baseline 时生成，用作逐文件标准化基准摘要
"""

import argparse
import csv
import hashlib
import json
import os
import platform
import re
import shutil
import subprocess
import sys
import time
import traceback
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import yaml


ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = ROOT.parent
CORE_PACKAGE = REPO_ROOT / "ymb-standardization-core"
if str(CORE_PACKAGE) not in sys.path:
    sys.path.insert(0, str(CORE_PACKAGE))

from ymb_standardization_core import core as standardize_core  # noqa: E402
from ymb_standardization_core.file_hints import load_file_hints, load_file_hints_for_path  # noqa: E402
from ymb_standardization_core.readers.routing.rule_loader import fingerprint_md5  # noqa: E402


SUPPORTED_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".pdf"}
GENERATED_NAME_MARKERS = (
    "__standardized",
    "__整合流水",
    "__打标流水",
    "__组合日余额",
    "__经营流水BI分析报告",
    "已清洗_待分析",
)
MATRIX_COLUMNS = [
    "银行",
    "账户类型(YAML)",
    "格式",
    "版本",
    "文件路径",
    "router类",
    "reader_id",
    "YAML指纹",
    "测试类",
    "测试日期",
    "测试结果",
    "期望行数",
    "本方户名",
    "本方账号",
    "创建时间≈修改时间",
    "创建人=修改人",
    "备注",
]


def iter_statement_files(root):
    """枚举原始流水候选文件，跳过管线产物和 Office 锁文件。"""
    root = Path(root)
    files = []
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        name = path.name
        lower = name.lower()
        if "_support_matrix_work" in path.parts:
            continue
        if lower in {"support_matrix.xlsx", "support_matrix.csv", "support_matrix.md", "baseline_summary.json"}:
            continue
        if path.suffix.lower() not in SUPPORTED_EXTENSIONS:
            continue
        if name.startswith("~$"):
            continue
        if any(marker.lower() in lower for marker in GENERATED_NAME_MARKERS):
            continue
        files.append(path)
    return sorted(files, key=lambda p: str(p.relative_to(root)).lower())


def file_sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _format_delta(seconds):
    seconds = int(abs(seconds))
    if seconds < 60:
        return f"{seconds}秒"
    minutes, sec = divmod(seconds, 60)
    if minutes < 60:
        return f"{minutes}分{sec}秒"
    hours, minutes = divmod(minutes, 60)
    return f"{hours}小时{minutes}分"


def _metadata_time_check(created, modified, tolerance_seconds=300):
    if not created or not modified:
        return "不适用（缺少创建/修改时间）"
    if created.tzinfo is not None:
        created = created.astimezone(timezone.utc).replace(tzinfo=None)
    if modified.tzinfo is not None:
        modified = modified.astimezone(timezone.utc).replace(tzinfo=None)
    delta = abs((modified - created).total_seconds())
    result = "是" if delta <= tolerance_seconds else "否"
    return f"{result}（差{_format_delta(delta)}）"


def _metadata_author_check(creator, modified_by):
    creator = str(creator or "").strip()
    modified_by = str(modified_by or "").strip()
    if not creator or not modified_by:
        return "是（缺少创建人/修改人，按一致处理）"
    result = "是" if creator == modified_by else "否"
    return f"{result}（创建人:{creator}；修改人:{modified_by}）"


def _parse_pdf_datetime(value):
    text = str(value or "").strip()
    if text.startswith("D:"):
        text = text[2:]
    match = re.match(r"(\d{4})(\d{2})(\d{2})(\d{2})?(\d{2})?(\d{2})?", text)
    if not match:
        return None
    parts = [int(part) if part else default for part, default in zip(match.groups(), [1, 1, 1, 0, 0, 0])]
    try:
        return datetime(*parts)
    except ValueError:
        return None


def _hints_for_metadata(path, hints_root=None):
    if hints_root is not None:
        file_hints = load_file_hints(hints_root)
    else:
        file_hints = load_file_hints_for_path(path)
    return file_hints.for_file(path), file_hints.audit_for_file(path)


def metadata_checks(path, hints_root=None):
    """读取文档元数据，输出支持矩阵用的两项一致性核验。"""
    path = Path(path)
    ext = path.suffix.lower()
    hints, _hints_audit = _hints_for_metadata(path, hints_root=hints_root)
    open_password = hints.get("open_password") or None
    try:
        if ext in {".xlsx", ".xlsm"}:
            import openpyxl

            from ymb_standardization_core.readers.input_router import _maybe_decrypted_office_file

            with _maybe_decrypted_office_file(str(path), open_password=open_password) as source:
                wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
            props = wb.properties
            return {
                "创建时间≈修改时间": _metadata_time_check(props.created, props.modified),
                "创建人=修改人": _metadata_author_check(props.creator, props.lastModifiedBy),
            }
        if ext == ".xls":
            import xlrd

            from ymb_standardization_core.readers.input_router import _maybe_decrypted_office_file

            with _maybe_decrypted_office_file(str(path), open_password=open_password) as source:
                book = xlrd.open_workbook(source, on_demand=True)
            creator = getattr(book, "user_name", "") or ""
            return {
                "创建时间≈修改时间": "是（XLS 未提供创建/修改时间，按一致处理）",
                "创建人=修改人": "是（XLS 仅提供创建人:%s，按一致处理）" % creator if creator else "是（XLS 未提供创建人/修改人，按一致处理）",
            }
        if ext == ".pdf":
            from ymb_standardization_core.readers.router import _open_pdf

            with _open_pdf(path, open_password=open_password) as pdf:
                metadata = pdf.metadata or {}
            created = _parse_pdf_datetime(metadata.get("CreationDate"))
            modified = _parse_pdf_datetime(metadata.get("ModDate"))
            return {
                "创建时间≈修改时间": _metadata_time_check(created, modified),
                "创建人=修改人": "是（PDF 未提供创建人/修改人，按一致处理）",
            }
    except Exception as exc:
        msg = exc.__class__.__name__
        return {
            "创建时间≈修改时间": f"核验失败（{msg}）",
            "创建人=修改人": f"核验失败（{msg}）",
        }
    return {
        "创建时间≈修改时间": "不适用（不支持的格式）",
        "创建人=修改人": "不适用（不支持的格式）",
    }


def git_sha():
    try:
        return subprocess.check_output(
            ["git", "-C", str(REPO_ROOT), "rev-parse", "HEAD"],
            text=True,
            encoding="utf-8",
            stderr=subprocess.DEVNULL,
        ).strip()
    except Exception:
        return ""


def test_class_for_fingerprint(fingerprint_id):
    mapping = {
        "md5:e833fbf4a2171d66315c5a3bda64711c": "test_jxrcb_pdf_route.py",
        "md5:37399b38ddd3572cc70fc6f8b9be2900": "test_kasikorn_pdf_route.py",
        "md5:69c7df7286e238aef80ae49938fd397a": "test_zhejiang_qyrcb_pdf_route.py",
    }
    return mapping.get(fingerprint_id or "", "")


def _load_route_rule_index():
    rules = {}
    for name in ("excel_rules.yaml", "pdf_rules.yaml"):
        path = CORE_PACKAGE / "ymb_standardization_core" / "readers" / "routing" / name
        try:
            items = yaml.safe_load(path.read_text(encoding="utf-8")) or []
        except Exception:
            items = []
        for item in items:
            fingerprint_id = item.get("id")
            if fingerprint_id:
                rules[fingerprint_id] = item
    return rules


ROUTE_RULE_INDEX = _load_route_rule_index()


def yaml_fingerprint_summary(fingerprint_id):
    """返回支持矩阵中可读的 YAML 指纹配置摘要。"""
    if not fingerprint_id:
        return ""
    if fingerprint_id in {"", "ambiguous_router_match"}:
        return "无 YAML 指纹"
    rule = ROUTE_RULE_INDEX.get(fingerprint_id)
    if not rule:
        return "未找到 YAML 规则"

    parts = []
    fingerprint = rule.get("fingerprint") or {}
    identity_any = fingerprint.get("identity", {}).get("any") or []
    columns_all = (fingerprint.get("columns") or {}).get("all") or {}
    metadata_all = fingerprint.get("metadata", {}).get("all") or {}
    style_all = fingerprint.get("style", {}).get("all") or []
    date_any = fingerprint.get("date_format", {}).get("any") or []

    if identity_any:
        parts.append(f"身份:{len(identity_any)}")
    if columns_all:
        parts.append(f"列标记:{len(columns_all)}")
    if metadata_all:
        parts.append("元数据:" + ",".join(metadata_all.keys()))
    if style_all:
        style_labels = [str(item.get("text", "")).strip() for item in style_all if isinstance(item, dict)]
        parts.append("样式:" + ",".join([label for label in style_labels if label][:3]))
    if date_any:
        parts.append("日期:" + ",".join(date_any))
    return "；".join(parts) if parts else "YAML 规则未配置指纹"


def template_from_mapping(image):
    fingerprint_id = image.get("fingerprint_id") or image.get("命中模板") or ""
    evidence = image.get("route_evidence") or {}
    marker = evidence.get("bank_marker") or image.get("命中模板") or ""
    if fingerprint_id:
        return marker or fingerprint_id
    return marker


def normalize_bank_name(bank, fingerprint_id="", template=""):
    """将映射结果里的简称或模板标记统一成支持矩阵里的银行全称。"""
    text = " ".join(str(part or "") for part in (bank, fingerprint_id, template))
    rules = [
        ("md5:69c7df7286e238aef80ae49938fd397a", "浙江庆元农商银行"),
        ("庆元农商银行", "浙江庆元农商银行"),
        ("md5:e833fbf4a2171d66315c5a3bda64711c", "江西农商银行"),
        ("江西·农商银行", "江西农商银行"),
        ("江西农商", "江西农商银行"),
        ("md5:37399b38ddd3572cc70fc6f8b9be2900", "开泰银行（Kasikorn Bank）"),
        ("Kasikorn", "开泰银行（Kasikorn Bank）"),
        ("农业银行", "中国农业银行"),
        ("农行", "中国农业银行"),
        ("工商银行", "中国工商银行"),
        ("工行", "中国工商银行"),
        ("建设银行", "中国建设银行"),
        ("建行", "中国建设银行"),
        ("邮储银行", "中国邮政储蓄银行"),
        ("邮政储蓄", "中国邮政储蓄银行"),
        ("招商银行", "招商银行"),
        ("浦发银行", "上海浦东发展银行"),
        ("长沙银行", "长沙银行"),
        ("三湘银行", "湖南三湘银行"),
        ("上饶银行", "上饶银行"),
        ("微信支付", "微信支付"),
        ("支付宝", "支付宝"),
    ]
    for needle, full_name in rules:
        if needle in text:
            return full_name
    return str(bank or "").strip() or "未识别"


def support_matrix_bank_name(image, fingerprint_id="", template=""):
    route_bank = (ROUTE_RULE_INDEX.get(fingerprint_id) or {}).get("bank") if fingerprint_id else ""
    if route_bank and route_bank not in {"未识别", "未知"}:
        return normalize_bank_name(route_bank, fingerprint_id, template)
    source = image.get("开户行识别来源") or ""
    if source == "文件名":
        return normalize_bank_name("", fingerprint_id, template)
    return normalize_bank_name(image.get("确认银行") or image.get("开户行") or "", fingerprint_id, template)


def yaml_account_type(fingerprint_id):
    return (ROUTE_RULE_INDEX.get(fingerprint_id) or {}).get("account_type", "") if fingerprint_id else ""


def support_matrix_fingerprint_id(fingerprint_id):
    """支持矩阵展示模板身份：必须来自 YAML 顶层 id。

    id 是 fingerprint 节点规范化后的 md5；reader_id 只表示读取策略。
    """
    if not fingerprint_id:
        return ""
    rule = ROUTE_RULE_INDEX.get(fingerprint_id)
    if not rule:
        raise ValueError(f"missing route rule for fingerprint_id: {fingerprint_id}")
    fingerprint = rule.get("fingerprint") or {}
    rule_fingerprint_id = str(rule.get("id") or "").strip()
    if not rule_fingerprint_id:
        raise ValueError(f"missing id for fingerprint_id: {fingerprint_id}")
    expected = fingerprint_md5(fingerprint)
    if rule_fingerprint_id != expected:
        raise ValueError(f"fingerprint id mismatch: {rule_fingerprint_id} != {expected}")
    return rule_fingerprint_id


def load_support_matrix_rows(matrix_path):
    """读取 support_matrix.xlsx，作为已支持样本与指纹归属的事实源。"""
    wb = load_workbook(matrix_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(value or "").strip() for value in rows[0]]
    records = []
    for values in rows[1:]:
        record = {}
        for idx, col in enumerate(header):
            if not col:
                continue
            record[col] = "" if idx >= len(values) or values[idx] is None else str(values[idx]).strip()
        if any(record.values()):
            records.append(record)
    return records


def support_matrix_files_for_fingerprint(matrix_path, fingerprint_id):
    rows = load_support_matrix_rows(matrix_path)
    files = []
    for row in rows:
        row_fingerprint_id = row.get("router类", "")
        if row_fingerprint_id != fingerprint_id:
            continue
        if row.get("测试结果") != "PASS":
            continue
        if not row.get("YAML指纹") or row.get("YAML指纹") == "无 YAML 指纹":
            continue
        files.append(row.get("文件路径", ""))
    return [path for path in files if path]


def read_csv_summary(csv_path):
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.DictReader(f))
    income_count = sum(1 for row in rows if str(row.get("收入金额", "")).strip())
    expense_count = sum(1 for row in rows if str(row.get("支出金额", "")).strip())
    first_time = rows[0].get("交易时间", "") if rows else ""
    last_time = rows[-1].get("交易时间", "") if rows else ""
    return {
        "row_count": len(rows),
        "income_count": income_count,
        "expense_count": expense_count,
        "first_time": first_time,
        "last_time": last_time,
        "columns": list(rows[0].keys()) if rows else [],
    }


def _new_record_and_baseline(path, root, today):
    relative_path = path.relative_to(root).as_posix()
    metadata_check = metadata_checks(path)
    _hints, hints_audit = _hints_for_metadata(path)
    record = {
        "银行": "",
        "账户类型(YAML)": "",
        "格式": path.suffix.lower().lstrip("."),
        "版本": "",
        "文件路径": relative_path,
        "router类": "",
        "reader_id": "",
        "YAML指纹": "",
        "测试类": "",
        "测试日期": today,
        "测试结果": "FAIL",
        "期望行数": "",
        "本方户名": "",
        "本方账号": "",
        "创建时间≈修改时间": metadata_check["创建时间≈修改时间"],
        "创建人=修改人": metadata_check["创建人=修改人"],
        "备注": "",
    }
    baseline = {
        "file_path": relative_path,
        "sha256": file_sha256(path),
        "status": "FAIL",
        "error": "",
        "mapping": {},
        "csv_summary": {},
        "file_hints": hints_audit,
    }
    return record, baseline


def _populate_record_from_standardized_outputs(record, baseline, csv_path, json_path):
    with open(json_path, encoding="utf-8") as f:
        mapping = json.load(f)
    image = mapping.get("文件画像", {})
    summary = read_csv_summary(csv_path)

    fingerprint_id = image.get("fingerprint_id") or ""
    fingerprint_id = support_matrix_fingerprint_id(fingerprint_id) if fingerprint_id else ""
    template = template_from_mapping(image)
    bank_name = support_matrix_bank_name(image, fingerprint_id, template)
    record.update({
        "银行": bank_name,
        "账户类型(YAML)": yaml_account_type(fingerprint_id),
        "版本": template,
        "router类": fingerprint_id,
        "reader_id": image.get("reader_id") or "",
        "YAML指纹": yaml_fingerprint_summary(fingerprint_id),
        "测试类": test_class_for_fingerprint(fingerprint_id),
        "测试结果": "PASS",
        "期望行数": str(summary["row_count"]),
        "本方户名": image.get("本方名称") or "",
        "本方账号": image.get("本方账户") or "",
        "备注": "开户行仅由文件名推断，支持矩阵未采信" if image.get("开户行识别来源") == "文件名" else "",
    })
    baseline.update({
        "status": "PASS",
        "mapping": {
            "bank": record["银行"],
            "yaml_account_type": record["账户类型(YAML)"],
            "fingerprint_id": fingerprint_id,
            "reader_id": record["reader_id"],
            "template": record["版本"],
            "yaml_fingerprint": record["YAML指纹"],
            "created_modified_check": record["创建时间≈修改时间"],
            "creator_modifier_check": record["创建人=修改人"],
            "account_name": record["本方户名"],
            "account_no": record["本方账号"],
        },
        "csv_summary": summary,
    })


def audit_one_file(path, root, output_work_dir, today):
    record, baseline = _new_record_and_baseline(path, root, today)
    work = output_work_dir / hashlib.sha1(str(path).encode("utf-8")).hexdigest()[:16]
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True, exist_ok=True)

    try:
        csv_path, json_path, _report = standardize_core.standardize(str(path), out_dir=str(work))
        _populate_record_from_standardized_outputs(record, baseline, csv_path, json_path)
    except Exception as exc:
        msg = str(exc).strip() or exc.__class__.__name__
        record["备注"] = msg
        baseline["error"] = msg
        baseline["traceback"] = traceback.format_exc()
    return record, baseline


def write_xlsx(path, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "support_matrix"
    ws.append(MATRIX_COLUMNS)
    for row in rows:
        ws.append([row.get(col, "") for col in MATRIX_COLUMNS])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "银行": 22,
        "账户类型(YAML)": 16,
        "格式": 10,
        "版本": 28,
        "文件路径": 72,
        "router类": 32,
        "reader_id": 28,
        "YAML指纹": 48,
        "测试类": 34,
        "测试日期": 14,
        "测试结果": 12,
        "期望行数": 12,
        "本方户名": 26,
        "本方账号": 26,
        "创建时间≈修改时间": 28,
        "创建人=修改人": 40,
        "备注": 60,
    }
    for idx, col in enumerate(MATRIX_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 16)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def write_json(path, payload):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def build_outputs(testdata_root, output_dir, write_baseline=False, sleep_seconds=0.5):
    testdata_root = Path(testdata_root)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")
    run_id = datetime.now().strftime("%Y%m%dT%H%M%S")
    work_root = output_dir / "_support_matrix_work" / run_id
    work_root.mkdir(parents=True, exist_ok=True)

    records = []
    baselines = []
    files = iter_statement_files(testdata_root)
    for index, path in enumerate(files, 1):
        record, baseline = audit_one_file(path, testdata_root, work_root, today)
        records.append(record)
        baselines.append(baseline)
        if sleep_seconds and index < len(files):
            time.sleep(float(sleep_seconds))

    support_xlsx = output_dir / "support_matrix.xlsx"

    write_xlsx(support_xlsx, records)
    baseline_json = None
    if write_baseline:
        baseline_json = output_dir / "baseline_summary.json"
        write_json(baseline_json, {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "git_sha": git_sha(),
            "python": sys.version,
            "platform": platform.platform(),
            "testdata_root": str(testdata_root),
            "file_count": len(records),
            "pass_count": sum(1 for row in records if row["测试结果"] == "PASS"),
            "fail_count": sum(1 for row in records if row["测试结果"] != "PASS"),
            "records": baselines,
        })
    return support_xlsx, baseline_json


def _source_filename_from_standardized_csv(csv_path):
    try:
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                source = (row.get("来源文件名") or "").strip()
                if source:
                    return source
    except Exception:
        return ""
    return ""


def _client_name_from_package_artifact(path):
    parts = Path(path).parts
    try:
        package_index = parts.index("_package_work")
    except ValueError:
        return ""
    if package_index + 1 >= len(parts):
        return ""
    package_dir_name = parts[package_index + 1]
    match = re.match(r"^\d{3}_(.+)$", package_dir_name)
    return match.group(1) if match else package_dir_name


def _artifact_source_stem_and_suffix(csv_path):
    stem = Path(csv_path).name.removesuffix("__standardized.csv")
    for suffix in ("pdf", "xlsx", "xls", "xlsm"):
        marker = f"__{suffix}"
        if stem.endswith(marker):
            return stem[:-len(marker)], suffix
    return stem, ""


def _match_original_file(csv_path, testdata_root, files_by_client_and_name):
    client = _client_name_from_package_artifact(csv_path)
    source = _source_filename_from_standardized_csv(csv_path)
    if source:
        match = files_by_client_and_name.get((client, source))
        if match:
            return match
    stem, suffix = _artifact_source_stem_and_suffix(csv_path)
    candidates = [
        path for (candidate_client, _name), path in files_by_client_and_name.items()
        if candidate_client == client
        and path.stem == stem
        and (not suffix or path.suffix.lower() == f".{suffix}")
    ]
    return candidates[0] if len(candidates) == 1 else None


def build_outputs_from_standardized_artifacts(testdata_root, package_work_root, output_dir, write_baseline=False):
    testdata_root = Path(testdata_root)
    package_work_root = Path(package_work_root)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y-%m-%d")

    files = iter_statement_files(testdata_root)
    files_by_relative = {path.relative_to(testdata_root).as_posix(): path for path in files}
    files_by_client_and_name = {
        (path.relative_to(testdata_root).parts[0], path.name): path
        for path in files
        if len(path.relative_to(testdata_root).parts) >= 2
    }

    records_by_relative = {}
    baselines_by_relative = {}
    for csv_path in sorted(package_work_root.rglob("*__standardized.csv")):
        json_path = csv_path.with_name(csv_path.name.replace("__standardized.csv", "__mapping.json"))
        original = _match_original_file(csv_path, testdata_root, files_by_client_and_name)
        if not original or not json_path.exists():
            continue
        relative = original.relative_to(testdata_root).as_posix()
        record, baseline = _new_record_and_baseline(original, testdata_root, today)
        try:
            _populate_record_from_standardized_outputs(record, baseline, csv_path, json_path)
        except Exception as exc:
            msg = str(exc).strip() or exc.__class__.__name__
            record["备注"] = msg
            baseline["error"] = msg
            baseline["traceback"] = traceback.format_exc()
        records_by_relative[relative] = record
        baselines_by_relative[relative] = baseline

    for relative, path in files_by_relative.items():
        if relative in records_by_relative:
            continue
        record, baseline = _new_record_and_baseline(path, testdata_root, today)
        msg = "未生成标准化产物（客户交付物阶段失败或文件被跳过）"
        record["备注"] = msg
        baseline["error"] = msg
        records_by_relative[relative] = record
        baselines_by_relative[relative] = baseline

    records = [records_by_relative[relative] for relative in sorted(files_by_relative)]
    baselines = [baselines_by_relative[relative] for relative in sorted(files_by_relative)]
    support_xlsx = output_dir / "support_matrix.xlsx"
    write_xlsx(support_xlsx, records)

    baseline_json = None
    if write_baseline:
        baseline_json = output_dir / "baseline_summary.json"
        write_json(baseline_json, {
            "generated_at": datetime.now().isoformat(timespec="seconds"),
            "git_sha": git_sha(),
            "python": sys.version,
            "platform": platform.platform(),
            "testdata_root": str(testdata_root),
            "file_count": len(records),
            "pass_count": sum(1 for row in records if row["测试结果"] == "PASS"),
            "fail_count": sum(1 for row in records if row["测试结果"] != "PASS"),
            "source": "standardized_artifacts",
            "package_work_root": str(package_work_root),
            "records": baselines,
        })
    return support_xlsx, baseline_json


def main(argv=None):
    parser = argparse.ArgumentParser(description="生成 testdata 支持矩阵")
    parser.add_argument("--testdata-root", default=str(ROOT / "testdata"))
    parser.add_argument("--output-dir", default=str(ROOT / "testdata"))
    parser.add_argument("--write-baseline", action="store_true", help="同时生成 baseline_summary.json 回归基准")
    parser.add_argument("--sleep-seconds", type=float, default=0.5, help="每处理完一个文件后的暂停秒数，默认 0.5 秒用于降低持续发热")
    args = parser.parse_args(argv)
    support_xlsx, baseline_json = build_outputs(
        args.testdata_root,
        args.output_dir,
        write_baseline=args.write_baseline,
        sleep_seconds=args.sleep_seconds,
    )
    print(f"support_matrix_xlsx={support_xlsx}")
    if baseline_json:
        print(f"baseline_summary_json={baseline_json}")


if __name__ == "__main__":
    main()
