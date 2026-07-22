#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
package_deliverable.py — 生成「<客户名>_已清洗_待分析.xlsx」单文件交付物

把一个授信客户名下「多主体（企业/个人）× 多银行账户 × 多份机械拆分的同构文件」
整合成**一份** Excel，供信贷员与审批官单文件流转、信贷分析。该文件确保：
  1. 已标准化：各银行各格式统一为中文标准字段（standardize.py）；
  2. 多账户多主体已整合为一：合并、按账户去重/余额校验、自有账户互转标记（integrate.py）；
     并嵌入「虚拟账户余额」——逐笔时点的组合总余额（virtual balance records，portfolio_balance.py 口径）；
  3. 交易类型已打标：资金用途三级标签（tag.py）。

输入三种方式（三选一）：
  A. 单文件夹：--folder 客户文件夹    （夹内所有原始文件视为同一客户，本方名称/账户自动嗅探，
                                       拆分文件按账户自动归并；自动跳过夹内的流水线产物文件）
  B. 多主体：--subject "主体名:文件夹或文件[:对公|个人]"  可重复（一个主体名下可多账户多文件）
  C. 复用产物：--reuse 产物目录或 *__打标流水.csv/*__整合流水.csv
     —— 已用 run_pipeline 跑过的，直接拿现成的整合/打标产物组装交付物，**跳过重复跑标准化/整合/打标**。
     与 A/B 走同一套尾段逻辑，交付物口径（列序/虚拟账户余额/去重统计）完全一致。

用法：
  python package_deliverable.py --client 客户名 --folder 客户文件夹 [--out-dir DIR]
  python package_deliverable.py --client 客户名 \
      --subject "甲公司:/path/甲" --subject "张三:/path/张三:个人" [--out-dir DIR]
  python package_deliverable.py --client 客户名 --reuse 客户文件夹/_标准化产物 [--out-dir DIR]

输出：
  <out-dir>/<客户名>_已清洗_待分析.xlsx
  （中间标准化产物落在 <out-dir>/_工作区/，可留存追溯）
"""
import argparse, glob, json, os, sys, time

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import pandas as pd
import standardize as S
import integrate as I
import tag as T
import portfolio_balance as PB
from stage_contracts import yaml_route_summary

# 格式初筛 / 产物识别 / 非流水排除 统一由 standardize 提供（S.screen_files / S.NotABankStatement）。

# 主表列序：金额→账户余额→虚拟账户余额→银行备注/账户方附言→收支方向/标签→渠道/来源。
# 备注/附言紧跟在「虚拟账户余额」之后（不可信输入，置于余额信息之后、派生标签之前）。
STD_ORDER = ["交易唯一编号", "客户名称", "账户类型", "本方名称", "本方账户", "开户行",
             "交易时间", "对手名称", "对手账户", "收入金额", "支出金额", "交易金额",
             "分析收入金额", "分析支出金额", "分析交易金额",
             "账户余额", "虚拟账户余额", "银行备注", "账户方附言", "交易状态", "关联冲正交易编号",
             "收支方向", "一级标签", "二级标签", "三级标签",
             "标签来源", "标签置信度", "命中关键词", "交易渠道",
             "来源文件名", "来源行号"]


def _analysis_amount(tagged, analysis_col, raw_col):
    source = analysis_col if analysis_col in tagged.columns else raw_col
    return pd.to_numeric(tagged.get(source), errors="coerce").fillna(0)


def _folder_files(path):
    files = []
    for root, _, names in os.walk(path):
        for name in names:
            files.append(os.path.join(root, name))
    return sorted(files)


def gather_subjects(args):
    """返回 ([(主体名, [(文件, 类型), ...]), ...], 跳过清单[(文件名, 原因)])。
    目录会经 S.screen_files 初筛：跳过图片/Word 等非流水格式与流水线产物（原因记入跳过清单）。"""
    subjects, skipped = [], []
    if args.subject:
        for spec in args.subject:
            parts = spec.split(":")
            name = parts[0]
            path = parts[1]
            ctype = parts[2] if len(parts) > 2 else None
            if os.path.isdir(path):
                files, sk = S.screen_files(_folder_files(path))
                skipped += sk
            elif os.path.isfile(path):
                files = [path]   # 显式单文件：尊重用户，交给 standardize 进一步判定是否流水
            else:
                files = []
            subjects.append((name, [(f, ctype) for f in files]))
    else:
        files, sk = S.screen_files(_folder_files(args.folder))
        skipped += sk
        # 单文件夹：主体名暂用客户名（本方名称由 standardize 自动嗅探/按行填充区分各主体）
        subjects.append((args.client, [(f, args.account_type) for f in files]))
    return subjects, skipped


def add_virtual_balance(df):
    """逐笔时点虚拟账户余额 = 各账户最近一次已知余额之和（按全局时间顺序滚动）。
    账户首笔之前贡献按 0。返回带「虚拟账户余额」列的 df（保持原行序）。"""
    d = df.copy()
    d["__t"] = pd.to_datetime(d["交易时间"], errors="coerce", format="mixed")
    d["__bal"] = pd.to_numeric(d.get("账户余额"), errors="coerce")
    d["__order"] = range(len(d))   # 整合后的行序（账户内为正确时序）；同时刻多笔用它兜底，不用来源行号（倒序文件会反向）
    order = d.sort_values(["__t", "__order"], kind="stable").index
    last = {}
    vbal = {}
    for idx in order:
        acct = d.at[idx, "本方账户"]
        b = d.at[idx, "__bal"]
        if pd.notna(b):
            last[acct] = b
        vbal[idx] = round(sum(v for v in last.values() if pd.notna(v)), 2)
    d["虚拟账户余额"] = d.index.map(vbal)
    return d.drop(columns=["__t", "__bal", "__order"])


def build_workbook(client, tagged, daily, irep, srep, pbrep, out_path, skipped=None):
    """组装单文件 xlsx（多 sheet）。skipped 为被自动排除的非流水/无法解析文件清单。"""
    skipped = skipped or []
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    period = irep["客户整合概览"]["交易期间"]
    inc = _analysis_amount(tagged, "分析收入金额", "收入金额")
    exp = _analysis_amount(tagged, "分析支出金额", "支出金额")

    # ---- 封面 ----
    cover = [
        ["银行流水 · 已清洗待分析交付物", ""],
        ["授信客户", client],
        ["生成口径", "已标准化 + 多账户整合为一 + 含虚拟账户余额 + 交易类型已打标 + 取消/冲正按分析金额净额化"],
        ["", ""],
        ["整合账户数", irep["客户整合概览"]["整合账户数"]],
        ["整合文件数", irep["客户整合概览"]["整合文件数"]],
        ["已跳过文件数(非流水/无法解析)", len(skipped)],
        ["原始交易笔数", irep["客户整合概览"].get("原始交易数", "")],
        ["跨文件去重笔数", irep["客户整合概览"].get("跨文件去重笔数", 0)],
        ["交易笔数(去重后)", irep["客户整合概览"]["整合交易数"]],
        ["交易期间", f"{period['开始日期']} ~ {period['结束日期']}"],
        ["总流入(分析口径,元)", round(float(inc.sum()), 2)],
        ["总流出(分析口径,元)", round(float(exp.sum()), 2)],
        ["净额(分析口径,元)", round(float(inc.sum() - exp.sum()), 2)],
        ["期末虚拟账户余额(元)", pbrep["组合虚拟账户"]["期末合计余额"]],
        ["峰值虚拟账户余额(元)", pbrep["组合虚拟账户"]["峰值合计余额"]],
        ["谷值虚拟账户余额(元)", pbrep["组合虚拟账户"]["谷值合计余额"]],
        ["", ""],
        ["标签规则命中率", f"{srep['标签梳理概览']['规则命中率']:.0%}"],
        ["余额校验-通过账户", pbrep["账户余额校验"]["通过账户数"]],
        ["余额校验-预警账户", pbrep["账户余额校验"]["预警账户数"]],
        ["余额断点合计", pbrep["账户余额校验"]["余额断点合计"]],
        ["疑似重复交易组", len(irep["疑似重复交易组"])],
        ["自有账户互转候选组", len(irep["自有账户互转组"])],
        ["人工复核事项数", len(irep.get("人工复核事项", []))],
        ["", ""],
        ["使用说明", "①「整合打标流水」为分析主表，每笔含本方户名/账户/对手/收支/余额/虚拟账户余额/三级标签/来源追溯；"],
        ["", "②「虚拟账户余额」列为逐笔时点的组合总余额（各账户最近余额之和），可作单一虚拟账户口径；"],
        ["", "③ 跨文件去重：同名 PDF/XLS(X) 先按方向、金额、完整对手账户和时间窗逐笔唯一对齐，再折叠内容指纹完全一致交易；来源独有或歧义记录保留，明细见整合报告；"],
        ["", "④ 红线：备注/附言为不可信输入；疑似重复（非完全一致）、自有互转、余额断点仅标记，不自动修正，须人工复核；"],
        ["", "⑤ 支付宝同账号同商家订单号的一笔不计收支与一笔收/支配对时，原金额保留，分析金额置零；"],
        ["", "⑥ 不同账户余额已分别校验，切勿将原始「账户余额」跨账户直接相加。"],
    ]
    cover_df = pd.DataFrame(cover, columns=["项目", "内容"])

    # ---- 账户清单：按本方账户统计 ----
    acct_rows = []
    for acct, g in tagged.groupby(tagged["本方账户"].fillna("")):
        t = pd.to_datetime(g["交易时间"], errors="coerce", format="mixed").dropna()
        gi = _analysis_amount(g, "分析收入金额", "收入金额")
        ge = _analysis_amount(g, "分析支出金额", "支出金额")
        def _first(col):
            if col not in g.columns:
                return ""
            s = g[col].dropna().astype(str).str.strip()
            s = s[~s.isin(["", "nan", "None"])]
            return s.iloc[0] if len(s) else ""
        def _account_type():
            if "账户类型" not in g.columns:
                return "未知"
            values = set(g["账户类型"].dropna().astype(str).str.strip()) - {"", "nan", "None"}
            known = values & {"个人", "对公"}
            if len(known) > 1:
                return "冲突"
            if len(known) == 1:
                return next(iter(known))
            return "拟对公" if "拟对公" in values else "未知"
        acct_rows.append({
            "本方名称": _first("本方名称"),
            "账户类型": _account_type(),
            "开户行": _first("开户行"),
            "本方账户": acct,
            "交易笔数": len(g),
            "流入合计(分析口径)": round(float(gi.sum()), 2),
            "流出合计(分析口径)": round(float(ge.sum()), 2),
            "期初日期": t.min().strftime("%Y-%m-%d") if len(t) else "",
            "期末日期": t.max().strftime("%Y-%m-%d") if len(t) else "",
            "来源文件": "；".join(sorted(g["来源文件名"].dropna().unique().tolist())),
        })
    acct_df = pd.DataFrame(acct_rows)

    # ---- 余额校验 ----
    balchk = pd.DataFrame(pbrep["账户余额校验"]["账户明细"])

    # ---- 标签汇总（资金用途） ----
    tg = tagged.copy()
    tg["__in"] = _analysis_amount(tg, "分析收入金额", "收入金额")
    tg["__out"] = _analysis_amount(tg, "分析支出金额", "支出金额")
    tagsum = (tg.groupby(["收支方向", "一级标签", "二级标签", "三级标签"])
                .agg(笔数=("交易唯一编号", "size"), 收入合计=("__in", "sum"), 支出合计=("__out", "sum"))
                .reset_index().sort_values(["收支方向", "笔数"], ascending=[True, False]))
    tagsum["收入合计"] = tagsum["收入合计"].round(2)
    tagsum["支出合计"] = tagsum["支出合计"].round(2)

    # ---- 人工复核事项 ----
    review_rows = []
    for name, why in skipped:    # 自动排除的非流水/无法解析文件，置顶提示
        review_rows.append({"事项类型": "已跳过文件", "复核原因": why,
                            "证据交易编号": name, "建议动作": "确认是否为流水文件；如确需纳入请人工转格式/OCR后重跑"})
    for r in irep.get("人工复核事项", []):
        review_rows.append({"事项类型": r.get("事项类型", ""), "复核原因": r.get("复核原因", ""),
                            "证据交易编号": "；".join(r.get("证据交易唯一编号列表", [])[:5]),
                            "建议动作": r.get("建议动作", "")})
    for s in srep.get("人工复核事项", [])[:30]:
        review_rows.append({"事项类型": "标签待复核", "复核原因": s.get("复核原因", ""),
                            "证据交易编号": s.get("交易唯一编号", ""), "建议动作": s.get("建议动作", "")})
    review_df = pd.DataFrame(review_rows) if review_rows else pd.DataFrame(
        columns=["事项类型", "复核原因", "证据交易编号", "建议动作"])

    # ---- 主流水（列排序） ----
    flow = tagged.copy()
    for c in STD_ORDER:
        if c not in flow.columns:
            flow[c] = ""
    flow = flow[[c for c in STD_ORDER if c in flow.columns]]

    def _apply_workbook_styles(wb):
        """在 ExcelWriter 持有的 workbook 上直接设置样式，避免保存后再二次加载。"""
        head_fill = PatternFill("solid", fgColor="1F4E78")
        head_font = Font(color="FFFFFF", bold=True)
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = head_fill
                cell.font = head_font
                cell.alignment = Alignment(vertical="center")
            # 列宽自适应（粗略）
            for col in ws.columns:
                width = 12
                letter = get_column_letter(col[0].column)
                for cell in col[:200]:
                    v = cell.value
                    if v is not None:
                        width = max(width, min(48, int(len(str(v)) * 1.6) + 2))
                ws.column_dimensions[letter].width = width
            ws.sheet_view.showGridLines = True

        # 主流水：金额列数字格式
        ws = wb["整合打标流水"]
        money_cols = {
            "收入金额", "支出金额", "交易金额",
            "分析收入金额", "分析支出金额", "分析交易金额",
            "账户余额", "虚拟账户余额",
        }
        header = {c.value: c.column for c in ws[1]}
        for name, col in header.items():
            if name in money_cols:
                letter = get_column_letter(col)
                for cell in ws[letter][1:]:
                    cell.number_format = "#,##0.00"

    # ---- 写出 ----
    # 样式必须在 writer 关闭前直接应用到 xw.book；否则需要重新打开 xlsx 再保存一遍，stage_4 会明显变慢。
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        cover_df.to_excel(xw, sheet_name="封面与说明", index=False)
        flow.to_excel(xw, sheet_name="整合打标流水", index=False)
        if not daily.empty:
            daily.to_excel(xw, sheet_name="组合日余额(虚拟账户)", index=False)
        acct_df.to_excel(xw, sheet_name="账户清单", index=False)
        balchk.to_excel(xw, sheet_name="余额校验", index=False)
        tagsum.to_excel(xw, sheet_name="标签汇总", index=False)
        review_df.to_excel(xw, sheet_name="人工复核事项", index=False)
        _apply_workbook_styles(xw.book)


def _safe(name):
    return "".join(c if c not in '\\/:*?"<>|' else "_" for c in name)


def run(client, args):
    import shutil
    out_dir = args.out_dir or os.getcwd()
    # 每个客户独立工作区，且运行前清空，避免读到其它客户/上次运行的标准化残留
    work = os.path.join(out_dir, "_工作区", _safe(client))
    if os.path.isdir(work):
        shutil.rmtree(work)
    os.makedirs(work, exist_ok=True)

    subjects, skipped = gather_subjects(args)
    print(f"=== 客户「{client}」：{len(subjects)} 个主体 ===")

    # 阶段一：逐文件标准化。本方名称只来自文件证据。
    n_files = 0
    file_routes = {}
    for subj, files in subjects:
        for f, ctype in files:
            try:
                csv_path, _json_path, rep = S.standardize(
                    f, out_dir=work, account_type=ctype)
                file_routes[os.path.basename(csv_path)] = yaml_route_summary(rep)
                n_files += 1
                st = rep["标准化统计"]
                print(f"  [OK] [{subj}] {os.path.basename(f)} -> {st['交易笔数']} 笔（{st['金额结构']}）")
            except S.SourceFormatQualityError as e:
                raise RuntimeError(
                    f"阶段一 QC 未通过：{os.path.basename(f)}：{e.reason}"
                ) from e
            except S.NotABankStatement as e:
                skipped.append((os.path.basename(f), e.reason))
                print(f"  [SKIP] [{subj}] {os.path.basename(f)}：{e.reason}")
            except Exception as e:
                skipped.append((os.path.basename(f), f"解析失败：{e}"))
                print(f"  ! {os.path.basename(f)} 失败：{e}")
            sleep_seconds = getattr(args, "sleep_seconds", 0)
            if sleep_seconds:
                time.sleep(sleep_seconds)

    if skipped:
        print(f"  [已跳过 {len(skipped)} 个非流水/无法解析文件]")
    if n_files == 0:
        detail = "；".join(f"{n}（{w}）" for n, w in skipped) or "目录内无候选文件"
        sys.exit(f"客户「{client}」无可处理的银行流水文件。已跳过：{detail}")

    # 阶段二：整合（全部主体/账户合并为一）
    int_csv, int_json, irep = I.integrate(
        client, [work], out_dir=work, file_routes=file_routes)
    print(f"  整合：{irep['客户整合概览']['整合账户数']} 账户 / {irep['客户整合概览']['整合交易数']} 笔")

    # 阶段三：打标
    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    rules = os.path.join(here, "assets", "tag_rules.csv")
    tag_csv, _, srep = T.tag(int_csv, rules, out_dir=work)
    tagged = pd.read_csv(tag_csv, dtype=str)

    return finalize_deliverable(client, int_csv, tagged, irep, srep, work, out_dir, skipped)


def finalize_deliverable(client, flow_csv, tagged, irep, srep, work, out_dir, skipped=None):
    """从（已整合/已打标的）流水组装单文件交付物：补算组合余额、回填客户/虚拟账户余额列后写 xlsx。
    raw（run）与 reuse（run_reuse）两条路径共用此尾段，保证交付物口径完全一致。"""
    # 组合（虚拟账户）余额 + 余额校验（轻量、确定性，始终重算以保证与当前数据一致）
    _, _, pbrep = PB.run(flow_csv, out_dir=work)
    daily_path = os.path.join(work, os.path.basename(flow_csv).replace(".csv", "") + "__组合日余额.csv")
    daily = pd.read_csv(daily_path) if os.path.exists(daily_path) else pd.DataFrame()

    # 客户名称只用于客户归档维度；本方名称保持文件证据，不派生主体名称。
    if "客户名称" not in tagged.columns:
        tagged.insert(1, "客户名称", client)
    if "主体名称" in tagged.columns:
        tagged = tagged.drop(columns=["主体名称"])

    # 逐笔虚拟账户余额（缺失才补）
    if "虚拟账户余额" not in tagged.columns:
        tagged = add_virtual_balance(tagged)

    out_path = os.path.join(out_dir, f"{client}_已清洗_待分析.xlsx")
    build_workbook(client, tagged, daily, irep, srep, pbrep, out_path, skipped)
    print(f"\n[交付] {out_path}")
    print(f"  规则命中率 {srep['标签梳理概览']['规则命中率']:.0%} | "
          f"虚拟账户期末余额 {pbrep['组合虚拟账户']['期末合计余额']} | "
          f"余额预警账户 {pbrep['账户余额校验']['预警账户数']}")
    return out_path


# ---- 复用已有流水线产物（避免重复跑整条流水线） --------------------------------
def _glob_first(folder, pat):
    hits = sorted(glob.glob(os.path.join(folder, pat)))
    return hits[0] if hits else None


def _load_json(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _build_tag_report_from_df(tagged):
    """复用的打标流水缺标签报告时，从数据现算最小标签概览（够封面/汇总用）。"""
    n = len(tagged)
    src = tagged["标签来源"].astype(str) if "标签来源" in tagged.columns else pd.Series([], dtype=str)
    hit = int((src == "规则库").sum())
    return {"标签梳理概览": {"交易总数": int(n), "规则命中数量": hit,
                          "兜底其他类数量": int(n - hit), "规则命中率": round(hit / max(1, n), 3)},
            "人工复核事项": []}


def _build_integrate_report_from_df(client, tagged):
    """复用流水缺整合报告时，从数据现算最小整合概览。"""
    t = pd.to_datetime(tagged["交易时间"], errors="coerce", format="mixed").dropna()
    return {"客户整合概览": {
                "客户名称": client,
                "整合文件数": int(tagged["来源文件名"].nunique()) if "来源文件名" in tagged.columns else 0,
                "整合账户数": int(tagged["本方账户"].nunique()) if "本方账户" in tagged.columns else 0,
                "原始交易数": int(len(tagged)), "跨文件去重笔数": 0, "整合交易数": int(len(tagged)),
                "交易期间": {"开始日期": t.min().strftime("%Y-%m-%d") if len(t) else "",
                           "结束日期": t.max().strftime("%Y-%m-%d") if len(t) else ""}},
            "疑似重复交易组": [], "自有账户互转组": [], "人工复核事项": []}


def run_reuse(client, reuse_path, out_dir):
    """复用 run_pipeline 已有产物直接组装交付物，跳过重复计算。
    reuse_path 可为产物目录，或直接给某个 *__打标流水.csv / *__整合流水.csv。
    缺失的报告 json 会从数据现算最小版本；组合余额始终重算以保证一致。"""
    import shutil
    work = os.path.join(out_dir, "_工作区", _safe(client))
    if os.path.isdir(work):
        shutil.rmtree(work)
    os.makedirs(work, exist_ok=True)

    tagged_csv = integrated_csv = None
    if os.path.isdir(reuse_path):
        src_dir = reuse_path
        tagged_csv = _glob_first(src_dir, "*__打标流水.csv")
        integrated_csv = _glob_first(src_dir, "*__整合流水.csv")
    elif os.path.isfile(reuse_path) and reuse_path.endswith("__打标流水.csv"):
        tagged_csv, src_dir = reuse_path, os.path.dirname(reuse_path)
    elif os.path.isfile(reuse_path) and reuse_path.endswith("__整合流水.csv"):
        integrated_csv, src_dir = reuse_path, os.path.dirname(reuse_path)
    else:
        sys.exit(f"--reuse 未找到可复用的整合/打标产物（需 *__打标流水.csv 或 *__整合流水.csv）：{reuse_path}")

    here = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    rules = os.path.join(here, "assets", "tag_rules.csv")

    if tagged_csv:
        print(f"=== 客户「{client}」：复用打标产物 {os.path.basename(tagged_csv)}（跳过 标准化/整合/打标）===")
        flow_csv = tagged_csv
        tagged = pd.read_csv(tagged_csv, dtype=str)
        tag_json = _glob_first(src_dir, "*__标签报告.json")
        srep = _load_json(tag_json) if tag_json else _build_tag_report_from_df(tagged)
    else:
        print(f"=== 客户「{client}」：复用整合流水 {os.path.basename(integrated_csv)}（跳过 标准化/整合，重新打标）===")
        flow_csv, _, srep = T.tag(integrated_csv, rules, out_dir=work)
        tagged = pd.read_csv(flow_csv, dtype=str)

    int_json = _glob_first(src_dir, "*__整合报告.json")
    irep = _load_json(int_json) if int_json else _build_integrate_report_from_df(client, tagged)

    return finalize_deliverable(client, flow_csv, tagged, irep, srep, work, out_dir)


def main():
    ap = argparse.ArgumentParser(description="生成 <客户名>_已清洗_待分析.xlsx 单文件交付物")
    ap.add_argument("--client", required=True)
    ap.add_argument("--folder")
    ap.add_argument("--subject", action="append")
    ap.add_argument("--reuse",
                    help="复用 run_pipeline 已有产物（产物目录，或 *__打标流水.csv/*__整合流水.csv），"
                         "跳过重复跑标准化/整合/打标，直接组装交付物")
    ap.add_argument("--account-type", choices=["对公", "个人", "未知"])
    ap.add_argument("--out-dir")
    ap.add_argument("--sleep-seconds", type=float, default=0,
                    help="每处理完一个候选流水文件后的暂停秒数")
    args = ap.parse_args()
    if not args.folder and not args.subject and not args.reuse:
        ap.error("需提供 --folder、--subject 或 --reuse 之一")
    if args.reuse:
        run_reuse(args.client, args.reuse, args.out_dir or os.getcwd())
    else:
        run(args.client, args)


if __name__ == "__main__":
    main()

