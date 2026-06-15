#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""银行流水反欺诈核查 —— 主入口。

依据《惠裕贷银行流水反欺诈特征体系与人工核查机制 V1.0》实现：
  L1 文件取证（可选 --raw-folder）/ L2 数理一致性 / L3 行为模式 /
  L4 经营真实性锚点 / L5 隐性负债与民间借贷 / L6 完整性与交叉验证，
  A/B/C 三级规则引擎计分，人工核查工单（含行级证据、动作字典 D01–D10、可解除标准）。

输入：bank-statement-standardization 的标准化交付物
  （含「整合打标流水」表的 .xlsx，或 *__打标流水.csv / *__整合流水.csv）。
输出：<客户名>__反欺诈核查.xlsx + <客户名>__反欺诈报告.json。

核心原则（与文档一致）：机器不下最终结论——本脚本产出带证据的核查工单，由人工确认定性。
"""

import argparse
import csv
import json
import math
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from glob import glob

# ---------------------------------------------------------------------------
# 阈值集中配置（文档第三部分「初始阈值建议」，冷启动值，上线后按 4.4 机制校准）
# ---------------------------------------------------------------------------
THRESHOLDS = {
    "F2-03_结息偏差": 0.20,          # 单周期复算偏差>20% 触发（A）
    # 未指定 --rate 时逐一尝试的活期挂牌利率（覆盖 2015–2026 各档：2024–2026 降息后多为 0.05%–0.10%）
    "F2-03_候选利率": [0.0005, 0.0010, 0.0015, 0.0020, 0.0025, 0.0030, 0.0035],
    "F2-04_MAD": 0.015,              # 首位 Benford MAD
    "F2-04_最少笔数": 300,
    "F2-04_最短区间天": 90,
    "F2-05_整数占比": 0.60,          # 行业 P95 缺失时的保守默认值（可按行业调整）
    "F2-05_奇异尾数占比": 0.10,
    "F2-06_CV低": 0.05,
    "F2-06_CV高": 2.0,
    "F3-01_金额偏差": 0.05,          # |入-出|/入 < 5%
    "F3-01_小额线": 5000.0,
    "F3-01_报送线": [50000.0, 500000.0],   # 贴近大额报送线 = [0.9×线, 线)
    "F3-01_月内次数": 3,
    "F3-02_沉淀率": 0.02,
    "F3-03_回流窗口天": 7,
    "F3-03_金额偏差": 0.10,
    "F3-04_对倒对手数": 5,
    "F3-04_平衡度": 0.80,            # 月内同对手 min(入,出)/max(入,出) ≥ 0.8 视为对倒
    "F3-05_冲量倍数": 2.0,
    "F3-06_非营业占比": 0.20,
    "F4-01_断缴月数": 2,
    "F4-01_骤降幅度": 0.50,
    "F4-03_税负率": 0.03,            # 默认行业综合税负率，--tax-rate 覆盖
    "F4-03_偏差": 0.30,
    "F4-07_存在性项数": 2,
    "F5-02_借入门槛": 500000.0,
    "F5-02_月息区间": (0.015, 0.035),
    "F5-02_持续期数": 3,
    "F5-03_缺口比例": 0.30,
    "F5-04_倒贷指数": 6.0,
    "F5-05_借款收入占比": 0.30,
    "F6-02_覆盖率": 0.60,
    "F6-03_相关系数": 0.90,
    "F6-03_最短重叠天": 60,
    "F6-04_匹配率": 0.40,
    "F6-04_金额窗口": 0.05,
    "F6-04_日期窗口天": 15,
    "进入人审分数": 25,
    "共振条数": 3,
}

# B 级软规则权重（文档 4.2，专家赋权三档制）
B_WEIGHTS = {
    "F3-01": 15, "F3-03": 15, "F4-03": 15, "F5-01": 15, "F5-02": 15, "F6-03": 15,
    "F2-02": 10, "F3-02": 10, "F3-04": 10, "F3-05": 10, "F4-01": 10, "F4-02": 10,
    "F4-04": 10, "F5-03": 10, "F5-04": 10, "F5-05": 10, "F6-02": 10, "F6-05": 10,
    "F1-02": 5, "F1-03": 5, "F2-04": 5, "F2-05": 5, "F4-05": 5,
}

# 规则 → 核查动作映射（文档附录 B）与可解除标准（文档 5.4 / 5.3）
ACTIONS = {
    "F1-01": ("D01 柜面打印带章回单核对", "要求客户经柜面或银企直连重新获取流水",
              "重新获取的官方渠道流水与入件流水逐笔一致方可解除"),
    "F1-04": ("D01 柜面打印带章回单核对", "要求客户经柜面或银企直连重新获取流水",
              "官方渠道重取流水验证一致方可解除"),
    "F2-01": ("D01 柜面带章回单核对", "D05 补充账户流水",
              "带章回单与入件流水逐笔一致，或断点可由缺失区间补件解释"),
    "F2-02": ("D01 柜面带章回单核对", "D05 补充账户流水",
              "补件后结息次数与覆盖区间推算一致"),
    "F2-03": ("D01 柜面带章回单核对", "加查 F2-05 尾数特征",
              "带章回单逐笔一致且结息金额与回单一致方可解除"),
    "F2-04": ("D01 柜面带章回单核对", "D02 发票逐笔比对",
              "样本量或交易结构可解释分布偏离（如定价惯例导致整数集中）"),
    "F2-05": ("D01 柜面带章回单核对", "D02 发票逐笔比对",
              "行业定价惯例可解释整数/尾数分布"),
    "F3-01": ("D08 资金管理协议/合同", "D06 面谈支付对象",
              "提供真实有效的集团资金管理协议且资金路径与协议一致"),
    "F3-02": ("D04 实地走访", "D02 发票逐笔比对",
              "现场产能与流水规模匹配，低沉淀可由经营模式（如代收代付）解释"),
    "F3-03": ("D07 征信与放款路径比对", "D10 反洗钱评估",
              "贷前：环路资金可由真实购销关系解释；贷后回流一经确认即认定挪用，不可解除"),
    "F3-04": ("D08 资金管理协议/合同", "D06 面谈支付对象",
              "提供真实有效的集团资金管理协议且资金路径与协议一致（定时、定向、工商关联）"),
    "F3-05": ("D02 发票逐笔比对", "D03 对手函证",
              "冲量区间入账可与发票/合同逐笔对应且事后未反向流出"),
    "F4-01": ("D04 实地走访（水电表、现场盘点）", "D09 税务数据核对",
              "现场产能与流水规模匹配；断缴可由账户切换/预存代扣解释并提供凭证"),
    "F4-02": ("D04 实地走访", "D09 税务数据核对",
              "提供工资表与社保缴纳凭证，与申报人数口径一致"),
    "F4-03": ("D09 税务申报数据核对", "D02 发票逐笔比对",
              "申报数据与流水推算口径差异可合理解释（如免税项目、核定征收）"),
    "F4-04": ("D03 电话/函证核实交易对手", "D06 面谈支付对象",
              "对手确认真实购销关系并可提供合同"),
    "F4-05": ("D04 实地走访", "D02 发票逐笔比对",
              "可由经营模式转变、大客户订单节奏等真实原因解释季节背离"),
    "F5-01": ("D06 面谈询问指定对私支付对象", "D07 征信明细逐项比对",
              "能说明资金性质且与征信/合同印证；已在征信或申报负债中体现的降级为已披露负债"),
    "F5-02": ("D06 面谈询问指定对私支付对象", "D07 征信明细逐项比对",
              "提供借款合同且该借贷已在征信或申报负债中体现，降级为已披露负债处理"),
    "F5-03": ("D07 征信明细逐项比对", "D05 补充流水",
              "缺口可由已披露负债（含申报民间借贷）解释"),
    "F5-04": ("D07 征信明细逐项比对", "D06 面谈支付对象",
              "滚动借贷可由正常经营周转性融资解释且无过桥序列"),
    "F5-05": ("D07 征信明细逐项比对", "D06 面谈支付对象",
              "借款用途明确、与经营扩张匹配且已在征信体现"),
    "F6-01": ("D05 要求补充指定账户流水", "（拒不补充直接升级 A 级 R-A04）",
              "补齐后整体校验通过；拒不补充升级 A 级拒贷/中止"),
    "F6-02": ("D02 发票逐笔比对", "D05 补充账户流水",
              "补充账户流水或现金日记账后覆盖率回升至阈值以上"),
    "F6-03": ("D01 柜面带章回单核对", "D05 补充账户流水",
              "官方渠道重取后相关性消失，或可由真实关联交易（如集团内购销）解释"),
    "F6-04": ("D02 发票逐笔比对", "D03 对手函证",
              "补充发票后匹配率回升至阈值以上"),
    "F6-05": ("D09 税务申报数据核对", "D02 发票逐笔比对",
              "申报数据与流水推算口径差异可合理解释；注意区分偏差方向（虚增收入 vs 隐匿收入）"),
}

STD_SHEETS = ["整合打标流水", "标准化流水主表", "主表", "流水", "打标流水", "整合流水"]

CORP_HINTS = ("公司", "有限", "银行", "支行", "分行", "合作社", "商行", "信用社", "厂",
              "店", "部", "中心", "集团", "合伙", "事务所", "学校", "医院", "局", "委员会")
EDITOR_BLACKLIST = re.compile(
    r"photoshop|illustrator|美图|gimp|canva|wps|microsoft\s*word|万兴|wondershare|"
    r"foxit\s*(phantom|editor)|福昕.*编辑|ilovepdf|sejda|smallpdf|nitro|pdfelement|pdf\s*editor",
    re.I)

UTILITY_RE = re.compile(r"水费|电费|燃气|煤气|自来水|电力|水电")
TAX_RE = re.compile(r"税|国库|TIPS", re.I)
INTEREST_RE = re.compile(r"结息|利息")
MISC_CATEGORIES = {
    "财务/管理软件": re.compile(r"软件服务|财务软件|用友|金蝶|ERP|畅捷通", re.I),
    "物流快递": re.compile(r"物流|快递|运费|货运|运输"),
    "平台佣金": re.compile(r"佣金|平台服务费|技术服务费"),
    "房租物业": re.compile(r"房租|租金|物业"),
    "办公耗材": re.compile(r"办公用品|文具|耗材|打印"),
}
BORROW_L2 = {"银行借款", "民间借贷", "负债", "其他筹资类", "消费金融"}


# ---------------------------------------------------------------------------
# 基础工具
# ---------------------------------------------------------------------------
def norm_name(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    return re.sub(r"[\s()（）·*＊]", "", s)


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("，", "").replace("¥", "").strip()
    if not s or s in ("-", "—"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def parse_dt(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip().replace("/", "-").replace(".", "-")
    m = re.match(r"(\d{4})-(\d{1,2})-(\d{1,2})[ T]?(\d{1,2})?:?(\d{1,2})?:?(\d{1,2})?", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                            int(m.group(4) or 0), int(m.group(5) or 0), int(m.group(6) or 0))
        except ValueError:
            return None
    m = re.match(r"(\d{4})(\d{2})(\d{2})(\d{2})?(\d{2})?(\d{2})?", s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)),
                            int(m.group(4) or 0), int(m.group(5) or 0), int(m.group(6) or 0))
        except ValueError:
            return None
    return None


def month_key(d):
    return f"{d.year:04d}-{d.month:02d}"


def month_range(d1, d2):
    out, y, m = [], d1.year, d1.month
    while (y, m) <= (d2.year, d2.month):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            y, m = y + 1, 1
    return out


def pearson(xs, ys):
    n = len(xs)
    if n < 3:
        return None
    mx, my = sum(xs) / n, sum(ys) / n
    sx = math.sqrt(sum((x - mx) ** 2 for x in xs))
    sy = math.sqrt(sum((y - my) ** 2 for y in ys))
    if sx == 0 or sy == 0:
        return None
    return sum((x - mx) * (y - my) for x, y in zip(xs, ys)) / (sx * sy)


def is_personal(cp_name, cp_acct):
    name = norm_name(cp_name)
    if not name:
        return False
    if any(h in name for h in CORP_HINTS):
        return False
    if len(name) <= 4:
        return True
    return bool(cp_acct and re.sub(r"\D", "", str(cp_acct)).startswith("62"))


# ---------------------------------------------------------------------------
# 数据加载
# ---------------------------------------------------------------------------
COLMAP = {  # 标准字段名 → 行内键
    "交易唯一编号": "txid", "客户名称": "client", "主体名称": "subject", "账户类型": "acct_type",
    "本方名称": "self_name", "本方账户": "acct", "开户行": "bank", "交易时间": "time",
    "对手名称": "cp", "对手账户": "cp_acct", "收入金额": "in", "支出金额": "out",
    "交易金额": "amt", "账户余额": "bal", "银行备注": "memo1", "账户方附言": "memo2",
    "收支方向": "dir", "一级标签": "l1", "二级标签": "l2", "三级标签": "l3",
    "来源文件名": "src", "来源行号": "line",
}


def _rows_from_table(header, data_rows):
    idx = {}
    for i, h in enumerate(header):
        h = str(h or "").strip()
        if h in COLMAP:
            idx[COLMAP[h]] = i
    if "bal" not in idx or ("in" not in idx and "amt" not in idx):
        return None
    rows = []
    for raw in data_rows:
        if raw is None or all(v in (None, "") for v in raw):
            continue
        g = lambda k: raw[idx[k]] if k in idx and idx[k] < len(raw) else None
        dt = parse_dt(g("time"))
        amt_in, amt_out, amt = to_num(g("in")), to_num(g("out")), to_num(g("amt"))
        if amt is None:
            amt = (amt_in or 0.0) - (amt_out or 0.0) if (amt_in is not None or amt_out is not None) else None
        if amt_in is None and amt is not None and amt > 0:
            amt_in = amt
        if amt_out is None and amt is not None and amt < 0:
            amt_out = -amt
        memo = " ".join(str(g(k) or "") for k in ("memo1", "memo2")).strip()
        rows.append({
            "dt": dt, "date": dt.date() if dt else None,
            "subject": str(g("subject") or g("self_name") or "").strip(),
            "self_name": str(g("self_name") or "").strip(),
            "acct": str(g("acct") or "").strip(), "acct_type": str(g("acct_type") or "").strip(),
            "bank": str(g("bank") or "").strip(),
            "cp": str(g("cp") or "").strip(), "cp_acct": str(g("cp_acct") or "").strip(),
            "in": amt_in or 0.0, "out": amt_out or 0.0, "amt": amt, "bal": to_num(g("bal")),
            "memo": memo, "l1": str(g("l1") or "").strip(), "l2": str(g("l2") or "").strip(),
            "l3": str(g("l3") or "").strip(),
            "src": str(g("src") or "").strip(), "line": str(g("line") or "").strip(),
        })
    return rows


def load_rows(path):
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        order = [n for c in STD_SHEETS for n in names if c in n] + names
        for sn in dict.fromkeys(order):
            ws = wb[sn]
            it = ws.iter_rows(values_only=True)
            try:
                header = next(it)
            except StopIteration:
                continue
            rows = _rows_from_table(list(header), it)
            if rows:
                return rows
        raise SystemExit(f"【错误】{path} 中未找到含标准字段的流水工作表")
    with open(path, newline="", encoding="utf-8-sig") as f:
        rd = csv.reader(f)
        try:
            header = next(rd)
        except StopIteration:
            raise SystemExit(f"【错误】{path} 为空")
        rows = _rows_from_table(header, list(rd))
        if not rows:
            raise SystemExit(f"【错误】{path} 缺少标准字段（账户余额 + 收入/交易金额）")
        return rows


def pick_input(input_path):
    if os.path.isfile(input_path):
        return input_path
    pats = ["*__打标流水.csv", "*__整合流水.csv", "*已清洗*待分析*.xlsx", "*标准化*.xlsx", "*.xlsx", "*.csv"]
    for p in pats:
        hits = [h for h in sorted(glob(os.path.join(input_path, "**", p), recursive=True))
                if "反欺诈" not in h and "BI分析" not in h and not os.path.basename(h).startswith("~")]
        if hits:
            return hits[0]
    raise SystemExit(f"【错误】{input_path} 下未找到标准化产物")


# ---------------------------------------------------------------------------
# 公共口径
# ---------------------------------------------------------------------------
class Ctx:
    """跑批上下文：成员名单、经营/借款口径、组合日余额等公共派生数据。"""

    def __init__(self, rows, args):
        self.rows = rows
        self.args = args
        self.members = {norm_name(n) for r in rows for n in (r["subject"], r["self_name"]) if n}
        self.self_accts = {re.sub(r"\D", "", r["acct"]) for r in rows if r["acct"]}
        self.self_accts.discard("")
        self.self_last4 = {a[-4:] for a in self.self_accts if len(a) >= 4}
        dts = [r["date"] for r in rows if r["date"]]
        if not dts:
            raise SystemExit("【错误】流水缺少可解析的交易时间")
        self.d_min, self.d_max = min(dts), max(dts)
        self.span_days = (self.d_max - self.d_min).days + 1
        self.months = month_range(self.d_min, self.d_max)
        self.labeled_ratio = (sum(1 for r in rows if r["l1"] and r["l1"] != "其他类") / len(rows)) if rows else 0.0

    def is_internal(self, r):
        return norm_name(r["cp"]) in self.members if r["cp"] else False

    def is_interest(self, r):
        return r["l2"] == "存款结息" or bool(INTEREST_RE.search(r["memo"])) and "贷款利息" not in r["memo"]

    def is_oper_in(self, r):
        """经营性入账：打标可用时取经营类收入；否则启发式（剔内部/结息/借款类关键词）。"""
        if r["in"] <= 0 or self.is_internal(r):
            return False
        if self.labeled_ratio >= 0.5:
            return r["l1"] == "经营类"
        return not self.is_interest(r) and not re.search(r"借款|贷款|融资|投资款", r["memo"])

    def is_borrow_in(self, r):
        if r["in"] <= 0 or self.is_internal(r):
            return False
        if self.labeled_ratio >= 0.5:
            return r["l1"] == "筹资类" and (not r["l2"] or r["l2"] in BORROW_L2)
        return bool(re.search(r"借款|贷款|拆借|过桥|转贷", r["memo"]))

    def is_repay_out(self, r):
        if r["out"] <= 0 or self.is_internal(r):
            return False
        if self.labeled_ratio >= 0.5:
            return r["l1"] == "筹资类"
        return bool(re.search(r"还款|还贷|还本|付息|归还", r["memo"]))

    # ---- 组合日余额（各账户日末余额前向填充后逐日求和，口径同 portfolio_balance.py）
    def daily_portfolio_balance(self):
        per_acct_daily = {}
        for r in self.rows:
            if r["date"] and r["bal"] is not None:
                per_acct_daily.setdefault(r["acct"] or r["src"], {})[r["date"]] = r["bal"]
        days = [self.d_min + timedelta(days=i) for i in range(self.span_days)]
        total = {d: 0.0 for d in days}
        for daily in per_acct_daily.values():
            last = None
            first_day = min(daily)
            for d in days:
                if d in daily:
                    last = daily[d]
                if last is not None and d >= first_day:
                    total[d] += last
        return total


def ev(r, note=""):
    """行级证据（文档 5.2：精确到日期、金额、对手、摘要、所在文件与行号）。"""
    return {
        "交易时间": r["dt"].strftime("%Y-%m-%d %H:%M:%S") if r["dt"] else "",
        "主体名称": r["subject"], "本方账户": r["acct"],
        "对手名称": r["cp"], "收入金额": r["in"] or "", "支出金额": r["out"] or "",
        "账户余额": r["bal"], "摘要": r["memo"][:80],
        "来源文件名": r["src"], "来源行号": r["line"], "备注": note,
    }


def feat(code, name, level, triggered, summary, metrics=None, evidence=None,
         pending=None, snapshot=""):
    a = ACTIONS.get(code, ("", "", ""))
    return {
        "编号": code, "名称": name, "级别": level, "命中": bool(triggered),
        "权重": B_WEIGHTS.get(code, 0) if level.startswith("B") else 0,
        "结论摘要": summary, "关键指标": metrics or {},
        "证据": (evidence or [])[:60], "证据总数": len(evidence or []),
        "待外部数据": pending or "", "计算过程快照": snapshot,
        "首选核查动作": a[0], "次选核查动作": a[1], "可解除标准": a[2],
    }


# ---------------------------------------------------------------------------
# L2 数理一致性层
# ---------------------------------------------------------------------------
def f2_01(ctx):
    """余额连续性试算：按账户、按输入行序（上游已按余额连续性还原，绝不按时间重排）。"""
    breaks, per_acct = [], defaultdict(int)
    by_acct = defaultdict(list)
    for r in ctx.rows:
        if r["bal"] is not None and r["amt"] is not None:
            by_acct[(r["subject"], r["acct"])].append(r)
    checked = 0
    for key, rs in by_acct.items():
        prev = None
        for r in rs:
            if prev is not None:
                checked += 1
                if abs(prev + r["amt"] - r["bal"]) > 0.01:
                    breaks.append(ev(r, f"上笔余额{prev:.2f}+本笔{r['amt']:.2f}≠{r['bal']:.2f}"))
                    per_acct[key[1]] += 1
            prev = r["bal"]
    n = len(breaks)
    snap = f"按 主体+本方账户 分组、按输入行序逐笔校验 {checked} 个衔接点，断点 {n} 个；" \
           f"断点账户分布：{dict(per_acct) or '无'}。注意：跨月/跨文件边界断点可能为对账单区间缺失。"
    return feat("F2-01", "余额连续性试算", "A", n >= 1,
                f"断点 {n} 个" + ("（立案必查 R-A01）" if n else "，校验通过"),
                {"断点数": n, "校验衔接点": checked}, breaks, snapshot=snap)


def _interest_rows(ctx):
    return [r for r in ctx.rows if r["in"] > 0 and ctx.is_interest(r) and not ctx.is_internal(r)
            and (r["l2"] == "存款结息" or re.search(r"结息|存款利息", r["memo"]))]


def _expected_interest_dates(d1, d2):
    out = []
    for y in range(d1.year, d2.year + 1):
        for m in (3, 6, 9, 12):
            d = date(y, m, 21)
            if d1 <= d <= d2:
                out.append(d)
    return out


def f2_02(ctx):
    by_acct = defaultdict(list)
    for r in ctx.rows:
        if r["date"] and r["bal"] is not None:
            by_acct[(r["subject"], r["acct"])].append(r)
    details, bad = [], False
    inter = defaultdict(list)
    for r in _interest_rows(ctx):
        inter[(r["subject"], r["acct"])].append(r)
    evid = []
    for key, rs in by_acct.items():
        ds = [r["date"] for r in rs]
        exp = _expected_interest_dates(min(ds), max(ds))
        act = inter.get(key, [])
        if not exp:
            continue
        ok = len(act) >= len(exp)
        if not ok:
            bad = True
            evid.extend(ev(r) for r in act)
        details.append({"账户": key[1], "主体": key[0], "应结次数": len(exp),
                        "实结次数": len(act), "判定": "一致" if ok else "缺结息"})
    snap = "按账户覆盖区间推算季结息日（3/6/9/12 月 21 日），与摘要含结息/存款利息的实际入账比对：" + \
           json.dumps(details, ensure_ascii=False)
    return feat("F2-02", "结息次数检查", "B", bad,
                "存在账户应结≠实结" if bad else "各账户结息次数与区间推算一致",
                {"账户明细": details}, evid, snapshot=snap)


def f2_03(ctx):
    """结息金额复算：应结利息≈Σ(每日日终余额)×活期挂牌利率/360，按结息周期归集比对。
    未指定 --rate 时对候选挂牌利率逐一尝试，全部偏差>20% 才触发，降低利率档不确定性误伤。"""
    rates = [ctx.args.rate] if ctx.args.rate else THRESHOLDS["F2-03_候选利率"]
    by_acct = defaultdict(list)
    for r in ctx.rows:
        if r["date"] and r["bal"] is not None:
            by_acct[(r["subject"], r["acct"])].append(r)
    inter = defaultdict(list)
    for r in _interest_rows(ctx):
        inter[(r["subject"], r["acct"])].append(r)
    periods, evid, triggered = [], [], False
    for key, rs in by_acct.items():
        daily = {}
        for r in rs:
            daily[r["date"]] = r["bal"]
        d0, d1 = min(daily), max(daily)
        days, last, filled = [], None, {}
        d = d0
        while d <= d1:
            if d in daily:
                last = daily[d]
            filled[d] = last
            d += timedelta(days=1)
        for ir in sorted(inter.get(key, []), key=lambda r: r["date"]):
            end = ir["date"]
            start = end - timedelta(days=91)
            q_start = date(end.year if end.month > 3 else end.year - 1,
                           {3: 12, 6: 3, 9: 6, 12: 9}.get(end.month, end.month), 21)
            start = max(start, q_start)
            if start < d0:   # 周期未被流水完整覆盖，跳过避免误报
                continue
            sum_bal = sum(filled[start + timedelta(days=i)] or 0.0
                          for i in range((end - start).days))
            best = None
            for rate in rates:
                calc = sum_bal * rate / 360.0
                dev = abs(ir["in"] - calc) / max(calc, 0.01)
                if best is None or dev < best[1]:
                    best = (rate, dev, calc)
            rate, dev, calc = best
            bad = dev > THRESHOLDS["F2-03_结息偏差"]
            periods.append({"账户": key[1], "结息日": str(end), "实际结息": round(ir["in"], 2),
                            "复算值": round(calc, 2), "最优利率": rate,
                            "偏差": round(dev, 4), "判定": "异常" if bad else "正常"})
            if bad:
                triggered = True
                evid.append(ev(ir, f"复算{calc:.2f} vs 实际{ir['in']:.2f}，偏差{dev:.0%}"))
    snap = ("逐结息周期复算（Σ每日日终余额×利率/360，余额前向填充，仅评估流水完整覆盖的周期；"
            f"候选利率 {rates}）：" + json.dumps(periods, ensure_ascii=False))
    note = "" if periods else "；流水区间内无可评估的完整结息周期"
    return feat("F2-03", "结息金额复算", "A", triggered,
                ("存在结息周期复算偏差>20%（立案必查 R-A02，同比例放大流水的典型破绽）" if triggered
                 else "已评估周期复算偏差均在容差内" + note),
                {"评估周期数": len(periods), "异常周期数": sum(1 for p in periods if p["判定"] == "异常")},
                evid, snapshot=snap)


def f2_04(ctx):
    pool0 = [r for r in ctx.rows if (ctx.is_oper_in(r) or
             (r["out"] > 0 and not ctx.is_internal(r) and r["l1"] == "经营类"))]
    amt_cnt = Counter(round(r["in"] or r["out"], 2) for r in pool0)
    pool = [r for r in pool0 if r["l2"] not in ("人力成本", "存款结息")
            and not re.search(r"租金|房租|还贷|结息", r["memo"])
            and amt_cnt[round(r["in"] or r["out"], 2)] < 5]
    n = len(pool)
    applicable = n >= THRESHOLDS["F2-04_最少笔数"] and ctx.span_days >= THRESHOLDS["F2-04_最短区间天"]
    mad, digit_dist = None, {}
    over_digits = []
    if applicable:
        firsts = Counter()
        for r in pool:
            s = re.sub(r"[^1-9]", "", f"{(r['in'] or r['out']):.2f}")
            if s:
                firsts[int(s[0])] += 1
        tot = sum(firsts.values())
        mad = 0.0
        for d in range(1, 10):
            obs = firsts.get(d, 0) / tot
            exp = math.log10(1 + 1 / d)
            digit_dist[d] = {"实际": round(obs, 4), "理论": round(exp, 4)}
            mad += abs(obs - exp)
            if obs - exp > 0.02:
                over_digits.append(d)
        mad /= 9.0
    trig = applicable and mad is not None and mad > THRESHOLDS["F2-04_MAD"]
    evid = []
    if trig and over_digits:
        evid = [ev(r, "首位数字超额样本") for r in pool
                if int(re.sub(r"[^1-9]", "", f"{(r['in'] or r['out']):.2f}")[0:1] or 0) in over_digits][:60]
    snap = (f"剔除固定金额交易（工资/租金/还贷/结息及同金额≥5次）后样本 {n} 笔，区间 {ctx.span_days} 天；"
            + (f"首位 MAD={mad:.4f}，超额数字={over_digits}，分布={json.dumps(digit_dist, ensure_ascii=False)}"
               if applicable else "不满足适用前提（笔数≥300 且区间≥3 个月），不评估"))
    return feat("F2-04", "Benford 分布检验", "B", trig,
                (f"首位 MAD={mad:.4f}>0.015，数字{over_digits}显著超额" if trig
                 else ("分布未见显著偏离" if applicable else "样本量/区间不满足适用前提，未评估")),
                {"样本笔数": n, "MAD": round(mad, 4) if mad is not None else None}, evid, snapshot=snap)


def f2_05(ctx):
    pool = [r for r in ctx.rows if ctx.is_oper_in(r) and not ctx.is_interest(r)]
    n = len(pool)
    if n < 30:
        return feat("F2-05", "整数与尾数分布", "B", False, "经营性入账样本不足 30 笔，未评估",
                    {"样本笔数": n})
    ints = [r for r in pool if r["in"] >= 100 and abs(r["in"] % 100) < 0.005]
    odd = [r for r in pool if round(r["in"] * 100) % 10 != 0]
    ir, orr = len(ints) / n, len(odd) / n
    trig = ir > THRESHOLDS["F2-05_整数占比"] or orr > THRESHOLDS["F2-05_奇异尾数占比"]
    snap = (f"经营性入账 {n} 笔：整百及以上整数 {len(ints)} 笔（{ir:.1%}，参考阈值>{THRESHOLDS['F2-05_整数占比']:.0%}，"
            f"行业 P95 缺失时的保守默认）；分位非零『奇异尾数』{len(odd)} 笔（{orr:.1%}，阈值>10%，"
            "疑似乘系数放大产生的非自然小数）。")
    return feat("F2-05", "整数与尾数分布", "B", trig,
                f"整数占比 {ir:.1%}、奇异尾数占比 {orr:.1%}" + ("，超阈值" if trig else "，正常"),
                {"整数占比": round(ir, 4), "奇异尾数占比": round(orr, 4)},
                ([ev(r) for r in (ints if ir > THRESHOLDS['F2-05_整数占比'] else odd)[:40]] if trig else []),
                snapshot=snap)


def f2_06(ctx):
    monthly = defaultdict(float)
    for r in ctx.rows:
        if ctx.is_oper_in(r) and r["date"]:
            monthly[month_key(r["date"])] += r["in"]
    xs = [monthly.get(m, 0.0) for m in ctx.months]
    if len(xs) < 3 or sum(xs) == 0:
        return feat("F2-06", "入账平稳度（变异系数）", "C", False, "月份数不足，未评估", {})
    mu = sum(xs) / len(xs)
    cv = math.sqrt(sum((x - mu) ** 2 for x in xs) / len(xs)) / mu if mu else None
    flag = cv is not None and (cv < THRESHOLDS["F2-06_CV低"] or cv > THRESHOLDS["F2-06_CV高"])
    return feat("F2-06", "入账平稳度（变异系数）", "C", flag,
                f"月度经营性入账 CV={cv:.3f}" + ("（异常低提示等额做账 / 异常高提示冲量，留痕观察）" if flag else ""),
                {"CV": round(cv, 4), "月均经营入账": round(mu, 2)},
                snapshot=f"月度经营性入账={json.dumps({m: round(monthly.get(m, 0), 2) for m in ctx.months}, ensure_ascii=False)}；与营销响应模型『入账资金平稳度』共用口径。")


# ---------------------------------------------------------------------------
# L3 行为模式层
# ---------------------------------------------------------------------------
def _related_cps(ctx):
    """关联对手：内部成员 + 双向高频互转对手（与借款人互转≥3 笔且双向都有）。"""
    pair = defaultdict(lambda: [0, 0])
    for r in ctx.rows:
        cp = norm_name(r["cp"])
        if not cp:
            continue
        if r["in"] > 0:
            pair[cp][0] += 1
        if r["out"] > 0:
            pair[cp][1] += 1
    related = set(ctx.members)
    for cp, (i, o) in pair.items():
        if i >= 3 and o >= 3:
            related.add(cp)
    return related


def _amount_band_hit(x):
    if x < THRESHOLDS["F3-01_小额线"]:
        return True
    for line in THRESHOLDS["F3-01_报送线"]:
        if 0.9 * line <= x < line:
            return True
    return False


def f3_01(ctx):
    related = _related_cps(ctx)
    by_acct = defaultdict(list)
    for r in ctx.rows:
        if r["date"]:
            by_acct[(r["subject"], r["acct"])].append(r)
    pairs, per_month = [], Counter()
    for key, rs in by_acct.items():
        ins = [r for r in rs if r["in"] > 0 and norm_name(r["cp"]) in related and _amount_band_hit(r["in"])]
        outs = [r for r in rs if r["out"] > 0 and norm_name(r["cp"]) in related]
        used = set()
        for ri in ins:
            for j, ro in enumerate(outs):
                if j in used or ro["date"] < ri["date"]:
                    continue
                if (ro["date"] - ri["date"]).days > 1:
                    continue
                if abs(ri["in"] - ro["out"]) / ri["in"] < THRESHOLDS["F3-01_金额偏差"]:
                    used.add(j)
                    pairs.append((ri, ro))
                    per_month[month_key(ri["date"])] += 1
                    break
    hot = {m: c for m, c in per_month.items() if c >= THRESHOLDS["F3-01_月内次数"]}
    trig = bool(hot)
    evid = []
    for ri, ro in pairs[:30]:
        evid.append(ev(ri, "快进"))
        evid.append(ev(ro, f"快出（T+{(ro['date']-ri['date']).days}，差{abs(ri['in']-ro['out'])/ri['in']:.1%}）"))
    snap = (f"T+0/T+1 配对、|入−出|/入<5%、对手限关联人（内部成员+双向互转≥3 笔对手）、"
            f"金额<5000 或贴近 5万/50万报送线；配对 {len(pairs)} 组，月分布 {dict(per_month)}，"
            f"命中月份（≥3 次/月）：{hot or '无'}")
    return feat("F3-01", "快进快出匹配", "B", trig,
                f"命中月份 {list(hot)}，合计配对 {len(pairs)} 组" if trig else f"配对 {len(pairs)} 组，未达月内 3 次阈值",
                {"配对组数": len(pairs), "命中月份": hot}, evid, snapshot=snap)


def f3_02(ctx):
    daily = ctx.daily_portfolio_balance()
    avg_bal = sum(daily.values()) / len(daily) if daily else 0.0
    oper_in = sum(r["in"] for r in ctx.rows if ctx.is_oper_in(r))
    repay = sum(r["out"] for r in ctx.rows if ctx.is_repay_out(r))
    n_mon = max(len(ctx.months), 1)
    m_in, m_repay = oper_in / n_mon, repay / n_mon
    rate = avg_bal / m_in if m_in else None
    cover = avg_bal / m_repay if m_repay else None
    trig = rate is not None and rate < THRESHOLDS["F3-02_沉淀率"]
    snap = (f"组合日均余额（各账户日末余额前向填充逐日求和，日历日加权）={avg_bal:,.2f}；"
            f"月均经营性流入={m_in:,.2f}；沉淀率={rate if rate is None else round(rate, 4)}；"
            f"月均还款额={m_repay:,.2f}，偿债余力比（日均余额/月均还款，参考 1:1）="
            f"{cover if cover is None else round(cover, 2)}")
    return feat("F3-02", "资金沉淀率", "B", trig,
                (f"沉淀率 {rate:.2%} < 2%（流量大、沉淀低，养流水典型形态）" if trig
                 else (f"沉淀率 {rate:.2%}" if rate is not None else "无经营性流入，未评估")),
                {"日均余额": round(avg_bal, 2), "月均经营性流入": round(m_in, 2),
                 "沉淀率": round(rate, 4) if rate is not None else None,
                 "偿债余力比": round(cover, 2) if cover is not None else None}, snapshot=snap)


def f3_03(ctx):
    """资金回流（简化口径）：同一外部对手 流出→7 日内等额(±10%)流入（A→B→A 直接往返）。
    仅凭单方流水只能看到本方边，跨外部多节点环路需外部账户数据，已在快照中声明。
    贷后模式（--loan-date）：贷款发放后命中的回流按 R-A06 升 A。"""
    win, tol = THRESHOLDS["F3-03_回流窗口天"], THRESHOLDS["F3-03_金额偏差"]
    by_cp = defaultdict(lambda: {"in": [], "out": []})
    for r in ctx.rows:
        cp = norm_name(r["cp"])
        if not cp or ctx.is_internal(r) or not r["date"]:
            continue
        if r["in"] > 0:
            by_cp[cp]["in"].append(r)
        if r["out"] > 0:
            by_cp[cp]["out"].append(r)
    loan_date = parse_dt(ctx.args.loan_date).date() if ctx.args.loan_date else None
    pairs, post_loan = [], []
    for cp, g in by_cp.items():
        used = set()
        for ro in g["out"]:
            for k, ri in enumerate(g["in"]):
                if k in used:
                    continue
                dd = (ri["date"] - ro["date"]).days
                if 0 < dd <= win and abs(ri["in"] - ro["out"]) / ro["out"] <= tol:
                    used.add(k)
                    pairs.append((ro, ri))
                    if loan_date and ro["date"] >= loan_date:
                        post_loan.append((ro, ri))
                    break
    evid = []
    for ro, ri in pairs[:25]:
        evid.append(ev(ro, "流出"))
        evid.append(ev(ri, f"{(ri['date']-ro['date']).days} 日内等额回流"))
    is_a = bool(post_loan)
    trig = len(pairs) > 0
    snap = (f"简化口径：仅检测『流出→{win} 日内同一外部对手等额(±{tol:.0%})流入』直接往返 "
            f"{len(pairs)} 组；跨外部多节点环路（A→B→C→A）需对手账户数据/图计算平台，无法仅凭单方流水识别。"
            + (f" 贷后模式：贷款发放日 {loan_date} 之后回流 {len(post_loan)} 组（R-A06 认定挪用评估）。" if loan_date else ""))
    return feat("F3-03", "资金回流闭环", "A" if is_a else "B", trig,
                ("贷款发放后存在资金回流（R-A06：认定挪用，触发提前收回条款评估并上报）" if is_a else
                 (f"贷前检出 {len(pairs)} 组短窗口等额往返" if trig else "未检出短窗口等额往返")),
                {"往返组数": len(pairs), "贷后回流组数": len(post_loan)}, evid, snapshot=snap)


def f3_04(ctx):
    seen_before = defaultdict(set)
    monthly_cp = defaultdict(lambda: defaultdict(lambda: [0.0, 0.0]))
    for r in ctx.rows:
        if not r["date"] or ctx.is_internal(r) or not r["cp"]:
            continue
        cp = norm_name(r["cp"])
        mk = month_key(r["date"])
        monthly_cp[mk][cp][0] += r["in"]
        monthly_cp[mk][cp][1] += r["out"]
    hits, evid = {}, []
    history = set()
    for mk in sorted(monthly_cp):
        new_pong = []
        for cp, (i, o) in monthly_cp[mk].items():
            if i > 0 and o > 0 and min(i, o) / max(i, o) >= THRESHOLDS["F3-04_平衡度"] and cp not in history:
                new_pong.append((cp, i, o))
        if len(new_pong) >= THRESHOLDS["F3-04_对倒对手数"]:
            hits[mk] = [c for c, _, _ in new_pong]
            for cp, _, _ in new_pong[:8]:
                evid += [ev(r, f"{mk} 新增对倒对手") for r in ctx.rows
                         if r["date"] and month_key(r["date"]) == mk and norm_name(r["cp"]) == cp][:4]
        history.update(monthly_cp[mk].keys())
    trig = bool(hits)
    snap = (f"按月统计与外部对手的双向进出（平衡度 min/max≥{THRESHOLDS['F3-04_平衡度']}），"
            f"月内新增对倒对手≥{THRESHOLDS['F3-04_对倒对手数']} 个即命中：{hits or '无'}。"
            "与集团资金归集区分：归集具有定时、定向特征且对手有工商关联，提供资金管理协议可解除（D08）。")
    return feat("F3-04", "星型对倒结构", "B", trig,
                f"命中月份 {list(hits)}" if trig else "未见单月新增≥5 个对倒对手",
                {"命中月份": {k: len(v) for k, v in hits.items()}}, evid[:60], snapshot=snap)


def f3_05(ctx):
    apply_d = parse_dt(ctx.args.apply_date).date() if ctx.args.apply_date else ctx.d_max
    proxy = "" if ctx.args.apply_date else "（未提供申贷日，以流水末日代估）"
    w_in = sum(r["in"] for r in ctx.rows if ctx.is_oper_in(r) and r["date"]
               and apply_d - timedelta(days=30) < r["date"] <= apply_d)
    base_rows = [r for r in ctx.rows if ctx.is_oper_in(r) and r["date"]
                 and apply_d - timedelta(days=210) < r["date"] <= apply_d - timedelta(days=30)]
    base_days = min(180, max((apply_d - timedelta(days=30) - ctx.d_min).days, 0))
    if base_days < 60:
        return feat("F3-05", "时点冲量", "B", False,
                    "申贷前基期不足 2 个月，未评估" + proxy, {})
    base_monthly = sum(r["in"] for r in base_rows) / (base_days / 30.0)
    mult = (w_in / base_monthly) if base_monthly else None
    after_rows = [r for r in ctx.rows if r["date"] and r["date"] > apply_d]
    fallback = ""
    if after_rows and mult and mult >= THRESHOLDS["F3-05_冲量倍数"]:
        after_in = sum(r["in"] for r in after_rows if ctx.is_oper_in(r))
        after_days = (ctx.d_max - apply_d).days or 1
        after_monthly = after_in / (after_days / 30.0)
        fallback = f"；申贷后月均经营入账 {after_monthly:,.0f}（{'回落' if after_monthly < base_monthly else '未见回落'}）"
    trig = mult is not None and mult >= THRESHOLDS["F3-05_冲量倍数"]
    evid = [ev(r) for r in ctx.rows if ctx.is_oper_in(r) and r["date"]
            and apply_d - timedelta(days=30) < r["date"] <= apply_d][:40] if trig else []
    snap = (f"申贷日 {apply_d}{proxy}：前 30 日经营入账 {w_in:,.2f}，前 6 个月月均 {base_monthly:,.2f}，"
            f"冲量倍数 {mult and round(mult, 2)}{fallback}。事后回落需贷后数据持续观察。")
    return feat("F3-05", "时点冲量", "B", trig,
                f"申贷前 30 天入账为前 6 月均值 {mult:.1f} 倍" + fallback if trig
                else (f"冲量倍数 {mult:.2f}，未达 2 倍" if mult is not None else "基期无经营入账，未评估"),
                {"冲量倍数": round(mult, 2) if mult else None}, evid, snapshot=snap)


def f3_06(ctx):
    timed = [r for r in ctx.rows if r["dt"] and r["acct_type"] == "对公"
             and not (r["dt"].hour == 0 and r["dt"].minute == 0 and r["dt"].second == 0)]
    if len(timed) < 30:
        return feat("F3-06", "非营业时间交易占比", "C", False,
                    "带时点的对公交易不足 30 笔，未评估（时点覆盖率不足）", {"带时点笔数": len(timed)})
    night = [r for r in timed if r["dt"].hour >= 22 or r["dt"].hour < 6 or r["dt"].weekday() >= 5]
    ratio = len(night) / len(timed)
    flag = ratio > THRESHOLDS["F3-06_非营业占比"]
    snap = (f"对公带时点交易 {len(timed)} 笔，非工作时段（22:00–6:00 及周末，节假日按周末近似）{len(night)} 笔，"
            f"占比 {ratio:.1%}。批量造假脚本常忽略真实交易的时间分布。")
    return feat("F3-06", "非营业时间交易占比", "C", flag,
                f"非营业时间占比 {ratio:.1%}" + ("（留痕观察）" if flag else ""),
                {"占比": round(ratio, 4), "带时点笔数": len(timed)},
                [ev(r) for r in night[:30]] if flag else [], snapshot=snap)


# ---------------------------------------------------------------------------
# L4 经营真实性锚点层
# ---------------------------------------------------------------------------
def f4_01(ctx):
    util = [r for r in ctx.rows if r["out"] > 0 and
            (UTILITY_RE.search(r["memo"]) or UTILITY_RE.search(r["cp"]) or UTILITY_RE.search(r["l3"]))]
    monthly = defaultdict(float)
    for r in util:
        if r["date"]:
            monthly[month_key(r["date"])] += r["out"]
    if not util:
        return feat("F4-01", "水电费连续性", "C", False,
                    "流水中未发现水电燃气代扣记录（可能由其他账户代缴）——留痕，并建议实地走访核实（D04）",
                    {"水电费笔数": 0}, pending="如有其他缴费账户请补充流水")
    mks = sorted(monthly)
    full = month_range(datetime.strptime(mks[0] + "-01", "%Y-%m-%d").date(), ctx.d_max)
    gaps, run = [], 0
    for m in full:
        if monthly.get(m, 0.0) <= 0:
            run += 1
        else:
            if run >= THRESHOLDS["F4-01_断缴月数"]:
                gaps.append(run)
            run = 0
    if run >= THRESHOLDS["F4-01_断缴月数"]:
        gaps.append(run)
    vals = [monthly[m] for m in mks]
    drop = False
    if len(vals) >= 4:
        head = sum(vals[:-2]) / (len(vals) - 2)
        tail = sum(vals[-2:]) / 2
        drop = head > 0 and (head - tail) / head > THRESHOLDS["F4-01_骤降幅度"]
    trig = bool(gaps) or drop
    level = "A" if (trig and ctx.args.stage == "贷后监控") else "B"
    snap = (f"逐月水电燃气支出序列={json.dumps({m: round(monthly.get(m, 0), 2) for m in full}, ensure_ascii=False)}；"
            f"断缴段（≥2 月）={gaps or '无'}；近 2 月较前期均值骤降={'是' if drop else '否'}。"
            "刚性支出不会说谎：断缴/骤降提示停产或流水非经营主账户。")
    return feat("F4-01", "水电费连续性", level, trig,
                ("存在断缴≥2 个月或骤降>50%" + ("（贷后 A 级）" if level == "A" else "") if trig
                 else "水电费缴纳连续"),
                {"覆盖月数": len(mks), "断缴段": gaps, "骤降": drop},
                [ev(r) for r in util[:40]], snapshot=snap)


def f4_02(ctx):
    wages = [r for r in ctx.rows if r["out"] > 0 and
             (r["l2"] == "人力成本" or re.search(r"工资|薪|代发", r["memo"]))]
    social = [r for r in ctx.rows if r["out"] > 0 and re.search(r"社保|社会保险|公积金|养老保险|医疗保险", r["memo"] + r["cp"])]
    n_mon = max(len(ctx.months), 1)
    m_wage = sum(r["out"] for r in wages) / n_mon
    m_soc = sum(r["out"] for r in social) / n_mon
    headcount_est = round(m_soc / 1500.0, 1) if m_soc else None   # 人均月社保约 1500 元的粗估，仅供比对
    declared = ctx.args.declared_headcount
    trig, msg = False, ""
    if declared:
        if headcount_est is not None and headcount_est < declared * 0.5:
            trig, msg = True, f"社保反推人数 {headcount_est} < 申报人数 {declared} 的 50%"
        per_capita = m_wage / declared if declared else None
        if per_capita is not None and per_capita < ctx.args.min_wage:
            trig, msg = True, (msg + "；" if msg else "") + f"人均月工资 {per_capita:,.0f} 低于最低工资 {ctx.args.min_wage:,.0f}"
    snap = (f"月均工资类支出 {m_wage:,.2f}（{len(wages)} 笔），月均社保/公积金代扣 {m_soc:,.2f}（{len(social)} 笔），"
            f"社保反推人数≈{headcount_est}（按人均 1500 元/月粗估）。"
            + ("" if declared else "未提供申报员工人数（--declared-headcount），无法比对，转待外部数据。"))
    return feat("F4-02", "工资社保匹配", "B", trig,
                msg if trig else (f"月均工资 {m_wage:,.0f}、社保反推人数≈{headcount_est}"
                                  + ("，与申报匹配" if declared else "（待申报人数比对）")),
                {"月均工资支出": round(m_wage, 2), "月均社保": round(m_soc, 2),
                 "社保反推人数": headcount_est, "申报人数": declared},
                [ev(r) for r in (wages + social)[:30]],
                pending=None if declared else "需申报员工人数（--declared-headcount）", snapshot=snap)


def f4_03(ctx):
    tax = [r for r in ctx.rows if r["out"] > 0 and not ctx.is_internal(r) and
           (r["l2"] == "税收成本" or TAX_RE.search(r["cp"]) or re.search(r"增值税|税款|缴税|完税", r["memo"]))]
    total_tax = sum(r["out"] for r in tax)
    oper_in = sum(r["in"] for r in ctx.rows if ctx.is_oper_in(r))
    if not tax:
        return feat("F4-03", "税费反推销售额", "C", False,
                    "流水中无税费缴纳记录（本身即线索：或从他户缴税、或规模未达起征、或停业）——留痕并建议 D09 税务数据核对",
                    {"税费笔数": 0}, pending="需税务申报数据核对（D09）")
    rate = ctx.args.tax_rate
    est_sales = total_tax / rate
    dev = abs(est_sales - oper_in) / max(est_sales, oper_in) if max(est_sales, oper_in) else None
    third = ""
    if ctx.args.declared_revenue:
        dev2 = abs(est_sales - ctx.args.declared_revenue) / max(est_sales, ctx.args.declared_revenue)
        third = f"；与申报收入 {ctx.args.declared_revenue:,.0f} 偏差 {dev2:.0%}"
    trig = dev is not None and dev > THRESHOLDS["F4-03_偏差"]
    snap = (f"区间税费合计 {total_tax:,.2f}，按税负率 {rate:.2%} 反推销售规模 {est_sales:,.2f}，"
            f"流水经营性入账 {oper_in:,.2f}，三角偏差 {dev and round(dev, 4)}{third}。"
            "税负率为行业综合默认值，请按借款人实际行业征收率用 --tax-rate 校正后复核。")
    return feat("F4-03", "税费反推销售额", "B", trig,
                f"税费反推销售与流水入账偏差 {dev:.0%} > 30%" + third if trig
                else f"偏差 {dev:.0%}，在容差内" + third,
                {"税费合计": round(total_tax, 2), "反推销售额": round(est_sales, 2),
                 "经营性入账": round(oper_in, 2), "偏差": round(dev, 4) if dev is not None else None},
                [ev(r) for r in tax[:30]], snapshot=snap)


def f4_04(ctx, kw_dict):
    fin_kw = [k for c, k, _ in kw_dict if c == "融资机构类"]
    hits = defaultdict(list)
    for r in ctx.rows:
        if ctx.is_internal(r) or not r["cp"]:
            continue
        for kw in fin_kw:
            if kw in r["cp"]:
                hits[r["cp"]].append(r)
                break
    trig = bool(hits)
    evid = [ev(r, f"融资类对手：{cp}") for cp, rs in hits.items() for r in rs[:5]]
    snap = (f"对手名称命中融资类专用字典（典当/小贷/融资租赁/保理/担保/投资资管/创投）："
            f"{ {cp: len(rs) for cp, rs in hits.items()} or '无'}。"
            "前 20 大对手的行业契合度需工商行业打标数据，转待外部数据。")
    return feat("F4-04", "对手行业契合度（融资类字典部分）", "B", trig,
                f"命中融资类对手 {len(hits)} 个" if trig else "未命中融资类对手字典",
                {"命中对手": {cp: len(rs) for cp, rs in hits.items()}}, evid,
                pending="对手行业契合度需工商行业打标数据", snapshot=snap)


def f4_05(ctx):
    monthly = defaultdict(float)
    for r in ctx.rows:
        if ctx.is_oper_in(r) and r["date"]:
            monthly[month_key(r["date"])] += r["in"]
    return feat("F4-05", "季节性背离", "B", False,
                "需行业季节模板（含春节错位调整）方可计算相关系数，本次仅输出月度入账序列供人工比对",
                {"月度经营入账": {m: round(monthly.get(m, 0), 2) for m in ctx.months}},
                pending="需行业季节模板（第三批上线项）")


def f4_06(ctx):
    cp_in = defaultdict(list)
    for r in ctx.rows:
        if ctx.is_oper_in(r) and r["cp"] and r["date"]:
            cp_in[norm_name(r["cp"])].append(r["date"])
    tops = sorted(cp_in.items(), key=lambda kv: -len(kv[1]))[:10]
    stats = {}
    for cp, ds in tops:
        ds = sorted(ds)
        gaps = [(b - a).days for a, b in zip(ds, ds[1:]) if (b - a).days > 0]
        if len(gaps) >= 3:
            gaps.sort()
            stats[cp] = {"回款次数": len(ds), "间隔中位数天": gaps[len(gaps) // 2],
                         "P25": gaps[len(gaps) // 4], "P75": gaps[3 * len(gaps) // 4]}
    return feat("F4-06", "账期一致性", "C", False,
                "已输出核心对手回款间隔分布（中位数/P25/P75），与尽调申报账期的比对需人工/外部数据",
                {"核心对手回款间隔": stats}, pending="需尽调申报账期比对")


def f4_07(ctx):
    found = {}
    evid = []
    for cat, pat in MISC_CATEGORIES.items():
        rs = [r for r in ctx.rows if r["out"] > 0 and not ctx.is_internal(r)
              and (pat.search(r["memo"]) or pat.search(r["cp"]))]
        if len(rs) >= 2:
            found[cat] = len(rs)
            evid.extend(ev(r, cat) for r in rs[:4])
    score = len(found)
    flag = score < THRESHOLDS["F4-07_存在性项数"]
    snap = (f"经营杂项刚性支出存在性（≥2 笔记 1 项）：{found or '无'}，评分 {score} 项。"
            "造假者几乎从不伪造此类明细；评分低提示流水非真实经营主账户。")
    return feat("F4-07", "经营杂项支出存在性", "C", flag,
                f"存在性评分 {score} 项" + ("（<2 项，留痕观察）" if flag else ""),
                {"存在性评分": score, "明细": found}, evid, snapshot=snap)


# ---------------------------------------------------------------------------
# L5 隐性负债与民间借贷层
# ---------------------------------------------------------------------------
def load_keywords(path):
    out = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            out.append((row["类别"].strip(), row["关键词"].strip(), row.get("使用约束", "")))
    return out


def f5_01(ctx, kw_dict):
    """筹资关键词命中：按文档要求结合方向、对手类型、规律性三要素，禁止裸关键词触发。
    征信比对（命中而征信无对应才定性）属人审动作 D07，工单中已注明。"""
    cats = {"筹资偿债类", "民间借贷场景类", "养流水嫌疑类"}
    hits = defaultdict(list)   # (类别, 关键词, 对手) -> rows
    for r in ctx.rows:
        if ctx.is_internal(r) or not r["memo"]:
            continue
        for cat, kw, _ in kw_dict:
            if cat not in cats or kw not in r["memo"]:
                continue
            if kw == "利息" and ctx.is_interest(r):
                continue   # 排除存款结息
            if cat == "民间借贷场景类" and not is_personal(r["cp"], r["cp_acct"]):
                continue   # 对私场景约束
            hits[(cat, kw, norm_name(r["cp"]) or "（无对手）")].append(r)
    qualified = {k: v for k, v in hits.items()
                 if len(v) >= 2 or max((x["in"] or x["out"]) for x in v) >= 100000
                 or k[0] == "养流水嫌疑类"}
    trig = bool(qualified)
    evid = [ev(r, f"{k[0]}/{k[1]}") for k, rs in qualified.items() for r in rs[:5]][:60]
    snap = ("摘要命中筹资字典（附录 A），三要素过滤：方向+对手类型（民间场景限对私）+规律性"
            "（同对手≥2 笔或单笔≥10 万，养流水嫌疑词单笔即记）。命中分布："
            + json.dumps({f"{k[0]}|{k[1]}|{k[2]}": len(v) for k, v in qualified.items()},
                         ensure_ascii=False)
            + "。定性需征信比对：命中且征信无对应机构借贷记录才触发（人审执行 D07）。")
    return feat("F5-01", "筹资关键词命中", "B", trig,
                f"命中 {len(qualified)} 组筹资/民间借贷/养流水关键词（需征信比对确认 D07）" if trig
                else "未命中筹资字典（含三要素过滤）",
                {"命中组数": len(qualified)}, evid,
                pending="需征信报告比对（命中且征信无对应才最终定性）", snapshot=snap)


def f5_02(ctx):
    lo, hi = THRESHOLDS["F5-02_月息区间"]
    big_ins = [r for r in ctx.rows if r["in"] >= THRESHOLDS["F5-02_借入门槛"]
               and is_personal(r["cp"], r["cp_acct"]) and not ctx.is_internal(r)]
    hits, evid = [], []
    for bi in big_ins:
        cp = norm_name(bi["cp"])
        pays = [r for r in ctx.rows if r["out"] > 0 and norm_name(r["cp"]) == cp
                and r["date"] and bi["date"] and r["date"] > bi["date"]]
        monthly = defaultdict(list)
        for p in pays:
            monthly[month_key(p["date"])].append(p)
        seq = []
        for mk in sorted(monthly):
            amts = [p["out"] for p in monthly[mk]]
            seq.append((mk, sum(amts), [p["date"].day for p in monthly[mk]]))
        if len(seq) < THRESHOLDS["F5-02_持续期数"]:
            continue
        ratios = [s[1] / bi["in"] for s in seq]
        inband = [r_ for r_ in ratios if lo <= r_ <= hi]
        days = [d for s in seq for d in s[2]]
        day_spread = (max(days) - min(days)) if days else 99
        if len(inband) >= THRESHOLDS["F5-02_持续期数"] and day_spread <= 10:
            hits.append({"对手": bi["cp"], "借入本金": bi["in"], "借入日": str(bi["date"]),
                         "持续期数": len(inband),
                         "月付比率": [round(r_, 4) for r_ in ratios[:6]]})
            evid.append(ev(bi, "疑似对私借入本金"))
            evid += [ev(p, "规律性付息") for mk in sorted(monthly) for p in monthly[mk]][:12]
    trig = bool(hits)
    snap = (f"对私大额入账（≥{THRESHOLDS['F5-02_借入门槛']:,.0f}）后，向同一自然人按月固定日"
            f"（日波动≤10 天）支付、月付/本金落入 [{lo:.1%},{hi:.1%}] 且持续≥3 期：")
    snap += json.dumps(hits, ensure_ascii=False) if hits else "无命中"
    return feat("F5-02", "对私借入＋规律性付息", "B", trig,
                f"检出 {len(hits)} 组高息民间借贷特征（月息 1.5%–3.5%）" if trig
                else "未检出对私借入+规律付息组合",
                {"命中组": hits}, evid, snapshot=snap)


def f5_03(ctx):
    repay_m = sum(r["out"] for r in ctx.rows if ctx.is_repay_out(r)) / max(len(ctx.months), 1)
    due = ctx.args.credit_monthly_due
    if due is None:
        return feat("F5-03", "偿债流量缺口", "B", False,
                    f"流水推算月偿债支出 {repay_m:,.2f}；未提供征信月应还额（--credit-monthly-due），无法计算缺口",
                    {"流水月偿债支出": round(repay_m, 2)},
                    pending="需征信报告月应还额（D07）")
    gap = repay_m - due
    trig = due >= 0 and gap > THRESHOLDS["F5-03_缺口比例"] * max(due, 1.0)
    snap = (f"流水推算月偿债支出（还本＋付息类流出月均）={repay_m:,.2f}，征信月应还={due:,.2f}，"
            f"缺口={gap:,.2f}。缺口为正且显著即隐性负债敞口估计值，直接进入授信扣减项。")
    return feat("F5-03", "偿债流量缺口", "B", trig,
                f"隐性负债敞口估计 {gap:,.0f} 元/月（>征信月应还×30%），建议作为授信扣减项" if trig
                else f"缺口 {gap:,.0f} 元/月，在容差内",
                {"月偿债支出": round(repay_m, 2), "征信月应还": due, "缺口": round(gap, 2)},
                snapshot=snap)


def f5_04(ctx):
    """倒贷指数 = 区间借款类入账总额 / 区间借贷资金平均余额。
    平均余额以『累计借入−累计归还（下限 0）』的逐日序列近似（流水内无本金台账）。"""
    events = []
    for r in ctx.rows:
        if not r["date"]:
            continue
        if ctx.is_borrow_in(r):
            events.append((r["date"], r["in"]))
        elif ctx.is_repay_out(r):
            events.append((r["date"], -r["out"]))
    total_in = sum(a for _, a in events if a > 0)
    if total_in <= 0:
        return feat("F5-04", "倒贷指数", "B", False, "区间内无借款类入账，未评估", {})
    events.sort()
    bal, day_sum, prev = 0.0, 0.0, ctx.d_min
    for d, a in events:
        day_sum += bal * (d - prev).days
        bal = max(bal + a, 0.0)
        prev = d
    day_sum += bal * (ctx.d_max - prev).days
    avg_bal = day_sum / max(ctx.span_days, 1)
    idx = total_in / max(avg_bal, 1.0)
    trig = idx > THRESHOLDS["F5-04_倒贷指数"]
    snap = (f"借款类入账合计 {total_in:,.2f}；借贷资金平均余额（累计借入−累计归还逐日近似）{avg_bal:,.2f}；"
            f"倒贷指数 {idx:.2f}（阈值>6）。过桥序列（他行贷款到期前后对私大额入→还贷→新贷→等额转出）"
            "需征信贷款到期日数据，人审 D07 时核对。")
    return feat("F5-04", "倒贷指数", "B", trig,
                f"倒贷指数 {idx:.1f} > 6（高频滚动借贷）" if trig else f"倒贷指数 {idx:.2f}",
                {"借款类入账": round(total_in, 2), "平均借贷余额": round(avg_bal, 2),
                 "倒贷指数": round(idx, 2)},
                [ev(r) for r in ctx.rows if ctx.is_borrow_in(r)][:30],
                pending="过桥序列识别需征信贷款到期日", snapshot=snap)


def f5_05(ctx):
    borrow = sum(r["in"] for r in ctx.rows if ctx.is_borrow_in(r))
    oper = sum(r["in"] for r in ctx.rows if ctx.is_oper_in(r))
    if borrow + oper <= 0:
        return feat("F5-05", "借款类收入占比", "B", False, "无可评估入账", {})
    ratio = borrow / oper if oper else float("inf")
    trig = ratio > THRESHOLDS["F5-05_借款收入占比"]
    snap = f"借款类入账 {borrow:,.2f} / 经营性入账 {oper:,.2f} = {ratio if ratio != float('inf') else '∞'}"
    return feat("F5-05", "借款类收入占比", "B", trig,
                (f"借款类收入占比 {min(ratio, 99):.0%} > 30%（经营造血不足、依赖融资续命）" if trig
                 else f"借款类收入占比 {ratio:.0%}"),
                {"借款类入账": round(borrow, 2), "经营性入账": round(oper, 2),
                 "占比": round(ratio, 4) if ratio != float("inf") else None},
                [ev(r) for r in ctx.rows if ctx.is_borrow_in(r)][:30], snapshot=snap)


# ---------------------------------------------------------------------------
# L6 完整性与交叉验证层
# ---------------------------------------------------------------------------
def f6_01(ctx):
    hidden = defaultdict(list)
    for r in ctx.rows:
        cp = norm_name(r["cp"])
        if not cp or cp not in ctx.members:
            continue
        acct_digits = re.sub(r"\D", "", r["cp_acct"] or "")
        if not acct_digits:
            if re.search(r"跨行.*本人|本人.*跨行", r["memo"]):
                hidden[("（摘要线索）", r["cp"])].append(r)
            continue
        if acct_digits in ctx.self_accts:
            continue
        if len(acct_digits) >= 4 and acct_digits[-4:] in ctx.self_last4 and len(acct_digits) < 10:
            continue   # 掩码账号后四位能对上已知账户，视为同户
        hidden[(r["cp_acct"], r["cp"])].append(r)
    trig = bool(hidden)
    evid = [ev(r, f"疑似隐藏账户 {k[0]}") for k, rs in hidden.items() for r in rs[:6]][:60]
    accts = [{"对手账户": k[0], "户名": k[1], "笔数": len(rs),
              "金额合计": round(sum(x["in"] + x["out"] for x in rs), 2)} for k, rs in hidden.items()]
    snap = ("对手户名＝借款人/关联主体本人、但对手账号不在已提供账户清单内（掩码账号按后四位比对豁免）："
            + json.dumps(accts, ensure_ascii=False)
            + "。处置：发现即要求限期补充该账户流水（D05）；拒不补充＝A 级 R-A04 拒贷/中止。")
    return feat("F6-01", "隐藏账户发现", "A", trig,
                f"发现 {len(hidden)} 个疑似未提供账户，要求限期补充流水（拒不补充→R-A04）" if trig
                else "未发现本人名下未提供账户线索",
                {"疑似隐藏账户": accts}, evid, snapshot=snap)


def _load_invoices(path):
    """发票明细：灵活映射列名（购买方/客户/对手→对手；价税合计/金额→金额；开票日期/日期→日期）。"""
    rows = []
    if path.lower().endswith((".xlsx", ".xlsm")):
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = [str(h or "") for h in next(it)]
        data = list(it)
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            rd = csv.reader(f)
            header = next(rd)
            data = list(rd)
    def find(pats):
        for i, h in enumerate(header):
            if any(p in str(h) for p in pats):
                return i
        return None
    ci = find(["购买方", "客户名称", "对手", "购方"])
    ai = find(["价税合计", "金额"])
    di = find(["开票日期", "日期"])
    if ci is None or ai is None or di is None:
        raise SystemExit("【错误】发票文件缺少可识别的 购买方/金额/开票日期 列")
    for raw in data:
        if raw is None or len(raw) <= max(ci, ai, di):
            continue
        amt, dt = to_num(raw[ai]), parse_dt(raw[di])
        if amt and amt > 0 and dt:
            rows.append({"cp": str(raw[ci] or "").strip(), "amt": amt, "date": dt.date()})
    return rows


def f6_02(ctx, invoices):
    oper = sum(r["in"] for r in ctx.rows if ctx.is_oper_in(r))
    sales = None
    src = ""
    if invoices:
        sales = sum(i["amt"] for i in invoices
                    if ctx.d_min <= i["date"] <= ctx.d_max)
        src = "区间内销项发票合计"
    elif ctx.args.declared_revenue:
        sales = ctx.args.declared_revenue
        src = "申报收入"
    if not sales:
        return feat("F6-02", "流水覆盖率", "B", False,
                    "未提供发票明细或申报收入，无法计算覆盖率",
                    {"经营性入账": round(oper, 2)}, pending="需发票销售额或申报收入")
    cov = oper / sales
    trig = cov < THRESHOLDS["F6-02_覆盖率"]
    snap = f"流水经营性入账 {oper:,.2f} / {src} {sales:,.2f} = 覆盖率 {cov:.1%}（阈值<60%）"
    return feat("F6-02", "流水覆盖率", "B", trig,
                (f"覆盖率 {cov:.1%} < 60%（提示账户不全或大量现金交易，分析结论不可靠）" if trig
                 else f"覆盖率 {cov:.1%}"),
                {"覆盖率": round(cov, 4), "经营性入账": round(oper, 2), "销售额": round(sales, 2)},
                snapshot=snap)


def f6_03(ctx):
    daily_in = defaultdict(lambda: defaultdict(float))
    for r in ctx.rows:
        if r["date"] and r["in"] > 0 and not ctx.is_internal(r):
            daily_in[(r["subject"], r["acct"])][r["date"]] += r["in"]
    keys = list(daily_in)
    hits, detail = [], []
    for a in range(len(keys)):
        for b in range(a + 1, len(keys)):
            da, db = daily_in[keys[a]], daily_in[keys[b]]
            lo, hi = max(min(da), min(db)), min(max(da), max(db))
            if (hi - lo).days < THRESHOLDS["F6-03_最短重叠天"]:
                continue
            days = [lo + timedelta(days=i) for i in range((hi - lo).days + 1)]
            xs = [da.get(d, 0.0) for d in days]
            ys = [db.get(d, 0.0) for d in days]
            if sum(1 for x in xs if x) < 20 or sum(1 for y in ys if y) < 20:
                continue
            r_ = pearson(xs, ys)
            if r_ is not None:
                detail.append({"账户A": keys[a][1], "账户B": keys[b][1],
                               "重叠天数": len(days), "相关系数": round(r_, 4)})
                if r_ > THRESHOLDS["F6-03_相关系数"]:
                    hits.append(detail[-1])
    trig = bool(hits)
    snap = ("账户间日经营性入账序列相关性（剔内部互转，重叠≥60 天且双方有效入账日≥20）："
            + json.dumps(detail, ensure_ascii=False)
            + "。相关性异常高提示同模板批量生成。")
    return feat("F6-03", "多份流水相关性", "B", trig,
                f"{len(hits)} 对账户日入账序列相关系数>0.9（疑似同模板生成）" if trig
                else ("各账户入账序列相关性正常" if detail else "无满足重叠条件的账户对，未评估"),
                {"账户对明细": detail, "命中": hits}, snapshot=snap)


def f6_04(ctx, invoices):
    if not invoices:
        return feat("F6-04", "发票逐笔交叉", "A", False,
                    "未提供发票明细（--invoice），无法逐笔匹配；该特征是破解『同比例放大＋假尽调收入』的决定性手段，建议补充",
                    {}, pending="需销项发票明细（D02）")
    tol, wind = THRESHOLDS["F6-04_金额窗口"], THRESHOLDS["F6-04_日期窗口天"]
    inv_in = [i for i in invoices
              if ctx.d_min - timedelta(days=wind) <= i["date"] <= ctx.d_max + timedelta(days=wind)]
    ins = [r for r in ctx.rows if r["in"] > 0 and not ctx.is_internal(r) and r["date"]]
    used = set()
    matched, miss = [], []
    for inv in inv_in:
        cpn = norm_name(inv["cp"])
        ok = None
        for k, r in enumerate(ins):
            if k in used:
                continue
            rcp = norm_name(r["cp"])
            if not rcp or (cpn not in rcp and rcp not in cpn):
                continue
            if abs(r["in"] - inv["amt"]) / inv["amt"] > tol:
                continue
            if abs((r["date"] - inv["date"]).days) > wind:
                continue
            ok = k
            break
        if ok is not None:
            used.add(ok)
            matched.append(inv)
        else:
            miss.append(inv)
    rate = len(matched) / len(inv_in) if inv_in else None
    trig = rate is not None and rate < THRESHOLDS["F6-04_匹配率"]
    evid = [{"交易时间": str(i["date"]), "对手名称": i["cp"], "收入金额": "", "支出金额": "",
             "账户余额": "", "摘要": f"发票 {i['amt']:,.2f} 未匹配到入账", "主体名称": "", "本方账户": "",
             "来源文件名": "发票明细", "来源行号": "", "备注": "未匹配发票"} for i in miss[:40]]
    snap = (f"区间内销项发票 {len(inv_in)} 张，按 对手＋金额±5%＋日期±{wind} 日 逐笔匹配入账："
            f"匹配 {len(matched)} 张，匹配率 {rate and round(rate, 4)}。"
            "同比例放大可骗过总额核对，但逐笔匹配率必然崩塌——匹配率<40% 即 R-A05 立案必查。")
    return feat("F6-04", "发票逐笔交叉", "A", trig,
                (f"发票逐笔匹配率 {rate:.0%} < 40%（R-A05 立案必查，同比例放大决定性证据）" if trig
                 else (f"匹配率 {rate:.0%}" if rate is not None else "区间内无可匹配发票")),
                {"发票张数": len(inv_in), "匹配张数": len(matched),
                 "匹配率": round(rate, 4) if rate is not None else None}, evid, snapshot=snap)


def f6_05(ctx):
    return feat("F6-05", "三表勾稽偏差", "B", False,
                "需财务报表（收入/采购/费用科目）勾稽，本工具不读取报表，转人工/外部数据；"
                "注意偏差方向：流水<报表收入提示虚增收入，流水>报表提示隐匿收入避税",
                {}, pending="需财务报表数据（D09）")


# ---------------------------------------------------------------------------
# L1 文件取证层（可选，--raw-folder；仅适用于客户自提供流水）
# ---------------------------------------------------------------------------
def l1_forensics(raw_folder):
    findings, sig_notes, skipped = [], [], []
    pdfs = sorted(glob(os.path.join(raw_folder, "**", "*.pdf"), recursive=True))
    xlsxs = sorted(glob(os.path.join(raw_folder, "**", "*.xlsx"), recursive=True))
    try:
        from pypdf import PdfReader
    except ImportError:
        PdfReader = None
    for p in pdfs:
        base = os.path.basename(p)
        if PdfReader is None:
            skipped.append(f"{base}：未安装 pypdf，无法读取 PDF 元数据")
            continue
        try:
            rd = PdfReader(p)
            meta = rd.metadata or {}
            producer = str(meta.get("/Producer", "") or "")
            creator = str(meta.get("/Creator", "") or "")
            author = str(meta.get("/Author", "") or "")
            cdate, mdate = str(meta.get("/CreationDate", "") or ""), str(meta.get("/ModDate", "") or "")
            blob = f"{producer} {creator} {author}"
            if EDITOR_BLACKLIST.search(blob):
                findings.append({"文件": base, "级别": "A", "问题": "元数据命中编辑器名单（R-A03）",
                                 "Producer": producer, "Creator": creator, "Author": author})
            elif cdate and mdate and mdate > cdate:
                findings.append({"文件": base, "级别": "C", "问题": "修改时间晚于创建时间（留痕；落款时间需人工核对）",
                                 "CreationDate": cdate, "ModDate": mdate})
            root = rd.trailer.get("/Root", {})
            acro = root.get("/AcroForm") if hasattr(root, "get") else None
            has_sig = False
            if acro and acro.get("/SigFlags"):
                has_sig = True
            sig_notes.append({"文件": base, "数字签名": "存在（建议专用工具验签证书链与签名后修改）" if has_sig
                              else "无签名（仅记录，C 级）"})
        except Exception as e:    # 损坏/加密文件如实记录
            findings.append({"文件": base, "级别": "B", "问题": f"PDF 无法解析：{e}"})
    for p in xlsxs:
        base = os.path.basename(p)
        if "反欺诈" in base or "BI分析" in base or "已清洗" in base:
            continue
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True)
            pr = wb.properties
            findings.append({"文件": base, "级别": "C",
                             "问题": "Excel 元数据留痕（银行系统导出通常为 POI/系统名；人名/办公软件创建需关注）",
                             "creator": str(pr.creator or ""), "lastModifiedBy": str(pr.lastModifiedBy or ""),
                             "created": str(pr.created or ""), "modified": str(pr.modified or "")})
        except Exception as e:
            skipped.append(f"{base}：{e}")
    a_hits = [f_ for f_ in findings if f_["级别"] == "A"]
    f101 = feat("F1-01", "文件元数据检查", "A", bool(a_hits),
                (f"{len(a_hits)} 个文件元数据命中编辑器名单（R-A03 立案必查）" if a_hits
                 else f"已检 PDF {len(pdfs)}/Excel {len(xlsxs)} 个，未命中编辑器名单"),
                {"发现": findings, "跳过": skipped},
                snapshot="编辑器名单：Photoshop/美图/GIMP/Canva/WPS/Word/万兴/福昕编辑/iLovePDF/Sejda/Smallpdf/Nitro/PDFelement 等；"
                         ".xls 旧格式与图片型文件不在本工具元数据检查范围，需人工取证。")
    f104 = feat("F1-04", "数字签名验签", "A/C", False,
                "完整验签（证书链有效性、签名后修改检测）需专用验签工具；本工具仅记录签名存在性",
                {"签名记录": sig_notes}, pending="带签名 PDF 需专用工具验签；验签失败＝A 级")
    return [f101, f104]


# ---------------------------------------------------------------------------
# 规则引擎计分与工单
# ---------------------------------------------------------------------------
def score_engine(features):
    a_hits = [f_ for f_ in features if f_["命中"] and f_["级别"].startswith("A")]
    b_hits = [f_ for f_ in features if f_["命中"] and f_["级别"] == "B"]
    c_flags = [f_ for f_ in features if f_["命中"] and f_["级别"] == "C"]
    score = sum(f_["权重"] for f_ in b_hits)
    resonance = len(b_hits) >= THRESHOLDS["共振条数"]
    to_review = bool(a_hits) or score >= THRESHOLDS["进入人审分数"] or resonance
    if a_hits:
        verdict = "A 级硬规则命中：立案必查（相关申请冻结流转）"
    elif to_review:
        verdict = ("三条及以上 B 级共振：升级处理，转入人工核查队列" if resonance
                   else f"B 级累计 {score} 分 ≥ {THRESHOLDS['进入人审分数']} 分：转入人工核查队列")
    else:
        verdict = f"B 级累计 {score} 分，未达人审阈值：通过反欺诈初筛（C 级留痕 {len(c_flags)} 项照常落库）"
    return {"A级命中": a_hits, "B级命中": b_hits, "C级留痕": c_flags,
            "B级总分": score, "共振": resonance, "进入人审": to_review, "结论建议": verdict}


def build_worksheet(client, args, ctx, result, features):
    today = datetime.now()
    hit_list = result["A级命中"] + sorted(result["B级命中"], key=lambda f_: -f_["权重"])
    deadline = "A 级：T+2 个工作日（期间相关申请冻结流转）" if result["A级命中"] else "B 级：T+5 个工作日"
    actions = []
    for f_ in hit_list:
        if f_["首选核查动作"] and f_["首选核查动作"] not in [a[1] for a in actions]:
            actions.append((f_["编号"], f_["首选核查动作"], f_["次选核查动作"]))
    return {
        "工单编号": f"FZ-{today:%Y%m%d}-001",
        "客户与申请信息": f"客户名称：{client}；产品：惠裕贷；申请编号/客户号/申请金额：待补录",
        "触发环节": args.stage,
        "命中规则清单": [
            {"规则": f_["编号"], "名称": f_["名称"], "级别": f_["级别"],
             "得分": f_["权重"] if f_["级别"] == "B" else "A级直触", "摘要": f_["结论摘要"]}
            for f_ in hit_list],
        "总分与级别": {"B级累计得分": result["B级总分"], "A级直触": bool(result["A级命中"]),
                       "多条共振升级": result["共振"]},
        "证据明细": "见『证据明细』工作表（行级：日期、金额、对手、摘要、来源文件与行号；统计类规则附计算过程快照）",
        "建议核查动作": [{"关联规则": c, "首选": a, "次选": b} for c, a, b in actions],
        "可解除标准": [{"规则": f_["编号"], "可解除标准": f_["可解除标准"]} for f_ in hit_list],
        "处理人与时限": f"指派客户经理/风险经理（待补录）；{deadline}",
        "核查结论": "（人工填写）欺诈确认 / 解释成立予以解除 / 无法核实（按从严原则处理）",
        "结论依据": "（人工填写）核查动作执行记录、客户提供材料清单、影像编号",
        "标签回流": "结论标签回流特征库，用于规则精度统计（Wilson 95% 置信下界）与权重重估",
    }


# ---------------------------------------------------------------------------
# 输出
# ---------------------------------------------------------------------------
def write_xlsx(path, client, ctx, features, result, worksheet, pending_items):
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    H = Font(bold=True, size=11, color="FFFFFF")
    TITLE = Font(bold=True, size=14)
    FILL = PatternFill("solid", fgColor="1F4E79")
    RED = PatternFill("solid", fgColor="C00000")
    YEL = PatternFill("solid", fgColor="BF8F00")
    GRN = PatternFill("solid", fgColor="2E7D32")
    WRAP = Alignment(wrap_text=True, vertical="top")

    def header_row(ws, row, headers, widths=None):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.font, c.fill = H, FILL
        if widths:
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

    # 1 总览
    ws = wb.active
    ws.title = "反欺诈总览"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    ws["A1"] = f"银行流水反欺诈核查报告 —— {client}"
    ws["A1"].font = TITLE
    rows = [
        ("触发环节", ctx.args.stage),
        ("流水区间", f"{ctx.d_min} 至 {ctx.d_max}（{ctx.span_days} 天 / {len(ctx.months)} 个月）"),
        ("关联主体", "、".join(sorted({r['subject'] for r in ctx.rows if r['subject']})) or "—"),
        ("账户数 / 交易笔数", f"{len({(r['subject'], r['acct']) for r in ctx.rows})} / {len(ctx.rows)}"),
        ("打标覆盖率", f"{ctx.labeled_ratio:.0%}（低于 50% 时经营/借款口径退化为关键词启发式）"),
        ("A 级硬规则命中", f"{len(result['A级命中'])} 条：" + "；".join(
            f"{f_['编号']} {f_['名称']}" for f_ in result["A级命中"]) if result["A级命中"] else "0 条"),
        ("B 级累计得分", f"{result['B级总分']} 分（进入人审阈值 {THRESHOLDS['进入人审分数']} 分）"),
        ("B 级共振", f"{len(result['B级命中'])} 条命中" + ("（≥3 条，共振升级）" if result["共振"] else "")),
        ("C 级留痕", f"{len(result['C级留痕'])} 项（不触发报警，落贷后监控特征池与建模标签池）"),
        ("结论建议", result["结论建议"]),
        ("重要声明", "机器不下最终结论：本报告产出物是带证据的核查工单，由人工确认定性。"
                     "全部阈值为冷启动建议值，应按 Wilson 95% 置信下界机制持续校准。"),
    ]
    r = 3
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=str(v))
        c.alignment = WRAP
        if k == "结论建议":
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = RED if result["A级命中"] else (YEL if result["进入人审"] else GRN)
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="B 级得分明细").font = Font(bold=True)
    r += 1
    header_row(ws, r, ["规则", "名称", "权重", "结论摘要"])
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 90
    for f_ in sorted(result["B级命中"], key=lambda x: -x["权重"]):
        r += 1
        for i, v in enumerate([f_["编号"], f_["名称"], f_["权重"], f_["结论摘要"]], 1):
            ws.cell(row=r, column=i, value=v).alignment = WRAP

    # 2 规则命中明细
    ws = wb.create_sheet("规则命中明细")
    ws.sheet_view.showGridLines = False
    header_row(ws, 1, ["编号", "特征名称", "级别", "权重", "是否命中", "结论摘要", "待外部数据", "关键指标"],
               [8, 22, 6, 6, 8, 60, 30, 60])
    for i, f_ in enumerate(features, 2):
        vals = [f_["编号"], f_["名称"], f_["级别"], f_["权重"] or "",
                "命中" if f_["命中"] else "未命中", f_["结论摘要"], f_["待外部数据"],
                json.dumps(f_["关键指标"], ensure_ascii=False)[:1500]]
        for j, v in enumerate(vals, 1):
            ws.cell(row=i, column=j, value=v).alignment = WRAP
        if f_["命中"]:
            ws.cell(row=i, column=5).fill = RED if f_["级别"].startswith("A") else YEL
            ws.cell(row=i, column=5).font = Font(bold=True, color="FFFFFF")

    # 3 证据明细（行级化 + 计算过程快照）
    ws = wb.create_sheet("证据明细")
    ws.sheet_view.showGridLines = False
    cols = ["规则", "交易时间", "主体名称", "本方账户", "对手名称", "收入金额", "支出金额",
            "账户余额", "摘要", "来源文件名", "来源行号", "备注"]
    header_row(ws, 1, cols, [8, 19, 18, 22, 22, 12, 12, 12, 36, 24, 8, 30])
    r = 1
    for f_ in features:
        if not (f_["命中"] and (f_["证据"] or f_["计算过程快照"])):
            continue
        r += 1
        c = ws.cell(row=r, column=1, value=f"{f_['编号']} {f_['名称']} —— 计算过程快照：{f_['计算过程快照']}")
        c.alignment = WRAP
        c.font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        for e in f_["证据"]:
            r += 1
            vals = [f_["编号"], e.get("交易时间"), e.get("主体名称"), e.get("本方账户"), e.get("对手名称"),
                    e.get("收入金额"), e.get("支出金额"), e.get("账户余额"), e.get("摘要"),
                    e.get("来源文件名"), e.get("来源行号"), e.get("备注")]
            for j, v in enumerate(vals, 1):
                ws.cell(row=r, column=j, value=v).alignment = WRAP
        if f_["证据总数"] > len(f_["证据"]):
            r += 1
            ws.cell(row=r, column=1,
                    value=f"……共 {f_['证据总数']} 条证据，全部明细见 JSON 报告").font = Font(italic=True)

    # 4 核查工单
    ws = wb.create_sheet("核查工单")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 120
    if worksheet:
        ws["A1"] = "人工核查工单（字段口径见文档 5.1；证据包四要素缺一不得派单）"
        ws["A1"].font = TITLE
        r = 3
        for k, v in worksheet.items():
            ws.cell(row=r, column=1, value=k).font = Font(bold=True)
            txt = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False, indent=1)
            c = ws.cell(row=r, column=2, value=txt)
            c.alignment = WRAP
            r += 1
    else:
        ws["A1"] = "未达派单条件（无 A 级命中、B 级未达 25 分且未共振）——C 级留痕照常落库"

    # 5 C级留痕
    ws = wb.create_sheet("C级留痕")
    ws.sheet_view.showGridLines = False
    header_row(ws, 1, ["编号", "特征名称", "是否异常", "结论摘要", "关键指标"], [8, 24, 8, 50, 80])
    r = 1
    for f_ in features:
        if not f_["级别"].startswith("C") and f_["级别"] != "A/C":
            continue
        r += 1
        for j, v in enumerate([f_["编号"], f_["名称"], "异常" if f_["命中"] else "正常",
                               f_["结论摘要"], json.dumps(f_["关键指标"], ensure_ascii=False)[:1500]], 1):
            ws.cell(row=r, column=j, value=v).alignment = WRAP

    # 6 待外部数据核查
    ws = wb.create_sheet("待外部数据核查")
    ws.sheet_view.showGridLines = False
    header_row(ws, 1, ["编号", "特征名称", "所需外部数据", "对应核查动作"], [8, 26, 50, 40])
    for i, f_ in enumerate(pending_items, 2):
        for j, v in enumerate([f_["编号"], f_["名称"], f_["待外部数据"], f_["首选核查动作"]], 1):
            ws.cell(row=i, column=j, value=v).alignment = WRAP

    wb.save(path)


def main():
    ap = argparse.ArgumentParser(description="银行流水反欺诈核查（消费 bank-statement-standardization 交付物）")
    ap.add_argument("--input", required=True, help="标准化交付物 .xlsx / 打标流水 .csv / 所在文件夹")
    ap.add_argument("--client", default="", help="授信客户归档名（默认取数据内客户名/主体名）")
    ap.add_argument("--stage", default="贷前尽调", choices=["贷前尽调", "审批中", "贷后监控"])
    ap.add_argument("--apply-date", default="", help="申贷日 YYYY-MM-DD（F3-05 时点冲量）")
    ap.add_argument("--rate", type=float, default=None, help="活期挂牌利率（年，如 0.002）；缺省时按候选档逐一尝试")
    ap.add_argument("--tax-rate", type=float, default=THRESHOLDS["F4-03_税负率"], help="行业综合税负率（F4-03）")
    ap.add_argument("--credit-monthly-due", type=float, default=None, help="征信月应还额（F5-03）")
    ap.add_argument("--declared-revenue", type=float, default=None, help="申报收入（F4-03/F6-02）")
    ap.add_argument("--declared-headcount", type=int, default=None, help="申报员工人数（F4-02）")
    ap.add_argument("--min-wage", type=float, default=1900.0, help="当地月最低工资（F4-02）")
    ap.add_argument("--invoice", default="", help="销项发票明细文件（F6-02/F6-04）")
    ap.add_argument("--raw-folder", default="", help="客户自提供原始流水文件夹（启用 L1 文件取证）")
    ap.add_argument("--loan-date", default="", help="贷款发放日（贷后 F3-03 回流 R-A06）")
    ap.add_argument("--out-dir", default="", help="输出目录，默认与输入同级")
    args = ap.parse_args()

    src = pick_input(args.input)
    print(f"[输入] {src}")
    rows = load_rows(src)
    print(f"[载入] {len(rows)} 笔交易")
    ctx = Ctx(rows, args)
    client = args.client or next((r["subject"] for r in rows if r["subject"]), "客户")

    kw_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets", "financing_keywords.csv")
    kw_dict = load_keywords(kw_path)
    invoices = _load_invoices(args.invoice) if args.invoice else None
    if invoices is not None:
        print(f"[发票] {len(invoices)} 张销项发票")

    features = []
    if args.raw_folder:
        features += l1_forensics(args.raw_folder)
    features += [f2_01(ctx), f2_02(ctx), f2_03(ctx), f2_04(ctx), f2_05(ctx), f2_06(ctx),
                 f3_01(ctx), f3_02(ctx), f3_03(ctx), f3_04(ctx), f3_05(ctx), f3_06(ctx),
                 f4_01(ctx), f4_02(ctx), f4_03(ctx), f4_04(ctx, kw_dict), f4_05(ctx),
                 f4_06(ctx), f4_07(ctx),
                 f5_01(ctx, kw_dict), f5_02(ctx), f5_03(ctx), f5_04(ctx), f5_05(ctx),
                 f6_01(ctx), f6_02(ctx, invoices), f6_03(ctx), f6_04(ctx, invoices), f6_05(ctx)]

    result = score_engine(features)
    worksheet = build_worksheet(client, args, ctx, result, features) if result["进入人审"] else None
    pending = [f_ for f_ in features if f_["待外部数据"]]

    out_dir = args.out_dir or os.path.dirname(os.path.abspath(src))
    os.makedirs(out_dir, exist_ok=True)
    xlsx_path = os.path.join(out_dir, f"{client}__反欺诈核查.xlsx")
    json_path = os.path.join(out_dir, f"{client}__反欺诈报告.json")
    write_xlsx(xlsx_path, client, ctx, features, result, worksheet, pending)
    payload = {
        "客户名称": client, "触发环节": args.stage,
        "流水区间": [str(ctx.d_min), str(ctx.d_max)], "交易笔数": len(rows),
        "打标覆盖率": round(ctx.labeled_ratio, 4),
        "评分": {"B级总分": result["B级总分"], "A级命中数": len(result["A级命中"]),
                 "B级命中数": len(result["B级命中"]), "共振": result["共振"],
                 "进入人审": result["进入人审"], "结论建议": result["结论建议"]},
        "核查工单": worksheet, "特征明细": features,
        "阈值版本": THRESHOLDS, "B级权重": B_WEIGHTS,
    }
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1, default=str)

    print(f"[结论] {result['结论建议']}")
    for f_ in result["A级命中"]:
        print(f"  [A] {f_['编号']} {f_['名称']}：{f_['结论摘要']}")
    for f_ in sorted(result["B级命中"], key=lambda x: -x["权重"]):
        print(f"  [B{f_['权重']:>3}] {f_['编号']} {f_['名称']}：{f_['结论摘要']}")
    print(f"[产出] {xlsx_path}")
    print(f"[产出] {json_path}")


if __name__ == "__main__":
    main()
